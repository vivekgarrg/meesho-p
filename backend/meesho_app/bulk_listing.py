"""
Bulk listing sheet generation — turns one filled-in product form into a
ready-to-upload Meesho bulk-listing `.xlsx`, built from Meesho's own category
template rather than a hand-rolled one, so every dropdown/validation Meesho's
importer enforces is already satisfied.

Nothing about a template is hardcoded per category: `parse_template()` reads
whatever `.xlsx` it's given (an upload, or the one bundled "quick start" file)
and derives the field list itself — see the module docstring on
`parse_template` for exactly what it relies on. That's what lets the same
code serve any Meesho category, not just the one example we started from.
"""

import random
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

TEMPLATES_DIR = Path(__file__).resolve().parent / "bulk_listing_templates"

# Bundled "quick start" templates — go through the exact same parser as an
# upload, they just don't require picking a file first.
BUILT_IN_TEMPLATES = {
    "puja_articles": {
        "label": "Puja Articles (Home & Kitchen / Home Decor / Puja Accessories)",
        "file": TEMPLATES_DIR / "puja_articles.xlsx",
    },
}

# Not a fixed row count any more — one listing per photo given, however many
# that is. This just guards against a mistaken paste (e.g. the wrong file's
# contents) turning into a many-thousand-row workbook.
MAX_ROWS = 300

# Which detected roles are per-row (varying) rather than shared across all 4
# rows. Everything else on the sheet — including fields we still recognise by
# role, like Country of Origin — is shared. Group ID is per-row (not simply
# shared like Country of Origin) because its whole point is to sometimes
# *not* be the same on every row — see plan_group_ids: "unique listing" wants
# a different group per row, "variation listing" wants one group for all.
_FIXED_PER_ROW_ROLES = {"title", "sku", "style", "group_id"}
_IMAGE_ROLE = re.compile(r"^image_(\d+)$")


class _PerRowRoles:
    """
    The roles that differ from one listing to the next.

    A plain set no longer works: categories carry however many image columns
    they carry — some have four, some fifteen — so image roles are matched by
    shape rather than listed. Kept callable as `role in PER_ROW_ROLES` because
    that is how every caller already reads.
    """

    def __contains__(self, role):
        return role in _FIXED_PER_ROW_ROLES or bool(role and _IMAGE_ROLE.match(role))

    def __iter__(self):
        return iter(_FIXED_PER_ROW_ROLES)


PER_ROW_ROLES = _PerRowRoles()

_RANGE_RE = re.compile(r"^(?:'([^']+)'|([A-Za-z0-9_]+))!\$?([A-Z]+)\$(\d+):\$?([A-Z]+)\$(\d+)$")
_SQREF_START_RE = re.compile(r"^([A-Z]+)(\d+)")


def category_choices():
    return [{"id": key, "label": cfg["label"]} for key, cfg in BUILT_IN_TEMPLATES.items()]


def built_in_path(key):
    cfg = BUILT_IN_TEMPLATES.get(key)
    if not cfg:
        raise KeyError(key)
    return cfg["file"]


def _detect_fill_sheet(wb):
    """The sheet actually meant to be filled in — prefer one named like
    "...-Fill this" (Meesho's own convention); else whichever sheet carries
    the most data validations (the fill sheet is always the heavily-
    constrained one); else just the first sheet."""
    for name in wb.sheetnames:
        if "fill" in name.lower():
            return name
    best, best_count = wb.sheetnames[0], -1
    for name in wb.sheetnames:
        n = len(wb[name].data_validations.dataValidation)
        if n > best_count:
            best, best_count = name, n
    return best


def _first_nonempty(ws, row):
    for ci in range(1, ws.max_column + 1):
        v = ws.cell(row=row, column=ci).value
        if v not in (None, ""):
            return str(v).strip()
    return None


def _sqref_start(dv):
    first_range = str(dv.sqref).split()[0]
    m = _SQREF_START_RE.match(first_range)
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


def _column_validations(wb, ws):
    """
    {column_letter: {"options": [...] or None, "start_row": int or None,
    "custom_formula": str or None}} — `start_row` is the first row that
    column's own validation applies to (used to find where real data starts
    without hardcoding a row number), `options` is the resolved dropdown
    list for list-type validations, read from wherever the validation's own
    formula actually points (not hand-transcribed, so it can't drift from
    what the file really allows), `custom_formula` is a "custom"-type
    validation's raw Excel formula (e.g. the price-shaped constraint that
    guards Meesho Price — see `_money_constraint`).
    """
    out = {}
    for dv in ws.data_validations.dataValidation:
        col, row = _sqref_start(dv)
        if not col:
            continue
        entry = out.setdefault(col, {"options": None, "start_row": None, "custom_formula": None})
        if row is not None and (entry["start_row"] is None or row < entry["start_row"]):
            entry["start_row"] = row
        if dv.type == "custom" and entry["custom_formula"] is None and dv.formula1:
            entry["custom_formula"] = dv.formula1.strip()
        if dv.type == "list" and entry["options"] is None and dv.formula1:
            raw = dv.formula1.strip()
            m = _RANGE_RE.match(raw)
            if m:
                src_sheet = m.group(1) or m.group(2)
                src_col, r1, r2 = m.group(3), int(m.group(4)), int(m.group(6))
                src_ws = wb[src_sheet] if src_sheet in wb.sheetnames else ws
                ci = openpyxl.utils.column_index_from_string(src_col)
                values = [
                    str(src_ws.cell(row=r, column=ci).value)
                    for r in range(r1, r2 + 1)
                    if src_ws.cell(row=r, column=ci).value not in (None, "")
                ]
                entry["options"] = values
            elif raw.startswith('"') and raw.endswith('"'):
                # An inline list ("A,B,C") rather than a range reference.
                entry["options"] = [v.strip() for v in raw[1:-1].split(",") if v.strip()]
    return out


def _parse_label(cell_text):
    """
    Meesho's own field-name row packs the short label and a long description
    into one cell as `"\\n\\n<Field Name>\\n\\n<description>\\n"` — the first
    non-blank line, once split on blank lines, is the label.
    """
    if not cell_text:
        return ""
    parts = [p.strip() for p in str(cell_text).split("\n\n") if p.strip()]
    return parts[0] if parts else str(cell_text).strip()


def _field_key(label, seen):
    base = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or "field"
    n = seen.get(base, 0)
    seen[base] = n + 1
    return base if n == 0 else f"{base}_{n + 1}"


_NUMBER_HINT = re.compile(
    r"price|mrp|weight|quantity|breadth|height|length|inventory|gst|%", re.I
)
_TEXTAREA_HINT = re.compile(r"description|address", re.I)

# Meesho guards its money-shaped columns (Meesho Price, MRP, and — in the
# newer templates — the unlabelled shadow input beside them) with a custom
# Excel validation shaped like `AND(MOD(F5*100,1)=0,F5>0,F5<10000000)`:
# positive, at most 2 decimal places (the *100 test is the "no more than
# paise" check), under some ceiling. Recognising this formula shape — instead
# of hardcoding it to "Meesho Price" — means any column Meesho guards the
# same way gets the same real validation, not just the one we've seen fail.
_MONEY_CONSTRAINT_RE = re.compile(
    r"MOD\([A-Z]+\d+\*100,\s*1\)\s*=\s*0.*?<\s*(\d+(?:\.\d+)?)", re.I
)


def _money_constraint(formula):
    """The ceiling from a price-shaped custom validation formula, or None if
    `formula` isn't shaped like one."""
    if not formula:
        return None
    m = _MONEY_CONSTRAINT_RE.search(formula)
    return float(m.group(1)) if m else None


def validate_money(value, money_max):
    """
    True if `value` satisfies Meesho's own price-shaped constraint: a
    positive number, at most 2 decimal places, strictly under `money_max`.
    Mirrors the exact rule in the template's own validation formula (see
    `_money_constraint`) so a value that would pass Meesho's own check is
    never rejected here, and — the point of this function — a value that
    would *fail* Meesho's check never reaches the sheet in the first place.
    """
    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return False
    if number <= 0 or number >= Decimal(str(money_max)):
        return False
    return (number * 100) == (number * 100).to_integral_value()


def _detect_role(label):
    l = label.lower()
    if l.startswith("product name"):
        return "title"
    if l == "sku id" or l.startswith("sku id"):
        return "sku"
    if "style id" in l or "product id" in l:
        return "style"
    # Any image column, however many the category has — "Image 1 (Front)",
    # "Image 2" … "Image 17". Capped at 1-4 before, which silently ignored
    # every slot past the fourth.
    m = re.match(r"image\s*(\d+)", l)
    if m:
        return f"image_{int(m.group(1))}"
    if "wrong" in l and "defective" in l:
        return "wrong_defective_price"
    if l.startswith("country of origin"):
        return "country_of_origin"
    if l.startswith("importer name"):
        return "importer_name"
    if l.startswith("importer address"):
        return "importer_address"
    if l.startswith("importer pincode"):
        return "importer_pincode"
    if l.startswith("group id"):
        return "group_id"
    return None


def _field_type(label, has_options):
    if has_options:
        return "select"
    if _TEXTAREA_HINT.search(label):
        return "textarea"
    if _NUMBER_HINT.search(label):
        return "number"
    return "text"


def load_workbook(source):
    """`source`: a path (str/Path) or a file-like object positioned at 0.
    Raises ValueError (not openpyxl's own exception types) so callers can
    turn a bad upload straight into a 400 without knowing openpyxl's API."""
    try:
        return openpyxl.load_workbook(source, data_only=False)
    except Exception as exc:
        raise ValueError(f"Could not read that as an Excel file: {exc}") from exc


def _attach_mirror_columns(ws, fields, col_validations):
    """
    Meesho's newer category templates insert extra "system" columns right
    after some fields — e.g. next to the labelled, Compulsory "Meesho Price"
    column, a plain "Enter Meesho Price" column with its own per-row numeric
    validation but *no* row2/row3 header of its own, so the normal scan above
    never sees it as a field. Meesho's importer reads that unlabelled cell,
    not the labelled one — so a value that's clearly filled in the labelled
    column can still come back "Meesho Price is empty" from Meesho's side,
    because the cell it actually checks was never written.

    This finds those columns after the fact: any column with real per-row
    validation that wasn't already claimed as a field, whose row4 hint text
    ("Enter Meesho Price") names an *already-detected* field ("Meesho
    Price") — matched by checking whether that field's key is a substring of
    the hint, slugified the same way field labels are. Matched columns are
    recorded on `mirror_columns` so build_workbook writes the same value into
    both places; unmatched ones are left alone rather than guessed at, since
    a wrongly-invented required field is worse than not handling a new one.
    """
    claimed = {f["column"] for f in fields}
    for col, info in col_validations.items():
        if col in claimed or not info.get("start_row"):
            continue
        ci = openpyxl.utils.column_index_from_string(col)
        hint = ws.cell(row=4, column=ci).value
        if not hint or "system generated" in str(hint).lower():
            continue
        hint_slug = re.sub(r"[^a-z0-9]+", "_", str(hint).lower()).strip("_")
        for f in fields:
            # The length floor keeps a short, generic key (e.g. "type") from
            # matching hint text that happens to contain it by coincidence.
            if len(f["key"]) >= 4 and f["key"] in hint_slug:
                f["mirror_columns"].append(col)
                break


def parse_template(source):
    """
    `source`: an already-loaded `openpyxl.Workbook` (see `load_workbook`
    below) — callers that also need to *write* into the same file (i.e.
    `generate`) load it once and pass the same object to both, rather than
    reading the upload twice.

    Returns `{category_label, sheet_name, data_start_row, fields: [...]}`.
    Each field is `{key, label, column, required, type, options, role}` —
    `role` is one of PER_ROW_ROLES, a handful of other recognised shared
    roles (see `_detect_role`), or None for an ordinary shared field.

    Raises ValueError if `source` doesn't look like a Meesho bulk template
    (no per-column data validation found on the detected sheet at all).
    """
    wb = source
    sheet_name = _detect_fill_sheet(wb)
    ws = wb[sheet_name]
    category_label = _first_nonempty(ws, row=1) or sheet_name

    col_validations = _column_validations(wb, ws)
    if not col_validations:
        raise ValueError(
            "This doesn't look like a Meesho bulk-listing template — no per-column "
            "validation found on its fill-in sheet."
        )

    data_start_row = min(
        (info["start_row"] for info in col_validations.values() if info["start_row"]),
        default=5,
    )

    fields = []
    seen_keys = {}
    for ci in range(1, ws.max_column + 1):
        col = openpyxl.utils.get_column_letter(ci)
        row2 = str(ws.cell(row=2, column=ci).value or "").lower()
        # Only columns Meesho itself marks as a real field ("* Compulsory
        # Field" / "Optional Field") are fillable — this is what actually
        # excludes the "Field Names" header column and the "Meesho only,
        # don't fill" error-status columns, which otherwise still carry
        # descriptive text in row 3 and would slip through a label-only check.
        if "compulsory" not in row2 and "optional" not in row2:
            continue
        required = "compulsory" in row2
        label = _parse_label(ws.cell(row=3, column=ci).value) or f"Column {col}"
        info = col_validations.get(col, {})
        options = info.get("options") or []
        fields.append({
            "key": _field_key(label, seen_keys),
            "label": label,
            "column": col,
            "required": required,
            "type": _field_type(label, bool(options)),
            "options": options,
            "role": _detect_role(label),
            "mirror_columns": [],
            "money_max": None,
        })

    _attach_mirror_columns(ws, fields, col_validations)

    # A field's real constraint may live on its own column, its mirror
    # column, or both (Meesho Price carries it on the unlabelled shadow
    # column in newer templates, on the labelled one in older ones) — check
    # every column this field could be written to, not just its own.
    for f in fields:
        for col in [f["column"], *f["mirror_columns"]]:
            constraint = _money_constraint(col_validations.get(col, {}).get("custom_formula"))
            if constraint is not None:
                f["money_max"] = constraint
                if f["type"] != "number":
                    f["type"] = "number"
                break

    return {
        "category_label": category_label,
        "sheet_name": sheet_name,
        "data_start_row": data_start_row,
        "fields": fields,
    }


def find_field(spec, role=None, key=None):
    for f in spec["fields"]:
        if role is not None and f["role"] == role:
            return f
        if key is not None and f["key"] == key:
            return f
    return None


def image_slots(spec):
    """The template's image columns, in slot order (Image 1, 2, 3 … N)."""
    slots = []
    for f in spec["fields"]:
        m = _IMAGE_ROLE.match(f["role"] or "")
        if m:
            slots.append((int(m.group(1)), f["role"]))
    return [role for _, role in sorted(slots)]


def plan_images(urls, slot_count, rng=None):
    """
    Decide which link goes in which image slot — one listing per URL given.

    Row i's own photo (`urls[i]`) is that listing's front image, unchanged —
    a photo picked for a specific product stays its hero shot. The remaining
    slots are filled with a random sample of every *other* row's photo, so a
    seller who's given several photos (pasted fresh, or already dropped one
    per row into a sheet) gets galleries that differ from listing to listing
    instead of every row showing an identical set.

    Number of listings produced = len(urls) — there is no separate row-count
    input; the photos you give *are* the rows. Fewer photos than a category
    has image slots just leaves the spare slots empty.
    """
    rng = rng or random.Random()
    plans = []
    for i, own in enumerate(urls):
        pool = [u for j, u in enumerate(urls) if j != i]
        rng.shuffle(pool)
        plans.append([own, *pool][:slot_count])
    return plans


def extract_front_images(spec, wb):
    """
    Front-image URLs already sitting in an uploaded sheet's own data rows —
    the "Prefilled Sheet" flow: Meesho's own bulk image-link generator (or an
    earlier partial fill) already dropped one photo per row, and the point is
    to build fresh listings from those photos rather than ask the seller to
    retype links they've already placed in the sheet.

    Reads from `data_start_row` down the Image 1 (Front) column, stopping at
    the first empty cell — which naturally skips the "Watch Explainer Video"
    tutorial-link row sitting just above the real data, and stops exactly
    where the seller's own fill-in ends. Returns [] if the template has no
    Image 1 column, or its data rows are empty (a genuinely blank template).
    """
    front_field = find_field(spec, role="image_1")
    if not front_field:
        return []
    ws = wb[spec["sheet_name"]]
    urls = []
    r = spec["data_start_row"]
    while r <= ws.max_row:
        v = ws[f'{front_field["column"]}{r}'].value
        if v in (None, ""):
            break
        urls.append(str(v).strip())
        r += 1
    return urls


LISTING_TYPE_UNIQUE = "unique"
LISTING_TYPE_VARIATION = "variation"


def plan_group_ids(row_count, listing_type, base="Group"):
    """
    Group ID per row — Meesho's own mechanism for tying several rows together
    into one catalog ("Enter same group ID for the products which should be
    grouped in a single catalog").

    "unique listing": every row is its own catalog, so every row gets a
    different group — "Group 1", "Group 2", "Group 3" …
    "variation listing": every row is a variant of the *same* product, so
    every row gets the same group — "Group 1" for all.
    """
    base = (base or "Group").strip() or "Group"
    if listing_type == LISTING_TYPE_VARIATION:
        return [f"{base} 1"] * row_count
    return [f"{base} {i + 1}" for i in range(row_count)]


def resolve_wrong_defective_price(meesho_price, override):
    if override not in (None, ""):
        return override
    try:
        price = Decimal(str(meesho_price))
    except Exception:
        return ""
    return max(price - 20, Decimal("0"))


def resolve_importer_fields(country, name, address, pincode):
    """Mirrors Meesho's own template formula:
    `=IF(country="India","Not Required","")`."""
    if (country or "").strip().lower() == "india":
        return "Not Required", "Not Required", "Not Required"
    return name, address, pincode


_LEADING_ZERO = re.compile(r"^0\d")


def coerce_cell(field, value):
    """
    Turn a submitted value into what the cell should actually hold.

    Everything arrives from JSON as a string, and writing "499" into a price
    cell puts *text* there: Excel left-aligns it, flags "number stored as
    text", and the validation formulas Meesho ships inside the template do not
    read it as a value at all. That is why prices looked like they were not
    filling in — they were being written, just not as numbers.

    Leading zeros stay text on purpose. A code that happens to be numeric — an
    HSN, a pincode — stops meaning anything the moment 003924 becomes 3924.
    """
    if field.get("type") != "number":
        return value
    text = str(value).strip()
    if not text or _LEADING_ZERO.match(text):
        return value
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return value                      # let the sheet keep whatever was typed
    # Whole numbers as int so the cell reads "499", not "499.0".
    return int(number) if number == number.to_integral_value() else float(number)


def build_workbook(spec, wb, shared, rows):
    """
    Writes `rows` (dicts of `product_name`, `sku_id`, `style_id`,
    `group_id` (optional — see plan_group_ids), `images` = an ordered list,
    one URL per image slot, `overrides` (optional — field key -> value,
    this row's own replacement for whatever `shared` says) — as many rows
    as the caller gives) into the parsed sheet starting at
    `spec["data_start_row"]`. Every key in `shared` goes onto every row via
    the matching field's column, *unless* that row's own `overrides` names
    the same key, in which case the override wins for that row only —
    everything else about a listing still comes from the one shared form,
    a seller only has to touch what's actually different about a specific
    row. Everything else in the workbook — every other sheet, every other
    cell — is untouched.
    """
    ws = wb[spec["sheet_name"]]
    per_row_fields = {f["role"]: f for f in spec["fields"] if f["role"] in PER_ROW_ROLES}
    shared_fields = [f for f in spec["fields"] if f["role"] not in PER_ROW_ROLES]

    for i, row in enumerate(rows):
        r = spec["data_start_row"] + i

        if "title" in per_row_fields:
            ws[f'{per_row_fields["title"]["column"]}{r}'] = row["product_name"]
        if "sku" in per_row_fields:
            ws[f'{per_row_fields["sku"]["column"]}{r}'] = row["sku_id"]
        if "style" in per_row_fields:
            ws[f'{per_row_fields["style"]["column"]}{r}'] = row["style_id"]
        if "group_id" in per_row_fields and row.get("group_id"):
            ws[f'{per_row_fields["group_id"]["column"]}{r}'] = row["group_id"]

        for slot, url in zip(image_slots(spec), row["images"]):
            if slot in per_row_fields and url:
                ws[f'{per_row_fields[slot]["column"]}{r}'] = url

        overrides = row.get("overrides") or {}
        for f in shared_fields:
            value = overrides.get(f["key"], shared.get(f["key"]))
            if value not in (None, ""):
                cell_value = coerce_cell(f, value)
                ws[f'{f["column"]}{r}'] = cell_value
                # Some templates carry an unlabelled "shadow" copy of a field
                # that Meesho's importer reads instead of (or as well as) the
                # labelled cell — see _attach_mirror_columns. Keep them in sync.
                for mirror_col in f.get("mirror_columns", []):
                    ws[f'{mirror_col}{r}'] = cell_value

    return wb
