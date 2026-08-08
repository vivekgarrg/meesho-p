import pandas as pd
import numpy as np
from decimal import Decimal, InvalidOperation
from django.db import transaction, IntegrityError
from django.db.models import Sum, Count, Min, Max, ExpressionWrapper, F, DecimalField as DjDecimalField, Q as DQ, Prefetch
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from datetime import datetime, time, timedelta
import requests as http_requests
import base64
import io
import zipfile
import re
import concurrent.futures
from PIL import Image
from .helpers.helper import status_wise_summary, strip_html
from accounts.models import Business, User
from .permissions import get_authorized_business, accessible_businesses
from .helpers.label_pdf import extract_all_pages

from .models import OrderPayment, AdsCost, ReferralPayment, CompensationRecovery, FinalPrice, Order, ParentItemPrice, ParentPriceHistory, LabelOrder, PurchaseBill, PurchaseItem, BlockedCustomer, InventoryAdjustment, ConsumableItem, ConsumablePurchase, ConsumableUsage, InventoryLog, MeeshoInventory, MeeshoPriceUpdate, ExpenseInvoice, ExpenseInvoiceItem, TransportCharge, PackedStockEvent, EstimatedProfitOrder, ReturnDelivery, GstTransaction, GstInvoiceDetail, ScannedOrder, ListingTemplate, ClaimTicket, WorkerTask, WalletEntry, WalletSettlement, TaskListing, PlatformRate, TaskDocument
from .serializers import (
    OrderPaymentSerializer, AdsCostSerializer,
    ReferralPaymentSerializer, CompensationRecoverySerializer,
    FinalPriceSerializer,
    ParentItemPriceSerializer,
    ParentPriceHistorySerializer,
    OrderSerializer,
    LabelOrderSerializer,
    ReturnDeliverySerializer,
    ScannedOrderSerializer,
    ListingTemplateSerializer,
    ClaimTicketSerializer,
    WorkerTaskSerializer,
    WalletEntrySerializer,
    WalletSettlementSerializer,
    TaskListingSerializer,
    PlatformRateSerializer,
    TaskDocumentSerializer,
)


def safe_decimal(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError):
        return None


def safe_str(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    return str(val).strip()


def safe_date(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        return pd.to_datetime(val).date()
    except Exception:
        return None


def safe_datetime(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        dt = pd.to_datetime(val)
    except Exception:
        return None
    if dt is None or pd.isna(dt):
        return None
    # pandas Timestamp -> stdlib datetime
    if hasattr(dt, "to_pydatetime"):
        dt = dt.to_pydatetime()
    # With USE_TZ active, attach the project timezone so Django stores an
    # aware datetime instead of warning about a naive one.
    from django.conf import settings
    if getattr(settings, "USE_TZ", False) and timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def safe_int(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _workbook_from_upload(uploaded):
    """
    The spreadsheet to parse, unwrapping a .zip if that's what arrived.

    Returns (file_like, extracted_name). Meesho's payment export downloads as a
    zip, and making people unzip it first is a step that only exists because we
    didn't do it for them.

    The catch worth knowing: an .xlsx *is* a zip archive, so "does it start with
    PK" cannot decide this. The test is what's inside — an archive of exports
    contains members ending in .xlsx/.xls, while a workbook contains
    xl/workbook.xml and never another spreadsheet. Checking for members first is
    therefore unambiguous in both directions.
    """
    uploaded.seek(0)
    if uploaded.read(2) != b"PK":
        uploaded.seek(0)
        return uploaded, None          # .xls or anything else — hand it straight on
    uploaded.seek(0)

    try:
        archive = zipfile.ZipFile(uploaded)
        names = archive.namelist()
    except zipfile.BadZipFile:
        uploaded.seek(0)
        return uploaded, None          # let pandas produce the real complaint

    sheets = [
        n for n in names
        if n.lower().endswith((".xlsx", ".xls"))
        and not n.startswith("__MACOSX/")
        and not n.rsplit("/", 1)[-1].startswith(".")
    ]

    if not sheets:
        if "xl/workbook.xml" in names:
            uploaded.seek(0)
            return uploaded, None      # it's the workbook itself
        raise ValueError(
            "That zip has no Excel file in it. Upload the payment .xlsx, or the "
            "zip Meesho gave you with the sheet still inside."
        )

    # Several exports in one zip: the payment sheet is far and away the biggest,
    # so size is a better guess than filename, which Meesho changes freely.
    chosen = max(sheets, key=lambda n: archive.getinfo(n).file_size)
    try:
        data = archive.read(chosen)
    except RuntimeError as exc:
        if "password" in str(exc).lower() or "encrypted" in str(exc).lower():
            raise ValueError(
                "That zip is password-protected. Unzip it yourself and upload the "
                "Excel file directly."
            ) from exc
        raise
    return io.BytesIO(data), chosen


@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_excel(request, business_id):
    """
    Upload a Meesho payment Excel file, or the zip it came in.
    Parses all 4 sheets and inserts/updates rows in the DB.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        source, extracted = _workbook_from_upload(file)
    except ValueError as exc:
        return Response({"error": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    try:
        xl = pd.ExcelFile(source)
    except Exception as e:
        hint = f" (from {extracted} inside the zip)" if extracted else ""
        return Response({"error": f"Could not read Excel file{hint}: {e}"},
                        status=status.HTTP_400_BAD_REQUEST)

    results = {}

    # ── 1. Order Payments ───────────────────────────────────────────────────
    if "Order Payments" in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name="Order Payments", header=None, skiprows=2)
        df.columns = [
            "sub_order_no", "order_date", "dispatch_date", "product_name",
            "supplier_sku", "catalog_id", "order_source", "live_order_status",
            "product_gst_percent", "listing_price_incl_taxes", "quantity",
            "transaction_id", "payment_date", "final_settlement_amount",
            "price_type", "total_sale_amount", "total_sale_return_amount",
            "fixed_fee_revenue", "warehousing_fee", "return_premium",
            "return_premium_of_return", "meesho_commission_percentage",
            "meesho_commission_incl_gst", "meesho_gold_platform_fee",
            "meesho_mall_platform_fee", "fixed_fee_deduction",
            "warehousing_fee_deduction", "return_shipping_charge",
            "gst_compensation_prp_shipping", "shipping_charge_incl_gst",
            "other_support_service_charges", "waivers",
            "net_other_support_service_charges",
            "gst_on_net_other_support_service_charges",
            "tcs", "tds_rate_percent", "tds",
            "compensation", "claims", "recovery",
            "compensation_reason", "claims_reason", "recovery_reason",
        ]
        # Drop the formula-description row (row index 0 has letters like A, B, C...)
        df = df[df["sub_order_no"].notna()]
        df = df[~df["sub_order_no"].astype(str).str.match(r"^[A-Z\s\(\)\+\-\*\/]+$")]

        created = updated = 0
        with transaction.atomic():
            for _, row in df.iterrows():
                pk = safe_str(row["sub_order_no"])
                if not pk:
                    continue
                defaults = {
                    "order_date": safe_datetime(row.get("order_date")),
                    "dispatch_date": safe_date(row.get("dispatch_date")),
                    "product_name": safe_str(row.get("product_name")),
                    "supplier_sku": safe_str(row.get("supplier_sku")),
                    "catalog_id": safe_int(row.get("catalog_id")),
                    "order_source": safe_str(row.get("order_source")),
                    "live_order_status": safe_str(row.get("live_order_status")),
                    "product_gst_percent": safe_decimal(row.get("product_gst_percent")),
                    "listing_price_incl_taxes": safe_decimal(row.get("listing_price_incl_taxes")),
                    "quantity": safe_int(row.get("quantity")),
                    "transaction_id": safe_str(row.get("transaction_id")),
                    "payment_date": safe_date(row.get("payment_date")),
                    "final_settlement_amount": safe_decimal(row.get("final_settlement_amount")),
                    "price_type": safe_str(row.get("price_type")),
                    "total_sale_amount": safe_decimal(row.get("total_sale_amount")),
                    "total_sale_return_amount": safe_decimal(row.get("total_sale_return_amount")),
                    "fixed_fee_revenue": safe_decimal(row.get("fixed_fee_revenue")),
                    "warehousing_fee": safe_decimal(row.get("warehousing_fee")),
                    "return_premium": safe_decimal(row.get("return_premium")),
                    "return_premium_of_return": safe_decimal(row.get("return_premium_of_return")),
                    "meesho_commission_percentage": safe_decimal(row.get("meesho_commission_percentage")),
                    "meesho_commission_incl_gst": safe_decimal(row.get("meesho_commission_incl_gst")),
                    "meesho_gold_platform_fee": safe_decimal(row.get("meesho_gold_platform_fee")),
                    "meesho_mall_platform_fee": safe_decimal(row.get("meesho_mall_platform_fee")),
                    "fixed_fee_deduction": safe_decimal(row.get("fixed_fee_deduction")),
                    "warehousing_fee_deduction": safe_decimal(row.get("warehousing_fee_deduction")),
                    "return_shipping_charge": safe_decimal(row.get("return_shipping_charge")),
                    "gst_compensation_prp_shipping": safe_decimal(row.get("gst_compensation_prp_shipping")),
                    "shipping_charge_incl_gst": safe_decimal(row.get("shipping_charge_incl_gst")),
                    "other_support_service_charges": safe_decimal(row.get("other_support_service_charges")),
                    "waivers": safe_decimal(row.get("waivers")),
                    "net_other_support_service_charges": safe_decimal(row.get("net_other_support_service_charges")),
                    "gst_on_net_other_support_service_charges": safe_decimal(row.get("gst_on_net_other_support_service_charges")),
                    "tcs": safe_decimal(row.get("tcs")),
                    "tds_rate_percent": safe_decimal(row.get("tds_rate_percent")),
                    "tds": safe_decimal(row.get("tds")),
                    "compensation": safe_decimal(row.get("compensation")),
                    "claims": safe_decimal(row.get("claims")),
                    "recovery": safe_decimal(row.get("recovery")),
                    "compensation_reason": safe_str(row.get("compensation_reason")),
                    "claims_reason": safe_str(row.get("claims_reason")),
                    "recovery_reason": safe_str(row.get("recovery_reason")),
                }
                # Composite lookup: (sub_order_no, payment_date, live_order_status)
                # allows multiple rows per order (e.g. affiliate-fee adjustments
                # alongside the main delivery row).
                lookup_payment_date    = defaults.pop("payment_date")
                lookup_live_status     = defaults.pop("live_order_status")
                obj, was_created = OrderPayment.objects.update_or_create(
                    business=business,
                    sub_order_no=pk,
                    payment_date=lookup_payment_date,
                    live_order_status=lookup_live_status,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        results["order_payments"] = {"created": created, "updated": updated}

    # ── 2. Ads Cost ─────────────────────────────────────────────────────────
    if "Ads Cost" in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name="Ads Cost", header=None, skiprows=2)
        df.columns = [
            "deduction_duration", "deduction_date", "campaign_id",
            "ad_cost", "credits_waivers_discounts",
            "ad_cost_incl_credits_waivers", "gst", "total_ads_cost",
        ]
        df = df[df["deduction_date"].notna()]
        df = df[~df["deduction_date"].astype(str).str.strip().str.startswith("No data")]

        created = skipped = 0
        with transaction.atomic():
            for _, row in df.iterrows():
                _, was_created = AdsCost.objects.update_or_create(
                    business=business,
                    deduction_duration=safe_date(row.get("deduction_duration")),
                    deduction_date=safe_date(row.get("deduction_date")),
                    campaign_id=safe_str(row.get("campaign_id")),
                    defaults={
                        "ad_cost": safe_decimal(row.get("ad_cost")),
                        "credits_waivers_discounts": safe_decimal(row.get("credits_waivers_discounts")),
                        "ad_cost_incl_credits_waivers": safe_decimal(row.get("ad_cost_incl_credits_waivers")),
                        "gst": safe_decimal(row.get("gst")),
                        "total_ads_cost": safe_decimal(row.get("total_ads_cost")),
                    },
                )
                if was_created:
                    created += 1
                else:
                    skipped += 1
        results["ads_cost"] = {"created": created, "updated": skipped}

    # ── 3. Referral Payments ─────────────────────────────────────────────────
    if "Referral Payments" in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name="Referral Payments", header=None, skiprows=2)
        df.columns = [
            "reward_id", "payment_date", "store_name",
            "reason", "net_referral_amount", "taxes_gst_tds",
        ]
        df = df[df["reward_id"].notna()]
        df = df[~df["reward_id"].astype(str).str.startswith("No data")]

        created = updated = 0
        with transaction.atomic():
            for _, row in df.iterrows():
                pk = safe_str(row["reward_id"])
                if not pk:
                    continue
                obj, was_created = ReferralPayment.objects.update_or_create(
                    business=business,
                    reward_id=pk,
                    defaults={
                        "payment_date": safe_date(row.get("payment_date")),
                        "store_name": safe_str(row.get("store_name")),
                        "reason": safe_str(row.get("reason")),
                        "net_referral_amount": safe_decimal(row.get("net_referral_amount")),
                        "taxes_gst_tds": safe_decimal(row.get("taxes_gst_tds")),
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        results["referral_payments"] = {"created": created, "updated": updated}

    # ── 4. Compensation and Recovery ─────────────────────────────────────────
    if "Compensation and Recovery" in xl.sheet_names:
        df = pd.read_excel(xl, sheet_name="Compensation and Recovery", header=None, skiprows=2)
        df.columns = ["date", "program_name", "reason", "amount_incl_gst"]
        df = df[df["date"].notna()]
        df = df[~df["date"].astype(str).str.startswith("No data")]

        created = skipped = 0
        with transaction.atomic():
            for _, row in df.iterrows():
                # The sheet has no id column, so treat the full row
                # (date + program + reason + amount, scoped to the business)
                # as its natural key. Skip rows that already exist so
                # re-uploading the same sheet doesn't duplicate entries.
                lookup = dict(
                    business=business,
                    date=safe_date(row.get("date")),
                    program_name=safe_str(row.get("program_name")),
                    reason=safe_str(row.get("reason")),
                    amount_incl_gst=safe_decimal(row.get("amount_incl_gst")),
                )
                if CompensationRecovery.objects.filter(**lookup).exists():
                    skipped += 1
                    continue
                CompensationRecovery.objects.create(**lookup)
                created += 1
        results["compensation_recovery"] = {"created": created, "skipped": skipped}

    return Response({"success": True, "results": results}, status=status.HTTP_201_CREATED)


def _inc(d, key, value=None):
    """Increment dict[key] by value (default 1). Skips None keys."""
    if key is None:
        return
    if key not in d:
        d[key] = 0
    if value is not None:
        d[key] += value
    else:
        d[key] += 1


# ── Order status constants ─────────────────────────────────────────────────────

_DELIVERED_STATUSES = frozenset(["DELIVERED"])

_EXCHANGE_STATUSES = frozenset(["EXCHANGE", "EXCHANGED"])

_RETURN_STATUSES  = frozenset([
    "RETURN", "RETURNED", "PREMIUM_RETURN",
])
_RTO_STATUSES = frozenset([
    "RTO", "RTO_COMPLETE"
])
_SHIPPED_STATUSES = frozenset(["SHIPPED", "IN TRANSIT"])

# Canonical accumulator key mapping (DB value → frontend-visible bucket key)
_STATUS_KEY_NORM = {
    "rto_complete":   "rto",
    "premium_return": "return",
    "returned":       "return",      # order summary uses "Returned"; payments sheet uses "Return"
    "exchanged":      "exchange",    # order summary uses "Exchanged"; payments sheet uses "Exchange"
    "in transit":     "shipped",     # order summary uses "In Transit"; payments sheet uses "Shipped"
}

def _norm_key(raw_lower):
    return _STATUS_KEY_NORM.get(raw_lower, raw_lower)


# ── Row-level helpers ──────────────────────────────────────────────────────────

def _payment_date_key(row):
    """Sort key: payment_date asc, fallback to order_date, fallback to epoch."""
    import datetime as _dt
    pd = row.payment_date
    if pd is not None:
        return pd
    od = row.order_date
    return od.date() if od else _dt.date.min


def _classify_rows(payment_rows):
    rows = sorted(payment_rows, key=_payment_date_key, reverse=True)
    main_all = [r for r in rows if r.live_order_status]
    adj      = [r for r in rows if not r.live_order_status]
    
    _seen: dict = {}
    for r in main_all:
        _seen[r.live_order_status.upper()] = r
        _seen[r.live_order_status.upper()] = r
    main = list(_seen.values())


    _seen2: dict = {}
    for r in adj:
        if r.claims and r.claims > 0:
            _seen2["claim"] = r
        elif r.recovery_reason and r.recovery_reason.lower() == "affiliate fee":
            _seen2["affiliate"] = r
        else:
            _seen2["others"] = r
            _seen2["others"] = r
                
        
    all_known_statuses = (
    _DELIVERED_STATUSES
    | _SHIPPED_STATUSES
    | _RETURN_STATUSES
    | _RTO_STATUSES
    | _EXCHANGE_STATUSES
)

    rto_rows = [r for r in main if r.live_order_status.upper() in _RTO_STATUSES]
    return_rows = [r for r in main if r.live_order_status.upper() in _RETURN_STATUSES]
    shipped        = [r for r in main if r.live_order_status.upper() in _SHIPPED_STATUSES]
    delivered_rows = [r for r in main if r.live_order_status.upper() in _DELIVERED_STATUSES]
    exchange_rows = [r for r in main if r.live_order_status.upper() in _EXCHANGE_STATUSES]
    other_status_rows = [r for r in main if r.live_order_status.upper() not in all_known_statuses]
    
    claim_rows = _seen2.get("claim", [])
    affiliate_rows = _seen2.get("affiliate", [])
    other_adj_rows = _seen2.get("others", [])

    return {
        "rows":           rows,
        "main":           main,
        "adj":            adj,
        
        #status wise rows
        "shipped_rows":   shipped,
        "return_rows":    return_rows,
        "delivered_rows": delivered_rows,
        "rto_rows": rto_rows, 
        "exchange_rows": exchange_rows,
        "other_status_rows": other_status_rows,
        
        #adjustment rows
        "claim_rows": claim_rows,
        "affiliate_rows": affiliate_rows,
        "other_adj_rows": other_adj_rows
    }


def _pick_settlement_rows(c):
    """
    Settlement precedence (status-based):
      1. Any RETURN/RTO/CANCELLED → use only return rows  (item came back)
      2. DELIVERED/EXCHANGE/CLAIM, no return → use delivered rows  (item sold)
      3. Fallback → all non-shipped rows

    Returns (settlement_rows, is_delivered).
    is_delivered=True means item cost must be deducted from settlement.
    """
    if c["return_rows"]:
        return c["return_rows"], False
    if c["delivered_rows"]:
        return c["delivered_rows"], True
    return c["main"], False
    # return c["non_shipped"], False


def _sum_settlement(rows, include_shipped=False):
    """
    Sum final_settlement_amount across rows.

    SHIPPED rows represent an advance/provisional payment.  They should be
    included only when the order has *progressed past shipping* (i.e. has a
    DELIVERED / RETURN / RTO / EXCHANGE row), because in that case both the
    advance credit (SHIPPED row) and the final adjustment (e.g. RETURNED row)
    are real money movements that must both be counted.

    For orders still sitting in SHIPPED state (no final row yet) the SHIPPED
    amount is provisional and is excluded to avoid double-counting when the
    final row arrives later.
    """
    return sum(
        Decimal(r.final_settlement_amount or 0)
        for r in rows
        if include_shipped
        or not r.live_order_status
        or r.live_order_status.upper() not in _SHIPPED_STATUSES
    )


def _extract_claims(adj_rows):
    """Total claims money credited from Meesho (info-only, not in P&L)."""
    total = Decimal("0")
    for r in adj_rows:
        total += Decimal(r.claims or 0)
    return total

def _extract_return_fee(main_rows):
    """Total claims money credited from Meesho (info-only, not in P&L)."""
    total = Decimal("0")
    for r in main_rows:
        total += Decimal(r.return_shipping_charge or 0)
    return total


def _affiliate_total(adj_rows):
    raw = sum(Decimal(r.final_settlement_amount or 0)
              for r in adj_rows if r.recovery_reason == "Affiliate Fee")
    return raw if raw else Decimal("0")


# ── Per-order profit formula ───────────────────────────────────────────────────

def compute_order_net(payment_rows, sku_final_price, sku_packaging_price, quantity, unique_statuses, 
                      sku_item_price=None, sku_tax_percent=None):
    # list(["Cancelled", "Delivered", "Return", "RTO", "Shipped", "Exchange", "Unknown"])
    
    ZERO = Decimal("0")
    TWO  = Decimal("2")
    qty  = Decimal(str(quantity))

    purchase_cost  = Decimal(str(sku_item_price or 0)) * qty   # item_cost × qty
    packaging_cost = Decimal(str(sku_packaging_price or 0))    # fixed per order, not × qty
    # Tax applies to the purchase value, which already includes × qty — so the
    # rate must NOT be multiplied by qty a second time. (It used to be, which
    # squared the quantity and overstated cost on every multi-unit order.)
    tax_cost       = purchase_cost * Decimal(str(sku_tax_percent or 0)) / Decimal("100")

    c = _classify_rows(payment_rows)

    adj_rows      = c["adj"]
    main_rows     = c["main"]
    affiliate_fees = _affiliate_total(adj_rows)
    claims         = _extract_claims(adj_rows)
    return_shipping_fee = _extract_return_fee(main_rows)
    sub_order_no   = payment_rows[0].sub_order_no

    # Include the SHIPPED row's settlement when the order progressed past shipping —
    # both the advance credit (SHIPPED) and the final adjustment (RETURN/RTO/DELIVERED)
    # are real money movements and must be summed together.
    progressed_past_shipped = bool(
        c["return_rows"] or c["rto_rows"] or c["delivered_rows"]
        or c["exchange_rows"] or c["claim_rows"]
    )
    total_settlement = _sum_settlement(c["rows"], include_shipped=progressed_past_shipped)
    
    net = 0
    status = ""
    
    unknown_rows = c["other_adj_rows"]  or c["other_status_rows"]
    claimed_orders = c["claim_rows"]
    return_orders = c["return_rows"]
    rto_orders = c["rto_rows"]
    exchange_orders = c["exchange_rows"]
    delivered_orders = c["delivered_rows"]
    
    final_purchasing = 0
    
    if unknown_rows:
        final_purchasing  = 0
        net = 0 - final_purchasing
        status = unique_statuses[-1]
        
    elif claimed_orders:
        final_purchasing = purchase_cost + packaging_cost
        net = total_settlement - final_purchasing
        status = unique_statuses[0]
    elif exchange_orders:
        final_purchasing = purchase_cost + (packaging_cost * 2 ) + tax_cost
        net = total_settlement - final_purchasing
        status = unique_statuses[-2]
    elif return_orders:
        final_purchasing = 0 
        net = total_settlement - final_purchasing
        status = unique_statuses[-5]
    elif rto_orders:
        final_purchasing = 0 
        net = total_settlement - final_purchasing
        status = unique_statuses[-4]
    elif delivered_orders:
        final_purchasing =  purchase_cost + packaging_cost + tax_cost
        net = total_settlement - final_purchasing
        status = unique_statuses[-6]
    else:
        final_purchasing = 0 
        net = total_settlement - final_purchasing
        status = unique_statuses[-1]
        
    return {
        "net": net, 
        "status": status,
        "total_settlement": total_settlement,
        "purchase_cost": purchase_cost,
        "packaging_cost": packaging_cost,
        "tax_cost": tax_cost,
        "quantity": qty,
        "sub_order_no": sub_order_no,
        "affiliate_fees": affiliate_fees,
        "return_shipping_fee" : return_shipping_fee,
        "claims": claims,
        "final_purchase_cost": final_purchasing
    }


# ── SKU accumulator helpers ────────────────────────────────────────────────────

def _init_sku_bucket(key, loss_or_profit, sku):
    """Zero-initialise a status bucket (delivered/return/rto) if not yet seen."""
    field = f"{key}_{loss_or_profit}"
    if sku.get(field) is None:
        sku[field]                    = 0
        sku[f"{key}_purchase_cost"]   = 0
        sku[f"{key}_packaging_cost"]  = 0
        sku[f"{key}_count"]           = 0
        sku[f"{key}_quantity"]        = 0
        sku[f"{key}_final_purchase_cost"] = 0
        sku[f"{key}_tax_cost"] = 0
        sku[f"{key}_total_settlement"] = 0
        sku[f"{key}_payment_rows"] = []


def _ensure_status_buckets(sku, _unique_statuses):
    """Pre-initialise the five main buckets + other_net so no KeyError later."""
    _init_sku_bucket("delivered", "profit", sku)
    _init_sku_bucket("return",    "loss",   sku)
    _init_sku_bucket("rto",       "loss",   sku)
    _init_sku_bucket("exchange",  "net",    sku)
    _init_sku_bucket("claim",     "loss",   sku)
    _init_sku_bucket("unknown",     "loss",   sku)
    if sku.get("other_net") is None:
        sku["other_net"]   = 0
        sku["other_count"] = 0


def _inject_order_into_bucket(sku, result, net):
    status_upper = (result["status"] or "").upper()
    raw = _norm_key((result["status"] or "").lower())

    if raw == "claim" or raw == "exchange" :
        _inc(sku, f"{raw}_loss",         net)
    elif raw in ("return", "rto"):
        _inc(sku, f"{raw}_loss",           net)
    elif status_upper in _DELIVERED_STATUSES:
        _inc(sku, "delivered_profit",         net)
    else:
        _inc(sku, "other_net",   net)
        _inc(sku, "other_count")
        
    _inc(sku, f"{raw}_purchase_cost",  result["purchase_cost"])
    _inc(sku, f"{raw}_packaging_cost", result["packaging_cost"])
    _inc(sku, f"{raw}_count")
    _inc(sku, f"{raw}_quantity",       result["quantity"])
    _inc(sku, f"{raw}_final_purchase_cost", result["final_purchase_cost"])
    _inc(sku, f"{raw}_tax_cost", result["tax_cost"])
    _inc(sku, f"{raw}_total_settlement", result["total_settlement"])


# ── Per-SKU accumulator ────────────────────────────────────────────────────────

def accumulate_sku_profit(sku_id, obj, result, price_map, packaging_map, unique_statuses):
    if sku_id not in obj:
        obj[sku_id] = {
            "order_count":    0,
            "one_unit_price": price_map.get(sku_id, Decimal("0")),
            "total_purchase_cost":  0,
            "settled_amount":       0,
            "affiliate_adj":        0,
            "shipped_count":        0,
            "shipped_settlement":   0,
            "shipped_sale":         0,
            "shipped_expected_profit": 0,
        }

    sku = obj[sku_id]
    net = result["net"]

    _inc(sku, "affiliate_adj", result["affiliate_fees"])
    _ensure_status_buckets(sku, unique_statuses)
    _inc(sku, "order_count")
    
    # ── Settled order: route to status bucket (claim-first priority) ─────────
    _inject_order_into_bucket(sku, result, net)
    _inc(sku, "settled_amount", result["total_settlement"])
    
    if result.get("status") == "Claim":
        _inc(sku, "claims_total", result["claims"])

    sku["net_profit"] = (
        Decimal(sku.get("delivered_profit", 0)) +
        Decimal(sku.get("return_loss",      0)) +
        Decimal(sku.get("rto_loss",         0)) +
        Decimal(sku.get("exchange_loss",     0)) +
        Decimal(sku.get("claim_loss",       0))
    )
    sku["total_purchase_cost"] = Decimal(sku.get("delivered_final_purchase_cost", 0)) +  Decimal(sku.get("exchange_final_purchase_cost", 0)) +  Decimal(sku.get("claim_final_purchase_cost", 0))
    sku["total_tax_cost"] = Decimal(sku.get("delivered_tax_cost", 0)) +  Decimal(sku.get("exchange_tax_cost", 0)) +  Decimal(sku.get("claim_tax_cost", 0))
    sku["total_packaging_cost"] = Decimal(sku.get("delivered_packaging_cost", 0)) +  Decimal(sku.get("exchange_packaging_cost", 0)) +  Decimal(sku.get("claim_packaging_cost", 0))
    sku["total_packaging_cost_for_returns"] = Decimal(sku.get("return_packaging_cost", 0)) + Decimal(sku.get("rto_packaging_cost", 0))
    

@api_view(["GET"])
def available_months(request, business_id):
    """
    Returns distinct order months (YYYY-MM) newest first.
    Primary source: Order.order_date (DateField, reliable).
    Falls back to OrderPayment.order_date if Order table is empty.
    """
    business = get_authorized_business(request, business_id)
    dates = list(Order.objects.filter(business=business).dates("order_date", "month", order="DESC"))
    if not dates:
        dates = list(
            OrderPayment.objects
            .filter(business=business)
            .exclude(order_date=None)
            .dates("order_date", "month", order="DESC")
        )
    return Response([d.strftime("%Y-%m") for d in dates])


@api_view(["GET"])
def unsettled_orders(request, business_id):
    """Orders in the Order table that have no matching OrderPayment record."""
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    page      = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 50))
    search    = request.GET.get("search", "")

    order_qs = Order.objects.filter(business=business)
    if date_from:
        order_qs = order_qs.filter(order_date__gte=date_from)
    if date_to:
        order_qs = order_qs.filter(order_date__lte=date_to)
    if search:
        order_qs = order_qs.filter(
            DQ(sub_order_no__icontains=search) |
            DQ(sku__icontains=search) |
            DQ(product_name__icontains=search)
        )

    settled_nos = set(OrderPayment.objects.filter(business=business).values_list("sub_order_no", flat=True).distinct())
    # Deduplicate to latest status per sub_order_no before listing
    latest_qs    = Order.latest_per_order(base_qs=order_qs)
    unsettled_qs = (
        latest_qs
        .exclude(sub_order_no__in=settled_nos)
        .exclude(reason_for_credit_entry__iexact="cancelled")
        .order_by("-order_date")
    )

    total       = unsettled_qs.count()
    total_value = unsettled_qs.aggregate(v=Sum("supplier_discounted_price"))["v"] or 0

    start = (page - 1) * page_size
    items = unsettled_qs[start: start + page_size]
    return Response({
        "total": total,
        "total_value": float(total_value),
        "page": page,
        "page_size": page_size,
        "results": OrderSerializer(items, many=True).data,
    })


@api_view(["GET"])
def unscheduled_payments(request, business_id):
    """Expected ("unscheduled") payments for shipped-but-not-yet-settled orders.

    These are orders whose settlement hasn't landed yet (no OrderPayment row).
    For each we estimate the *gross* payout from the order's own discounted
    price and compare it against the SKU's landed cost (FinalPrice) to surface
    a rough margin — so low-margin products can be spotted and stopped before
    the money comes in. This is an estimate, not the final settlement (Meesho
    deductions are not applied here).

    Query params: month (YYYY-MM) or date_from/date_to, search, threshold
    (per-unit margin below which a row/product is flagged; default 50),
    page, page_size.
    """
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    month     = request.GET.get("month", "")
    search    = request.GET.get("search", "")
    page      = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 50))
    try:
        threshold = Decimal(str(request.GET.get("threshold", "50")))
    except Exception:
        threshold = Decimal("50")

    if month:
        # Convert YYYY-MM to an inclusive date range.
        y, m = (int(x) for x in month.split("-"))
        date_from = f"{y:04d}-{m:02d}-01"
        date_to = f"{y + (m // 12):04d}-{(m % 12) + 1:02d}-01"  # first of next month (exclusive-ish)

    order_qs = Order.objects.filter(business=business)
    if date_from:
        order_qs = order_qs.filter(order_date__gte=date_from)
    if date_to:
        order_qs = order_qs.filter(order_date__lt=date_to) if month else order_qs.filter(order_date__lte=date_to)
    if search:
        order_qs = order_qs.filter(
            DQ(sub_order_no__icontains=search) |
            DQ(sku__icontains=search) |
            DQ(product_name__icontains=search)
        )

    settled_nos = set(OrderPayment.objects.filter(business=business).values_list("sub_order_no", flat=True).distinct())
    unsettled_qs = (
        Order.latest_per_order(base_qs=order_qs)
        .exclude(sub_order_no__in=settled_nos)
        .exclude(reason_for_credit_entry__iexact="cancelled")
        .order_by("-order_date")
    )

    # SKU -> per-unit landed cost (final_price already bakes in item price,
    # tax and packaging; fall back to item_price when final_price is unset).
    cost_by_sku = _SkuMap()
    for fp in FinalPrice.objects.filter(business=business).only("sku_id", "final_price", "item_price"):
        cost_by_sku[fp.sku_id] = fp.final_price if fp.final_price is not None else fp.item_price

    def estimate(order):
        qty = order.quantity or 1
        expected = (order.supplier_discounted_price or order.supplier_listed_price or Decimal("0")) * qty
        unit_cost = cost_by_sku.get(order.sku)
        has_cost = unit_cost is not None
        cost = (unit_cost or Decimal("0")) * qty
        margin = expected - cost if has_cost else None
        margin_pu = (margin / qty) if (has_cost and qty) else None
        return expected, unit_cost, cost, margin, margin_pu, has_cost

    # ── Per-product aggregation across the WHOLE unsettled set ──
    agg = {}
    for o in unsettled_qs.only("sku", "product_name", "quantity",
                               "supplier_discounted_price", "supplier_listed_price").iterator():
        expected, unit_cost, cost, margin, margin_pu, has_cost = estimate(o)
        key = o.sku or "—"
        a = agg.setdefault(key, {
            "sku": key, "product_name": o.product_name or "",
            "order_count": 0, "total_quantity": 0,
            "total_expected": Decimal("0"), "total_cost": Decimal("0"),
            "total_margin": Decimal("0"), "has_cost": has_cost,
        })
        a["order_count"] += 1
        a["total_quantity"] += (o.quantity or 1)
        a["total_expected"] += expected
        if has_cost:
            a["total_cost"] += cost
            a["total_margin"] += margin
            a["has_cost"] = True

    by_product = []
    for a in agg.values():
        qty = a["total_quantity"] or 1
        avg_pu = (a["total_margin"] / qty) if a["has_cost"] else None
        by_product.append({
            "sku": a["sku"],
            "product_name": a["product_name"],
            "order_count": a["order_count"],
            "total_quantity": a["total_quantity"],
            "total_expected": float(a["total_expected"]),
            "total_cost": float(a["total_cost"]) if a["has_cost"] else None,
            "total_margin": float(a["total_margin"]) if a["has_cost"] else None,
            "avg_margin_per_unit": float(avg_pu) if avg_pu is not None else None,
            "has_cost": a["has_cost"],
            "is_low_margin": bool(a["has_cost"] and avg_pu is not None and avg_pu < threshold),
        })
    # Low-margin (and priced) products first, worst margin at the top.
    by_product.sort(key=lambda p: (p["avg_margin_per_unit"] is None, p["avg_margin_per_unit"] if p["avg_margin_per_unit"] is not None else 0))

    total = unsettled_qs.count()
    total_expected = sum((p["total_expected"] for p in by_product), 0.0)
    at_risk_margin = sum((p["total_margin"] for p in by_product if p["is_low_margin"] and p["total_margin"] is not None), 0.0)
    low_product_count = sum(1 for p in by_product if p["is_low_margin"])

    # ── Page of order rows ──
    start = (page - 1) * page_size
    results = []
    for o in unsettled_qs[start: start + page_size]:
        expected, unit_cost, cost, margin, margin_pu, has_cost = estimate(o)
        results.append({
            "sub_order_no": o.sub_order_no,
            "order_date": o.order_date.strftime("%Y-%m-%d") if o.order_date else None,
            "sku": o.sku,
            "product_name": o.product_name,
            "size": o.size,
            "quantity": o.quantity,
            "status": o.reason_for_credit_entry or "SHIPPED",
            "expected_payout": float(expected),
            "unit_cost": float(unit_cost) if has_cost else None,
            "cost": float(cost) if has_cost else None,
            "est_margin": float(margin) if has_cost else None,
            "est_margin_per_unit": float(margin_pu) if margin_pu is not None else None,
            "has_cost": has_cost,
            "is_low_margin": bool(has_cost and margin_pu is not None and margin_pu < threshold),
        })

    return Response({
        "total": total,
        "total_expected": total_expected,
        "at_risk_margin": at_risk_margin,
        "low_product_count": low_product_count,
        "threshold": float(threshold),
        "page": page,
        "page_size": page_size,
        "results": results,
        "by_product": by_product,
    })


@api_view(["GET"])
def payment_mismatch(request, business_id):
    """
    Bi-directional mismatch:
      - orders_no_payment: Orders in Order table with no matching OrderPayment row
      - payments_no_order: OrderPayment rows whose sub_order_no has no Order row
    """
    from collections import defaultdict

    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to",   "")
    page      = int(request.GET.get("page",      1))
    page_size = int(request.GET.get("page_size", 50))
    view      = request.GET.get("view", "orders")  # "orders" | "payments"

    # ── Orders with no payment ────────────────────────────────────────────────
    order_qs = Order.objects.filter(business=business)
    if date_from:
        order_qs = order_qs.filter(order_date__gte=date_from)
    if date_to:
        order_qs = order_qs.filter(order_date__lte=date_to)

    payment_sub_nos = set(OrderPayment.objects.filter(business=business).values_list("sub_order_no", flat=True).distinct())
    latest_orders   = Order.latest_per_order(base_qs=order_qs)
    orders_no_pay   = (
        latest_orders
        .exclude(sub_order_no__in=payment_sub_nos)
        .exclude(reason_for_credit_entry__iexact="cancelled")
        .order_by("-order_date")
    )
    onp_agg = orders_no_pay.aggregate(count=Count("sub_order_no"), total_value=Sum("supplier_discounted_price"))

    # ── Payments with no order ────────────────────────────────────────────────
    pay_qs = OrderPayment.objects.filter(business=business)
    if date_from:
        pay_qs = pay_qs.filter(order_date__date__gte=date_from)
    if date_to:
        pay_qs = pay_qs.filter(order_date__date__lte=date_to)

    order_sub_nos = set(Order.objects.filter(business=business).values_list("sub_order_no", flat=True).distinct())
    orphan_pays   = pay_qs.exclude(sub_order_no__in=order_sub_nos).order_by("-order_date")
    pno_agg       = orphan_pays.aggregate(
        count=Count("sub_order_no", distinct=True),
        total_settlement=Sum("final_settlement_amount"),
    )

    summary = {
        "orders_no_payment": {
            "count":       onp_agg["count"] or 0,
            "total_value": float(onp_agg["total_value"] or 0),
        },
        "payments_no_order": {
            "count":            pno_agg["count"] or 0,
            "total_settlement": float(pno_agg["total_settlement"] or 0),
        },
    }

    start = (page - 1) * page_size
    if view == "payments":
        total   = pno_agg["count"] or 0
        items   = orphan_pays[start: start + page_size]
        results = OrderPaymentSerializer(items, many=True).data
    else:
        total   = onp_agg["count"] or 0
        items   = orders_no_pay[start: start + page_size]
        results = OrderSerializer(items, many=True).data

    return Response({
        **summary,
        "view":      view,
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "results":   results,
    })


@api_view(["GET"])
def return_claims_detail(request, business_id):
    """
    Grouped view of all RETURN / RTO / CLAIM orders.
    Returns one summary row per sub_order_no + all sub-payment rows.
    """
    from collections import defaultdict

    business = get_authorized_business(request, business_id)
    date_from     = request.GET.get("date_from",  "")
    date_to       = request.GET.get("date_to",    "")
    page          = int(request.GET.get("page",      1))
    page_size     = int(request.GET.get("page_size", 30))
    search        = request.GET.get("search",    "")
    status_filter = request.GET.get("status",    "")

    RETURN_STATUSES = ["RETURN", "RETURNED", "RTO", "RTO_COMPLETE"]

    base_qs = OrderPayment.objects.filter(business=business)
    if date_from:
        base_qs = base_qs.filter(order_date__date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(order_date__date__lte=date_to)

    # Find qualifying sub_order_nos (have a return/RTO status OR a claims row)
    qualifying_ids = (
        base_qs
        .filter(DQ(live_order_status__in=RETURN_STATUSES) | DQ(claims__gt=0))
        .values_list("sub_order_no", flat=True)
        .distinct()
    )

    if search:
        qualifying_ids = (
            OrderPayment.objects
            .filter(
                business=business,
                sub_order_no__in=qualifying_ids,
            )
            .filter(
                DQ(sub_order_no__icontains=search) |
                DQ(supplier_sku__icontains=search) |
                DQ(product_name__icontains=search)
            )
            .values_list("sub_order_no", flat=True)
            .distinct()
        )

    # Claim-first: orders with any claims row are CLAIM, not RETURN or RTO
    _claimed_sub_orders = (
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=qualifying_ids, claims__gt=0)
        .values_list("sub_order_no", flat=True)
    )
    if status_filter == "return":
        # Pure returns: have RETURN status AND no claim payment
        qualifying_ids = (
            OrderPayment.objects
            .filter(business=business, sub_order_no__in=qualifying_ids, live_order_status__in=["RETURN", "RETURNED"])
            .exclude(sub_order_no__in=_claimed_sub_orders)
            .values_list("sub_order_no", flat=True).distinct()
        )
    elif status_filter == "rto":
        # Pure RTOs: have RTO status AND no claim payment
        qualifying_ids = (
            OrderPayment.objects
            .filter(business=business, sub_order_no__in=qualifying_ids, live_order_status__in=["RTO", "RTO_COMPLETE"])
            .exclude(sub_order_no__in=_claimed_sub_orders)
            .values_list("sub_order_no", flat=True).distinct()
        )
    elif status_filter == "claim":
        qualifying_ids = _claimed_sub_orders

    qualifying_list = list(qualifying_ids)
    total = len(qualifying_list)

    # Paginate sub_order_nos
    start          = (page - 1) * page_size
    page_sub_orders = qualifying_list[start: start + page_size]

    # Fetch all payment rows for this page's orders
    all_rows = (
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=page_sub_orders)
        # payment_date sorts ascending so the Return row (latest payment) is last
        .order_by("sub_order_no", "payment_date", "live_order_status")
    )

    groups: dict = defaultdict(list)
    for row in all_rows:
        groups[row.sub_order_no].append(row)

    results = []
    for sub_order_no in page_sub_orders:
        rows = groups.get(sub_order_no, [])

        # Use the same status-based precedence as compute_order_net
        c = _classify_rows(rows)
        settlement_rows, _ = _pick_settlement_rows(c)
        # Total settlement = status rows + adj rows (claims are now in P&L for all types)
        net_settlement = sum(float(r.final_settlement_amount or 0) for r in settlement_rows + c["adj"])
        total_claims     = sum(float(r.claims or 0) for r in rows)
        total_commission = sum(float(r.meesho_commission_incl_gst or 0) for r in rows)
        total_tcs        = sum(float(r.tcs or 0) for r in rows)
        total_tds        = sum(float(r.tds or 0) for r in rows)
        
        latest_status = (settlement_rows[-1].live_order_status
                         if settlement_rows else None)
        first = rows[0] if rows else None

        # Claim-first classification for display
        all_statuses = {(r.live_order_status or "").upper() for r in rows}
        if total_claims > 0:
            order_type = "CLAIM"
        elif all_statuses & {"RETURN", "RETURNED"}:
            order_type = "RETURN"
        elif all_statuses & {"RTO", "RTO_COMPLETE"}:
            order_type = "RTO"
        else:
            order_type = "DELIVERED"

        results.append({
            "sub_order_no":    sub_order_no,
            "sku":             first.supplier_sku if first else None,
            "product_name":    first.product_name if first else None,
            "order_date":      str(first.order_date.date()) if first and first.order_date else None,
            "latest_status":   latest_status,
            "order_type":      order_type,
            "net_settlement":  round(net_settlement,   2),
            "total_claims":    round(total_claims,     2),
            "total_commission": round(total_commission, 2),
            "total_tcs":       round(total_tcs,        2),
            "total_tds":       round(total_tds,        2),
            "payment_count":   len(rows),
            "rows":            OrderPaymentSerializer(rows, many=True).data,
        })

    return Response({
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "results":   results,
    })


@api_view(["GET"])
def claimed_orders(request, business_id):
    """
    Orders where Meesho paid a claim (claims > 0 in any payment row).
    Supports: date_from, date_to, page, page_size,
              status (all/return/rto), sku, view (orders/sku)
    """
    from collections import defaultdict

    business = get_authorized_business(request, business_id)
    date_from     = request.GET.get("date_from",  "")
    date_to       = request.GET.get("date_to",    "")
    page          = int(request.GET.get("page",      1))
    page_size     = int(request.GET.get("page_size", 30))
    status_filter = request.GET.get("status", "")   # "" / "return" / "rto"
    sku_filter    = request.GET.get("sku",    "")
    view_mode     = request.GET.get("view",   "orders")  # "orders" / "sku"

    RETURN_STATUSES = ["RETURN", "RETURNED", "RTO", "RTO_COMPLETE", "PREMIUM_RETURN"]

    base_qs = OrderPayment.objects.filter(business=business)
    if date_from:
        base_qs = base_qs.filter(order_date__date__gte=date_from)
    if date_to:
        base_qs = base_qs.filter(order_date__date__lte=date_to)

    # All sub_order_nos that have at least one claims > 0 row
    claimed_qs = base_qs.filter(claims__gt=0).values_list("sub_order_no", flat=True).distinct()

    # Status filter: narrow to orders that also have the matching status row
    if status_filter == "return":
        claimed_qs = (
            OrderPayment.objects
            .filter(business=business, sub_order_no__in=claimed_qs, live_order_status__in=["RETURN", "RETURNED"])
            .values_list("sub_order_no", flat=True).distinct()
        )
    elif status_filter == "rto":
        claimed_qs = (
            OrderPayment.objects
            .filter(business=business, sub_order_no__in=claimed_qs, live_order_status__in=["RTO", "RTO_COMPLETE"])
            .values_list("sub_order_no", flat=True).distinct()
        )

    # SKU filter
    if sku_filter:
        claimed_qs = (
            OrderPayment.objects
            .filter(business=business, sub_order_no__in=claimed_qs, supplier_sku__icontains=sku_filter)
            .values_list("sub_order_no", flat=True).distinct()
        )

    claimed_list = list(claimed_qs)
    total = len(claimed_list)

    # Summary stats across ALL matching orders (not just this page)
    all_rows_qs = OrderPayment.objects.filter(business=business, sub_order_no__in=claimed_list)
    total_claim_amount = float(
        all_rows_qs.aggregate(t=Sum("claims"))["t"] or 0
    )
    return_count = (
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=claimed_list, live_order_status__in=["RETURN", "RETURNED"])
        .values("sub_order_no").distinct().count()
    )
    rto_count = (
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=claimed_list, live_order_status__in=["RTO", "RTO_COMPLETE"])
        .values("sub_order_no").distinct().count()
    )

    # ── SKU-wise view ─────────────────────────────────────────────────────────
    if view_mode == "sku":
        sku_map: dict = defaultdict(lambda: {
            "sku": None, "product_name": None,
            "total_claims": 0.0, "order_count": 0,
            "return_count": 0, "rto_count": 0, "other_count": 0,
        })
        # Aggregate per (sub_order_no, sku) then roll up to sku
        rows_for_agg = (
            all_rows_qs
            .values("sub_order_no", "supplier_sku", "product_name", "live_order_status", "claims")
        )
        order_sku: dict = {}   # sub_order_no → sku
        order_claims: dict = defaultdict(float)
        order_status: dict = {}  # sub_order_no → primary status
        for r in rows_for_agg:
            so = r["sub_order_no"]
            if so not in order_sku:
                order_sku[so] = r["supplier_sku"]
                sku_map[r["supplier_sku"]]["sku"] = r["supplier_sku"]
                sku_map[r["supplier_sku"]]["product_name"] = r["product_name"]
            if r["claims"]:
                order_claims[so] += float(r["claims"])
            st = (r["live_order_status"] or "").upper()
            if st in ("RETURN",):
                order_status[so] = "return"
            elif st in ("RTO", "RTO_COMPLETE"):
                if order_status.get(so) != "return":
                    order_status[so] = "rto"
            elif st and order_status.get(so) not in ("return", "rto"):
                order_status[so] = "other"

        for so, sku in order_sku.items():
            b = sku_map[sku]
            b["order_count"] += 1
            b["total_claims"] = round(b["total_claims"] + order_claims.get(so, 0), 2)
            st = order_status.get(so, "other")
            b[f"{st}_count"] += 1

        sku_rows = sorted(sku_map.values(), key=lambda x: -x["total_claims"])
        start = (page - 1) * page_size
        return Response({
            "total": len(sku_rows), "page": page, "page_size": page_size,
            "total_claim_amount": round(total_claim_amount, 2),
            "return_count": return_count, "rto_count": rto_count,
            "results": sku_rows[start: start + page_size],
        })

    # ── Order-by-order view ───────────────────────────────────────────────────
    start           = (page - 1) * page_size
    page_sub_orders = claimed_list[start: start + page_size]

    page_rows = (
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=page_sub_orders)
        .order_by("sub_order_no", "payment_date", "live_order_status")
    )
    groups: dict = defaultdict(list)
    for row in page_rows:
        groups[row.sub_order_no].append(row)

    results = []
    for so in page_sub_orders:
        rows = groups.get(so, [])
        c = _classify_rows(rows)
        settlement_rows, _ = _pick_settlement_rows(c)
        net_settlement = sum(float(r.final_settlement_amount or 0) for r in settlement_rows + c["adj"])
        order_claims_total = sum(float(r.claims or 0) for r in rows)
        first = rows[0] if rows else None

        statuses = [r.live_order_status for r in rows if r.live_order_status]
        _STATUS_PREF = {"RETURN": 0, "RTO": 1, "RTO_COMPLETE": 1, "PREMIUM_RETURN": 2,
                        "DELIVERED": 3, "EXCHANGE": 4, "CLAIM": 5, "CANCELLED": 6}
        primary_status = (
            min(statuses, key=lambda s: _STATUS_PREF.get(s.upper(), 99))
            if statuses else None
        )

        results.append({
            "sub_order_no":    so,
            "sku":             first.supplier_sku if first else None,
            "product_name":    first.product_name if first else None,
            "order_date":      str(first.order_date.date()) if first and first.order_date else None,
            "primary_status":  primary_status,
            "net_settlement":  round(net_settlement, 2),
            "total_claims":    round(order_claims_total, 2),
            "quantity":        first.quantity if first else None,
            "rows":            OrderPaymentSerializer(rows, many=True).data,
        })

    return Response({
        "total":              total,
        "page":               page,
        "page_size":          page_size,
        "total_claim_amount": round(total_claim_amount, 2),
        "return_count":       return_count,
        "rto_count":          rto_count,
        "results":            results,
    })


def _build_payout_breakdown(qs):
    """
    Raw per-status settlement for payout CSV reconciliation.
    Groups payment rows by live_order_status and returns count + settlement
    in the same format as the Meesho monthly payout summary.
    """
    from django.db.models import Count, Sum as DSum

    STATUS_LABELS = {
        "delivered":  "Delivered",
        "return":     "Returned",
        "returned":   "Returned",
        "rto":        "RTO",
        "rto_complete": "RTO",
        "exchange":   "Exchanged",
        "exchanged":  "Exchanged",
        "cancelled":  "Cancelled",
        "shipped":    "Shipped (in-transit)",
        "in transit": "Shipped (in-transit)",
    }

    # Status rows (non-null)
    status_rows = (
        qs.exclude(DQ(live_order_status__isnull=True) | DQ(live_order_status=""))
        .values("live_order_status")
        .annotate(
            order_count=Count("sub_order_no", distinct=True),
            settlement=DSum("final_settlement_amount"),
            return_shipping=DSum("return_shipping_charge"),
        )
    )

    buckets = {}
    for row in status_rows:
        raw = (row["live_order_status"] or "").lower()
        label = STATUS_LABELS.get(raw, row["live_order_status"])
        if label not in buckets:
            buckets[label] = {"count": 0, "settlement": Decimal("0"), "return_shipping": Decimal("0")}
        buckets[label]["count"]           += row["order_count"]
        buckets[label]["settlement"]      += Decimal(str(row["settlement"] or 0))
        buckets[label]["return_shipping"] += Decimal(str(row["return_shipping"] or 0))

    # Adj rows — claims and other (affiliate, pickup, etc.)
    adj_rows = qs.filter(DQ(live_order_status__isnull=True) | DQ(live_order_status=""))
    claim_agg = adj_rows.filter(claims__gt=0).aggregate(
        count=Count("sub_order_no", distinct=True),
        total_claims=DSum("claims"),
        settlement=DSum("final_settlement_amount"),
    )
    other_adj_agg = adj_rows.filter(
        DQ(claims__isnull=True) | DQ(claims=0)
    ).values("recovery_reason").annotate(
        count=Count("id"),
        settlement=DSum("final_settlement_amount"),
    )

    result = []
    order_priority = ["Delivered", "Returned", "RTO", "Exchanged", "Cancelled", "Shipped (in-transit)"]
    for label in order_priority:
        if label in buckets:
            b = buckets[label]
            result.append({
                "type": label,
                "count": b["count"],
                "settlement": round(float(b["settlement"]), 2),
                "return_shipping": round(float(b["return_shipping"]), 2),
            })
    # Remaining statuses not in priority list
    for label, b in buckets.items():
        if label not in order_priority:
            result.append({
                "type": label,
                "count": b["count"],
                "settlement": round(float(b["settlement"]), 2),
                "return_shipping": round(float(b["return_shipping"]), 2),
            })

    # Claims
    if claim_agg["count"]:
        result.append({
            "type": "Claims Accepted",
            "count": claim_agg["count"],
            "settlement": round(float(claim_agg["settlement"] or 0), 2),
            "claims_amount": round(float(claim_agg["total_claims"] or 0), 2),
        })

    # Other adj rows (affiliate fees, manual pickup, etc.)
    for row in other_adj_agg:
        reason = row["recovery_reason"] or "Unknown Adj"
        result.append({
            "type": f"Adj: {reason}",
            "count": row["count"],
            "settlement": round(float(row["settlement"] or 0), 2),
        })

    return result



@api_view(["GET"])
def profit_summary(request, business_id):
    """
    Calculate overall Meesho profit.
    Accepts date_from / date_to (YYYY-MM-DD).
    Filters OrderPayment via Order.order_date join; falls back to
    OrderPayment.order_date__date if Order table has no records for range.
    """
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")

    qs = OrderPayment.objects.filter(business=business)


    if date_from or date_to:
        ord_qs = Order.objects.filter(business=business)
        if date_from:
            # qs = qs.filter(order_date__gte=date_from)
            ord_qs = ord_qs.filter(order_date__gte=date_from)
        if date_to:
            # qs = qs.filter(order_date__lte=date_to)
            ord_qs = ord_qs.filter(order_date__lte=date_to)
        # Use a DB subquery — avoids loading thousands of IDs into Python memory
        if ord_qs.exists():
            qs = qs.filter(sub_order_no__in=ord_qs.values("sub_order_no"))
        else:
            # Fallback: filter OrderPayment.order_date directly (DateTimeField)
            if date_from:
                qs = qs.filter(order_date__date__gte=date_from)
            if date_to:
                qs = qs.filter(order_date__date__lte=date_to)

    # Load pricing once (include item_price + tax_percent for tax cost calculation)
    _fp_all        = list(FinalPrice.objects.filter(business=business).only("sku_id", "final_price", "packaging_cost", "parent_id", "item_price", "tax_percent"))
    price_map      = _SkuMap((fp.sku_id, fp.final_price    or Decimal("0")) for fp in _fp_all)
    packaging_map  = _SkuMap((fp.sku_id, fp.packaging_cost or Decimal("0")) for fp in _fp_all)
    item_price_map = _SkuMap((fp.sku_id, fp.item_price     or Decimal("0")) for fp in _fp_all)
    tax_map        = _SkuMap((fp.sku_id, fp.tax_percent    or 0)            for fp in _fp_all)
    sku_parent_map = _SkuMap((fp.sku_id, fp.parent_id) for fp in _fp_all if fp.parent_id)
    # Any casing of a SKU -> the spelling stored in pricing, so orders that spell
    # the same SKU differently accumulate into one row instead of two.
    canonical_sku  = _SkuMap((fp.sku_id, fp.sku_id) for fp in _fp_all)
    
    # Key parent maps by the parent's surrogate id, matching FinalPrice.parent_id
    # and ParentPriceHistory.parent_id (both now store the integer id).
    _fp_parent     = list(ParentItemPrice.objects.filter(business=business).only("item_id", "item_price", "tax_percent", "packaging_cost", "final_price"))
    parent_price_map      = {fp.id: fp.final_price    or Decimal("0") for fp in _fp_parent}
    parent_packaging_map  = {fp.id: fp.packaging_cost or Decimal("0") for fp in _fp_parent}
    parent_item_price_map = {fp.id: fp.item_price     or Decimal("0") for fp in _fp_parent}
    parent_tax_map        = {fp.id: fp.tax_percent    or 0            for fp in _fp_parent}

    # Build date-effective price history: {parent_id: [(effective_from, final_price, packaging_cost, item_price, tax_percent), ...]}
    from collections import defaultdict
    _hist = list(
        ParentPriceHistory.objects
        .filter(business=business)
        .values("parent_id", "effective_from", "final_price", "packaging_cost", "item_price", "tax_percent")
        .order_by("effective_from")
    )
    _parent_histories = defaultdict(list)
    for h in _hist:
        _parent_histories[h["parent_id"]].append((
            h["effective_from"],
            h["final_price"]    or Decimal("0"),
            h["packaging_cost"] or Decimal("0"),
            h["item_price"]     or Decimal("0"),
            h["tax_percent"]    or 0,
        ))

    def get_eff_price(sku_id, order_date):
        """
        (final_price, packaging_cost, item_price, tax_percent) for a SKU.

        Priority is the parent's price, then the SKU's own — in that order and
        regardless of whether an order date is known. Previously the parent was
        only consulted when an order_date was present, so any order without one
        silently fell back to SKU pricing even when a parent existed.

        A parent counts as "priced" only if it actually carries an item price;
        an empty parent record falls through to the SKU so a missing parent
        price can't zero out a cost.
        """
        pid = sku_parent_map.get(sku_id)

        if pid:
            # Dated parent price history wins when we know when the order was placed.
            if order_date:
                od = order_date.date() if hasattr(order_date, "date") else order_date
                applicable = [
                    (d, fp, pkg, ip, tax)
                    for d, fp, pkg, ip, tax in _parent_histories[pid] if d <= od
                ]
                if applicable:
                    _, fp, pkg, ip, tax = applicable[-1]
                    if ip:
                        return fp, pkg, ip, tax

            # Otherwise the parent's current price.
            p_item = parent_item_price_map.get(pid, Decimal("0"))
            if p_item:
                return (
                    parent_price_map.get(pid,      Decimal("0")),
                    parent_packaging_map.get(pid,  Decimal("0")),
                    p_item,
                    parent_tax_map.get(pid, 0),
                )

        # No parent, or the parent has no price of its own — use the SKU.
        return (
            price_map.get(sku_id,      Decimal("0")),
            packaging_map.get(sku_id,  Decimal("0")),
            item_price_map.get(sku_id, Decimal("0")),
            tax_map.get(sku_id, 0),
        )

    # Pre-fetch distinct statuses ONCE so _ensure_status_buckets can pre-initialise keys
    unique_statuses = list(["Claim","Cancelled", "Delivered", "Return", "RTO", "Shipped", "Exchange", "Unknown"])

    _FIELDS = (
        "sub_order_no", "supplier_sku", "quantity",
        "final_settlement_amount", "live_order_status",
        "recovery_reason", "claims", "return_shipping_charge",
        "order_date", "payment_date",
    )

    # Group ALL payment rows by sub_order_no
    order_groups = defaultdict(list)
    for payment in qs.only(*_FIELDS).order_by("-payment_date"):
        order_groups[payment.sub_order_no].append(payment)

    order_wise_profit   = {}
    missing_sku         = []
    orders_with_price   = 0
    orders_missing_price = 0
    
    for sub_no, payments in order_groups.items():
        order = (
            Order.objects
            .filter(business=business, sub_order_no=sub_no)
            .values("sub_order_no", "sku", "quantity")
            .first()
            ) or {}
        primary = next((p for p in payments if p.live_order_status), payments[0])
        sku = order.get("sku") or primary.supplier_sku
        qty = order.get("quantity") or primary.quantity

        if not sku or sku not in price_map:
            missing_sku.append(sku)
            orders_missing_price += 1
            continue

        eff_price, eff_pkg, eff_item_price, eff_tax_pct = get_eff_price(sku, primary.order_date)
        result = compute_order_net(payments, eff_price, eff_pkg, qty, unique_statuses,eff_item_price, eff_tax_pct, )
        accumulate_sku_profit(canonical_sku.get(sku, sku), order_wise_profit, result, price_map, packaging_map, unique_statuses)
        orders_with_price += 1


    missing_sku          = list(set(missing_sku))
    total_loss           = sum(Decimal(str(v.get("return_loss",      0) or 0)) for v in order_wise_profit.values())
    total_rto_loss       = sum(Decimal(str(v.get("rto_loss",         0) or 0)) for v in order_wise_profit.values())
    total_exchange_net   = sum(Decimal(str(v.get("exchange_loss",     0) or 0)) for v in order_wise_profit.values())
    total_other          = sum(Decimal(str(v.get("other_net",        0) or 0)) for v in order_wise_profit.values())
    total_purchase_cost  = sum(Decimal(str(v.get("total_purchase_cost",  0) or 0)) for v in order_wise_profit.values())
    total_packaging_cost = sum(Decimal(str(v.get("total_packaging_cost", 0) or 0)) for v in order_wise_profit.values())
    total_packaging_cost_for_return = sum(Decimal(str(v.get("total_packaging_cost_for_returns", 0) or 0)) for v in order_wise_profit.values())
    total_tax_cost = sum(Decimal(str(v.get("total_tax_cost", 0) or 0)) for v in order_wise_profit.values())

    # Counts per outcome (for rate calculations)
    delivered_summary = status_wise_summary(order_wise_profit, "delivered")
    exchanged_summary = status_wise_summary(order_wise_profit, "exchange")
    rto_summary =  status_wise_summary(order_wise_profit, "rto")
    return_summary =  status_wise_summary(order_wise_profit, "return")
    claim_summary = status_wise_summary(order_wise_profit, "claim")
    unknown_summary = status_wise_summary(order_wise_profit, "unknown")
    
    order_status_summary = {
        "delivered_summary":delivered_summary,
        "exchanged_summary":exchanged_summary,
        "rto_summary": rto_summary,
        "return_summary":return_summary,
        "claim_summary":claim_summary,
        "unknown_summary": unknown_summary
    }
    
    
    
    total_return_count     = sum(v.get("return_count",     0) for v in order_wise_profit.values())
    total_rto_count        = sum(v.get("rto_count",        0) for v in order_wise_profit.values())
    total_exchange_count   = sum(v.get("exchange_count",   0) for v in order_wise_profit.values())
    total_other_count      = sum(v.get("other_count",      0) for v in order_wise_profit.values())
    total_claim_count      = sum(v.get("claim_count",      0) for v in order_wise_profit.values())
    total_claim_loss       = sum(Decimal(str(v.get("claim_loss",         0) or 0)) for v in order_wise_profit.values())
    total_claim_purchase_cost  = sum(Decimal(str(v.get("claim_purchase_cost",     0) or 0)) for v in order_wise_profit.values())
    total_exchange_pkg_cost    = sum(Decimal(str(v.get("exchange_packaging_cost",  0) or 0)) for v in order_wise_profit.values())
    total_return_pkg_cost      = sum(Decimal(str(v.get("return_packaging_cost",   0) or 0)) for v in order_wise_profit.values())
    total_rto_pkg_cost         = sum(Decimal(str(v.get("rto_packaging_cost",      0) or 0)) for v in order_wise_profit.values())

    # Aggregate affiliate fees — blank-status rows, forced negative (always a cost)
    adj_qs = qs.filter(
    (DQ(live_order_status__isnull=True) | DQ(live_order_status="")) &
    DQ(recovery_reason="Affiliate Fee")
    )
    raw_aff = adj_qs.aggregate(total=Sum("final_settlement_amount"))["total"] or Decimal("0")
    total_affiliate_fee = -abs(raw_aff)

    # Approved claims (positive claims field = money credited to supplier)
    total_claims = (
        qs.filter(claims__gt=0).aggregate(total=Sum("claims"))["total"] or Decimal("0")
    )

    ads_qs = AdsCost.objects.filter(business=business)
    if date_from:
        ads_qs = ads_qs.filter(deduction_date__gte=date_from)
    if date_to:
        ads_qs = ads_qs.filter(deduction_date__lte=date_to)
    raw_ads = ads_qs.aggregate(total=Sum("total_ads_cost"))["total"] or Decimal("0")
    # Ads cost is always a deduction from seller earnings. Meesho's Excel stores it
    # as a positive spend amount. Force negative so the formula: order_net_pl + ads
    # always subtracts the ad spend regardless of how it's stored.
    ads = -abs(raw_ads)

    ref_qs = ReferralPayment.objects.filter(business=business)
    if date_from:
        ref_qs = ref_qs.filter(payment_date__gte=date_from)
    if date_to:
        ref_qs = ref_qs.filter(payment_date__lte=date_to)
    referral = ref_qs.aggregate(total=Sum("net_referral_amount"))["total"] or Decimal("0")

    comp_qs = CompensationRecovery.objects.filter(business=business)
    if date_from:
        comp_qs = comp_qs.filter(date__gte=date_from)
    if date_to:
        comp_qs = comp_qs.filter(date__lte=date_to)
    comp_recovery = comp_qs.aggregate(total=Sum("amount_incl_gst"))["total"] or Decimal("0")

    # SHIPPED rows are PROVISIONAL — exclude them from ALL financial aggregates.
    # Rationale:
    #   • Shipped-only orders: still in transit, payment not yet earned.
    #   • Orders that progressed SHIPPED → DELIVERED/RETURN/RTO: the final
    #     status row has the actual settlement; SHIPPED row was just a preview.
    # Adjustment rows (live_order_status NULL/empty) are claims/affiliate fees
    # and must be INCLUDED — Django's exclude() leaves NULLs in.
    settled_qs = qs.exclude(
        DQ(live_order_status__iexact="shipped") | DQ(live_order_status__iexact="in transit")
    )

    agg = settled_qs.aggregate(
        revenue=Sum("final_settlement_amount"),
        gross_revenue=Sum("total_sale_amount"),
        total_commission=Sum("meesho_commission_incl_gst"),
        total_tcs=Sum("tcs"),
        total_tds=Sum("tds"),
        total_forward_shipping=Sum("shipping_charge_incl_gst"),
        total_return_shipping=Sum("return_shipping_charge"),
    )

    revenue                  = agg["revenue"]                  or Decimal("0")
    gross_revenue            = agg["gross_revenue"]            or Decimal("0")
    total_commission         = agg["total_commission"]         or Decimal("0")
    total_tcs                = agg["total_tcs"]                or Decimal("0")
    total_tds                = agg["total_tds"]                or Decimal("0")
    total_forward_shipping   = agg["total_forward_shipping"]   or Decimal("0")
    total_return_shipping    = agg["total_return_shipping"]    or Decimal("0")
    total_shipping_cost      = total_forward_shipping + total_return_shipping

    # ── Transportation charges ────────────────────────────────────────────────
    # A business-level overhead: it can't be attributed to an individual order,
    # so it is applied to the final bottom line rather than to Order Net P&L
    # (which the per-SKU figures reconcile against). Deducted only when this
    # business has the switch on in its profile; the total is always reported so
    # the figure stays visible either way.
    transport_qs = TransportCharge.objects.filter(business=business)
    if date_from:
        transport_qs = transport_qs.filter(date__gte=date_from)
    if date_to:
        transport_qs = transport_qs.filter(date__lte=date_to)
    transport_total = transport_qs.aggregate(t=Sum("amount"))["t"] or Decimal("0")

    profile = getattr(business, "profile", None)
    deduct_transport = bool(profile and profile.deduct_transport_charges)
    transport_deducted = transport_total if deduct_transport else Decimal("0")

    # Order-level result: settlement less item cost. `ads` is already negative.
    net_profit_loss = revenue - total_purchase_cost
    net_revenue = revenue - total_purchase_cost + ads + comp_recovery - transport_deducted

    # Ensure sum(sku.net_profit) == revenue - total_purchase_cost by absorbing the gap
    # (settlement from missing-SKU orders) into a synthetic __unattributed__ bucket.
    total_sku_net = sum(Decimal(str(v.get("net_profit", 0) or 0)) for v in order_wise_profit.values())
    unattributed_gap = (revenue - total_purchase_cost) - total_sku_net
    if abs(unattributed_gap) > Decimal("0.01"):
        order_wise_profit["__unattributed__"] = {
            "net_profit": float(unattributed_gap),
            "order_count": orders_missing_price,
            "sku_id": "__unattributed__",
            "one_unit_price": 0,
            "total_purchase_cost": 0,
        }

    # Keep raw DB aggregate as informational (includes claims adj rows) — NOT used in P&L
    raw_settlement = revenue

    # Tax summary — what % of gross revenue was withheld as tax
    total_tax_withheld = total_tcs + total_tds
    tax_rate_pct = round(float(total_tax_withheld / gross_revenue * 100), 2) if gross_revenue else 0.0
    commission_rate_pct = round(float(total_commission / gross_revenue * 100), 2) if gross_revenue else 0.0
    total_deduction_rate_pct = round(float((total_tax_withheld + total_commission) / gross_revenue * 100), 2) if gross_revenue else 0.0



    return Response({
        "gross_revenue":          round(gross_revenue, 2),
        "total_settled":          round(revenue, 2), 
        "total_purchase_cost":    round(total_purchase_cost, 2),
        "total_packaging_cost":   round(total_packaging_cost, 2),
        "total_packaging_cost_for_returns": round(total_packaging_cost_for_return, 2),
        "total_tax_cost":         round(total_tax_cost, 2),
        "total_transport_charges":   round(transport_total, 2),
        "transport_charges_deducted": deduct_transport,
        "net_profit_loss": round(float(net_profit_loss), 2),
        "order_summary": order_status_summary,
        "total_loss":             round(total_loss, 2),
        "total_rto_loss":         round(total_rto_loss, 2),
        "total_exchange_net":     round(total_exchange_net, 2),
        "total_exchange_count":   total_exchange_count,
        "total_exchange_pkg_cost": round(total_exchange_pkg_cost, 2),
        "total_other":            round(total_other, 2),
        "net_revenue":            round(net_revenue, 2),
        "sku_wise_profit":        order_wise_profit,
        "orders_with_price":      orders_with_price,
        "orders_missing_price":   orders_missing_price,
        "orders_missing_sku":     len(missing_sku),
        "missing_sku":            missing_sku,
        "total_ads_cost":               round(ads, 2),
        "total_referral_income":        round(referral, 2),
        "total_compensation_recovery":  round(comp_recovery, 2),
        "total_commission_paid":        round(total_commission, 2),
        "total_tcs":                    round(total_tcs, 2),
        "total_tds":                    round(total_tds, 2),
        "total_shipping_cost":          round(total_shipping_cost, 2),
        "total_forward_shipping":       round(total_forward_shipping, 2),
        "total_return_shipping":        round(total_return_shipping, 2),
        "net_settlement_revenue":       round(revenue, 2),
        "total_claims":                 round(total_claims, 2),
        "total_affiliate_fee":          round(total_affiliate_fee, 2),
        "total_pure_returns":           total_return_count,
        "total_other_count":            total_other_count,
        "total_claimed_orders":         qs.filter(claims__gt=0).values("sub_order_no").distinct().count(),
        "order_count":                  len(order_groups),
        "adjustment_count":             adj_qs.count(),
        "ads_campaigns":                ads_qs.count(),
        "referral_count":               ref_qs.count(),
        "compensation_recovery_count":  comp_qs.count(),
        "total_return_count":          total_return_count,
        "total_rto_count":             total_rto_count,
        "total_claim_count":           total_claim_count,
        "total_claim_loss":            round(total_claim_loss, 2),
        "total_claim_purchase_cost":   round(total_claim_purchase_cost, 2),
        "total_return_pkg_cost":       round(total_return_pkg_cost, 2),
        "total_rto_pkg_cost":          round(total_rto_pkg_cost, 2),
        "tax_summary": {
            "total_tax_withheld":       round(float(total_tax_withheld), 2),
            "tax_rate_pct":             tax_rate_pct,
            "commission_rate_pct":      commission_rate_pct,
            "total_deduction_rate_pct": total_deduction_rate_pct,
        },
        "payout_breakdown": _build_payout_breakdown(qs),
    })


@api_view(["GET"])
def profit_daily_summary(request, business_id):
    """
    Per-day profit: for each day orders shipped (Order.order_date), how many
    shipped, how each has settled so far (Delivered/RTO/Return/Exchange/Claim),
    how many are still awaiting a Meesho payment sheet, and the resulting net
    profit for that day's shipments.

    Deliberately settlement-based, not a "T+N days" estimate: an order counts
    toward its ship day's profit only once compute_order_net can classify it
    from real OrderPayment data, same as /profit/. A day's numbers fill in and
    firm up over the following days/weeks as more sheets are uploaded — a
    "shipped today" row legitimately starts as all no_payment_count and settles
    in over time, which is exactly what payment_delay looks like without having
    to model it as a guess.

    Uses its own copy of the SKU/parent pricing lookup (mirrors get_eff_price in
    profit_summary, same as _estimated_profit_pricing_lookup does for the CSV
    estimate) rather than refactoring that function, so this new endpoint can't
    regress the existing /profit/ figures. The actual per-order math is the same
    compute_order_net() used everywhere else — only the grouping key (ship day
    instead of SKU) differs.

    Accepts date_from / date_to (YYYY-MM-DD). Defaults to the last 30 days and
    caps the window at 92 days — a per-day breakdown over someone's entire order
    history isn't useful to render and would be a heavy per-order scan for every
    day in between.
    """
    from collections import defaultdict

    business = get_authorized_business(request, business_id)

    today = timezone.localdate()
    d_from = _parse_day(request.GET.get("date_from")) or (today - timedelta(days=29))
    d_to   = _parse_day(request.GET.get("date_to")) or today
    if d_to < d_from:
        d_from, d_to = d_to, d_from
    if (d_to - d_from).days > 92:
        d_from = d_to - timedelta(days=92)

    order_qs = Order.objects.filter(business=business, order_date__gte=d_from, order_date__lte=d_to)

    # Shipped count per day — same grouping dashboard_analytics uses.
    shipped_by_day = {
        row["order_date"]: row["n"]
        for row in order_qs.values("order_date").annotate(n=Count("sub_order_no", distinct=True))
    }
    # A sub-order's rows all share one order_date (see _resolve_ship_status), so
    # any row gives the day it shipped, however its status later progresses.
    order_date_map = dict(order_qs.values_list("sub_order_no", "order_date"))
    sku_qty_map = {
        row["sub_order_no"]: (row["sku"], row["quantity"])
        for row in order_qs.values("sub_order_no", "sku", "quantity")
    }

    # ── Pricing lookup — see docstring: a deliberate copy, not a shared helper. ──
    _fp_all = list(FinalPrice.objects.filter(business=business).only(
        "sku_id", "final_price", "packaging_cost", "parent_id", "item_price", "tax_percent"))
    price_map      = _SkuMap((fp.sku_id, fp.final_price    or Decimal("0")) for fp in _fp_all)
    packaging_map  = _SkuMap((fp.sku_id, fp.packaging_cost or Decimal("0")) for fp in _fp_all)
    item_price_map = _SkuMap((fp.sku_id, fp.item_price     or Decimal("0")) for fp in _fp_all)
    tax_map        = _SkuMap((fp.sku_id, fp.tax_percent    or 0)            for fp in _fp_all)
    sku_parent_map = _SkuMap((fp.sku_id, fp.parent_id) for fp in _fp_all if fp.parent_id)

    _fp_parent = list(ParentItemPrice.objects.filter(business=business).only(
        "item_id", "item_price", "tax_percent", "packaging_cost", "final_price"))
    parent_price_map      = {fp.id: fp.final_price    or Decimal("0") for fp in _fp_parent}
    parent_packaging_map  = {fp.id: fp.packaging_cost or Decimal("0") for fp in _fp_parent}
    parent_item_price_map = {fp.id: fp.item_price     or Decimal("0") for fp in _fp_parent}
    parent_tax_map        = {fp.id: fp.tax_percent    or 0            for fp in _fp_parent}

    _hist = list(
        ParentPriceHistory.objects.filter(business=business)
        .values("parent_id", "effective_from", "final_price", "packaging_cost", "item_price", "tax_percent")
        .order_by("effective_from")
    )
    _parent_histories = defaultdict(list)
    for h in _hist:
        _parent_histories[h["parent_id"]].append((
            h["effective_from"], h["final_price"] or Decimal("0"),
            h["packaging_cost"] or Decimal("0"), h["item_price"] or Decimal("0"),
            h["tax_percent"] or 0,
        ))

    def get_eff_price(sku_id, order_date):
        pid = sku_parent_map.get(sku_id)
        if pid:
            if order_date:
                od = order_date.date() if hasattr(order_date, "date") else order_date
                applicable = [
                    (d, fp, pkg, ip, tax)
                    for d, fp, pkg, ip, tax in _parent_histories[pid] if d <= od
                ]
                if applicable:
                    _, fp, pkg, ip, tax = applicable[-1]
                    if ip:
                        return fp, pkg, ip, tax
            p_item = parent_item_price_map.get(pid, Decimal("0"))
            if p_item:
                return (
                    parent_price_map.get(pid, Decimal("0")),
                    parent_packaging_map.get(pid, Decimal("0")),
                    p_item, parent_tax_map.get(pid, 0),
                )
        return (
            price_map.get(sku_id, Decimal("0")), packaging_map.get(sku_id, Decimal("0")),
            item_price_map.get(sku_id, Decimal("0")), tax_map.get(sku_id, 0),
        )

    unique_statuses = ["Claim", "Cancelled", "Delivered", "Return", "RTO", "Shipped", "Exchange", "Unknown"]
    STATUS_TO_KEY = {
        "Delivered": "delivered", "Return": "return", "RTO": "rto",
        "Exchange": "exchange", "Claim": "claim",
        "Unknown": "unknown", "Cancelled": "unknown",
    }
    BUCKET_KEYS = ["delivered", "return", "rto", "exchange", "claim", "unknown"]

    payments = OrderPayment.objects.filter(
        business=business, sub_order_no__in=order_date_map.keys()
    ).only(
        "sub_order_no", "supplier_sku", "quantity", "final_settlement_amount",
        "live_order_status", "recovery_reason", "claims", "return_shipping_charge",
        "order_date", "payment_date",
    ).order_by("-payment_date")

    order_groups = defaultdict(list)
    for payment in payments:
        order_groups[payment.sub_order_no].append(payment)

    days = {}

    def day_bucket(d):
        if d not in days:
            days[d] = {
                "date": d.isoformat(), "shipped_count": shipped_by_day.get(d, 0),
                "delivered_count": 0, "return_count": 0, "rto_count": 0,
                "exchange_count": 0, "claim_count": 0, "unknown_count": 0,
                "settled_count": 0, "no_payment_count": 0, "missing_price_count": 0,
                "net_profit": Decimal("0"),
            }
        return days[d]

    # Every day with a shipment appears even if nothing on it has settled yet.
    for d in shipped_by_day:
        day_bucket(d)

    for sub_no, group in order_groups.items():
        d = order_date_map.get(sub_no)
        if d is None:
            continue
        bucket = day_bucket(d)
        sku, qty = sku_qty_map.get(sub_no, (None, None))
        primary = next((p for p in group if p.live_order_status), group[0])
        sku = sku or primary.supplier_sku
        qty = qty or primary.quantity

        if not sku or sku not in price_map:
            bucket["missing_price_count"] += 1
            bucket["unknown_count"] += 1
            continue

        eff_price, eff_pkg, eff_item_price, eff_tax_pct = get_eff_price(sku, primary.order_date)
        result = compute_order_net(group, eff_price, eff_pkg, qty, unique_statuses, eff_item_price, eff_tax_pct)
        key = STATUS_TO_KEY.get(result["status"], "unknown")
        bucket[f"{key}_count"] += 1
        if key != "unknown":
            bucket["settled_count"] += 1
            bucket["net_profit"] += result["net"]

    for d, count in shipped_by_day.items():
        bucket = day_bucket(d)
        accounted = sum(bucket[f"{k}_count"] for k in BUCKET_KEYS)
        bucket["no_payment_count"] = max(0, count - accounted)

    ordered = sorted(days.values(), key=lambda r: r["date"])
    for row in ordered:
        row["net_profit"] = round(float(row["net_profit"]), 2)

    totals = {key: sum(r[key] for r in ordered) for key in (
        "shipped_count", "delivered_count", "return_count", "rto_count",
        "exchange_count", "claim_count", "unknown_count", "no_payment_count",
        "settled_count", "missing_price_count",
    )}
    totals["net_profit"] = round(sum(r["net_profit"] for r in ordered), 2)

    return Response({
        "date_from": d_from.isoformat(),
        "date_to": d_to.isoformat(),
        "days": ordered,
        "totals": totals,
    })


# ── Estimated Profit from an uploaded Order Summary CSV ───────────────────────

_ESTIMATED_PROFIT_EXCLUDED_STATUSES = {"in transit", "ready to ship"}
_ESTIMATED_PROFIT_BUCKETS = ["delivered", "return", "rto", "exchange", "claim", "other"]
_ESTIMATED_PROFIT_STATUS_LABELS = {
    "delivered": "Delivered", "return": "Returned", "rto": "RTO",
    "exchange": "Exchanged", "claim": "Claim", "other": "Other / Cancelled",
}


def _estimated_profit_pricing_lookup(business):
    """
    Build a get_price(sku_id, order_date) -> (item_price, packaging_cost, tax_percent)
    closure using the same effective-date SKU/parent pricing as /profit/, so the
    uploaded-CSV estimate uses identical unit economics to the Overview tab.
    """
    from collections import defaultdict

    fp_all = list(FinalPrice.objects.filter(business=business).only(
        "sku_id", "item_price", "packaging_cost", "tax_percent", "parent_id"))
    item_price_map = _SkuMap((fp.sku_id, fp.item_price or Decimal("0")) for fp in fp_all)
    packaging_map  = _SkuMap((fp.sku_id, fp.packaging_cost or Decimal("0")) for fp in fp_all)
    tax_map        = _SkuMap((fp.sku_id, fp.tax_percent or 0) for fp in fp_all)
    sku_parent_map = _SkuMap((fp.sku_id, fp.parent_id) for fp in fp_all if fp.parent_id)
    # Canonical keys, so membership must be tested with _sku_key() too.
    known_skus     = set(item_price_map.keys())
    # Any casing of a SKU -> the spelling stored in pricing, so the per-SKU
    # aggregation shows one row per SKU rather than one per spelling.
    canonical_sku  = _SkuMap((fp.sku_id, fp.sku_id) for fp in fp_all)

    parent_qs = ParentItemPrice.objects.filter(business=business).only(
        "id", "item_price", "tax_percent", "packaging_cost")
    parent_item_price_map = {p.id: p.item_price or Decimal("0") for p in parent_qs}
    parent_packaging_map  = {p.id: p.packaging_cost or Decimal("0") for p in parent_qs}
    parent_tax_map        = {p.id: p.tax_percent or 0 for p in parent_qs}

    hist = list(
        ParentPriceHistory.objects.filter(business=business)
        .values("parent_id", "effective_from", "packaging_cost", "item_price", "tax_percent")
        .order_by("effective_from")
    )
    parent_histories = defaultdict(list)
    for h in hist:
        parent_histories[h["parent_id"]].append((
            h["effective_from"], h["item_price"] or Decimal("0"),
            h["packaging_cost"] or Decimal("0"), h["tax_percent"] or 0,
        ))

    def get_price(sku_id, order_date):
        if order_date is not None and hasattr(order_date, "date"):
            order_date = order_date.date()
        # Parent price first, SKU price only as a fallback — and a parent with no
        # price of its own must not zero out the cost.
        pid = sku_parent_map.get(sku_id)
        if pid:
            applicable = [
                (d, ip, pkg, tax) for d, ip, pkg, tax in parent_histories[pid]
                if order_date and d <= order_date
            ]
            if applicable:
                _, ip, pkg, tax = applicable[-1]
                if ip:
                    return ip, pkg, tax
            p_item = parent_item_price_map.get(pid, Decimal("0"))
            if p_item:
                return (
                    p_item,
                    parent_packaging_map.get(pid, Decimal("0")),
                    parent_tax_map.get(pid, 0),
                )
        return (
            item_price_map.get(sku_id, Decimal("0")),
            packaging_map.get(sku_id, Decimal("0")),
            tax_map.get(sku_id, 0),
        )

    return get_price, known_skus, canonical_sku


def _to_num(v):
    return round(float(v), 2) if isinstance(v, Decimal) else v


def _compute_estimated_profit_summary(business, date_from=None, date_to=None):
    """
    Live aggregate over persisted EstimatedProfitOrder rows — cost/profit is
    recomputed against *current* SKU pricing every call, so adding a missing
    SKU's price (or correcting one) is reflected immediately without needing
    to re-upload the sheet.

    Cost formula (matches Overview / compute_order_net): Delivered and an
    approved Claim both deduct item cost×qty + tax + packaging; Exchanged
    deducts the same but with packaging doubled. Returned, RTO, and a rejected
    (or otherwise non-approved) Claim deduct nothing — the payout stands as-is.

    Those payout-only rows (Returned / RTO / non-approved Claim) only represent
    a real loss when Meesho actually deducted money back — i.e. the settlement
    is negative or zero. A positive settlement doesn't fit that pattern, so the
    row is dropped from the estimate entirely, the same as an In Transit / Ready
    To Ship row.
    """
    get_price, known_skus, canonical_sku = _estimated_profit_pricing_lookup(business)

    qs = EstimatedProfitOrder.objects.filter(business=business)
    if date_from:
        qs = qs.filter(order_date__gte=date_from)
    if date_to:
        qs = qs.filter(order_date__lte=date_to)

    missing_price_skus_seen = []
    missing_price_payout = Decimal("0")
    missing_price_count = 0
    total_order_value = Decimal("0")   # informational: Price × Qty, not used in P&L
    non_loss_payout_excluded_count = 0
    non_loss_payout_excluded_sum = Decimal("0")

    bucket_totals = {
        b: {"count": 0, "gross": Decimal("0"), "cost": Decimal("0"), "net": Decimal("0")}
        for b in _ESTIMATED_PROFIT_BUCKETS
    }
    sku_agg = {}

    def _bump_sku(sku_id, bucket, payout_value, final_cost, net, one_unit_price):
        sku_id = canonical_sku.get(sku_id, sku_id)
        if sku_id not in sku_agg:
            sku_agg[sku_id] = {
                "sku_id": sku_id, "order_count": 0,
                "delivered_count": 0, "return_count": 0, "rto_count": 0,
                "exchange_count": 0, "claim_count": 0, "other_count": 0,
                "gross_payout": Decimal("0"), "total_cost": Decimal("0"), "net_profit": Decimal("0"),
                "one_unit_price": one_unit_price,
            }
        s = sku_agg[sku_id]
        s["order_count"] += 1
        s[f"{bucket}_count"] += 1
        s["gross_payout"] += payout_value
        s["total_cost"]  += final_cost
        s["net_profit"]  += net

    for row in qs.iterator():
        sku_id = row.sku_id
        order_status = (row.order_status or "").lower()
        claim_status = (row.claim_status or "").strip()
        claim_status_lower = claim_status.lower()
        payout_value = row.payout_value or Decimal("0")
        qty_d = Decimal(str(row.quantity or 0))

        is_claim = bool(claim_status)
        claim_approved = is_claim and claim_status_lower == "approved"

        # Returned / RTO / a Claim that isn't explicitly "approved" (rejected, or
        # any other non-blank claim status) all deduct nothing — payout stands
        # as-is, and only counts when it's actually a loss (<= 0).
        payout_only_bucket = None
        if is_claim and not claim_approved:
            payout_only_bucket = "claim"
        elif not is_claim and order_status == "returned":
            payout_only_bucket = "return"
        elif not is_claim and order_status == "rto":
            payout_only_bucket = "rto"

        if payout_only_bucket:
            if payout_value > 0:
                non_loss_payout_excluded_count += 1
                non_loss_payout_excluded_sum += payout_value
                continue
            final_cost = Decimal("0")
            net = payout_value
            total_order_value += (row.price or Decimal("0")) * qty_d
            bt = bucket_totals[payout_only_bucket]
            bt["count"] += 1
            bt["gross"] += payout_value
            bt["cost"]  += final_cost
            bt["net"]   += net
            _bump_sku(sku_id, payout_only_bucket, payout_value, final_cost, net, Decimal("0"))
            continue

        # Remaining rows deduct cost and need pricing: Delivered, Exchanged, an
        # approved Claim.
        if _sku_key(sku_id) not in known_skus:
            missing_price_skus_seen.append(sku_id)
            missing_price_payout += payout_value
            missing_price_count += 1
            continue

        item_price, packaging_cost, tax_pct = get_price(sku_id, row.order_date)
        item_price = Decimal(str(item_price or 0))
        packaging_cost = Decimal(str(packaging_cost or 0))
        tax_pct = Decimal(str(tax_pct or 0))

        purchase_cost = item_price * qty_d
        # purchase_cost already carries × qty, so the rate must not be
        # multiplied by qty again — same fix as compute_order_net().
        tax_cost = purchase_cost * tax_pct / Decimal("100")

        if claim_approved:
            bucket = "claim"
            final_cost = purchase_cost + tax_cost + packaging_cost
        elif order_status == "exchanged":
            bucket = "exchange"
            final_cost = purchase_cost + tax_cost + (packaging_cost * Decimal("2"))
        elif order_status == "delivered":
            bucket = "delivered"
            final_cost = purchase_cost + tax_cost + packaging_cost
        else:
            bucket = "other"  # cancelled / anything else
            final_cost = Decimal("0")

        net = payout_value - final_cost
        total_order_value += (row.price or Decimal("0")) * qty_d

        bt = bucket_totals[bucket]
        bt["count"] += 1
        bt["gross"] += payout_value
        bt["cost"]  += final_cost
        bt["net"]   += net

        _bump_sku(sku_id, bucket, payout_value, final_cost, net, item_price)

    processed_count = sum(v["count"] for v in bucket_totals.values())
    total_gross = sum(v["gross"] for v in bucket_totals.values())
    total_cost  = sum(v["cost"]  for v in bucket_totals.values())
    total_net   = sum(v["net"]   for v in bucket_totals.values())

    sku_list = []
    for s in sku_agg.values():
        row_out = {k: _to_num(v) for k, v in s.items()}
        row_out["avg_profit_per_order"] = round(row_out["net_profit"] / row_out["order_count"], 2)
        sku_list.append(row_out)
    # Sorted by profit, best performers first — losses sink to the bottom.
    sku_list.sort(key=lambda r: r["net_profit"], reverse=True)

    status_breakdown = [
        {
            "status": _ESTIMATED_PROFIT_STATUS_LABELS[b],
            "bucket": b,
            "count":  bucket_totals[b]["count"],
            "gross":  _to_num(bucket_totals[b]["gross"]),
            "cost":   _to_num(bucket_totals[b]["cost"]),
            "net":    _to_num(bucket_totals[b]["net"]),
        }
        for b in _ESTIMATED_PROFIT_BUCKETS
    ]

    missing_price_sku_list = sorted(set(missing_price_skus_seen))

    return {
        "saved_order_count": qs.count(),
        "processed_count":  processed_count,
        "missing_price_count":      missing_price_count,
        "missing_price_skus":       missing_price_sku_list,
        "missing_price_payout_sum": _to_num(missing_price_payout),
        "non_loss_payout_excluded_count": non_loss_payout_excluded_count,
        "non_loss_payout_excluded_sum":   _to_num(non_loss_payout_excluded_sum),
        "totals": {
            "order_count":       processed_count,
            "gross_payout":      _to_num(total_gross),
            "total_cost":        _to_num(total_cost),
            "net_profit":        _to_num(total_net),
            "total_order_value": _to_num(total_order_value),
        },
        "status_breakdown": status_breakdown,
        "sku_wise":         sku_list,
        "top_profit_skus":  [r for r in sku_list[:10] if r["net_profit"] > 0],
        "top_loss_skus":    [r for r in sku_list[::-1][:10] if r["net_profit"] < 0],
    }


@api_view(["GET"])
def estimated_profit_summary(request, business_id):
    """Current live estimate over every previously-uploaded Order Summary row for this business."""
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "") or None
    date_to = request.GET.get("date_to", "") or None
    return Response(_compute_estimated_profit_summary(business, date_from, date_to))


@api_view(["POST"])
@parser_classes([MultiPartParser])
def estimated_profit_upload(request, business_id):
    """
    Upload a Meesho 'Order Summary' CSV (Sub orderId, Catalog ID, Quantity, Price,
    Order Date, Order Status, Payout Value, Payout Status, Claim Status, SKU ID).
    Rows are saved to EstimatedProfitOrder, keyed by (business, sub_order_no,
    order_status) — re-uploading the same sheet (or an updated export covering the
    same orders) updates those rows in place instead of duplicating them. The
    returned P&L is then computed live from every row saved so far — see
    _compute_estimated_profit_summary for the cost formula (item + packaging on
    Delivered/Claim, item + 2×packaging on Exchanged, no cost on Returned/RTO).

    In Transit / Ready To Ship rows are excluded (payment not yet settled) and are
    not saved. Delivered / Returned / RTO / Exchanged rows are included at face
    value; Cancelled (and any other status) lands in the "Other" bucket. A
    non-blank Claim Status overrides Order Status — claim-first.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = pd.read_csv(file, dtype=str)
    except Exception as e:
        return Response({"error": f"Unable to read CSV: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    df.columns = [c.strip() for c in df.columns]
    required_cols = {"Sub orderId", "Order Status", "Payout Value", "SKU ID", "Quantity", "Order Date"}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        return Response(
            {"error": f"Missing required columns: {', '.join(sorted(missing_cols))}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    excluded_count = 0
    invalid_rows = 0
    created = 0
    updated = 0

    with transaction.atomic():
        for _, row in df.iterrows():
            order_status_raw = safe_str(row.get("Order Status")) or ""
            if order_status_raw.lower() in _ESTIMATED_PROFIT_EXCLUDED_STATUSES:
                excluded_count += 1
                continue

            sub_order_no = safe_str(row.get("Sub orderId")) or ""
            sku_id = safe_str(row.get("SKU ID")) or ""
            try:
                qty = int(float(row.get("Quantity")))
            except (TypeError, ValueError):
                qty = None
            payout_value = safe_decimal(row.get("Payout Value"))

            if not sub_order_no or not sku_id or sku_id.lower() == "null" or not qty or payout_value is None:
                invalid_rows += 1
                continue

            try:
                order_date = pd.to_datetime(row.get("Order Date")).date()
            except Exception:
                order_date = None

            _, was_created = EstimatedProfitOrder.objects.update_or_create(
                business=business,
                sub_order_no=sub_order_no,
                order_status=order_status_raw,
                defaults={
                    "catalog_id":    safe_str(row.get("Catalog ID")),
                    "sku_id":        sku_id,
                    "quantity":      qty,
                    "price":         safe_decimal(row.get("Price")),
                    "order_date":    order_date,
                    "payout_value":  payout_value,
                    "payout_status": safe_str(row.get("Payout Status")),
                    "claim_status":  safe_str(row.get("Claim Status")) or "",
                },
            )
            if was_created:
                created += 1
            else:
                updated += 1

    summary = _compute_estimated_profit_summary(business)
    summary.update({
        "total_rows":     len(df),
        "excluded_count": excluded_count,
        "invalid_rows":   invalid_rows,
        "created":        created,
        "updated":        updated,
    })
    return Response(summary)


@api_view(["GET"])
def ads_sku_analysis(request, business_id):
    """
    Per-SKU breakdown of orders placed through Meesho Ads (OrderPayment.order_source
    == "Ad order") vs organic — order counts, delivery outcome, and profit, computed
    with the same P&L engine as /profit/. Also reports total ads spend for the period.

    Meesho's Ads Cost export only has a campaign_id (no SKU/catalog link), so ad
    spend can only be compared in aggregate against total ad-driven profit — not
    attributed to individual SKUs.
    """
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")

    qs = OrderPayment.objects.filter(business=business)
    if date_from or date_to:
        ord_qs = Order.objects.filter(business=business)
        if date_from:
            ord_qs = ord_qs.filter(order_date__gte=date_from)
        if date_to:
            ord_qs = ord_qs.filter(order_date__lte=date_to)
        if ord_qs.exists():
            qs = qs.filter(sub_order_no__in=ord_qs.values("sub_order_no"))
        else:
            if date_from:
                qs = qs.filter(order_date__date__gte=date_from)
            if date_to:
                qs = qs.filter(order_date__date__lte=date_to)

    # sub_order_nos where any row is flagged as coming through ads
    ad_sub_orders = set(qs.filter(order_source="Ad order").values_list("sub_order_no", flat=True))

    # Pricing lookups — mirrors profit_summary
    _fp_all        = list(FinalPrice.objects.filter(business=business).only("sku_id", "final_price", "packaging_cost", "parent_id", "item_price", "tax_percent"))
    price_map      = _SkuMap((fp.sku_id, fp.final_price    or Decimal("0")) for fp in _fp_all)
    packaging_map  = _SkuMap((fp.sku_id, fp.packaging_cost or Decimal("0")) for fp in _fp_all)
    item_price_map = _SkuMap((fp.sku_id, fp.item_price     or Decimal("0")) for fp in _fp_all)
    tax_map        = _SkuMap((fp.sku_id, fp.tax_percent    or 0)            for fp in _fp_all)
    sku_parent_map = _SkuMap((fp.sku_id, fp.parent_id) for fp in _fp_all if fp.parent_id)
    # Any casing of a SKU -> the spelling stored in pricing, so orders that spell
    # the same SKU differently accumulate into one row instead of two.
    canonical_sku  = _SkuMap((fp.sku_id, fp.sku_id) for fp in _fp_all)

    _fp_parent     = list(ParentItemPrice.objects.filter(business=business).only("item_id", "item_price", "tax_percent", "packaging_cost", "final_price"))
    parent_price_map      = {fp.id: fp.final_price    or Decimal("0") for fp in _fp_parent}
    parent_packaging_map  = {fp.id: fp.packaging_cost or Decimal("0") for fp in _fp_parent}
    parent_item_price_map = {fp.id: fp.item_price     or Decimal("0") for fp in _fp_parent}
    parent_tax_map        = {fp.id: fp.tax_percent    or 0            for fp in _fp_parent}
    parent_name_map        = {fp.id: fp.item_id for fp in _fp_parent}

    from collections import defaultdict
    _hist = list(
        ParentPriceHistory.objects
        .filter(business=business)
        .values("parent_id", "effective_from", "final_price", "packaging_cost", "item_price", "tax_percent")
        .order_by("effective_from")
    )
    _parent_histories = defaultdict(list)
    for h in _hist:
        _parent_histories[h["parent_id"]].append((
            h["effective_from"],
            h["final_price"]    or Decimal("0"),
            h["packaging_cost"] or Decimal("0"),
            h["item_price"]     or Decimal("0"),
            h["tax_percent"]    or 0,
        ))

    def get_eff_price(sku_id, order_date):
        """
        (final_price, packaging_cost, item_price, tax_percent) for a SKU.

        Priority is the parent's price, then the SKU's own — in that order and
        regardless of whether an order date is known. Previously the parent was
        only consulted when an order_date was present, so any order without one
        silently fell back to SKU pricing even when a parent existed.

        A parent counts as "priced" only if it actually carries an item price;
        an empty parent record falls through to the SKU so a missing parent
        price can't zero out a cost.
        """
        pid = sku_parent_map.get(sku_id)

        if pid:
            # Dated parent price history wins when we know when the order was placed.
            if order_date:
                od = order_date.date() if hasattr(order_date, "date") else order_date
                applicable = [
                    (d, fp, pkg, ip, tax)
                    for d, fp, pkg, ip, tax in _parent_histories[pid] if d <= od
                ]
                if applicable:
                    _, fp, pkg, ip, tax = applicable[-1]
                    if ip:
                        return fp, pkg, ip, tax

            # Otherwise the parent's current price.
            p_item = parent_item_price_map.get(pid, Decimal("0"))
            if p_item:
                return (
                    parent_price_map.get(pid,      Decimal("0")),
                    parent_packaging_map.get(pid,  Decimal("0")),
                    p_item,
                    parent_tax_map.get(pid, 0),
                )

        # No parent, or the parent has no price of its own — use the SKU.
        return (
            price_map.get(sku_id,      Decimal("0")),
            packaging_map.get(sku_id,  Decimal("0")),
            item_price_map.get(sku_id, Decimal("0")),
            tax_map.get(sku_id, 0),
        )

    unique_statuses = ["Claim", "Cancelled", "Delivered", "Return", "RTO", "Shipped", "Exchange", "Unknown"]

    _FIELDS = (
        "sub_order_no", "supplier_sku", "quantity",
        "final_settlement_amount", "live_order_status",
        "recovery_reason", "claims", "return_shipping_charge",
        "order_date", "payment_date",
    )

    order_groups = defaultdict(list)
    for payment in qs.only(*_FIELDS).order_by("-payment_date"):
        order_groups[payment.sub_order_no].append(payment)

    ad_profit  = {}
    org_profit = {}
    ad_monthly  = {}
    org_monthly = {}

    for sub_no, payments in order_groups.items():
        order = (
            Order.objects
            .filter(business=business, sub_order_no=sub_no)
            .values("sub_order_no", "sku", "quantity")
            .first()
        ) or {}
        primary = next((p for p in payments if p.live_order_status), payments[0])
        sku = order.get("sku") or primary.supplier_sku
        qty = order.get("quantity") or primary.quantity

        if not sku or sku not in price_map:
            continue

        eff_price, eff_pkg, eff_item_price, eff_tax_pct = get_eff_price(sku, primary.order_date)
        result = compute_order_net(payments, eff_price, eff_pkg, qty, unique_statuses, eff_item_price, eff_tax_pct)

        is_ad = sub_no in ad_sub_orders
        bucket = ad_profit if is_ad else org_profit
        accumulate_sku_profit(canonical_sku.get(sku, sku), bucket, result, price_map, packaging_map, unique_statuses)

        # Same order, grouped by month instead of SKU — reuses the identical
        # accumulator so the monthly trend stays consistent with the per-SKU totals.
        od = primary.order_date
        month_key = od.strftime("%Y-%m") if od else "Unknown"
        monthly_bucket = ad_monthly if is_ad else org_monthly
        accumulate_sku_profit(month_key, monthly_bucket, result, price_map, packaging_map, unique_statuses)

    def build_rows(bucket):
        rows = []
        for sku_id, v in bucket.items():
            del_qty    = float(v.get("delivered_quantity", 0) or 0)
            del_profit = float(v.get("delivered_profit", 0) or 0)
            rows.append({
                "sku_id":                sku_id,
                "parent_sku":            parent_name_map.get(sku_parent_map.get(sku_id)),
                "order_count":           v.get("order_count", 0),
                "delivered_count":       v.get("delivered_count", 0),
                "return_count":          v.get("return_count", 0),
                "rto_count":             v.get("rto_count", 0),
                "exchange_count":        v.get("exchange_count", 0),
                "claim_count":           v.get("claim_count", 0),
                "other_count":           v.get("other_count", 0),
                "delivered_quantity":    del_qty,
                "one_unit_price":        float(v.get("one_unit_price", 0) or 0),
                "delivered_profit":      round(del_profit, 2),
                "avg_profit_per_piece":  round(del_profit / del_qty, 2) if del_qty > 0 else None,
                "net_profit":            round(float(v.get("net_profit", 0) or 0), 2),
            })
        return rows

    ad_rows  = build_rows(ad_profit)
    org_by_sku = {r["sku_id"]: r for r in build_rows(org_profit)}

    # Organic comparison, attached per row so the UI can judge "did ads actually help".
    for r in ad_rows:
        o = org_by_sku.get(r["sku_id"])
        r["organic_avg_profit_per_piece"] = o["avg_profit_per_piece"] if o else None
        r["organic_delivered_count"]      = o["delivered_count"] if o else 0

    ad_rows.sort(key=lambda r: r["net_profit"])

    ads_qs = AdsCost.objects.filter(business=business)
    if date_from:
        ads_qs = ads_qs.filter(deduction_date__gte=date_from)
    if date_to:
        ads_qs = ads_qs.filter(deduction_date__lte=date_to)
    # Meesho's export can store this as a positive spend or a negative deduction
    # depending on the sheet — normalize to a positive "amount spent" figure.
    total_ads_spend = abs(float(ads_qs.aggregate(t=Sum("total_ads_cost"))["t"] or 0))

    total_ad_orders    = sum(r["order_count"]     for r in ad_rows)
    total_ad_delivered = sum(r["delivered_count"] for r in ad_rows)
    total_ad_net_profit = round(sum(r["net_profit"] for r in ad_rows), 2)
    net_benefit = round(total_ad_net_profit - total_ads_spend, 2)
    roi_pct = round((total_ad_net_profit / total_ads_spend) * 100, 1) if total_ads_spend > 0 else None

    total_org_orders    = sum(v.get("order_count", 0) for v in org_profit.values())
    total_org_delivered = sum(v.get("delivered_count", 0) for v in org_profit.values())

    # ── Monthly trend: ad orders/profit vs ad spend, month over month ──────────
    spend_by_month = {}
    for row in ads_qs.values("deduction_date", "total_ads_cost"):
        d = row["deduction_date"]
        if not d:
            continue
        mk = d.strftime("%Y-%m")
        spend_by_month[mk] = spend_by_month.get(mk, 0) + abs(float(row["total_ads_cost"] or 0))

    months = sorted(set(ad_monthly.keys()) | set(org_monthly.keys()) | set(spend_by_month.keys()))
    months = [m for m in months if m != "Unknown"]
    monthly_trend = []
    for m in months:
        a = ad_monthly.get(m, {})
        o = org_monthly.get(m, {})
        spend = round(spend_by_month.get(m, 0), 2)
        profit = round(float(a.get("net_profit", 0) or 0), 2)
        monthly_trend.append({
            "month":              m,
            "ad_orders":          a.get("order_count", 0),
            "ad_delivered":       a.get("delivered_count", 0),
            "ad_profit":          profit,
            "ads_spend":          spend,
            "net_benefit":        round(profit - spend, 2),
            "organic_orders":     o.get("order_count", 0),
            "organic_delivered":  o.get("delivered_count", 0),
            "organic_profit":     round(float(o.get("net_profit", 0) or 0), 2),
        })

    # ── Campaign spend breakdown (no SKU link in Meesho's export, cost only) ──
    campaign_agg = (
        ads_qs.values("campaign_id")
        .annotate(total_cost=Sum("total_ads_cost"), entries=Count("id"))
        .order_by("-total_cost")
    )
    campaigns = [{
        "campaign_id": c["campaign_id"] or "—",
        "total_cost":  round(abs(float(c["total_cost"] or 0)), 2),
        "entries":     c["entries"],
    } for c in campaign_agg][:25]

    return Response({
        "summary": {
            "total_ad_orders":       total_ad_orders,
            "total_ad_delivered":    total_ad_delivered,
            "total_ad_return":       sum(r["return_count"]    for r in ad_rows),
            "total_ad_rto":          sum(r["rto_count"]       for r in ad_rows),
            "total_ad_claim":        sum(r["claim_count"]     for r in ad_rows),
            "total_ad_net_profit":   total_ad_net_profit,
            "total_ads_spend":       round(total_ads_spend, 2),
            "net_benefit":           net_benefit,
            "roi_pct":               roi_pct,
            "sku_count":             len(ad_rows),
            "ad_delivery_rate_pct":  round(total_ad_delivered / total_ad_orders * 100, 1) if total_ad_orders else None,
            "organic_delivery_rate_pct": round(total_org_delivered / total_org_orders * 100, 1) if total_org_orders else None,
            "total_organic_orders": total_org_orders,
        },
        "results": ad_rows,
        "monthly_trend": monthly_trend,
        "campaigns": campaigns,
    })


@api_view(["GET"])
def orders_grouped(request, business_id):
    """
    Returns OrderPayment rows grouped by sub_order_no with a derived classified status.
    Filters: ?month=YYYY-MM, ?date_from=, ?date_to=, ?status=DELIVERED|RETURN|RTO|EXCHANGE|SHIPPED|UNKNOWN
    Also returns KPI summary for the filtered set.
    """
    from collections import defaultdict
    import calendar as _cal

    business = get_authorized_business(request, business_id)
    month      = request.GET.get("month", "").strip()
    date_from  = request.GET.get("date_from", "").strip()
    date_to    = request.GET.get("date_to", "").strip()
    status_f   = request.GET.get("status", "").strip().upper()
    page       = int(request.GET.get("page", 1))
    page_size  = int(request.GET.get("page_size", 50))
    search = request.GET.get("search", "").strip()

    qs = OrderPayment.objects.filter(business=business)

    if month:
        year, mo = (int(x) for x in month.split("-"))
        first_day = datetime(year, mo, 1).date()
        last_day  = datetime(year, mo, _cal.monthrange(year, mo)[1]).date()
        qs = qs.filter(
            order_date__gte=datetime.combine(first_day, time.min),
            order_date__lte=datetime.combine(last_day,  time.max),
        )
    else:
        if date_from:
            d = datetime.strptime(date_from, "%Y-%m-%d").date()
            qs = qs.filter(order_date__gte=datetime.combine(d, time.min))
        if date_to:
            d = datetime.strptime(date_to, "%Y-%m-%d").date()
            qs = qs.filter(order_date__lte=datetime.combine(d, time.max))
            
    if search:
        qs = qs.filter(
            DQ(sub_order_no__icontains=search) |
            DQ(supplier_sku__icontains=search) | 
            DQ(product_name__icontains=search)
        )    

    all_rows = list(qs.order_by("sub_order_no", "payment_date", "id"))

    # ── Group by sub_order_no ─────────────────────────────────────────────────
    groups: dict = defaultdict(list)
    for r in all_rows:
        groups[r.sub_order_no].append(r)

    def _effective_status(row_list):
        statuses = {(r.live_order_status or "").upper() for r in row_list}
        has_claim = any((r.claims or 0) > 0 for r in row_list)
        if has_claim:                                    return "CLAIM"
        if statuses & _EXCHANGE_STATUSES:                return "EXCHANGE"
        if statuses & _RETURN_STATUSES:                  return "RETURN"
        if statuses & _RTO_STATUSES:                     return "RTO"
        if statuses & _DELIVERED_STATUSES:               return "DELIVERED"
        if statuses & _SHIPPED_STATUSES:                 return "SHIPPED"
        return "UNKNOWN"

    def _group_settlement(row_list):
        statuses = {(r.live_order_status or "").upper() for r in row_list}
        progressed = bool(
            statuses & (_RETURN_STATUSES | _RTO_STATUSES | _DELIVERED_STATUSES | _EXCHANGE_STATUSES)
        )
        return float(_sum_settlement(row_list, include_shipped=progressed))

    # ── Classify every group ──────────────────────────────────────────────────
    classified = []
    for sub_no, rows in groups.items():
        eff = _effective_status(rows)
        first = rows[0]
        classified.append({
            "sub_order_no":   sub_no,
            "status":         eff,
            "settlement":     _group_settlement(rows),
            "sku":            first.supplier_sku,
            "catalog_id":     first.catalog_id,
            "product_name":   first.product_name,
            "order_date":     str(first.order_date.date()) if first.order_date else None,
            "row_count":      len(rows),
            "rows":           [],     # filled below for the requested page only
        })

    # ── KPI summary (all groups, before pagination) ───────────────────────────
    from collections import Counter
    status_counts = Counter(g["status"] for g in classified)
    status_settlement = {}
    for g in classified:
        status_settlement[g["status"]] = round(
            status_settlement.get(g["status"], 0) + g["settlement"], 2
        )
    kpi = {
        "total_groups":     len(classified),
        "total_settlement": round(sum(g["settlement"] for g in classified), 2),
        "by_status":        [
            {"status": s, "count": status_counts[s], "settlement": status_settlement.get(s, 0)}
            for s in ["DELIVERED", "RETURN", "RTO", "EXCHANGE", "CLAIM", "SHIPPED", "UNKNOWN"]
            if status_counts.get(s, 0) > 0
        ],
    }

    # ── Apply status filter ───────────────────────────────────────────────────
    if status_f:
        classified = [g for g in classified if g["status"] == status_f]

    # Sort newest first
    classified.sort(key=lambda g: g["order_date"] or "", reverse=True)

    total = len(classified)
    page_groups = classified[(page - 1) * page_size: page * page_size]

    # Attach full row detail only for page groups (avoid serializing thousands of rows)
    sub_nos_page = {g["sub_order_no"] for g in page_groups}
    detail_map: dict = defaultdict(list)
    for r in all_rows:
        if r.sub_order_no in sub_nos_page:
            detail_map[r.sub_order_no].append({
                "id":                          r.id,
                "sub_order_no":                r.sub_order_no,
                "live_order_status":           r.live_order_status,
                "payment_date":                str(r.payment_date) if r.payment_date else None,
                "order_date":                  str(r.order_date.date()) if r.order_date else None,
                "final_settlement_amount":     float(r.final_settlement_amount or 0),
                "total_sale_amount":           float(r.total_sale_amount or 0),
                "meesho_commission_incl_gst":  float(r.meesho_commission_incl_gst or 0),
                "tcs":                         float(r.tcs or 0),
                "tds":                         float(r.tds or 0),
                "claims":                      float(r.claims or 0),
                "return_shipping_charge":      float(r.return_shipping_charge or 0),
                "listing_price_incl_taxes":    float(r.listing_price_incl_taxes or 0),
            })

    for g in page_groups:
        g["rows"] = detail_map.get(g["sub_order_no"], [])

    return Response({
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "kpi":       kpi,
        "groups":    page_groups,
    })


@api_view(["GET"])
def order_payments_list(request, business_id):
    business = get_authorized_business(request, business_id)
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 50))
    status_filter = request.GET.get("status", "")
    sku_filter = request.GET.get("sku", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")

    qs = OrderPayment.objects.filter(business=business)
    if status_filter:
        qs = qs.filter(live_order_status__iexact=status_filter)
    if sku_filter:
        qs = qs.filter(DQ(supplier_sku__icontains=sku_filter) | DQ(sub_order_no__icontains=sku_filter))
    if date_from:
        date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        qs = qs.filter(order_date__gte=datetime.combine(date_from, time.min))
    if date_to:
        date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        qs = qs.filter(order_date__lte=datetime.combine(date_to, time.max))

    total = qs.count()
    start = (page - 1) * page_size
    items = qs[start: start + page_size]
    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": OrderPaymentSerializer(items, many=True).data,
    })


@api_view(["GET"])
def ads_cost_list(request, business_id):
    business = get_authorized_business(request, business_id)
    qs = AdsCost.objects.filter(business=business)
    return Response(AdsCostSerializer(qs, many=True).data)


@api_view(["GET"])
def referral_list(request, business_id):
    business = get_authorized_business(request, business_id)
    qs = ReferralPayment.objects.filter(business=business)
    return Response(ReferralPaymentSerializer(qs, many=True).data)


@api_view(["GET"])
def compensation_recovery_list(request, business_id):
    business = get_authorized_business(request, business_id)
    qs = CompensationRecovery.objects.filter(business=business)
    return Response(CompensationRecoverySerializer(qs, many=True).data)


@api_view(["GET"])
def order_status_breakdown(request, business_id):
    business = get_authorized_business(request, business_id)
    breakdown = (
        OrderPayment.objects
        .filter(business=business)
        .values("live_order_status")
        .annotate(count=Count("sub_order_no"), total_revenue=Sum("final_settlement_amount"))
        .order_by("-count")
    )
    return Response(list(breakdown))


def _compute_purchase_cost():
    """Sum final_price × quantity for orders matched by supplier_sku → sku_id."""
    price_map = _SkuMap(
        (fp.sku_id, fp.final_price or Decimal("0"))
        for fp in FinalPrice.objects.all()
    )
    total = Decimal("0")
    matched = 0
    missing = 0
    missing_list = []
    for order in OrderPayment.objects.only("supplier_sku", "quantity"):
        sku = order.supplier_sku
        qty = order.quantity or 1
        if sku and sku in price_map:
            if price_map[sku] > 0:
                total += price_map[sku] * qty
            matched += 1
        elif sku:
            missing += 1
            missing_list.append(sku)
    return total, matched, missing, list(set(missing_list))


@api_view(["GET", "POST"])
def final_price_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        search = request.GET.get("search", "")
        qs = FinalPrice.objects.filter(business=business)
        if search:
            qs = qs.filter(sku_id__icontains=search)
        items = qs
        return Response({
            "results": FinalPriceSerializer(items, many=True).data,
        })

    # Save = create-or-update. Posting a SKU that already has a price row used to
    # raise a duplicate-key IntegrityError and 500, which the UI could only show
    # as "Save failed". That bit hardest with a casing difference: MySQL compares
    # these case-insensitively (utf8mb4_unicode_ci), so saving
    # "brass_wooden_panchaarti_155555" collided with the stored
    # "BRASS_WOODEN_panchaarti_155555" even though no visible duplicate existed.
    sku_id = (request.data.get("sku_id") or "").strip()
    if not sku_id:
        return Response({"sku_id": ["This field is required."]},
                        status=status.HTTP_400_BAD_REQUEST)

    # Case-insensitive so the lookup finds the row the database would collide with.
    existing = FinalPrice.objects.filter(business=business, sku_id__iexact=sku_id).first()

    payload = dict(request.data)
    payload["sku_id"] = existing.sku_id if existing else sku_id   # keep the stored spelling

    serializer = FinalPriceSerializer(existing, data=payload, partial=bool(existing))
    serializer.is_valid(raise_exception=True)

    # Only touch the parent link when the caller actually sent one — the
    # SKU-analysis dialog doesn't, and it must not silently unlink the parent.
    extra = {"business": business}
    if "parent" in request.data:
        extra["parent"] = _resolve_parent(request.data.get("parent"), business)

    serializer.save(**extra)
    return Response(
        serializer.data,
        status=status.HTTP_200_OK if existing else status.HTTP_201_CREATED,
    )


def _resolve_parent(parent_item_id, business):
    """Resolve an incoming `parent` value (a parent item_id string, or falsy for
    unlink) to the business-scoped ParentItemPrice object (or None)."""
    if not parent_item_id:
        return None
    return ParentItemPrice.objects.filter(business=business, item_id=parent_item_id).first()


@api_view(["GET", "POST"])
def parent_price_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        search = request.GET.get("search", "")
        qs = ParentItemPrice.objects.filter(business=business)
        if search:
            qs = qs.filter(item_id__icontains=search)
        items = qs
        return Response({
            "results": ParentItemPriceSerializer(items, many=True).data,
        })

    serializer = ParentItemPriceSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save(business=business)
    return Response(serializer.data, status=status.HTTP_201_CREATED)



@api_view(["GET", "PUT", "PATCH", "DELETE"])
def final_price_detail(request, business_id, sku_id):
    business = get_authorized_business(request, business_id)
    try:
        obj = FinalPrice.objects.get(sku_id=sku_id, business=business)
    except FinalPrice.DoesNotExist:
        return Response({"error": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(FinalPriceSerializer(obj).data)

    if request.method == "DELETE":
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    partial = request.method == "PATCH"
    serializer = FinalPriceSerializer(obj, data=request.data, partial=partial)
    serializer.is_valid(raise_exception=True)
    # Only touch parent when the caller sent it (supports partial PATCH unlink
    # with {"parent": null} as well as re-parenting by item_id).
    if "parent" in request.data:
        serializer.save(parent=_resolve_parent(request.data.get("parent"), business))
    else:
        serializer.save()
    return Response(serializer.data)

@api_view(["GET", "PUT", "PATCH", "DELETE"])
def parent_price_detail(request, business_id, item_id):
    business = get_authorized_business(request, business_id)
    try:
        obj = ParentItemPrice.objects.get(item_id=item_id, business=business)
    except ParentItemPrice.DoesNotExist:
        return Response({"error": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        return Response(ParentItemPriceSerializer(obj).data)

    if request.method == "DELETE":
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    partial = request.method == "PATCH"
    serializer = ParentItemPriceSerializer(obj, data=request.data, partial=partial)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)

@api_view(["POST", "PUT"])
def parent_linking_to_sku(request, business_id):
    business = get_authorized_business(request, business_id)
    try:

        if request.method == "POST":
            serializer = ParentItemPriceSerializer(data=request.data)
        else:
            obj = ParentItemPrice.objects.get(item_id=request.data["item_id"], business=business)
            serializer = ParentItemPriceSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        parent = serializer.save(business=business)
    except Exception as e:
        return Response(
        {
            "message": f"Not Valid Parent {e}",
            "parent_id": request.data.get("item_id")
        },
        status=status.HTTP_400_BAD_REQUEST,
    )

    if request.method == "PUT":
        sku_ids = request.data.get("sku_ids", "")

        sku_ids = [
            sku.strip()
            for sku in sku_ids.split(",")
            if sku.strip()
        ]

        updated_count = FinalPrice.objects.filter(
            business=business, sku_id__in=sku_ids
        ).update(
            parent=parent,
            item_price=parent.item_price,
            tax_percent=parent.tax_percent,
            packaging_cost=parent.packaging_cost,
            final_price=parent.final_price,
        )

        return Response(
            {
                "message": f"{updated_count} SKU(s) linked successfully",
                "parent_id": parent.item_id,
                "sku_ids": sku_ids,
            },
            status=status.HTTP_200_OK,
        )
    else:
        return Response(
            {
                "message": "Parent Created successfully",
                "parent_id": parent.item_id
            },
            status=status.HTTP_200_OK,
        )
def _sync_parent_current_price(item_id, business):
    """After adding/deleting a history entry, keep ParentItemPrice + linked FinalPrices in sync.

    `item_id` is the parent's item_id string (scoped to the business).
    """
    history = (
        ParentPriceHistory.objects
        .filter(parent__item_id=item_id, business=business)
        .order_by("-effective_from").first()
    )
    if not history:
        return
    ParentItemPrice.objects.filter(item_id=item_id, business=business).update(
        item_price=history.item_price,
        tax_percent=history.tax_percent,
        packaging_cost=history.packaging_cost,
        final_price=history.final_price,
    )
    FinalPrice.objects.filter(parent__item_id=item_id, business=business).update(
        item_price=history.item_price,
        tax_percent=history.tax_percent,
        packaging_cost=history.packaging_cost,
        final_price=history.final_price,
    )


@api_view(["GET", "POST"])
def parent_price_history_list(request, business_id, item_id):
    business = get_authorized_business(request, business_id)
    try:
        parent = ParentItemPrice.objects.get(item_id=item_id, business=business)
    except ParentItemPrice.DoesNotExist:
        return Response({"error": "Parent not found."}, status=404)

    if request.method == "GET":
        return Response(ParentPriceHistorySerializer(parent.price_history.all(), many=True).data)

    data = request.data.copy()
    data["parent"] = parent.pk
    if "item_price" in data and "final_price" not in data:
        ip  = Decimal(str(data.get("item_price") or 0))
        tax = Decimal(str(data.get("tax_percent") or 0)) / 100
        pkg = Decimal(str(data.get("packaging_cost") or 0))
        data["final_price"] = str((ip + ip * tax + pkg).quantize(Decimal("0.01")))

    serializer = ParentPriceHistorySerializer(data=data)
    serializer.is_valid(raise_exception=True)
    serializer.save(business=business)
    _sync_parent_current_price(item_id, business)
    return Response(serializer.data, status=201)


@api_view(["DELETE"])
def parent_price_history_detail(request, business_id, item_id, pk):
    business = get_authorized_business(request, business_id)
    try:
        obj = ParentPriceHistory.objects.get(pk=pk, parent__item_id=item_id, business=business)
    except ParentPriceHistory.DoesNotExist:
        return Response({"error": "Not found."}, status=404)
    obj.delete()
    _sync_parent_current_price(item_id, business)
    return Response({"deleted": True})


@api_view(["GET"])
def unlinked_skus(request, business_id):
    """SKUs available to link to a parent.

    Two sources, merged:
      1. FinalPrice rows that have no parent yet (priced but unlinked).
      2. SKUs seen in the orders table that have no linked FinalPrice entry —
         i.e. brand-new SKUs that were never priced/linked ("new" SKUs).

    Each result carries `order_count` (how many orders reference the SKU) and
    `has_price` (whether a FinalPrice row already exists). Response shape stays
    backward compatible ({"results": [{sku_id, item_price, final_price, ...}]}).
    """
    business = get_authorized_business(request, business_id)
    q = request.GET.get("q", "").strip().lower()

    # Compared on the canonical key, not the raw string: a SKU stored as
    # "BRASS_WOODEN_x" is the same SKU as an order's "brass_wooden_x", and
    # matching case-sensitively left a freshly-linked SKU showing here forever.
    linked = {
        _sku_key(sku) for sku in
        FinalPrice.objects
        .filter(business=business, parent__isnull=False)
        .values_list("sku_id", flat=True)
    }

    results = {}

    # 1) Priced FinalPrice rows not yet attached to a parent.
    for row in (
        FinalPrice.objects
        .filter(business=business, parent__isnull=True)
        .values("id", "sku_id", "item_price", "final_price")
    ):
        results[_sku_key(row["sku_id"])] = {
            "id": row["id"],                 # stable row id — safe to key a UI row on
            "sku_id": row["sku_id"],         # the stored spelling
            "item_price": row["item_price"],
            "final_price": row["final_price"],
            "has_price": True,
            "order_count": 0,
        }

    # 2) SKUs from the orders table with no linked entry — the "new" SKUs.
    order_counts = (
        Order.objects
        .filter(business=business)
        .exclude(sku__isnull=True)
        .exclude(sku="")
        .values("sku")
        .annotate(n=Count("sku"))
    )
    for row in order_counts:
        sku = row["sku"]
        key = _sku_key(sku)
        if key in linked:
            continue
        if key in results:
            # Same SKU, possibly a different casing in the orders table — add the
            # order count to the existing entry instead of listing it twice.
            results[key]["order_count"] += row["n"]
        else:
            results[key] = {
                "id": None,                  # not priced yet, so no row id exists
                "sku_id": sku,
                "item_price": None,
                "final_price": None,
                "has_price": False,
                "order_count": row["n"],
            }

    out = list(results.values())
    if q:
        out = [r for r in out if q in r["sku_id"].lower()]
    # Most-ordered first, then alphabetical — surfaces the SKUs that matter most.
    out.sort(key=lambda r: (-r["order_count"], r["sku_id"]))
    return Response({"results": out})


@api_view(["POST"])
def link_sku_to_parent(request, business_id):
    """Link a single SKU to a parent, creating the FinalPrice row if it does not
    exist yet (e.g. a SKU that only appears in the orders table). The child
    inherits the parent's current pricing. Used by the drag-and-drop UI.

    Body: {"sku_id": "...", "parent_id": "..."}
    """
    business = get_authorized_business(request, business_id)
    sku_id = (request.data.get("sku_id") or "").strip()
    parent_id = (request.data.get("parent_id") or "").strip()

    if not sku_id:
        return Response({"error": "sku_id is required."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        parent = ParentItemPrice.objects.get(item_id=parent_id, business=business)
    except ParentItemPrice.DoesNotExist:
        return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

    result = _bulk_link_skus_to_parent(business=business, parent=parent, sku_ids=[sku_id])
    if result["failed"]:
        return Response({
            "error": "SKU belongs to another business and cannot be linked here.",
            "sku_id": sku_id,
            "parent_id": parent.item_id,
        }, status=status.HTTP_409_CONFLICT)

    return Response({
        "message": f"{sku_id} linked to {parent.item_id}",
        "sku_id": sku_id,
        "parent_id": parent.item_id,
        "created": result["created"] > 0,
    }, status=status.HTTP_200_OK)


@api_view(["POST"])
def create_parent_from_sku(request, business_id):
    """Create a brand-new parent group *from* an existing child/unlinked SKU and
    link that SKU to it in one atomic step.

    The new parent inherits its pricing from the SKU's existing FinalPrice row
    (if any) unless overridden in the request body. Used by the "New parent from
    this SKU" action in the pricing UI.

    Body: {"sku_id": "...", "parent_id": "..." (optional, defaults to sku_id),
           "item_price", "tax_percent", "packaging_cost", "final_price" (all optional)}
    """
    business = get_authorized_business(request, business_id)
    sku_id = (request.data.get("sku_id") or "").strip()
    parent_id = (request.data.get("parent_id") or sku_id).strip()

    if not sku_id:
        return Response({"error": "sku_id is required."}, status=status.HTTP_400_BAD_REQUEST)
    if not parent_id:
        return Response({"error": "parent_id is required."}, status=status.HTTP_400_BAD_REQUEST)
    if ParentItemPrice.objects.filter(business=business, item_id=parent_id).exists():
        return Response(
            {"error": f"A parent named '{parent_id}' already exists."},
            status=status.HTTP_409_CONFLICT,
        )

    existing = FinalPrice.objects.filter(business=business, sku_id=sku_id).first()

    def pick(key, default):
        v = request.data.get(key)
        return v if v not in (None, "") else default

    item_price = pick("item_price", existing.item_price if existing else 0)
    tax_percent = pick("tax_percent", (existing.tax_percent if existing else 0) or 0)
    packaging_cost = pick("packaging_cost", (existing.packaging_cost if existing else 0) or 0)
    final_price = pick("final_price", None)
    if final_price in (None, ""):
        ip = Decimal(str(item_price or 0))
        tax = Decimal(str(tax_percent or 0)) / 100
        pkg = Decimal(str(packaging_cost or 0))
        final_price = str((ip + ip * tax + pkg).quantize(Decimal("0.01")))

    with transaction.atomic():
        parent = ParentItemPrice.objects.create(
            business=business,
            item_id=parent_id,
            item_price=item_price or 0,
            tax_percent=tax_percent or 0,
            packaging_cost=packaging_cost or 0,
            final_price=final_price,
        )
        _bulk_link_skus_to_parent(business=business, parent=parent, sku_ids=[sku_id])

    return Response(ParentItemPriceSerializer(parent).data, status=status.HTTP_201_CREATED)


def _sku_key(sku_id):
    """
    Canonical form used whenever two SKU ids are compared in Python.

    The database compares sku_id case-insensitively (the schema uses
    utf8mb4_unicode_ci), so "Foo" and "foo" are the *same* SKU as far as the
    unique key on (business, sku_id) is concerned. Python's `==`/`in` are
    case-sensitive, and that mismatch caused two real bugs: a linked SKU kept
    appearing in the unlinked list, and linking one whose casing differed from
    the stored row tried to insert a duplicate. Comparing on this key keeps the
    application's idea of identity the same as the database's.
    """
    return (sku_id or "").strip().casefold()


class _SkuMap(dict):
    """
    A dict keyed by SKU id that matches keys the way the database does.

    Orders and pricing rows don't always agree on the casing of a SKU — Meesho's
    exports have changed spelling over time, and the DB's case-insensitive
    collation happily treats them as one SKU. Plain dicts don't, so a lookup of
    the order's spelling against a map keyed on the pricing row's spelling
    missed, and the order was costed at zero. Keying through _sku_key() makes the
    lookup agree with the database.

    Note keys()/items() yield the canonical (case-folded) form, so use the value
    from the database when a SKU needs to be displayed.
    """

    def __init__(self, source=()):
        super().__init__()
        pairs = source.items() if isinstance(source, dict) else source
        for k, v in pairs:
            self[k] = v

    def __setitem__(self, key, value):
        super().__setitem__(_sku_key(key), value)

    def __getitem__(self, key):
        return super().__getitem__(_sku_key(key))

    def __contains__(self, key):
        return super().__contains__(_sku_key(key))

    def get(self, key, default=None):
        return super().get(_sku_key(key), default)

    def setdefault(self, key, default=None):
        return super().setdefault(_sku_key(key), default)

    def pop(self, key, *args):
        return super().pop(_sku_key(key), *args)


def _normalize_sku_ids(raw_sku_ids):
    """Trim, drop blanks, and de-duplicate case-insensitively (first spelling wins)."""
    seen = set()
    out = []
    for sku in raw_sku_ids or []:
        s = (sku or "").strip()
        k = _sku_key(s)
        if not s or k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _bulk_link_skus_to_parent(*, business, parent, sku_ids):
    """Fast in-memory planning + bulk DB writes for linking SKUs to a parent."""
    sku_ids = _normalize_sku_ids(sku_ids)
    if not sku_ids:
        return {"requested": 0, "linked": 0, "created": 0, "updated": 0, "failed": 0, "failed_skus": []}

    # sku_id is unique per-business now, so a SKU existing in another business
    # is not a conflict — we only look at this business's rows.
    conflicts = []
    # Keyed on the canonical form: the DB lookup below is case-insensitive, so
    # keying on the stored spelling would miss a row that differs only in case
    # and then try to insert a duplicate.
    existing_business_rows = {
        _sku_key(row.sku_id): row
        for row in FinalPrice.objects.filter(business=business, sku_id__in=sku_ids)
    }

    to_create_ids = [sku for sku in sku_ids if _sku_key(sku) not in existing_business_rows]
    to_update = [
        existing_business_rows[_sku_key(sku)]
        for sku in sku_ids if _sku_key(sku) in existing_business_rows
    ]

    for row in to_update:
        row.parent = parent
        row.item_price = parent.item_price
        row.tax_percent = parent.tax_percent
        row.packaging_cost = parent.packaging_cost
        row.final_price = parent.final_price

    created = 0
    updated = 0
    with transaction.atomic():
        if to_update:
            FinalPrice.objects.bulk_update(
                to_update,
                ["parent", "item_price", "tax_percent", "packaging_cost", "final_price"],
                batch_size=500,
            )
            updated = len(to_update)

        if to_create_ids:
            to_create = [
                FinalPrice(
                    sku_id=sku,
                    business=business,
                    parent=parent,
                    item_price=parent.item_price,
                    tax_percent=parent.tax_percent,
                    packaging_cost=parent.packaging_cost,
                    final_price=parent.final_price,
                )
                for sku in to_create_ids
            ]
            try:
                FinalPrice.objects.bulk_create(to_create, batch_size=500)
                created = len(to_create)
            except IntegrityError:
                # Graceful fallback for race/conflict edges.
                for obj in to_create:
                    try:
                        FinalPrice.objects.create(
                            sku_id=obj.sku_id,
                            business=business,
                            parent=parent,
                            item_price=parent.item_price,
                            tax_percent=parent.tax_percent,
                            packaging_cost=parent.packaging_cost,
                            final_price=parent.final_price,
                        )
                        created += 1
                    except IntegrityError:
                        conflicts.append(obj.sku_id)

    linked = created + updated
    return {
        "requested": len(sku_ids),
        "linked": linked,
        "created": created,
        "updated": updated,
        "failed": len(conflicts),
        "failed_skus": sorted(set(conflicts)),
    }


@api_view(["POST"])
def bulk_link_skus_to_parent(request, business_id):
    """Bulk-link many SKUs to a parent in one fast request.

    Body:
      {
        "parent_id": "PARENT_1",
        "sku_ids": ["SKU1", "SKU2", ...]
      }
    """
    business = get_authorized_business(request, business_id)
    parent_id = (request.data.get("parent_id") or "").strip()
    sku_ids = request.data.get("sku_ids") or []

    if not parent_id:
        return Response({"error": "parent_id is required."}, status=status.HTTP_400_BAD_REQUEST)
    if not isinstance(sku_ids, list):
        return Response({"error": "sku_ids must be a list."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        parent = ParentItemPrice.objects.get(item_id=parent_id, business=business)
    except ParentItemPrice.DoesNotExist:
        return Response({"error": "Parent not found."}, status=status.HTTP_404_NOT_FOUND)

    result = _bulk_link_skus_to_parent(business=business, parent=parent, sku_ids=sku_ids)
    return Response({
        "message": f"Linked {result['linked']} SKU(s) to {parent_id}",
        "parent_id": parent_id,
        **result,
    }, status=status.HTTP_200_OK)


@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_final_price(request, business_id):
    """
    Upload an Excel or CSV sheet to upsert FinalPrice rows.
    Expected columns: sku_id, item_price, tax_percent, packaging_cost, final_price
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        if file.name.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(file)
        else:
            df = pd.read_csv(file)
    except Exception as e:
        return Response({"error": f"Could not read file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
    if "sku_id" not in df.columns:
        return Response(
            {"error": "Missing required column: sku_id"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created = updated = skipped = 0
    with transaction.atomic():
        for _, row in df.iterrows():
            pk = safe_str(row.get("sku_id"))
            if not pk:
                skipped += 1
                continue
            defaults = {}
            for col in ("item_price", "packaging_cost", "final_price", "tax_percent"):
                if col in df.columns:
                    if col == "tax_percent":
                        defaults[col] = safe_int(row.get(col))
                    else:
                        defaults[col] = safe_decimal(row.get(col))
            _, was_created = FinalPrice.objects.update_or_create(
                business=business, sku_id=pk, defaults=defaults
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return Response(
        {"success": True, "created": created, "updated": updated, "skipped": skipped},
        status=status.HTTP_201_CREATED,
    )


@api_view(["GET"])
def download_final_price(request, business_id):
    """
    Download all SKU pricing rows as an Excel sheet.

    The column headers normalize (lower-cased, spaces -> underscores) to
    exactly the columns `upload_final_price` expects, so a downloaded file
    can be edited and re-uploaded as-is. An empty DB yields a header-only
    template.
    """
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    business = get_authorized_business(request, business_id)
    qs = FinalPrice.objects.filter(business=business).select_related("parent").order_by("sku_id")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKU Pricing"

    headers = ["SKU ID", "Item Price", "Tax Percent", "Packaging Cost", "Final Price", "Parent ID"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for fp in qs:
        ws.append([
            fp.sku_id,
            fp.item_price,
            fp.tax_percent,
            fp.packaging_cost,
            fp.final_price,
            # export the parent's item_id (not the surrogate id) so the file
            # round-trips on import.
            fp.parent.item_id if fp.parent_id else None,
        ])

    col_widths = [30, 14, 14, 16, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="sku_pricing.xlsx"'
    return response


@api_view(["GET"])
def download_pricing_workbook(request, business_id):
    """
    Download parents AND child SKUs together in a single Excel workbook.

    Two sheets:
      - "Parents": Item ID, Item Price, Tax Percent, Packaging Cost, Final Price
      - "SKUs":    SKU ID, Item Price, Tax Percent, Packaging Cost, Final Price, Parent ID

    Column headers normalize to the columns `upload_pricing_workbook` expects,
    so a downloaded file can be edited and re-uploaded as-is. Empty tables
    yield header-only sheets (a blank template).
    """
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    business = get_authorized_business(request, business_id)
    parents = ParentItemPrice.objects.filter(business=business).order_by("item_id")
    skus = FinalPrice.objects.filter(business=business).select_related("parent").order_by("sku_id")

    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")

    def style_header(ws):
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    wb = openpyxl.Workbook()

    ws_p = wb.active
    ws_p.title = "Parents"
    ws_p.append(["Item ID", "Item Price", "Tax Percent", "Packaging Cost", "Final Price"])
    for p in parents:
        ws_p.append([p.item_id, p.item_price, p.tax_percent, p.packaging_cost, p.final_price])
    for i, w in enumerate([30, 14, 14, 16, 14], 1):
        ws_p.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    style_header(ws_p)

    ws_s = wb.create_sheet("SKUs")
    ws_s.append(["SKU ID", "Item Price", "Tax Percent", "Packaging Cost", "Final Price", "Parent ID"])
    for s in skus:
        # export the parent's item_id (not the surrogate id) so the file
        # round-trips on import.
        parent_item = s.parent.item_id if s.parent_id else None
        ws_s.append([s.sku_id, s.item_price, s.tax_percent, s.packaging_cost, s.final_price, parent_item])
    for i, w in enumerate([30, 14, 14, 16, 14, 20], 1):
        ws_s.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    style_header(ws_s)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="pricing_parents_and_skus.xlsx"'
    return response


@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_pricing_workbook(request, business_id):
    """
    Upload the combined workbook to upsert parents AND child SKUs in one go.

    Sheets are identified by their columns, not their names: a sheet with an
    `item_id` column is treated as parents; a sheet with a `sku_id` column as
    SKUs. Parents are upserted first so SKUs can link to them via `parent_id`.
    A single-sheet .csv is also accepted and routed the same way.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    def normalize(df):
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        return df

    try:
        if file.name.lower().endswith(".csv"):
            sheets = {"sheet1": pd.read_csv(file)}
        else:
            sheets = pd.read_excel(file, sheet_name=None)  # dict of all sheets
    except Exception as e:
        return Response({"error": f"Could not read file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

    sheets = {name: normalize(df) for name, df in sheets.items()}
    parent_sheets = [df for df in sheets.values() if "item_id" in df.columns]
    sku_sheets = [df for df in sheets.values() if "sku_id" in df.columns]

    if not parent_sheets and not sku_sheets:
        return Response(
            {"error": "No recognizable sheet found. Expected a column named 'item_id' (parents) or 'sku_id' (SKUs)."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    p_created = p_updated = s_created = s_updated = skipped = unlinked = 0

    with transaction.atomic():
        # Parents first, so SKUs can resolve parent_id against them.
        for df in parent_sheets:
            for _, row in df.iterrows():
                pk = safe_str(row.get("item_id"))
                if not pk:
                    skipped += 1
                    continue
                defaults = {}
                for col in ("item_price", "packaging_cost", "final_price", "tax_percent"):
                    if col in df.columns:
                        defaults[col] = safe_int(row.get(col)) if col == "tax_percent" else safe_decimal(row.get(col))
                _, created = ParentItemPrice.objects.update_or_create(
                    business=business, item_id=pk, defaults=defaults
                )
                p_created += 1 if created else 0
                p_updated += 0 if created else 1

        # Then SKUs, linking to parents by parent_id where possible.
        for df in sku_sheets:
            for _, row in df.iterrows():
                pk = safe_str(row.get("sku_id"))
                if not pk:
                    skipped += 1
                    continue
                defaults = {}
                for col in ("item_price", "packaging_cost", "final_price", "tax_percent"):
                    if col in df.columns:
                        defaults[col] = safe_int(row.get(col)) if col == "tax_percent" else safe_decimal(row.get(col))
                if "parent_id" in df.columns:
                    parent_id = safe_str(row.get("parent_id"))
                    if parent_id:
                        parent = ParentItemPrice.objects.filter(business=business, item_id=parent_id).first()
                        if parent:
                            defaults["parent"] = parent
                        else:
                            unlinked += 1  # parent_id given but no matching parent
                    else:
                        defaults["parent"] = None
                _, created = FinalPrice.objects.update_or_create(
                    business=business, sku_id=pk, defaults=defaults
                )
                s_created += 1 if created else 0
                s_updated += 0 if created else 1

    return Response(
        {
            "success": True,
            "parents_created": p_created,
            "parents_updated": p_updated,
            "skus_created": s_created,
            "skus_updated": s_updated,
            "skipped": skipped,
            "unlinked_parent_refs": unlinked,
        },
        status=status.HTTP_201_CREATED,
    )


@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_orders_csv(request, business_id):
    """
    Upload Orders CSV file and create/update Order records.
    """
    business = get_authorized_business(request, business_id)

    file = request.FILES.get("file")

    if not file:
        return Response(
            {"error": "No file provided"},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        df = pd.read_csv(file)
    except Exception as e:
        return Response(
            {"error": f"Unable to read CSV: {str(e)}"},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Clean column names
    df.columns = [col.strip() for col in df.columns]

    created = 0
    updated = 0

    with transaction.atomic():
        for _, row in df.iterrows():

            sub_order_no = str(row.get("Sub Order No", "")).strip()

            if not sub_order_no:
                continue

            defaults = {
                "reason_for_credit_entry": str(
                    row.get("Reason for Credit Entry", "")
                ).strip(),
                "catalog_id": int(row.get("Catalog ID"))
                if pd.notna(row.get("Catalog ID"))
                else None,
                "order_date": pd.to_datetime(
                    row.get("Order Date")
                ).date()
                if pd.notna(row.get("Order Date"))
                else None,
                "order_source": str(
                    row.get("Order source", "")
                ).strip(),
                "customer_state": str(
                    row.get("Customer State", "")
                ).strip(),
                "product_name": str(
                    row.get("Product Name", "")
                ).strip(),
                "sku": str(
                    row.get("SKU", "")
                ).strip(),
                "size": str(
                    row.get("Size", "")
                ).strip(),
                "quantity": int(row.get("Quantity", 0))
                if pd.notna(row.get("Quantity"))
                else 0,
                "supplier_listed_price": row.get(
                    "Supplier Listed Price (Incl. GST + Commission)",
                    0
                ),
                "supplier_discounted_price": row.get(
                    "Supplier Discounted Price (Incl GST and Commision)",
                    0
                ),
                "packet_id": str(
                    row.get("Packet Id", "")
                ).strip(),
            }

            # Composite lookup: (sub_order_no, status, order_date).
            # A new status change on the same order creates a new row so that
            # the full lifecycle history is preserved.
            lookup_status     = defaults.pop("reason_for_credit_entry")
            lookup_order_date = defaults.pop("order_date")
            _, was_created = Order.objects.update_or_create(
                business=business,
                sub_order_no=sub_order_no,
                reason_for_credit_entry=lookup_status,
                order_date=lookup_order_date,
                defaults=defaults,
            )

            if was_created:
                created += 1
            else:
                updated += 1

    return Response(
        {
            "success": True,
            "created": created,
            "updated": updated,
            "total_rows": len(df),
        },
        status=status.HTTP_201_CREATED,
    )


# ── Full Orders (Order model) ─────────────────────────────────────────────────

@api_view(["GET"])
def full_orders_list(request, business_id):
    """Paginated list of Order rows with date/status/sku/state/search/month filters."""
    business = get_authorized_business(request, business_id)
    page      = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 50))
    status_filter = request.GET.get("status", "")
    sku_filter    = request.GET.get("sku", "")
    state_filter  = request.GET.get("state", "")
    search        = request.GET.get("search", "")
    month         = request.GET.get("month", "")   # YYYY-MM
    date_from     = request.GET.get("date_from", "")
    date_to       = request.GET.get("date_to", "")

    # month shortcut overrides explicit date_from/date_to
    if month:
        try:
            y, m = int(month[:4]), int(month[5:7])
            import calendar
            last_day = calendar.monthrange(y, m)[1]
            date_from = f"{y:04d}-{m:02d}-01"
            date_to   = f"{y:04d}-{m:02d}-{last_day:02d}"
        except (ValueError, IndexError):
            pass

    qs = Order.objects.filter(business=business)
    if status_filter:
        qs = qs.filter(reason_for_credit_entry__iexact=status_filter)
    if sku_filter:
        qs = qs.filter(sku__icontains=sku_filter)
    if state_filter:
        qs = qs.filter(customer_state__iexact=state_filter)
    if search:
        qs = qs.filter(
            DQ(sub_order_no__icontains=search) |
            DQ(product_name__icontains=search) |
            DQ(sku__icontains=search)
        )
    if date_from:
        qs = qs.filter(order_date__gte=date_from)
    if date_to:
        qs = qs.filter(order_date__lte=date_to)

    total = qs.count()
    start = (page - 1) * page_size
    items = qs[start: start + page_size]
    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": OrderSerializer(items, many=True).data,
    })


@api_view(["GET"])
def full_orders_analytics(request, business_id):
    """Aggregate stats for Order model — drives the Orders tab summary cards."""
    business = get_authorized_business(request, business_id)
    date_from     = request.GET.get("date_from", "")
    date_to       = request.GET.get("date_to", "")
    status_filter = request.GET.get("status", "")
    month         = request.GET.get("month", "")

    if month:
        try:
            y, m = int(month[:4]), int(month[5:7])
            import calendar
            last_day = calendar.monthrange(y, m)[1]
            date_from = f"{y:04d}-{m:02d}-01"
            date_to   = f"{y:04d}-{m:02d}-{last_day:02d}"
        except (ValueError, IndexError):
            pass

    base = Order.objects.filter(business=business)
    if date_from:
        base = base.filter(order_date__gte=date_from)
    if date_to:
        base = base.filter(order_date__lte=date_to)
    if status_filter:
        base = base.filter(reason_for_credit_entry__iexact=status_filter)

    qs = Order.latest_per_order(base_qs=base)

    by_status = list(
        qs.values("reason_for_credit_entry")
        .annotate(
            count=Count("sub_order_no", distinct=True),
            total_listed=Sum("supplier_listed_price"),
            total_discounted=Sum("supplier_discounted_price"),
        )
    )

    by_state = list(
        qs.values("customer_state")
        .annotate(count=Count("sub_order_no", distinct=True))
        .order_by("-count")[:10]
    )

    by_sku = list(
        qs.values("sku")
        .annotate(count=Count("sub_order_no", distinct=True), total_qty=Sum("quantity"))
        .order_by("-count")[:15]
    )

    daily = list(
        qs.values("order_date")
        .annotate(count=Count("sub_order_no", distinct=True))
        .order_by("order_date")
    )

    # Available months — always from unfiltered base for the month picker
    all_months = list(
        Order.objects.filter(business=business).values_list("order_date", flat=True)
        .exclude(order_date__isnull=True)
        .order_by("order_date")
        .distinct()
    )
    months = sorted({str(d)[:7] for d in all_months})

    return Response({
        "total": qs.count(),
        "by_status": by_status,
        "by_state": by_state,
        "by_sku": by_sku,
        "daily": daily,
        "months": months,
    })


@api_view(["GET"])
def dashboard_analytics(request, business_id):
    """
    Primary filter: Order.order_date (reliable DateField).
    Then join OrderPayment by sub_order_no to find settled orders.
    Unsettled = orders in date range with no matching payment.
    """
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")

    # ── Step 1: filter Orders by placement date, then deduplicate to latest ──
    date_filtered_qs = Order.objects.filter(business=business)
    if date_from:
        date_filtered_qs = date_filtered_qs.filter(order_date__gte=date_from)
    if date_to:
        date_filtered_qs = date_filtered_qs.filter(order_date__lte=date_to)

    # latest_per_order ensures each sub_order_no appears once (most-recent status row)
    order_qs  = Order.latest_per_order(base_qs=date_filtered_qs)
    order_nos = set(order_qs.values_list("sub_order_no", flat=True))

    # ── Step 2: find payments for those orders ────────────────────────────────
    payment_qs = (
        OrderPayment.objects.filter(business=business, sub_order_no__in=order_nos)
        if order_nos
        else OrderPayment.objects.none()
    )

    payment_nos = set(payment_qs.values_list("sub_order_no", flat=True).distinct())
    matched     = order_nos & payment_nos
    match_rate  = round(len(matched) / len(order_nos) * 100, 1) if order_nos else 0.0

    # ── Order aggregates (on latest-per-order queryset) ───────────────────────
    order_by_status = list(
        order_qs.values("reason_for_credit_entry")
        .annotate(
            count=Count("sub_order_no", distinct=True),
            total_value=Sum("supplier_discounted_price"),
        )
    )

    order_daily = list(
        order_qs.values("order_date")
        .annotate(count=Count("sub_order_no", distinct=True))
        .order_by("order_date")
        .values("order_date", "count")
    )

    # ── Payment aggregates ────────────────────────────────────────────────────
    payment_agg = payment_qs.aggregate(
        total_settlement=Sum("final_settlement_amount"),
        total_sale=Sum("total_sale_amount"),
        total_commission=Sum("meesho_commission_incl_gst"),
        total_tcs=Sum("tcs"),
        total_tds=Sum("tds"),
        settled_count=Count("sub_order_no", distinct=True),
    )

    payment_by_status = list(
        payment_qs.values("live_order_status")
        .annotate(
            count=Count("sub_order_no", distinct=True),
            total_settlement=Sum("final_settlement_amount"),
            total_sale=Sum("total_sale_amount"),
        )
    )

    payment_daily = list(
        payment_qs.exclude(payment_date=None)
        .values("payment_date")
        .annotate(
            count=Count("sub_order_no", distinct=True),
            total=Sum("final_settlement_amount"),
        )
        .order_by("payment_date")
        .values("payment_date", "count", "total")
    )

    # ── Per-status crosswalk ──────────────────────────────────────────────────
    status_settlement = []
    for row in order_by_status:
        status_val = row["reason_for_credit_entry"]
        sub_nos = list(
            order_qs.filter(reason_for_credit_entry=status_val)
            .values_list("sub_order_no", flat=True)
        )
        agg = payment_qs.filter(DQ(sub_order_no__in=sub_nos), DQ(claims=0) | DQ(claims__isnull=True)).aggregate(
            total=Sum("final_settlement_amount"),
            
            count=Count("sub_order_no", distinct=True),
        )
        status_settlement.append({
            "status": status_val,
            "order_count": row["count"],
            "settled_count": agg["count"] or 0,
            "settlement_amount": float(agg["total"] or 0),
            "order_value": float(row["total_value"] or 0),
        })

    # ── Unsettled orders summary (latest status per order, no payment row) ────
    unsettled_qs  = order_qs.exclude(sub_order_no__in=payment_nos)
    unsettled_agg = unsettled_qs.aggregate(total_value=Sum("supplier_discounted_price"))

    return Response({
        "order_stats": {
            "total": len(order_nos),
            "by_status": order_by_status,
            "daily": order_daily,
        },
        "payment_stats": {
            "total": len(payment_nos),
            "by_status": payment_by_status,
            "daily": payment_daily,
            "total_settlement": float(payment_agg["total_settlement"] or 0),
            "total_sale": float(payment_agg["total_sale"] or 0),
            "total_commission": float(payment_agg["total_commission"] or 0),
            "total_tcs": float(payment_agg["total_tcs"] or 0),
            "total_tds": float(payment_agg["total_tds"] or 0),
            "settled_count": payment_agg["settled_count"] or 0,
        },
        "join_stats": {
            "matched_count": len(matched),
            "unmatched_count": len(order_nos) - len(matched),
            "match_rate": match_rate,
        },
        "status_settlement": status_settlement,
        "unsettled": {
            "count": unsettled_qs.count(),
            "total_value": float(unsettled_agg["total_value"] or 0),
        },
    })


# ── Meesho Labels PDF helpers ─────────────────────────────────────────────────

def _parse_label_page(label_text, full_text, tables):
    """
    Extract all structured fields from a single label page.

    label_text  – pdfplumber text from label region only (above TAX INVOICE line)
    full_text   – full page text (needed for order date buried in invoice section)
    tables      – pdfplumber tables from the full page
    """
    import re
    from datetime import datetime as _dt

    r = {
        "customer_name": "", "customer_address": "",
        "customer_city": "", "customer_state": "", "customer_pincode": "",
        "courier_name": "", "awb_number": "", "payment_type": "", "pickup_date": "",
        "sku": "", "size": "", "qty": 1, "color": "", "order_id": "", "order_date": None,
    }

    lines = [l.strip() for l in label_text.split("\n") if l.strip()]

    # ── Customer address section ──────────────────────────────────────────────
    ca_idx    = next((i for i, l in enumerate(lines) if "Customer Address" in l), -1)
    undel_idx = next((i for i, l in enumerate(lines) if "If undelivered" in l), -1)
    if ca_idx >= 0 and undel_idx > ca_idx + 1:
        addr = lines[ca_idx + 1 : undel_idx]
        if addr:
            r["customer_name"] = addr[0]
            r["customer_address"] = "\n".join(addr[1:])
            last = addr[-1]
            m = re.match(r"^(.*),\s*([^,]{3,}),\s*(\d{6})\s*$", last)
            if m:
                r["customer_city"]    = m.group(1).strip().rstrip(",")
                r["customer_state"]   = m.group(2).strip()
                r["customer_pincode"] = m.group(3)

    # ── Payment type ──────────────────────────────────────────────────────────
    r["payment_type"] = (
        "Prepaid" if re.search(r"\bPrepaid\b", label_text, re.I)
        else "COD" if re.search(r"\bCOD\b", label_text, re.I)
        else ""
    )

    # ── Courier name ──────────────────────────────────────────────────────────
    for c in ("Shadowfax", "Delhivery", "Xpress Bees", "Valmo", "BlueDart", "Ekart", "DTDC"):
        if re.search(r"\b" + re.escape(c) + r"\b", label_text, re.I):
            r["courier_name"] = c
            break

    # ── Pickup date (Valmo prints "Pickup DD/MM" on the label) ───────────────
    pm = re.search(r"Pickup\s+(\d{2}/\d{2})", label_text)
    if pm:
        r["pickup_date"] = pm.group(1)

    # ── AWB number (standalone string just above "Product Details") ───────────
    # Each courier has a distinctive format:
    #   Valmo:       VL + 13 digits
    #   Shadowfax:   SF + digits + optional letters (e.g. SF3451319161FPL)
    #   Delhivery:   16-digit number
    #   Xpress Bees: 13-15 digit number
    for pat in (
        r"\b(VL\d{10,16})\b",
        r"\b(SF[0-9A-Z]{8,18})\b",
        r"\b(\d{16})\b",
        r"\b(\d{13,15})\b",
    ):
        m = re.search(pat, label_text)
        if m:
            r["awb_number"] = m.group(1)
            break

    # ── Pass 1: Product Details via pdfplumber table extraction ──────────────
    # The table header is: SKU | Size | Qty | Color | Order No.
    for table in (tables or []):
        for ri, row in enumerate(table or []):
            hdr = [str(c or "").strip().upper() for c in row]
            if "SKU" in hdr and "QTY" in hdr and ri + 1 < len(table):
                dr   = table[ri + 1]
                hdr_ = hdr   # capture for closure
                dr_  = dr

                def _g(name, _h=hdr_, _d=dr_):
                    try:
                        return str(_d[_h.index(name)] or "").strip()
                    except (ValueError, IndexError):
                        return ""

                r["sku"]      = " ".join(_g("SKU").split())
                r["size"]     = _g("SIZE")
                r["color"]    = _g("COLOR")
                r["order_id"] = _g("ORDER NO.")
                try:
                    r["qty"] = max(1, int(_g("QTY")))
                except ValueError:
                    pass
                break
        if r["sku"]:   # break on sku found, not order_id (order_id may be empty)
            break

    # ── Pass 2: text fallback when table extraction finds nothing ─────────────
    # Handles PDFs where Product Details has no border lines (pdfplumber can't
    # detect borderless text as a table). Looks for the header line then data.
    if not r["sku"]:
        # Pattern: after "SKU  Size  Qty  Color  Order No." header, parse data line
        m = re.search(
            r"SKU\s+Size\s+Qty\s+Color\s+Order\s*No\.\s*\n(.+?)(?:\n|$)",
            label_text, re.IGNORECASE,
        )
        if m:
            data_line = m.group(1).strip()
            # Most Meesho SKUs have no spaces; size field is "Free Size" or "X cm"
            size_m = re.search(r"\s+(Free\s+Size|\d+\s*[Cc][Mm])\s+", data_line, re.I)
            if size_m:
                r["sku"]  = data_line[:size_m.start()].strip()
                r["size"] = size_m.group(1).strip()
                after     = data_line[size_m.end():].strip()
                parts     = after.split()
                if parts:
                    try:
                        r["qty"] = max(1, int(parts[0]))
                    except ValueError:
                        pass
                    if len(parts) >= 2:
                        r["color"]    = parts[1]
                    if len(parts) >= 3:
                        r["order_id"] = parts[-1]
            else:
                # Last-resort: treat the last token as the order ID
                parts = data_line.split()
                if len(parts) >= 2 and re.match(r"\d+_\d+", parts[-1]):
                    r["order_id"] = parts[-1]
                    r["sku"]      = parts[0]

    # ── Order date from invoice section ───────────────────────────────────────
    dm = re.search(r"Order Date\s+(\d{2}\.\d{2}\.\d{4})", full_text)
    if dm:
        try:
            r["order_date"] = _dt.strptime(dm.group(1), "%d.%m.%Y").date()
        except ValueError:
            pass

    return r


# ── Meesho Labels PDF upload ──────────────────────────────────────────────────

def _sku_business_owners(businesses, sku_ids, default):
    """
    Which business owns each SKU, across the ones this user can reach.

    Returns (owner_by_sku, ambiguous) where `ambiguous` lists SKUs priced by more
    than one business — the same SKU name genuinely exists in two catalogues, and
    there is no honest way to pick. Those fall back to `default` and are reported
    so the UI can show both business names rather than silently choosing.

    Ownership is read from FinalPrice (the pricing catalogue) because that is
    where a business declares "this SKU is mine"; label history would be
    circular, since it is what we are trying to write.
    """
    wanted = {s for s in sku_ids if s}
    if not wanted:
        return {}, []

    by_sku = {}
    for b in businesses:
        for sku in FinalPrice.objects.filter(
            business=b, sku_id__in=wanted
        ).values_list("sku_id", flat=True):
            by_sku.setdefault(sku, []).append(b)

    owner_by_sku, ambiguous = {}, []
    for sku, owners in by_sku.items():
        if len(owners) == 1:
            owner_by_sku[sku] = owners[0]
        else:
            ambiguous.append({"sku": sku, "businesses": [o.name for o in owners]})
            owner_by_sku[sku] = default
    return owner_by_sku, ambiguous


def _sku_parent_map_pooled(businesses, sku_ids):
    """
    Parent name per SKU, looked up across every reachable business.

    A label batch spanning businesses needs each SKU grouped under *its own*
    catalogue's parent; resolving only against the selected business would leave
    the other business's SKUs looking unparented and break the grouping the
    labels view is built around.
    """
    wanted = {s for s in sku_ids if s}
    if not wanted:
        return {}

    mapping = {}
    for b in businesses:
        rows = (
            FinalPrice.objects.filter(business=b, sku_id__in=wanted)
            .select_related("parent").values("sku_id", "parent__item_id")
        )
        for row in rows:
            # First business to claim a SKU wins; the selected business is first
            # in `businesses`, so its own catalogue always takes precedence.
            if row["sku_id"] not in mapping and row["parent__item_id"]:
                mapping[row["sku_id"]] = row["parent__item_id"]
    return mapping


@api_view(["POST"])
@parser_classes([MultiPartParser])
def upload_labels_pdf(request, business_id):
    """
    Upload a Meesho shipping labels PDF (one label per page).

    Each label page has two sections separated by "TAX INVOICE":
      - Upper half: shipping label with Product Details table (SKU, Size, Qty, Color, Order No.)
      - Lower half: tax invoice (excluded from cropped output)

    Strategy:
      1. Use pdfplumber table extraction to find the "Product Details" table and
         extract SKU + Qty precisely (handles multi-word/wrapped SKUs).
      2. Locate "TAX INVOICE" text to determine crop boundary per page.
      3. Use pypdf to crop each page to the label-only region and return as base64 PDF.
    """
    business = get_authorized_business(request, business_id)
    import io, re, base64

    try:
        import pdfplumber
    except ImportError:
        return Response({"error": "pdfplumber not installed. Run: pip install pdfplumber"},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    try:
        from pypdf import PdfReader, PdfWriter
        HAS_PYPDF = True
    except ImportError:
        HAS_PYPDF = False

    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
    if not file.name.lower().endswith(".pdf"):
        return Response({"error": "Only PDF files are accepted."}, status=status.HTTP_400_BAD_REQUEST)

    pdf_bytes = file.read()

    # ── Batch date: accept back-dated uploads ─────────────────────────────────
    # Caller can pass upload_date=YYYY-MM-DD in the form to record historical
    # label batches under a past date. Must be ≤ today; defaults to today.
    from datetime import date as _date_cls
    today = timezone.now().date()
    upload_date_str = request.data.get("upload_date", "").strip()
    if upload_date_str:
        try:
            candidate = _date_cls.fromisoformat(upload_date_str)
            if candidate <= today:
                today = candidate
        except ValueError:
            pass  # invalid format — fall back to today

    sku_data = {}
    page_details   = []
    crop_y_list    = []   # pdfplumber top-y for TAX INVOICE, per page
    pl_heights     = []   # pdfplumber page heights
    db_rows        = []   # parsed dicts ready for LabelOrder.update_or_create
    total_pages    = 0

    extract_meta = {}
    try:
        # Page extraction (pdfminer layout + table detection) is the expensive
        # part of an upload and is independent per page, so it runs across a
        # process pool for larger batches. Field parsing below is unchanged and
        # still runs here, serially — there is only one parser either way.
        extracted, extract_meta = extract_all_pages(pdf_bytes)
        total_pages = extract_meta["total_pages"]

        for page_num, pg in enumerate(extracted, 1):
            pl_heights.append(pg["height"])
            crop_y_list.append(pg["crop_y"])

            tables    = pg["tables"]
            full_text = pg["full_text"]

            # Derive label-only text by splitting at "TAX INVOICE"
            tax_pos    = full_text.upper().find("TAX INVOICE")
            label_text = full_text[:tax_pos].strip() if tax_pos > 0 else full_text

            # ── Parse all fields via helper ───────────────────────────────
            parsed = _parse_label_page(label_text, full_text, tables)

            # ── Accumulate SKU analytics ──────────────────────────────────
            sku, qty = parsed["sku"], parsed["qty"]
            if sku:
                if sku not in sku_data:
                    sku_data[sku] = {"count": 0, "total_qty": 0, "max_qty": 0, "high_qty_orders": 0}
                sku_data[sku]["count"]    += 1
                sku_data[sku]["total_qty"] += qty
                sku_data[sku]["max_qty"]   = max(sku_data[sku]["max_qty"], qty)
                if qty > 1:
                    sku_data[sku]["high_qty_orders"] += 1

            page_details.append({
                "page":             page_num,
                "sku":              sku,
                "qty":              qty,
                "courier":          parsed["courier_name"],
                "awb":              parsed["awb_number"],
                "order_id":         parsed["order_id"],
                "customer_name":    parsed["customer_name"],
                "customer_pincode": parsed["customer_pincode"],
                "customer_address": parsed["customer_address"],
                "customer_city":    parsed["customer_city"],
                "customer_state":   parsed["customer_state"],
            })

            # ── Queue DB row (only if we have an order_id) ────────────────
            if parsed["order_id"]:
                db_rows.append({
                    "order_id":        parsed["order_id"],
                    "customer_name":   parsed["customer_name"],
                    "customer_address":parsed["customer_address"],
                    "customer_city":   parsed["customer_city"],
                    "customer_state":  parsed["customer_state"],
                    "customer_pincode":parsed["customer_pincode"],
                    "courier_name":    parsed["courier_name"],
                    "awb_number":      parsed["awb_number"],
                    "payment_type":    parsed["payment_type"],
                    "pickup_date":     parsed["pickup_date"],
                    "sku":             parsed["sku"],
                    "size":            parsed["size"],
                    "qty":             parsed["qty"],
                    "color":           parsed["color"],
                    "order_date":      parsed["order_date"],
                    "uploaded_date":   today,
                })

    except Exception as exc:
        return Response({"error": f"Could not parse PDF: {exc}"},
                        status=status.HTTP_400_BAD_REQUEST)

    # ── Save parsed rows to DB ────────────────────────────────────────────────
    # Rule: uploaded_date is SET only when a record is FIRST created.
    # On re-upload the row's data (courier, AWB, SKU, address…) is refreshed
    # but uploaded_date stays untouched — so orders remain on their original
    # day and don't "move" when the same order appears in a later batch.
    #
    # Written in bulk: one SELECT for what already exists, then one bulk_create
    # and one bulk_update. The previous row-at-a-time get_or_create + update did
    # two queries per label, which dominated the non-parsing time of a large
    # batch. Semantics are unchanged.
    # ── Which business owns each SKU in this PDF ─────────────────────────────
    # One tray of labels covers several businesses, so a single upload has to
    # split itself. Ownership is decided by which business already prices the
    # SKU; a SKU nobody prices, or one priced by more than one business, stays
    # with the business the operator had selected — guessing between two owners
    # would file parcels somewhere they'd never be found.
    scope = accessible_businesses(request, include=business)
    sku_owner, ambiguous_skus = _sku_business_owners(
        scope, {r.get("sku") or "" for r in db_rows}, default=business
    )

    saved = updated = 0
    per_business = {}
    if db_rows:
        # Last occurrence wins if the same order appears twice in one PDF —
        # matching the old loop, where the later row's update ran last.
        by_id = {}
        for row in db_rows:
            by_id[row["order_id"]] = row

        # Group by owning business so each still gets one SELECT and one write.
        rows_by_owner = {}
        for oid, row in by_id.items():
            owner = sku_owner.get(row.get("sku") or "", business)
            rows_by_owner.setdefault(owner.pk, (owner, {}))[1][oid] = row

        with transaction.atomic():
            for owner, owner_rows in rows_by_owner.values():
                existing_ids = set(
                    LabelOrder.objects
                    .filter(business=owner, order_id__in=list(owner_rows.keys()))
                    .values_list("order_id", flat=True)
                )

                to_create, to_update = [], []
                for oid, row in owner_rows.items():
                    data = {k: v for k, v in row.items() if k not in ("order_id", "uploaded_date")}
                    if oid in existing_ids:
                        # uploaded_date deliberately absent: an order keeps the day
                        # it first arrived and must not move to a later batch.
                        to_update.append(LabelOrder(order_id=oid, business=owner, **data))
                    else:
                        to_create.append(LabelOrder(
                            order_id=oid, business=owner,
                            uploaded_date=row["uploaded_date"], **data,
                        ))

                if to_create:
                    LabelOrder.objects.bulk_create(to_create, batch_size=500)
                    saved += len(to_create)
                if to_update:
                    LabelOrder.objects.bulk_update(
                        to_update,
                        [
                            "customer_name", "customer_address", "customer_city",
                            "customer_state", "customer_pincode", "courier_name",
                            "awb_number", "payment_type", "pickup_date",
                            "sku", "size", "qty", "color", "order_date",
                        ],
                        batch_size=500,
                    )
                    updated += len(to_update)

                per_business[owner.name] = len(owner_rows)

    # ── Resolve parent NAME (ParentItemPrice.item_id) for each child SKU ─────────
    # Use the human parent name (not the surrogate parent_id) so the upload
    # preview/analytics group and label by parent instead of the variant SKU.
    # Pooled across businesses: a batch can contain another business's SKUs, and
    # those must still group under their own catalogue's parent.
    sku_to_parent = _SkuMap(
        _sku_parent_map_pooled(scope, list(sku_data.keys())).items()
    )

    # Enrich page_details with parent SKU and, when the label belongs to another
    # business, which one — so the operator sorting the printed stack can tell.
    for _pd in page_details:
        _sku = _pd.get("sku") or ""
        _pd["parent_sku"] = sku_to_parent.get(_sku)
        _owner = sku_owner.get(_sku)
        _pd["business_name"] = _owner.name if _owner and _owner.pk != business.pk else None

    # Build parent-level rank: aggregate child counts under each parent so the
    # most-dispatched parent group appears first in the cropped PDF.
    parent_count: dict = {}
    for _sku, _data in sku_data.items():
        _parent = sku_to_parent.get(_sku) or _sku
        parent_count[_parent] = parent_count.get(_parent, 0) + _data["count"]
    parent_rank = {
        _p: _r for _r, (_p, _) in enumerate(
            sorted(parent_count.items(), key=lambda x: -x[1])
        )
    }

    # ── Generate cropped PDF (label region only, sorted by parent→child count) ─
    # Sort: parent group (most labels first) → child SKU rank → original page idx
    sku_rank = {
        sku: rank
        for rank, (sku, _) in enumerate(
            sorted(sku_data.items(), key=lambda x: -x[1]["count"])
        )
    }

    cropped_pdf_b64 = None
    if HAS_PYPDF:
        try:
            reader = PdfReader(io.BytesIO(pdf_bytes))
            writer = PdfWriter()
            page_order = sorted(
                range(len(reader.pages)),
                key=lambda i: (
                    parent_rank.get(
                        sku_to_parent.get(
                            (page_details[i]["sku"] if i < len(page_details) else "") or ""
                        ) or (page_details[i]["sku"] if i < len(page_details) else "") or "",
                        len(parent_count),
                    ),
                    sku_rank.get(
                        page_details[i]["sku"] if i < len(page_details) else "",
                        len(sku_data),
                    ),
                    i,
                ),
            )
            for i in page_order:
                pg   = reader.pages[i]
                ph   = float(pg.mediabox.height)
                pw   = float(pg.mediabox.width)
                pl_h = pl_heights[i] if i < len(pl_heights) else ph
                scale = ph / pl_h if pl_h else 1.0
                crop  = crop_y_list[i] if i < len(crop_y_list) else ph * 0.54
                y_bot = max(0.0, ph - crop * scale)
                pg.cropbox.lower_left  = (0, y_bot)
                pg.cropbox.upper_right = (pw, ph)
                writer.add_page(pg)
            buf = io.BytesIO()
            writer.write(buf)
            cropped_pdf_b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
        except Exception:
            cropped_pdf_b64 = None

    # ── Build analytics response ──────────────────────────────────────────────
    sku_table = sorted(
        [{"sku": k, "parent_sku": sku_to_parent.get(k),
          "count": v["count"], "total_qty": v["total_qty"],
          "max_qty": v["max_qty"], "high_qty_orders": v["high_qty_orders"]}
         for k, v in sku_data.items()],
        key=lambda x: -x["count"],
    )
    total_labels = sum(v["count"] for v in sku_data.values())

    # ── Flag blocked customers found in this upload ───────────────────────────
    blocked_set = set(
        BlockedCustomer.objects.filter(business=business, is_active=True)
        .values_list("customer_name", "customer_pincode")
    )
    blocked_in_batch = []
    for _pd in page_details:
        _name    = _pd.get("customer_name", "")
        _pincode = _pd.get("customer_pincode", "")
        if _name and (_name, _pincode) in blocked_set:
            blocked_in_batch.append({
                "order_id":         _pd.get("order_id", ""),
                "customer_name":    _name,
                "customer_pincode": _pincode,
                "sku":              _pd.get("sku", ""),
            })

    # ── Repeated customers: find prior-history customers in this batch ────────
    # Build (name, pincode) → batch info map
    batch_customers: dict = {}
    batch_order_ids: set  = set()
    for _pd in page_details:
        _name = (_pd.get("customer_name") or "").strip().lower()
        _pincode = str(_pd.get("customer_pincode") or "").strip()
        _oid    = _pd.get("order_id", "")
        if _oid:
            batch_order_ids.add(_oid)
        if not _name or not _pincode:
            continue
        _key = (_name, _pincode)
        if _key not in batch_customers:
            batch_customers[_key] = {
                "customer_name":    _name,
                "customer_pincode": _pincode,
                "customer_address": _pd.get("customer_address", "") or "",
                "customer_city":    _pd.get("customer_city", "") or "",
                "customer_state":   _pd.get("customer_state", "") or "",
                "batch_orders": [],
            }
        batch_customers[_key]["batch_orders"].append({
            "order_id": _oid,
            "sku":      _pd.get("sku", ""),
            "qty":      _pd.get("qty", 1),
        })

    repeated_customers = []
    if batch_customers:
        _all_names    = [k[0] for k in batch_customers]
        _all_pincodes = [k[1] for k in batch_customers]

        # Prior orders = same customer, NOT in this batch
        _prior_rows = list(
            LabelOrder.objects
            .filter(business=business, customer_name__in=_all_names, customer_pincode__in=_all_pincodes)
            .exclude(order_id__in=list(batch_order_ids))
            .values("order_id", "customer_name", "customer_pincode",
                    "sku", "qty", "order_date", "uploaded_date")
        )

        # Group prior rows by customer key
        _prior_by_customer: dict = {}
        for _row in _prior_rows:
            _k = (_row["customer_name"], _row["customer_pincode"])
            _prior_by_customer.setdefault(_k, []).append(_row)

        # Only process customers that actually have prior history
        _prior_order_ids = [_r["order_id"] for _r in _prior_rows]
        _outcomes: dict  = {}
        if _prior_order_ids:
            for _o in (Order.objects
                       .filter(business=business, sub_order_no__in=_prior_order_ids)
                       .values("sub_order_no", "reason_for_credit_entry")):
                _outcomes.setdefault(_o["sub_order_no"], []).append(
                    _o["reason_for_credit_entry"]
                )

        _RETURN_S = {"RETURN", "RETURNED"}

        for _key, _prior in _prior_by_customer.items():
            if not _prior:
                continue
            print("Missing key:", repr(_key))
            print("Available keys:")
            # _bc = batch_customers[_key]

            # Resolve status for each prior order
            _enriched = []
            for _r in _prior:
                _sts = _outcomes.get(_r["order_id"], [])
                if "DELIVERED" in _sts:
                    _resolved = "DELIVERED"
                elif any(_s in _RETURN_S for _s in _sts):
                    _resolved = "RETURN"
                elif "RTO_COMPLETE" in _sts:
                    _resolved = "RTO"
                elif "CANCELLED" in _sts:
                    _resolved = "CANCELLED"
                else:
                    _resolved = "PENDING"
                _enriched.append({
                    "order_id":   _r["order_id"],
                    "sku":        _r["sku"] or "",
                    "qty":        _r["qty"] or 1,
                    "order_date": str(_r["order_date"]) if _r["order_date"] else "",
                    "status":     _resolved,
                })

            _enriched.sort(key=lambda x: x["order_date"], reverse=True)

            _total     = len(_enriched)
            _delivered = sum(1 for _o in _enriched if _o["status"] == "DELIVERED")
            _returned  = sum(1 for _o in _enriched if _o["status"] == "RETURN")
            _rto       = sum(1 for _o in _enriched if _o["status"] == "RTO")
            _ret_rate  = round(_returned / _total, 3) if _total > 0 else 0.0

            # SKU breakdown
            _sku_stats: dict = {}
            for _o in _enriched:
                _s_sku = _o["sku"] or "Unknown"
                _s     = _sku_stats.setdefault(_s_sku, {
                    "sku": _s_sku, "orders": 0, "qty": 0,
                    "delivered": 0, "returned": 0, "rto": 0,
                })
                _s["orders"] += 1
                _s["qty"]    += int(_o["qty"] or 1)
                if _o["status"] == "DELIVERED": _s["delivered"] += 1
                elif _o["status"] == "RETURN":  _s["returned"]  += 1
                elif _o["status"] == "RTO":     _s["rto"]       += 1

            repeated_customers.append({
                "customer_name":     _key[0],
                "customer_pincode":  _key[1],
                # "customer_address":  _bc["customer_address"],
                # "customer_city":     _bc["customer_city"],
                # "customer_state":    _bc["customer_state"],
                # "batch_orders":      _bc["batch_orders"],
                "prior_order_count": _total,
                "delivered":         _delivered,
                "returned":          _returned,
                "rto":               _rto,
                "return_rate":       _ret_rate,
                "risk_level":        _risk_level(_ret_rate, _returned),
                "is_blocked":        _key in blocked_set,
                "sku_breakdown":     sorted(_sku_stats.values(), key=lambda x: -x["orders"]),
                "orders":            _enriched[:50],
                "last_seen":         _enriched[0]["order_date"] if _enriched else "",
            })

        # Sort: blocked first, then high risk, then by return rate desc
        _rl_order = {"high": 0, "medium": 1, "low": 2}
        repeated_customers.sort(key=lambda r: (
            0 if r["is_blocked"] else 1,
            _rl_order.get(r["risk_level"], 3),
            -r["return_rate"],
        ))

    return Response({
        "success": True,
        "upload_date":      str(today),
        # How this batch split across businesses, and any SKU whose owner was
        # genuinely ambiguous (same name priced by two catalogues).
        "per_business":     per_business,
        "ambiguous_skus":   ambiguous_skus,
        "total_pages":      total_pages,
        "total_unique_skus":len(sku_data),
        "total_labels":     total_labels,
        "has_results":      total_labels > 0,
        "has_high_qty":     any(v["high_qty_orders"] > 0 for v in sku_data.values()),
        "sku_table":        sku_table,
        "page_details":     page_details,
        "cropped_pdf_b64":  cropped_pdf_b64,
        "db_saved":         saved,
        "db_updated":       updated,
        "blocked_customers_found": blocked_in_batch,
        "repeated_customers":      repeated_customers,
    })


# ── Label Orders — read endpoints ─────────────────────────────────────────────

@api_view(["GET"])
def label_orders_list(request, business_id):
    """
    Paginated LabelOrder list.
    Query params: date (YYYY-MM-DD), date_from, date_to, courier, payment_type,
                  page, page_size
    """
    business = get_authorized_business(request, business_id)
    page        = int(request.GET.get("page", 1))
    page_size   = int(request.GET.get("page_size", 50))
    date_single = request.GET.get("date", "")
    date_from   = request.GET.get("date_from", "")
    date_to     = request.GET.get("date_to", "")
    courier     = request.GET.get("courier", "")
    pay_type    = request.GET.get("payment_type", "")
    search      = request.GET.get("q", "").strip()
    sort        = (request.GET.get("sort") or "").strip()

    qs = LabelOrder.objects.filter(business=business)

    if date_single:
        qs = qs.filter(uploaded_date=date_single)
    else:
        if date_from:
            qs = qs.filter(uploaded_date__gte=date_from)
        if date_to:
            qs = qs.filter(uploaded_date__lte=date_to)

    if courier:
        qs = qs.filter(courier_name__iexact=courier)
    if pay_type:
        qs = qs.filter(payment_type__iexact=pay_type)
    if search:
        qs = qs.filter(
            DQ(sku__icontains=search) |
            DQ(order_id__icontains=search) |
            DQ(awb_number__icontains=search) |
            DQ(customer_name__icontains=search) |
            DQ(customer_city__icontains=search) |
            DQ(customer_pincode__icontains=search)
        )

    # Explicit sort, so the table's column headers can drive ordering. Unknown
    # keys fall through to the model's own default (uploaded_date, courier).
    _SORTS = {
        "sku":      ("sku", "order_id"),
        "-sku":     ("-sku", "order_id"),
        "courier":  ("courier_name", "sku"),
        "-courier": ("-courier_name", "sku"),
        "qty":      ("qty", "sku"),
        "-qty":     ("-qty", "sku"),
        "date":     ("uploaded_date", "courier_name"),
        "-date":    ("-uploaded_date", "courier_name"),
        "name":     ("customer_name", "order_id"),
        "-name":    ("-customer_name", "order_id"),
        "city":     ("customer_city", "customer_name"),
        "-city":    ("-customer_city", "customer_name"),
        "order":    ("order_id",),
        "-order":   ("-order_id",),
    }
    if sort in _SORTS:
        qs = qs.order_by(*_SORTS[sort])

    total = qs.count()
    start = (page - 1) * page_size
    items = list(qs[start : start + page_size])

    # Annotate blocked status: build lookup set from active blocked customers
    blocked_set = set(
        BlockedCustomer.objects.filter(business=business, is_active=True)
        .values_list("customer_name", "customer_pincode")
    )

    serialized = LabelOrderSerializer(items, many=True).data

    # Resolve each label's SKU to its parent name (ParentItemPrice.item_id) so
    # the UI can show the parent instead of the variant-level SKU. Bulk map to
    # avoid N+1; falls back to None when the SKU isn't linked to a parent.
    skus = {row.get("sku") for row in serialized if row.get("sku")}
    sku_to_parent = _SkuMap(
        FinalPrice.objects.filter(business=business, sku_id__in=skus)
        .select_related("parent")
        .values_list("sku_id", "parent__item_id")
    ) if skus else _SkuMap()

    for row in serialized:
        row["is_blocked"] = (row.get("customer_name", ""), row.get("customer_pincode", "")) in blocked_set
        row["parent_sku"] = sku_to_parent.get(row.get("sku")) or None

    return Response({
        "total":     total,
        "page":      page,
        "page_size": page_size,
        "results":   serialized,
    })


@api_view(["GET"])
def label_couriers_summary(request, business_id):
    """
    Courier-wise order count for a date range.
    Query params: date (single day), date_from, date_to
    No date params = return all records.
    Also returns the list of available uploaded dates for the date-picker.
    """
    business = get_authorized_business(request, business_id)
    date_single = request.GET.get("date", "")
    date_from   = request.GET.get("date_from", "")
    date_to     = request.GET.get("date_to", "")

    qs = LabelOrder.objects.filter(business=business)
    if date_single:
        qs = qs.filter(uploaded_date=date_single)
        filter_label = date_single
    elif date_from or date_to:
        if date_from:
            qs = qs.filter(uploaded_date__gte=date_from)
        if date_to:
            qs = qs.filter(uploaded_date__lte=date_to)
        filter_label = f"{date_from or '…'} → {date_to or '…'}"
    else:
        filter_label = "All"

    courier_rows = list(
        qs.values("courier_name")
        .annotate(
            count       = Count("order_id"),
            prepaid     = Count("order_id", filter=DQ(payment_type="Prepaid")),
            cod         = Count("order_id", filter=DQ(payment_type="COD")),
            total_items = Sum("qty"),
        )
        .order_by("-count")
    )

    sku_rows = list(
        qs.values("sku")
        .annotate(count=Count("order_id"), total_items=Sum("qty"))
        .order_by("-count")
    )

    available_dates = list(
        LabelOrder.objects
        .filter(business=business)
        .values_list("uploaded_date", flat=True)
        .distinct()
        .order_by("-uploaded_date")
    )

    return Response({
        "filter":          filter_label,
        "total":           qs.count(),
        "courier_summary": courier_rows,
        "sku_summary":     sku_rows,
        "available_dates": [str(d) for d in available_dates],
    })


@api_view(["GET"])
def label_duplicate_customers(request, business_id):
    """
    Find repeat customers by ADDRESS (city + state + pincode) — NOT by name.
    Returns addresses that appear in more than one LabelOrder.

    For each duplicate address, returns every order's full details so the
    frontend can display a complete history drawer.

    Query params: date_from, date_to (optional)
    """
    business = get_authorized_business(request, business_id)
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to",   "")

    # Step 1: addresses that appear in the selected period
    period_qs = LabelOrder.objects.filter(business=business).exclude(customer_pincode="").exclude(customer_city="")
    if date_from:
        period_qs = period_qs.filter(uploaded_date__gte=date_from)
    if date_to:
        period_qs = period_qs.filter(uploaded_date__lte=date_to)

    # Step 2: find which (pincode, city, state) tuples appear > 1 time all-time,
    # but only for addresses that show up in the requested period.
    period_addrs = set(
        period_qs.values_list("customer_pincode", "customer_city", "customer_state").distinct()
    )
    if not period_addrs:
        return Response({"total": 0, "results": []})

    from django.db.models import Q
    addr_filter = Q()
    for pin, city, state in period_addrs:
        addr_filter |= Q(customer_pincode=pin, customer_city=city, customer_state=state)

    all_time_qs = LabelOrder.objects.filter(addr_filter, business=business)

    # Aggregate per address
    addr_agg = list(
        all_time_qs
        .values("customer_pincode", "customer_city", "customer_state")
        .annotate(
            order_count=Count("order_id"),
            first_ordered=Min("uploaded_date"),
            last_ordered=Max("uploaded_date"),
        )
        .filter(order_count__gt=1)
        .order_by("-order_count")
    )

    results = []
    for agg in addr_agg:
        pin   = agg["customer_pincode"]
        city  = agg["customer_city"]
        state = agg["customer_state"]
        orders = list(
            LabelOrder.objects
            .filter(business=business, customer_pincode=pin, customer_city=city, customer_state=state)
            .values(
                "order_id", "customer_name", "customer_city", "customer_state",
                "customer_pincode", "uploaded_date", "courier_name", "awb_number",
                "payment_type", "sku", "qty", "is_packed",
            )
            .order_by("-uploaded_date")
        )
        results.append({
            "address_key":      f"{city}, {state} – {pin}",
            "customer_pincode": pin,
            "customer_city":    city,
            "customer_state":   state,
            "order_count":      agg["order_count"],
            "first_ordered":    str(agg["first_ordered"]),
            "last_ordered":     str(agg["last_ordered"]),
            "orders":           orders,
        })

    return Response({"total": len(results), "results": results})


@api_view(["GET"])
def label_customer_history(request, business_id):
    """
    Full order + settlement history for one customer.

    Query params (at least one required):
      name    – customer name (case-insensitive exact match)
      pincode – customer pincode (exact)

    Joins:
      LabelOrder  →  Order        (via order_id = sub_order_no) for delivery status
      LabelOrder  →  OrderPayment (via order_id = sub_order_no) for settlement
    """
    from datetime import date as _date, datetime as _datetime

    business = get_authorized_business(request, business_id)
    name    = request.GET.get("name", "").strip()
    pincode = request.GET.get("pincode", "").strip()
    address = request.GET.get("address", "").strip()

    if not name and not pincode and not address:
        return Response({"error": "Provide at least name, pincode or address."}, status=400)

    qs = LabelOrder.objects.filter(business=business)
    if name:
        qs = qs.filter(customer_name__iexact=name)
    if pincode:
        qs = qs.filter(customer_pincode=pincode)
    # Address narrows a same-name match down to one household. Optional and
    # additive, so existing name/pincode-only callers behave exactly as before.
    if address:
        qs = qs.filter(customer_address=address)

    label_list = list(qs.order_by("-uploaded_date").values())
    if not label_list:
        return Response({"orders": [], "summary": {}, "customer_name": name})

    order_ids = [r["order_id"] for r in label_list]

    # Join OrderPayment
    payment_map = {
        p.sub_order_no: {
            "live_order_status":  p.live_order_status or "",
            "settlement_amount":  float(p.final_settlement_amount or 0),
            "total_sale_amount":  float(p.total_sale_amount or 0),
            "payment_date":       str(p.payment_date) if p.payment_date else None,
        }
        for p in OrderPayment.objects.filter(business=business, sub_order_no__in=order_ids)
    }

    # Join Order (lifecycle status)
    status_map = {
        o.sub_order_no: o.reason_for_credit_entry
        for o in Order.objects.filter(business=business, sub_order_no__in=order_ids)
    }

    # ── Enrich and accumulate ────────────────────────────────────────────────
    delivered = rto = cancelled = settled_count = 0
    total_settlement = 0.0
    enriched = []

    for lo in label_list:
        oid             = lo["order_id"]
        pmt             = payment_map.get(oid)
        delivery_status = status_map.get(oid, "")
        is_settled      = oid in payment_map
        amt             = pmt["settlement_amount"] if pmt else None

        if delivery_status == "DELIVERED":
            delivered += 1
        elif delivery_status == "RTO_COMPLETE":
            rto += 1
        elif delivery_status == "CANCELLED":
            cancelled += 1

        if is_settled:
            settled_count    += 1
            total_settlement += (amt or 0)

        # Serialize dates to strings
        row = {}
        for k, v in lo.items():
            if isinstance(v, (_date, _datetime)):
                row[k] = str(v)
            else:
                row[k] = v

        row.update({
            "delivery_status":   delivery_status,
            "live_order_status": pmt["live_order_status"] if pmt else "",
            "settlement_amount": amt,
            "total_sale_amount": pmt["total_sale_amount"] if pmt else None,
            "payment_date":      pmt["payment_date"] if pmt else None,
            "is_settled":        is_settled,
        })
        enriched.append(row)

    first = label_list[0]
    total = len(label_list)

    return Response({
        "customer_name":    first.get("customer_name", ""),
        "customer_address": first.get("customer_address", ""),
        "customer_city":    first.get("customer_city", ""),
        "customer_state":   first.get("customer_state", ""),
        "customer_pincode": first.get("customer_pincode", ""),
        "summary": {
            "total_orders":     total,
            "total_items":      sum(r.get("qty", 1) for r in label_list),
            "delivered":        delivered,
            "rto":              rto,
            "cancelled":        cancelled,
            "pending":          total - delivered - rto - cancelled,
            "settled_count":    settled_count,
            "total_settlement": round(total_settlement, 2),
        },
        "orders": enriched,
    })


def _log_inventory(business, entity_type, entity_id, action, description,
                   parent_sku_id="", quantity_change=None, metadata=None):
    """Write one immutable audit-log entry."""
    InventoryLog.objects.create(
        business=business,
        entity_type=entity_type,
        entity_id=str(entity_id),
        action=action,
        parent_sku_id=parent_sku_id or "",
        quantity_change=quantity_change,
        description=description,
        metadata=metadata or {},
    )

# ── Purchases & Inventory ─────────────────────────────────────────────────────

def _purchase_bills_qs(business):
    """Bills with items and each item's parent preloaded (avoids N+1 when
    resolving the parent's item_id name for display)."""
    return PurchaseBill.objects.filter(business=business).prefetch_related(
        Prefetch("items", queryset=PurchaseItem.objects.select_related("parent_sku"))
    )


def _bill_to_dict(bill):
    """Serialize a PurchaseBill (with pre-fetched items) to a plain dict.

    parent_sku_id is exposed as the parent's item_id string (the human name),
    not the surrogate FK id, so the UI shows/round-trips the SKU name.
    """
    total = Decimal("0")
    items = []
    for item in bill.items.all():
        item_total = item.quantity * item.price_per_unit if not item.is_exchange else Decimal("0")
        total += item_total
        items.append({
            "id":                  item.id,
            "parent_sku_id":       item.parent_sku.item_id if item.parent_sku_id else None,
            "product_description": item.product_description,
            "quantity":            item.quantity,
            "price_per_unit":      str(item.price_per_unit),
            "is_exchange":         item.is_exchange,
            "total_amount":        str(item_total),
        })
    return {
        "id":           bill.id,
        "date":         str(bill.date),
        "seller_name":  bill.seller_name,
        "bill_number":  bill.bill_number,
        "notes":        bill.notes,
        "total_amount": str(total),
        "items":        items,
        "created_at":   bill.created_at.isoformat(),
    }


@api_view(["GET", "POST"])
def purchases_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        qs = _purchase_bills_qs(business).order_by("-date", "-created_at")
        date_from = request.GET.get("date_from")
        date_to   = request.GET.get("date_to")
        seller    = request.GET.get("seller", "").strip()
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if seller:
            qs = qs.filter(seller_name__icontains=seller)
        bills = [_bill_to_dict(b) for b in qs]
        return Response({"results": bills, "total": len(bills)})

    # POST — create new bill with nested items
    data = request.data
    with transaction.atomic():
        bill = PurchaseBill.objects.create(
            business=business,
            date=data["date"],
            seller_name=data["seller_name"],
            bill_number=data.get("bill_number", ""),
            notes=data.get("notes", ""),
        )
        for it in data.get("items", []):
            # parent_sku arrives as the parent's item_id string; resolve it to
            # the business-scoped ParentItemPrice object (None if unlinked/new).
            PurchaseItem.objects.create(
                business=business,
                bill=bill,
                parent_sku=_resolve_parent(it.get("parent_sku_id"), business),
                product_description=it.get("product_description", ""),
                quantity=int(it["quantity"]),
                price_per_unit=Decimal(str(it["price_per_unit"])),
                is_exchange=bool(it.get("is_exchange", False)),
            )
    bill.refresh_from_db()
    for it in bill.items.select_related("parent_sku").all():
        if it.parent_sku_id:
            parent_name = it.parent_sku.item_id
            _log_inventory(
                business,
                "PURCHASE", it.id, "CREATE",
                f"Added {it.quantity} units of {parent_name} from {bill.seller_name}",
                parent_sku_id=parent_name,
                quantity_change=it.quantity,
                metadata={"bill_id": bill.id, "price_per_unit": str(it.price_per_unit), "is_exchange": it.is_exchange},
            )
    return Response(_bill_to_dict(_purchase_bills_qs(business).get(pk=bill.pk)), status=201)


@api_view(["GET", "PUT", "DELETE"])
def purchase_detail(request, business_id, bill_id):
    business = get_authorized_business(request, business_id)
    try:
        bill = PurchaseBill.objects.prefetch_related("items").get(pk=bill_id, business=business)
    except PurchaseBill.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "GET":
        return Response(_bill_to_dict(bill))

    if request.method == "PUT":
        data = request.data
        with transaction.atomic():
            bill.date        = data.get("date", bill.date)
            bill.seller_name = data.get("seller_name", bill.seller_name)
            bill.bill_number = data.get("bill_number", bill.bill_number)
            bill.notes       = data.get("notes", bill.notes)
            bill.save()
            # Replace all items
            bill.items.all().delete()
            for it in data.get("items", []):
                PurchaseItem.objects.create(
                    business=business,
                    bill=bill,
                    parent_sku=_resolve_parent(it.get("parent_sku_id"), business),
                    product_description=it.get("product_description", ""),
                    quantity=int(it["quantity"]),
                    price_per_unit=Decimal(str(it["price_per_unit"])),
                    is_exchange=bool(it.get("is_exchange", False)),
                )
        bill = _purchase_bills_qs(business).get(pk=bill_id)
        return Response(_bill_to_dict(bill))

    # DELETE
    bill.delete()
    return Response({"message": "Deleted"})


@api_view(["GET"])
def purchase_pdf(request, business_id, bill_id):
    """Generate and stream a PDF receipt for one purchase bill."""
    from io import BytesIO
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    business = get_authorized_business(request, business_id)
    try:
        bill = PurchaseBill.objects.prefetch_related("items").get(pk=bill_id, business=business)
    except PurchaseBill.DoesNotExist:
        return HttpResponse("Not found", status=404)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )
    W = A4[0] - 40*mm  # usable width

    orange = colors.HexColor("#E8510A")
    gray50 = colors.HexColor("#F9FAFB")
    gray200 = colors.HexColor("#E5E7EB")
    gray700 = colors.HexColor("#374151")

    title_style  = ParagraphStyle("title",  fontSize=20, fontName="Helvetica-Bold", textColor=orange, spaceAfter=2)
    sub_style    = ParagraphStyle("sub",    fontSize=10, fontName="Helvetica",      textColor=colors.HexColor("#6B7280"), spaceAfter=4)
    label_style  = ParagraphStyle("lbl",    fontSize=9,  fontName="Helvetica-Bold", textColor=gray700)
    value_style  = ParagraphStyle("val",    fontSize=9,  fontName="Helvetica",      textColor=gray700)
    total_style  = ParagraphStyle("tot",    fontSize=12, fontName="Helvetica-Bold", textColor=orange, alignment=TA_RIGHT)

    elems = []

    # Header
    # Brand header: Rudam + bill title on same line
    brand_style = ParagraphStyle("brand", fontSize=26, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#E8510A"), spaceAfter=0)
    tagline_style = ParagraphStyle("tagline", fontSize=9, fontName="Helvetica",
                                   textColor=colors.HexColor("#9CA3AF"), spaceAfter=6)
    brand_tbl = Table(
        [[Paragraph("RUDAM", brand_style), Paragraph("PURCHASE BILL", title_style)]],
        colWidths=[W * 0.4, W * 0.6],
    )
    brand_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("ALIGN",  (1, 0), (1, 0),  "RIGHT"),
    ]))
    elems.append(brand_tbl)
    elems.append(Paragraph("Official Purchase Record", tagline_style))
    elems.append(HRFlowable(width=W, thickness=1, color=orange, spaceAfter=10))

    # Meta info table (left: labels, right: values)
    bill_no_display = bill.bill_number or f"#{bill.id}"
    meta = [
        [Paragraph("Bill No:", label_style),   Paragraph(bill_no_display, value_style),
         Paragraph("Date:", label_style),       Paragraph(str(bill.date), value_style)],
        [Paragraph("Seller:", label_style),     Paragraph(bill.seller_name, value_style),
         Paragraph("", label_style),            Paragraph("", value_style)],
    ]
    if bill.notes:
        meta.append([Paragraph("Notes:", label_style), Paragraph(bill.notes, value_style), Paragraph("", label_style), Paragraph("", value_style)])
    meta_tbl = Table(meta, colWidths=[22*mm, W/2 - 22*mm, 22*mm, W/2 - 22*mm])
    meta_tbl.setStyle(TableStyle([
        ("VALIGN",    (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elems.append(meta_tbl)
    elems.append(Spacer(1, 10))
    elems.append(HRFlowable(width=W, thickness=0.5, color=gray200, spaceAfter=8))

    # Items table
    col_w = [W * f for f in (0.20, 0.35, 0.10, 0.15, 0.15, 0.05)]
    hdr = [
        Paragraph("<b>SKU</b>", label_style),
        Paragraph("<b>Description</b>", label_style),
        Paragraph("<b>Qty</b>", label_style),
        Paragraph("<b>Price/Unit</b>", label_style),
        Paragraph("<b>Total</b>", label_style),
        Paragraph("<b>Exch</b>", label_style),
    ]
    rows = [hdr]
    grand_total = Decimal("0")
    for item in bill.items.select_related("parent_sku").all():
        line_total = item.quantity * item.price_per_unit if not item.is_exchange else Decimal("0")
        grand_total += line_total
        rows.append([
            Paragraph((item.parent_sku.item_id if item.parent_sku_id else None) or "—", value_style),
            Paragraph(item.product_description or "—", value_style),
            Paragraph(str(item.quantity), value_style),
            Paragraph(f"₹{item.price_per_unit}", value_style),
            Paragraph(f"₹{line_total}", value_style),
            Paragraph("Y" if item.is_exchange else "N", value_style),
        ])
    # Total row
    rows.append([
        Paragraph("", value_style), Paragraph("", value_style),
        Paragraph("", value_style), Paragraph("<b>TOTAL</b>", label_style),
        Paragraph(f"<b>₹{grand_total}</b>", label_style),
        Paragraph("", value_style),
    ])

    items_tbl = Table(rows, colWidths=col_w)
    items_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  gray50),
        ("GRID",          (0, 0), (-1, -2), 0.4, gray200),
        ("LINEABOVE",     (0, -1), (-1, -1), 1, orange),
        ("BACKGROUND",    (0, -1), (-1, -1), colors.HexColor("#FFF0EA")),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
    ]))
    elems.append(items_tbl)
    elems.append(Spacer(1, 14))

    # Grand total callout
    elems.append(Paragraph(f"Grand Total: ₹{grand_total}", total_style))
    elems.append(Spacer(1, 6))
    elems.append(HRFlowable(width=W, thickness=0.5, color=gray200))
    elems.append(Spacer(1, 8))
    elems.append(Paragraph("Rudam — Generated by Meesho Profit Tracker", sub_style))

    doc.build(elems)
    buf.seek(0)
    filename = f"purchase_bill_{bill.bill_number or bill.id}.pdf"
    resp = HttpResponse(buf.read(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@api_view(["GET"])
def inventory_view(request, business_id):
    """
    Compute current stock per parent SKU:
      current_stock = purchased_qty (non-exchange) - sold_qty (DELIVERED) + rto_qty (RTO_COMPLETE) + net_adjustments
    Also returns parent SKUs that have only adjustments (no purchases) so manual-only SKUs appear.
    """
    from django.db.models import ExpressionWrapper, F, DecimalField as DField

    business = get_authorized_business(request, business_id)

    # 1. Purchases per parent SKU
    purchase_agg = (
        PurchaseItem.objects
        .filter(business=business, is_exchange=False, parent_sku__isnull=False)
        .values("parent_sku_id")
        .annotate(
            purchased_qty=Sum("quantity"),
            purchase_value=Sum(
                ExpressionWrapper(F("quantity") * F("price_per_unit"), output_field=DField(max_digits=14, decimal_places=2))
            ),
        )
    )
    purchased_by_parent = {
        r["parent_sku_id"]: {"qty": r["purchased_qty"], "value": r["purchase_value"]}
        for r in purchase_agg
    }

    # 2. Manual adjustments per parent SKU
    adj_agg = (
        InventoryAdjustment.objects
        .filter(business=business)
        .values("parent_sku_id")
        .annotate(net_adj=Sum("quantity"))
    )
    adj_by_parent = {r["parent_sku_id"]: r["net_adj"] for r in adj_agg}

    # Merge: all parent SKUs that appear in either purchases or adjustments
    all_parent_ids = set(purchased_by_parent.keys()) | set(adj_by_parent.keys())
    if not all_parent_ids:
        return Response({"results": [], "total": 0})

    # Map surrogate parent id -> human item_id (name) for display.
    parent_name = dict(
        ParentItemPrice.objects.filter(business=business, id__in=all_parent_ids)
        .values_list("id", "item_id")
    )

    # 3. Map child SKU → parent SKU. Case-insensitive, because step 5 rolls up
    # using the *orders* spelling of the SKU, which doesn't always match the
    # pricing row's — a plain dict dropped those quantities from the parent.
    sku_to_parent = _SkuMap(
        FinalPrice.objects
        .filter(business=business, parent_id__in=all_parent_ids)
        .values_list("sku_id", "parent_id")
    )

    # 4. Sold qty (DELIVERED) and returned qty (RTO_COMPLETE) per child SKU
    child_skus = list(sku_to_parent.keys())
    delivered_by_sku = dict(
        Order.objects
        .filter(business=business, reason_for_credit_entry="DELIVERED", sku__in=child_skus)
        .values("sku")
        .annotate(qty=Sum("quantity"))
        .values_list("sku", "qty")
    )
    rto_by_sku = dict(
        Order.objects
        .filter(business=business, reason_for_credit_entry="RTO_COMPLETE", sku__in=child_skus)
        .values("sku")
        .annotate(qty=Sum("quantity"))
        .values_list("sku", "qty")
    )

    # 5. Roll up child → parent
    sold_by_parent = {}
    rto_by_parent  = {}
    for sku, qty in delivered_by_sku.items():
        parent = sku_to_parent.get(sku)
        if parent:
            sold_by_parent[parent] = sold_by_parent.get(parent, 0) + qty
    for sku, qty in rto_by_sku.items():
        parent = sku_to_parent.get(sku)
        if parent:
            rto_by_parent[parent] = rto_by_parent.get(parent, 0) + qty

    # 6. Last purchase date per parent SKU
    last_purchase = dict(
        PurchaseItem.objects
        .filter(business=business, parent_sku_id__in=all_parent_ids)
        .values("parent_sku_id")
        .annotate(last=Max("bill__date"))
        .values_list("parent_sku_id", "last")
    )

    results = []
    for parent_id in all_parent_ids:
        pdata    = purchased_by_parent.get(parent_id, {"qty": 0, "value": Decimal("0")})
        purchased = pdata["qty"]
        sold      = sold_by_parent.get(parent_id, 0)
        rto       = rto_by_parent.get(parent_id, 0)
        net_adj   = adj_by_parent.get(parent_id, 0)
        results.append({
            "sku_id":          parent_name.get(parent_id, str(parent_id)),
            "purchased_qty":   purchased,
            "sold_qty":        sold,
            "rto_qty":         rto,
            "adjustment_qty":  net_adj,
            "current_stock":   purchased - sold + rto + net_adj,
            "purchase_value":  str(pdata["value"] or 0),
            "last_purchase":   str(last_purchase.get(parent_id, "")),
        })

    results.sort(key=lambda r: r["current_stock"])
    return Response({"results": results, "total": len(results)})


@api_view(["GET", "POST"])
def inventory_adjustment_list(request, business_id):
    """List or create manual inventory adjustments for a parent SKU."""
    business = get_authorized_business(request, business_id)
    parent_sku_id = request.GET.get("parent_sku", "").strip() if request.method == "GET" else request.data.get("parent_sku_id", "").strip()

    if request.method == "GET":
        if not parent_sku_id:
            return Response({"error": "parent_sku required"}, status=400)
        adjs = InventoryAdjustment.objects.filter(business=business, parent_sku__item_id=parent_sku_id).order_by("-date", "-created_at")
        results = [{
            "id":         a.id,
            "quantity":   a.quantity,
            "reason":     a.reason,
            "reason_display": dict(InventoryAdjustment.REASON_CHOICES).get(a.reason, a.reason),
            "notes":      a.notes,
            "date":       str(a.date),
            "created_at": a.created_at.isoformat(),
        } for a in adjs]
        return Response({"results": results, "total": len(results)})

    # POST
    data = request.data
    parent_sku_id = data.get("parent_sku_id", "").strip()
    if not parent_sku_id:
        return Response({"error": "parent_sku_id required"}, status=400)
    try:
        qty = int(data.get("quantity", 0))
    except (TypeError, ValueError):
        return Response({"error": "quantity must be an integer"}, status=400)
    if qty == 0:
        return Response({"error": "quantity cannot be zero"}, status=400)

    parent = ParentItemPrice.objects.filter(item_id=parent_sku_id, business=business).first()
    if not parent:
        return Response({"error": f"Parent SKU '{parent_sku_id}' not found"}, status=404)

    adj = InventoryAdjustment.objects.create(
        business=business,
        parent_sku=parent,
        quantity=qty,
        reason=data.get("reason", "OTHER"),
        notes=data.get("notes", ""),
        date=data.get("date"),
    )
    _log_inventory(
        business,
        "ADJUSTMENT", adj.id, "CREATE",
        f"Manual adjustment: {'+' if adj.quantity >= 0 else ''}{adj.quantity} units of {parent_sku_id} ({adj.reason})",
        parent_sku_id=parent_sku_id,
        quantity_change=adj.quantity,
        metadata={"reason": adj.reason, "notes": adj.notes},
    )
    return Response({
        "id":         adj.id,
        "quantity":   adj.quantity,
        "reason":     adj.reason,
        "reason_display": dict(InventoryAdjustment.REASON_CHOICES).get(adj.reason, adj.reason),
        "notes":      adj.notes,
        "date":       str(adj.date),
        "created_at": adj.created_at.isoformat(),
    }, status=201)


@api_view(["PUT", "DELETE"])
def inventory_adjustment_detail(request, business_id, adj_id):
    """Update or delete a single inventory adjustment."""
    business = get_authorized_business(request, business_id)
    try:
        adj = InventoryAdjustment.objects.get(pk=adj_id, business=business)
    except InventoryAdjustment.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    apname = adj.parent_sku.item_id if adj.parent_sku_id else ""

    if request.method == "PUT":
        data = request.data
        try:
            qty = int(data.get("quantity", adj.quantity))
        except (TypeError, ValueError):
            return Response({"error": "quantity must be an integer"}, status=400)
        if qty == 0:
            return Response({"error": "quantity cannot be zero"}, status=400)
        adj.quantity = qty
        adj.reason   = data.get("reason", adj.reason)
        adj.notes    = data.get("notes", adj.notes)
        if data.get("date"):
            adj.date = data["date"]
        adj.save()
        _log_inventory(
            business,
            "ADJUSTMENT", adj.id, "UPDATE",
            f"Updated adjustment for {apname}: qty→{adj.quantity} reason={adj.reason}",
            parent_sku_id=apname,
            quantity_change=adj.quantity,
        )
        return Response({
            "id":           adj.id,
            "quantity":     adj.quantity,
            "reason":       adj.reason,
            "reason_display": dict(InventoryAdjustment.REASON_CHOICES).get(adj.reason, adj.reason),
            "notes":        adj.notes,
            "date":         str(adj.date),
            "created_at":   adj.created_at.isoformat(),
        })

    # DELETE
    _log_inventory(
        business,
        "ADJUSTMENT", adj.id, "DELETE",
        f"Deleted adjustment: {apname} {adj.quantity} ({adj.reason})",
        parent_sku_id=apname,
        quantity_change=-adj.quantity,
    )
    adj.delete()
    return Response({"message": "Deleted"})


@api_view(["POST"])
def inventory_create_sku(request, business_id):
    """Create a new Parent SKU directly from the inventory screen."""
    business = get_authorized_business(request, business_id)
    data = request.data
    sku_id = (data.get("sku_id") or "").strip()
    if not sku_id:
        return Response({"error": "sku_id is required"}, status=400)
    if ParentItemPrice.objects.filter(item_id=sku_id, business=business).exists():
        return Response({"error": f"Parent SKU '{sku_id}' already exists"}, status=409)

    with transaction.atomic():
        parent = ParentItemPrice.objects.create(
            business=business,
            item_id=sku_id,
            item_price=data.get("item_price") or None,
            tax_percent=data.get("tax_percent") or None,
            packaging_cost=data.get("packaging_cost") or None,
            final_price=data.get("final_price") or None,
        )
        _log_inventory(business, "SKU", parent.item_id, "CREATE", f"Created Parent SKU: {parent.item_id}", parent_sku_id=parent.item_id)
        # Optional: add initial stock as a purchase bill
        init_qty = int(data.get("initial_qty") or 0)
        if init_qty > 0:
            price_per_unit = Decimal(str(data.get("initial_price") or "0"))
            bill = PurchaseBill.objects.create(
                business=business,
                date=data.get("initial_date") or timezone.now().date(),
                seller_name=data.get("initial_seller") or "Opening Stock",
                bill_number="",
                notes="Initial stock — created from Inventory",
            )
            PurchaseItem.objects.create(
                business=business,
                bill=bill,
                parent_sku=parent,
                product_description="Opening stock",
                quantity=init_qty,
                price_per_unit=price_per_unit,
                is_exchange=False,
            )

    return Response({"sku_id": parent.item_id, "message": "Created"}, status=201)


# ── Inventory Labels / Barcodes & Packed Stock ────────────────────────────────

@api_view(["GET"])
def inventory_labels_list(request, business_id):
    """
    Parent SKUs available for label/barcode printing. The barcode encodes the
    SKU id itself (no separate code system) — this just lists SKUs plus their
    current packed-stock count so the tab can show what's already labelled.
    """
    business = get_authorized_business(request, business_id)
    search = request.GET.get("search", "").strip()
    qs = ParentItemPrice.objects.filter(business=business)
    if search:
        qs = qs.filter(item_id__icontains=search)
    parents = list(qs.values("id", "item_id").order_by("item_id"))
    parent_ids = [p["id"] for p in parents]

    packed_by_parent = dict(
        PackedStockEvent.objects.filter(business=business, parent_sku_id__in=parent_ids)
        .values("parent_sku_id").annotate(total=Sum("quantity"))
        .values_list("parent_sku_id", "total")
    )

    results = [{
        "sku_id":       p["item_id"],
        "code":         p["item_id"],
        "packed_stock": packed_by_parent.get(p["id"], 0) or 0,
    } for p in parents]

    return Response({"results": results, "total": len(results)})


@api_view(["GET", "POST"])
def packed_stock_list(request, business_id):
    """
    GET: current packed-stock total per parent SKU + recent scan/entry events.
    POST: record one scan/entry event against a SKU's printed barcode —
    {code, quantity, notes}. quantity may be negative to correct a mistake.
    """
    business = get_authorized_business(request, business_id)

    if request.method == "GET":
        totals = (
            PackedStockEvent.objects.filter(business=business)
            .values("parent_sku__item_id").annotate(total=Sum("quantity"))
        )
        events = list(
            PackedStockEvent.objects.filter(business=business)
            .select_related("parent_sku").order_by("-created_at")[:50]
        )
        return Response({
            "totals": [
                {"sku_id": t["parent_sku__item_id"], "packed_stock": t["total"]}
                for t in totals if t["total"]
            ],
            "events": [{
                "id":         e.id,
                "sku_id":     e.parent_sku.item_id,
                "quantity":   e.quantity,
                "notes":      e.notes,
                "created_at": e.created_at.isoformat(),
            } for e in events],
        })

    data = request.data
    code = (data.get("code") or "").strip()
    if not code:
        return Response({"error": "code is required"}, status=400)
    try:
        qty = int(data.get("quantity", 1))
    except (TypeError, ValueError):
        return Response({"error": "quantity must be an integer"}, status=400)
    if qty == 0:
        return Response({"error": "quantity cannot be zero"}, status=400)

    parent = ParentItemPrice.objects.filter(business=business, item_id=code).first()
    if not parent:
        return Response({"error": f"No SKU found for code '{code}'"}, status=404)

    event = PackedStockEvent.objects.create(
        business=business, parent_sku=parent, quantity=qty,
        code_scanned=code, notes=(data.get("notes") or "").strip(),
    )
    _log_inventory(
        business, "PACKED_STOCK", event.id, "CREATE",
        f"{'Packed' if qty > 0 else 'Removed'} {abs(qty)} units of {parent.item_id}",
        parent_sku_id=parent.item_id, quantity_change=qty,
    )
    total = (
        PackedStockEvent.objects.filter(business=business, parent_sku=parent)
        .aggregate(t=Sum("quantity"))["t"] or 0
    )
    return Response({
        "id": event.id, "sku_id": parent.item_id, "quantity": qty,
        "packed_stock": total, "created_at": event.created_at.isoformat(),
    }, status=201)


@api_view(["DELETE"])
def packed_stock_detail(request, business_id, event_id):
    """Undo a scan/entry event (deletes it and reverses the running total)."""
    business = get_authorized_business(request, business_id)
    try:
        event = PackedStockEvent.objects.get(pk=event_id, business=business)
    except PackedStockEvent.DoesNotExist:
        return Response({"error": "Not found"}, status=404)
    sku_id = event.parent_sku.item_id
    event.delete()
    _log_inventory(
        business, "PACKED_STOCK", event_id, "DELETE",
        f"Removed packed-stock entry for {sku_id}", parent_sku_id=sku_id,
    )
    return Response(status=204)


# ── Purchase item-level CRUD ──────────────────────────────────────────────────

@api_view(["GET"])
def purchase_sku_items(request, business_id):
    """All PurchaseItems for a given parent_sku, with bill metadata."""
    business = get_authorized_business(request, business_id)
    parent_sku = request.GET.get("parent_sku", "").strip()
    if not parent_sku:
        return Response({"error": "parent_sku required"}, status=400)
    items = (
        PurchaseItem.objects
        .filter(business=business, parent_sku__item_id=parent_sku)
        .select_related("bill")
        .order_by("-bill__date", "-bill__id")
    )
    results = []
    for it in items:
        results.append({
            "id":                  it.id,
            "bill_id":             it.bill_id,
            "bill_number":         it.bill.bill_number or f"#{it.bill_id}",
            "bill_date":           str(it.bill.date),
            "seller_name":         it.bill.seller_name,
            "product_description": it.product_description,
            "quantity":            it.quantity,
            "price_per_unit":      str(it.price_per_unit),
            "is_exchange":         it.is_exchange,
            "total_amount":        str(it.quantity * it.price_per_unit if not it.is_exchange else 0),
        })
    return Response({"results": results, "total": len(results)})


@api_view(["PUT", "DELETE"])
def purchase_item_detail(request, business_id, item_id):
    business = get_authorized_business(request, business_id)
    try:
        item = PurchaseItem.objects.select_related("bill").get(pk=item_id, business=business)
    except PurchaseItem.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    pname = item.parent_sku.item_id if item.parent_sku_id else ""

    if request.method == "PUT":
        data = request.data
        item.product_description = data.get("product_description", item.product_description)
        item.quantity            = int(data.get("quantity", item.quantity))
        item.price_per_unit      = Decimal(str(data.get("price_per_unit", item.price_per_unit)))
        item.is_exchange         = bool(data.get("is_exchange", item.is_exchange))
        item.save()
        _log_inventory(
            business,
            "PURCHASE", item.id, "UPDATE",
            f"Updated purchase item: {pname} qty→{item.quantity} price→{item.price_per_unit}",
            parent_sku_id=pname,
            quantity_change=item.quantity,
        )
        return Response({
            "id":             item.id,
            "quantity":       item.quantity,
            "price_per_unit": str(item.price_per_unit),
            "is_exchange":    item.is_exchange,
            "total_amount":   str(item.quantity * item.price_per_unit if not item.is_exchange else 0),
        })

    # DELETE
    bill = item.bill
    _log_inventory(
        business,
        "PURCHASE", item.id, "DELETE",
        f"Deleted purchase item: {pname} ×{item.quantity}",
        parent_sku_id=pname,
        quantity_change=-item.quantity,
    )
    item.delete()
    # If no items left on the bill, delete the bill too
    if not bill.items.exists():
        bill.delete()
    return Response({"message": "Item deleted"})


@api_view(["GET"])
def purchase_sku_monthly(request, business_id):
    """Monthly purchase aggregates for a parent SKU."""
    business = get_authorized_business(request, business_id)
    parent_sku = request.GET.get("parent_sku", "").strip()
    if not parent_sku:
        return Response({"error": "parent_sku required"}, status=400)
    from django.db.models.functions import TruncMonth
    from django.db.models import ExpressionWrapper as EW, F as Fld, DecimalField as DField
    rows = (
        PurchaseItem.objects
        .filter(business=business, parent_sku__item_id=parent_sku, is_exchange=False)
        .annotate(month=TruncMonth("bill__date"))
        .values("month")
        .annotate(
            total_qty=Sum("quantity"),
            total_value=Sum(
                EW(Fld("quantity") * Fld("price_per_unit"), output_field=DField(max_digits=14, decimal_places=2))
            ),
            bill_count=Count("bill_id", distinct=True),
        )
        .order_by("month")
    )
    results = [{
        "month":       r["month"].strftime("%Y-%m") if r["month"] else "",
        "label":       r["month"].strftime("%b %Y") if r["month"] else "",
        "total_qty":   r["total_qty"],
        "total_value": str(r["total_value"] or 0),
        "bill_count":  r["bill_count"],
    } for r in rows]
    return Response({"results": results, "sku_id": parent_sku})


# ── Consumable Items ──────────────────────────────────────────────────────────

@api_view(["GET", "POST"])
def consumable_items_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        items = ConsumableItem.objects.filter(business=business)
        results = []
        for item in items:
            purchased  = item.purchases.aggregate(t=Sum("quantity"))["t"] or Decimal("0")
            used       = item.usages.filter(event_type__in=["USE", "WASTE"]).aggregate(t=Sum("quantity"))["t"] or Decimal("0")
            total_spend = item.purchases.aggregate(
                t=Sum(ExpressionWrapper(F("quantity") * F("price_per_unit"), output_field=DjDecimalField(max_digits=14, decimal_places=2)))
            )["t"] or Decimal("0")
            last_purchase = item.purchases.aggregate(last=Max("date"))["last"]
            last_usage    = item.usages.aggregate(last=Max("date"))["last"]
            results.append({
                "id":               item.id,
                "name":             item.name,
                "category":         item.category,
                "category_display": item.get_category_display(),
                "unit":             item.unit,
                "notes":            item.notes,
                "purchased_qty":    float(purchased),
                "used_qty":         float(used),
                "current_stock":    float(purchased - used),
                "total_spend":      str(total_spend),
                "last_purchase":    str(last_purchase) if last_purchase else "",
                "last_usage":       str(last_usage) if last_usage else "",
            })
        return Response({"results": results, "total": len(results)})

    data = request.data
    if not data.get("name"):
        return Response({"error": "name is required"}, status=400)
    item = ConsumableItem.objects.create(
        business=business,
        name=data["name"],
        category=data.get("category", "OTHER"),
        unit=data.get("unit", "pieces"),
        notes=data.get("notes", ""),
    )
    _log_inventory(business, "CONSUMABLE_ITEM", item.id, "CREATE", f"Created consumable item: {item.name}")
    return Response({"id": item.id, "name": item.name, "category": item.category, "unit": item.unit}, status=201)


@api_view(["GET", "PUT", "DELETE"])
def consumable_item_detail(request, business_id, item_id):
    business = get_authorized_business(request, business_id)
    try:
        item = ConsumableItem.objects.get(pk=item_id, business=business)
    except ConsumableItem.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "GET":
        purchases = [{"id": p.id, "date": str(p.date), "quantity": float(p.quantity), "price_per_unit": str(p.price_per_unit), "total": str(p.quantity * p.price_per_unit), "seller_name": p.seller_name, "notes": p.notes} for p in item.purchases.order_by("-date")]
        usages    = [{"id": u.id, "date": str(u.date), "event_type": u.event_type, "event_display": u.get_event_type_display(), "quantity": float(u.quantity), "notes": u.notes} for u in item.usages.order_by("-date")]
        return Response({"id": item.id, "name": item.name, "category": item.category, "unit": item.unit, "notes": item.notes, "purchases": purchases, "usages": usages})

    if request.method == "PUT":
        old_name = item.name
        item.name     = request.data.get("name", item.name)
        item.category = request.data.get("category", item.category)
        item.unit     = request.data.get("unit", item.unit)
        item.notes    = request.data.get("notes", item.notes)
        item.save()
        _log_inventory(business, "CONSUMABLE_ITEM", item.id, "UPDATE", f"Updated consumable: {old_name} → {item.name}")
        return Response({"id": item.id, "name": item.name, "category": item.category, "unit": item.unit})

    name = item.name
    item.delete()
    _log_inventory(business, "CONSUMABLE_ITEM", item_id, "DELETE", f"Deleted consumable item: {name}")
    return Response({"message": "Deleted"})


@api_view(["GET", "POST"])
def consumable_purchases_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        item_id = request.GET.get("item_id")
        qs = ConsumablePurchase.objects.filter(business=business).select_related("item").order_by("-date")
        if item_id:
            qs = qs.filter(item_id=item_id)
        results = [{
            "id": p.id, "item_id": p.item_id, "item_name": p.item.name, "item_unit": p.item.unit,
            "date": str(p.date), "quantity": float(p.quantity),
            "price_per_unit": str(p.price_per_unit),
            "total_amount": str(p.quantity * p.price_per_unit),
            "seller_name": p.seller_name, "notes": p.notes,
        } for p in qs]
        return Response({"results": results})

    data = request.data
    try:
        item = ConsumableItem.objects.get(pk=data.get("item_id"), business=business)
    except ConsumableItem.DoesNotExist:
        return Response({"error": "Consumable item not found"}, status=404)

    p = ConsumablePurchase.objects.create(
        business=business,
        item=item,
        date=data["date"],
        quantity=Decimal(str(data.get("quantity", 1))),
        price_per_unit=Decimal(str(data.get("price_per_unit", 0))),
        seller_name=data.get("seller_name", ""),
        notes=data.get("notes", ""),
    )
    _log_inventory(
        business,
        "CONSUMABLE_PURCHASE", p.id, "CREATE",
        f"Purchased {float(p.quantity)} {item.unit} of {item.name} from {p.seller_name or 'unknown'}",
        quantity_change=int(p.quantity),
        metadata={"item_id": item.id, "item_name": item.name, "quantity": float(p.quantity), "price_per_unit": str(p.price_per_unit)},
    )
    return Response({"id": p.id, "item_id": p.item_id, "date": str(p.date), "quantity": float(p.quantity), "price_per_unit": str(p.price_per_unit)}, status=201)


@api_view(["PUT", "DELETE"])
def consumable_purchase_detail(request, business_id, purchase_id):
    business = get_authorized_business(request, business_id)
    try:
        p = ConsumablePurchase.objects.select_related("item").get(pk=purchase_id, business=business)
    except ConsumablePurchase.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "PUT":
        old_qty = float(p.quantity)
        data = request.data
        p.date           = data.get("date", p.date)
        p.quantity       = Decimal(str(data.get("quantity", p.quantity)))
        p.price_per_unit = Decimal(str(data.get("price_per_unit", p.price_per_unit)))
        p.seller_name    = data.get("seller_name", p.seller_name)
        p.notes          = data.get("notes", p.notes)
        p.save()
        _log_inventory(business, "CONSUMABLE_PURCHASE", p.id, "UPDATE",
            f"Updated {p.item.name} purchase: qty {old_qty} → {float(p.quantity)}", quantity_change=int(p.quantity) - int(old_qty))
        return Response({"id": p.id, "quantity": float(p.quantity), "price_per_unit": str(p.price_per_unit)})

    item_name = p.item.name
    qty = float(p.quantity)
    p.delete()
    _log_inventory(business, "CONSUMABLE_PURCHASE", purchase_id, "DELETE",
        f"Deleted purchase of {qty} {item_name}", quantity_change=-int(qty))
    return Response({"message": "Deleted"})


@api_view(["GET", "POST"])
def consumable_usages_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        item_id = request.GET.get("item_id")
        qs = ConsumableUsage.objects.filter(business=business).select_related("item").order_by("-date")
        if item_id:
            qs = qs.filter(item_id=item_id)
        results = [{
            "id": u.id, "item_id": u.item_id, "item_name": u.item.name, "item_unit": u.item.unit,
            "date": str(u.date), "event_type": u.event_type,
            "event_display": u.get_event_type_display(),
            "quantity": float(u.quantity), "notes": u.notes,
        } for u in qs]
        return Response({"results": results})

    data = request.data
    try:
        item = ConsumableItem.objects.get(pk=data.get("item_id"), business=business)
    except ConsumableItem.DoesNotExist:
        return Response({"error": "Consumable item not found"}, status=404)

    event_type = data.get("event_type", "USE")
    qty = Decimal(str(data.get("quantity", 1)))
    u = ConsumableUsage.objects.create(
        business=business,
        item=item, date=data["date"], event_type=event_type, quantity=qty,
        notes=data.get("notes", ""),
    )
    desc_map = {
        "USE":   f"Used {float(qty)} {item.unit} of {item.name}",
        "OPEN":  f"Opened new package of {item.name} ({float(qty)} {item.unit})",
        "WASTE": f"Wasted/damaged {float(qty)} {item.unit} of {item.name}",
    }
    _log_inventory(business, "CONSUMABLE_USAGE", u.id, "CREATE",
        desc_map.get(event_type, f"Used {float(qty)} of {item.name}"),
        quantity_change=-int(qty) if event_type in ["USE", "WASTE"] else 0,
        metadata={"item_id": item.id, "item_name": item.name, "event_type": event_type, "quantity": float(qty)})
    return Response({"id": u.id, "item_id": u.item_id, "date": str(u.date), "event_type": u.event_type, "quantity": float(u.quantity)}, status=201)


@api_view(["DELETE"])
def consumable_usage_detail(request, business_id, usage_id):
    business = get_authorized_business(request, business_id)
    try:
        u = ConsumableUsage.objects.select_related("item").get(pk=usage_id, business=business)
    except ConsumableUsage.DoesNotExist:
        return Response({"error": "Not found"}, status=404)
    item_name, qty, event_type = u.item.name, float(u.quantity), u.event_type
    u.delete()
    _log_inventory(business, "CONSUMABLE_USAGE", usage_id, "DELETE",
        f"Deleted usage record: {qty} of {item_name} ({event_type})")
    return Response({"message": "Deleted"})


# ── Inventory Audit Log ───────────────────────────────────────────────────────

@api_view(["GET"])
def inventory_logs_list(request, business_id):
    business = get_authorized_business(request, business_id)
    qs = InventoryLog.objects.filter(business=business)
    parent_sku  = request.GET.get("parent_sku",  "").strip()
    entity_type = request.GET.get("entity_type", "").strip()
    action      = request.GET.get("action",      "").strip()
    date_from   = request.GET.get("date_from",   "").strip()
    date_to     = request.GET.get("date_to",     "").strip()

    if parent_sku:  qs = qs.filter(parent_sku_id=parent_sku)
    if entity_type: qs = qs.filter(entity_type=entity_type)
    if action:      qs = qs.filter(action=action)
    if date_from:   qs = qs.filter(created_at__date__gte=date_from)
    if date_to:     qs = qs.filter(created_at__date__lte=date_to)

    page      = max(1, int(request.GET.get("page", 1)))
    page_size = max(1, min(200, int(request.GET.get("page_size", 50))))
    total     = qs.count()
    offset    = (page - 1) * page_size

    results = [{
        "id":              log.id,
        "entity_type":     log.entity_type,
        "entity_id":       log.entity_id,
        "action":          log.action,
        "parent_sku_id":   log.parent_sku_id,
        "quantity_change": log.quantity_change,
        "description":     log.description,
        "metadata":        log.metadata,
        "created_at":      log.created_at.isoformat(),
    } for log in qs[offset:offset + page_size]]

    return Response({"results": results, "total": total, "page": page, "page_size": page_size})


# ── Inventory Charts ──────────────────────────────────────────────────────────

@api_view(["GET"])
def inventory_charts(request, business_id):
    """Aggregated data for inventory dashboard charts."""
    import datetime
    from django.db.models.functions import TruncMonth
    from django.db.models import ExpressionWrapper as EW, F as Fld, DecimalField as DField2

    business = get_authorized_business(request, business_id)

    # ── 1. Current stock by SKU ───────────────────────────────────────────────
    purchase_agg = (
        PurchaseItem.objects.filter(business=business, is_exchange=False, parent_sku__isnull=False)
        .values("parent_sku_id")
        .annotate(
            purchased_qty=Sum("quantity"),
            purchase_value=Sum(EW(Fld("quantity") * Fld("price_per_unit"), output_field=DField2(max_digits=14, decimal_places=2))),
        )
    )
    purchased_by_parent = {r["parent_sku_id"]: {"qty": r["purchased_qty"], "value": float(r["purchase_value"] or 0)} for r in purchase_agg}

    adj_agg = {r["parent_sku_id"]: r["net_adj"] for r in InventoryAdjustment.objects.filter(business=business).values("parent_sku_id").annotate(net_adj=Sum("quantity"))}

    all_parents = set(purchased_by_parent.keys()) | set(adj_agg.keys())
    parent_name = dict(ParentItemPrice.objects.filter(business=business, id__in=all_parents).values_list("id", "item_id"))
    sku_to_parent = _SkuMap(FinalPrice.objects.filter(business=business, parent_id__in=all_parents).values_list("sku_id", "parent_id"))
    child_skus = list(sku_to_parent.keys())

    del_by_sku = dict(Order.objects.filter(business=business, reason_for_credit_entry="DELIVERED", sku__in=child_skus).values("sku").annotate(q=Sum("quantity")).values_list("sku", "q"))
    rto_by_sku = dict(Order.objects.filter(business=business, reason_for_credit_entry="RTO_COMPLETE", sku__in=child_skus).values("sku").annotate(q=Sum("quantity")).values_list("sku", "q"))

    sold_by_parent = {}
    rto_by_parent  = {}
    for sku, qty in del_by_sku.items():
        p = sku_to_parent.get(sku)
        if p: sold_by_parent[p] = sold_by_parent.get(p, 0) + qty
    for sku, qty in rto_by_sku.items():
        p = sku_to_parent.get(sku)
        if p: rto_by_parent[p] = rto_by_parent.get(p, 0) + qty

    stock_by_sku = []
    for parent_id in all_parents:
        pdata   = purchased_by_parent.get(parent_id, {"qty": 0, "value": 0})
        current = pdata["qty"] - sold_by_parent.get(parent_id, 0) + rto_by_parent.get(parent_id, 0) + adj_agg.get(parent_id, 0)
        status  = "out" if current <= 0 else "low" if current <= 3 else "instock"
        stock_by_sku.append({"sku_id": parent_name.get(parent_id, str(parent_id)), "current_stock": current, "purchase_value": pdata["value"], "status": status})
    stock_by_sku.sort(key=lambda x: -x["current_stock"])

    stock_status = {
        "instock": sum(1 for s in stock_by_sku if s["status"] == "instock"),
        "low":     sum(1 for s in stock_by_sku if s["status"] == "low"),
        "out":     sum(1 for s in stock_by_sku if s["status"] == "out"),
    }

    # ── 2. Monthly purchases last 12 months ───────────────────────────────────
    twelve_ago = datetime.date.today() - datetime.timedelta(days=365)
    monthly_qs = (
        PurchaseItem.objects.filter(business=business, is_exchange=False, bill__date__gte=twelve_ago)
        .annotate(month=TruncMonth("bill__date"))
        .values("month")
        .annotate(total_qty=Sum("quantity"), total_value=Sum(EW(Fld("quantity") * Fld("price_per_unit"), output_field=DField2(max_digits=14, decimal_places=2))), bill_count=Count("bill_id", distinct=True))
        .order_by("month")
    )
    monthly_purchases = [{"month": r["month"].strftime("%Y-%m"), "label": r["month"].strftime("%b %y"), "total_qty": r["total_qty"], "total_value": float(r["total_value"] or 0), "bill_count": r["bill_count"]} for r in monthly_qs if r["month"]]

    # ── 3. Consumable monthly spend last 6 months ─────────────────────────────
    six_ago = datetime.date.today() - datetime.timedelta(days=180)
    consumable_monthly = []
    if ConsumableItem.objects.filter(business=business).exists():
        cons_qs = (
            ConsumablePurchase.objects.filter(business=business, date__gte=six_ago)
            .annotate(month=TruncMonth("date"))
            .values("month", "item__category")
            .annotate(total_spend=Sum(EW(Fld("quantity") * Fld("price_per_unit"), output_field=DField2(max_digits=14, decimal_places=2))))
            .order_by("month")
        )
        by_month = {}
        for r in cons_qs:
            key = r["month"].strftime("%Y-%m") if r["month"] else ""
            if not key: continue
            if key not in by_month:
                by_month[key] = {"month": key, "label": r["month"].strftime("%b %y"), "total": 0}
            by_month[key][r["item__category"]] = float(r["total_spend"] or 0)
            by_month[key]["total"] += float(r["total_spend"] or 0)
        consumable_monthly = sorted(by_month.values(), key=lambda x: x["month"])

    # ── 4. Top SKUs by purchase value ─────────────────────────────────────────
    top_by_value = sorted(stock_by_sku, key=lambda x: -x["purchase_value"])[:10]

    return Response({
        "stock_by_sku":       stock_by_sku[:20],
        "top_by_value":       top_by_value,
        "stock_status":       stock_status,
        "monthly_purchases":  monthly_purchases,
        "consumable_monthly": consumable_monthly,
        "total_skus":         len(stock_by_sku),
        "total_stock":        sum(s["current_stock"] for s in stock_by_sku),
        "total_value":        sum(s["purchase_value"] for s in stock_by_sku),
    })


# ── Label packing ──────────────────────────────────────────────────────────────

@api_view(["PATCH"])
def label_order_pack(request, business_id, order_id):
    """Toggle or set is_packed on a LabelOrder."""
    business = get_authorized_business(request, business_id)
    try:
        lo = LabelOrder.objects.get(pk=order_id, business=business)
    except LabelOrder.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    force = request.data.get("packed")   # None = toggle, True/False = set
    if force is None:
        lo.is_packed = not lo.is_packed
    else:
        lo.is_packed = bool(force)

    lo.packed_at = timezone.now() if lo.is_packed else None
    lo.save(update_fields=["is_packed", "packed_at"])
    return Response({"order_id": order_id, "is_packed": lo.is_packed})


@api_view(["POST"])
def label_bulk_pack(request, business_id):
    """Mark a list of order_ids as packed=True/False."""
    business = get_authorized_business(request, business_id)
    order_ids = request.data.get("order_ids", [])
    packed    = bool(request.data.get("packed", True))
    packed_at = timezone.now() if packed else None
    updated   = LabelOrder.objects.filter(business=business, order_id__in=order_ids).update(
        is_packed=packed, packed_at=packed_at
    )
    return Response({"updated": updated, "packed": packed})


@api_view(["GET"])
def label_unpacked(request, business_id):
    """List unpacked LabelOrders, optionally filtered by uploaded_date."""
    business = get_authorized_business(request, business_id)
    qs = LabelOrder.objects.filter(business=business, is_packed=False)
    date = request.GET.get("date", "").strip()
    if date:
        qs = qs.filter(uploaded_date=date)
    results = list(qs.values(
        "order_id", "customer_name", "customer_pincode",
        "customer_city", "customer_state",
        "sku", "qty", "courier_name", "uploaded_date", "order_date",
    ))
    return Response({"results": results, "total": len(results)})


# ── Fraud Customers & Blocked Customers ──────────────────────────────────────

def _risk_level(return_rate, return_count):
    """Return 'high' / 'medium' / 'low' based on RETURN behaviour (not RTO)."""
    if return_rate >= 0.5 or return_count >= 3:
        return "high"
    if return_rate >= 0.25 or return_count >= 2:
        return "medium"
    return "low"


@api_view(["GET"])
def fraud_customers(request, business_id):
    """
    Per-customer analysis: identity = customer_name + customer_pincode.
    Groups orders by same name + same address, returns full order history,
    per-SKU breakdown, and return-based risk level.
    """
    business = get_authorized_business(request, business_id)
    min_orders = int(request.GET.get("min_orders", 2))

    # 1. Load every LabelOrder row with full detail in one query
    all_label_rows = list(
        LabelOrder.objects
        .filter(business=business)
        .exclude(customer_name="")
        .values(
            "order_id", "customer_name", "customer_pincode",
            "customer_address", "customer_city", "customer_state",
            "sku", "qty", "order_date",
        )
    )
    if not all_label_rows:
        return Response({"results": [], "total": 0})

    # 2. Group by (name, pincode); build per-order detail list
    customer_orders = {}
    all_order_ids   = []
    for lo in all_label_rows:
        key = (lo["customer_name"], lo["customer_pincode"])
        customer_orders.setdefault(key, []).append({
            "order_id":   lo["order_id"],
            "sku":        lo["sku"] or "",
            "qty":        lo["qty"] or 1,
            "order_date": str(lo["order_date"]) if lo["order_date"] else "",
            "address":    lo["customer_address"] or "",
            "city":       lo["customer_city"] or "",
            "state":      lo["customer_state"] or "",
            "status":     "PENDING",
        })
        all_order_ids.append(lo["order_id"])

    # 3. Outcomes from Order table  (DELIVERED / RETURN / RETURNED / RTO_COMPLETE / CANCELLED)
    outcome_map = {}
    for row in Order.objects.filter(business=business, sub_order_no__in=all_order_ids).values("sub_order_no", "reason_for_credit_entry"):
        outcome_map.setdefault(row["sub_order_no"], []).append(row["reason_for_credit_entry"])

    # 4. Claims lookup
    claimed_ids = set(
        OrderPayment.objects
        .filter(business=business, sub_order_no__in=all_order_ids, claims__isnull=False)
        .exclude(claims=0)
        .values_list("sub_order_no", flat=True)
        .distinct()
    )

    # 5. Blocked lookup
    blocked_set = set(
        BlockedCustomer.objects.filter(business=business, is_active=True)
        .values_list("customer_name", "customer_pincode")
    )

    RETURN_STATUSES = {"RETURN", "RETURNED"}

    results = []
    for (name, pincode), orders in customer_orders.items():
        # Deduplicate order_ids (same order can appear twice if multi-item)
        seen_ids: dict = {}
        for o in orders:
            oid = o["order_id"]
            if oid not in seen_ids:
                seen_ids[oid] = o
        unique_orders = list(seen_ids.values())
        if len(unique_orders) < min_orders:
            continue

        # Resolve statuses
        enriched = []
        for o in unique_orders:
            statuses = outcome_map.get(o["order_id"], [])
            if "DELIVERED" in statuses:
                resolved = "DELIVERED"
            elif any(s in RETURN_STATUSES for s in statuses):
                resolved = "RETURN"
            elif "RTO_COMPLETE" in statuses:
                resolved = "RTO"
            elif "CANCELLED" in statuses:
                resolved = "CANCELLED"
            else:
                resolved = "PENDING"
            enriched.append({**o, "status": resolved})

        enriched.sort(key=lambda x: x["order_date"], reverse=True)

        delivered = sum(1 for o in enriched if o["status"] == "DELIVERED")
        returned  = sum(1 for o in enriched if o["status"] == "RETURN")
        rto       = sum(1 for o in enriched if o["status"] == "RTO")
        cancelled = sum(1 for o in enriched if o["status"] == "CANCELLED")
        pending   = sum(1 for o in enriched if o["status"] == "PENDING")
        total     = len(enriched)
        claim_count = sum(1 for o in enriched if o["order_id"] in claimed_ids)

        return_rate = round(returned / total, 3) if total > 0 else 0.0
        rto_rate    = round(rto    / total, 3) if total > 0 else 0.0

        # Per-SKU breakdown
        sku_stats: dict = {}
        for o in enriched:
            sku = o["sku"] or "Unknown"
            s   = sku_stats.setdefault(sku, {"sku": sku, "orders": 0, "qty": 0,
                                              "delivered": 0, "returned": 0, "rto": 0})
            s["orders"] += 1
            s["qty"]    += int(o["qty"] or 1)
            if o["status"] == "DELIVERED": s["delivered"] += 1
            elif o["status"] == "RETURN":  s["returned"]  += 1
            elif o["status"] == "RTO":     s["rto"]       += 1

        sample = enriched[0] if enriched else {}
        results.append({
            "customer_name":    name,
            "customer_pincode": pincode,
            "customer_address": sample.get("address", ""),
            "customer_city":    sample.get("city", ""),
            "customer_state":   sample.get("state", ""),
            "total_orders":     total,
            "delivered":        delivered,
            "returned":         returned,
            "rto":              rto,
            "cancelled":        cancelled,
            "pending":          pending,
            "return_rate":      return_rate,
            "rto_rate":         rto_rate,
            "claim_count":      claim_count,
            "risk_level":       _risk_level(return_rate, returned),
            "last_order":       enriched[0]["order_date"] if enriched else "",
            "is_blocked":       (name, pincode) in blocked_set,
            "sku_breakdown":    sorted(sku_stats.values(), key=lambda x: -x["orders"]),
            "orders":           enriched[:60],
        })

    order_map = {"high": 0, "medium": 1, "low": 2}
    results.sort(key=lambda r: (order_map[r["risk_level"]], -r["return_rate"], -r["returned"]))

    risk_filter = request.GET.get("risk", "")
    if risk_filter in ("high", "medium", "low"):
        results = [r for r in results if r["risk_level"] == risk_filter]

    return Response({"results": results, "total": len(results)})


@api_view(["GET", "POST"])
def blocked_customers_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "GET":
        qs = BlockedCustomer.objects.filter(business=business, is_active=True).order_by("-blocked_at")
        results = [{
            "id":               bc.id,
            "id":               bc.id,
            "customer_name":    bc.customer_name,
            "customer_pincode": bc.customer_pincode,
            "customer_city":    bc.customer_city,
            "customer_state":   bc.customer_state,
            "customer_address": getattr(bc, "customer_address", ""),
            "reason":           bc.reason,
            "blocked_at":       bc.blocked_at.isoformat(),
        } for bc in qs]
        return Response({"results": results, "total": len(results)})

    # POST — block a customer
    data = request.data
    name    = (data.get("customer_name") or "").strip()
    pincode = (data.get("customer_pincode") or "").strip()
    if not name or not pincode:
        return Response({"error": "customer_name and customer_pincode are required."}, status=400)

    bc, created = BlockedCustomer.objects.update_or_create(
        business=business,
        customer_name=name,
        customer_pincode=pincode,
        defaults={
            "customer_city":  data.get("customer_city", ""),
            "customer_state": data.get("customer_state", ""),
            "reason":         data.get("reason", ""),
            "is_active":      True,
        },
    )
    return Response({
        "id":            bc.id,
        "customer_name": bc.customer_name,
        "created":       created,
    }, status=201 if created else 200)


@api_view(["DELETE", "PATCH"])
def blocked_customer_detail(request, business_id, bc_id):
    business = get_authorized_business(request, business_id)
    try:
        bc = BlockedCustomer.objects.get(pk=bc_id, business=business)
    except BlockedCustomer.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "DELETE":
        # Soft-unblock
        bc.is_active = False
        bc.save()
        return Response({"message": "Unblocked"})

    # PATCH — update reason
    bc.reason = request.data.get("reason", bc.reason)
    bc.save()
    return Response({"id": bc.id, "reason": bc.reason})


@api_view(["DELETE"])
def inventory_delete_sku(request, business_id, sku_id):
    """Delete all purchase items for a parent SKU; also delete bills that become empty."""
    business = get_authorized_business(request, business_id)
    items = PurchaseItem.objects.filter(business=business, parent_sku_id=sku_id)
    bill_ids = list(items.values_list("bill_id", flat=True).distinct())
    count = items.count()
    _log_inventory(business, "SKU", sku_id, "DELETE", f"Deleted all {count} purchase entries for {sku_id}", parent_sku_id=sku_id, quantity_change=-count)
    items.delete()
    for bid in bill_ids:
        if not PurchaseItem.objects.filter(bill_id=bid, business=business).exists():
            PurchaseBill.objects.filter(pk=bid, business=business).delete()
    return Response({"deleted": True})


# ── Product Photography AI ────────────────────────────────────────────────────

_PHOTO_STYLES = [
    {"index": 0, "name": "White Studio",    "prompt": "professional product photography, clean white background, soft box studio lighting, crisp shadows, commercial quality, sharp focus"},
    {"index": 1, "name": "Marble Luxury",   "prompt": "product on white marble surface, luxury brand aesthetic, soft diffused natural light, minimalist high-end photography, elegant"},
    {"index": 2, "name": "Dark Dramatic",   "prompt": "product photography, dark charcoal background, dramatic moody cinematic lighting, premium contrast shadows, bold artistic"},
    {"index": 3, "name": "Wooden Table",    "prompt": "product on rustic light wooden table, warm cozy interior, soft golden natural light, lifestyle photography, homely aesthetic"},
    {"index": 4, "name": "Nature Outdoor",  "prompt": "product in lush green outdoor nature setting, bokeh foliage background, golden hour sunlight, fresh organic aesthetic"},
    {"index": 5, "name": "Pastel Flat Lay", "prompt": "product flat lay overhead shot, soft pastel pink background, Instagram aesthetic, minimalist styling, clean composition"},
    {"index": 6, "name": "Navy Premium",    "prompt": "product on deep navy blue background, premium luxury brand aesthetic, subtle gradient, professional commercial shot, rich tones"},
    {"index": 7, "name": "Botanical",       "prompt": "product surrounded by tropical leaves and flowers, botanical garden aesthetic, lush vibrant greens, artistic lifestyle shot"},
    {"index": 8, "name": "Warm Gradient",   "prompt": "product photography, warm orange and amber gradient background, sunset tones, lifestyle brand shot, soft glowing light"},
    {"index": 9, "name": "Minimal Concrete","prompt": "product on smooth concrete surface, Scandinavian minimalist aesthetic, cool neutral tones, editorial photography, clean modern look"},
]


def _resize_for_sdxl(image_bytes):
    """Resize and pad image to 1024x1024 for SDXL img2img."""
    img = Image.open(io.BytesIO(image_bytes)).convert("RGBA")
    img.thumbnail((1024, 1024), Image.LANCZOS)
    bg = Image.new("RGBA", (1024, 1024), (255, 255, 255, 255))
    offset = ((1024 - img.width) // 2, (1024 - img.height) // 2)
    bg.paste(img, offset, img)
    final = bg.convert("RGB")
    buf = io.BytesIO()
    final.save(buf, format="PNG")
    return buf.getvalue()


def _call_stability(api_key, image_bytes, style):
    try:
        resp = http_requests.post(
            "https://api.stability.ai/v1/generation/stable-diffusion-xl-1024-v1-0/image-to-image",
            headers={"Authorization": f"Bearer {api_key}", "Accept": "application/json"},
            files={"init_image": ("product.png", image_bytes, "image/png")},
            data={
                "image_strength": 0.40,
                "text_prompts[0][text]": style["prompt"],
                "text_prompts[0][weight]": 1,
                "text_prompts[1][text]": "blurry, low quality, watermark, text, logo, deformed, ugly, distorted",
                "text_prompts[1][weight]": -1,
                "cfg_scale": 7,
                "steps": 30,
                "samples": 1,
            },
            timeout=120,
        )
        if resp.status_code == 200:
            img_b64 = resp.json()["artifacts"][0]["base64"]
            return {"index": style["index"], "name": style["name"], "image": img_b64, "status": "ok"}
        return {"index": style["index"], "name": style["name"], "status": "error", "error": resp.text[:300]}
    except Exception as exc:
        return {"index": style["index"], "name": style["name"], "status": "error", "error": str(exc)}


@api_view(["POST"])
@parser_classes([MultiPartParser])
def generate_product_images(request, business_id):
    business = get_authorized_business(request, business_id)
    image_file = request.FILES.get("image")
    api_key = request.data.get("api_key", "").strip()

    if not image_file:
        return Response({"error": "No image file provided."}, status=400)
    if not api_key:
        return Response({"error": "Stability AI API key is required."}, status=400)

    try:
        resized_bytes = _resize_for_sdxl(image_file.read())
    except Exception as exc:
        return Response({"error": f"Could not process image: {exc}"}, status=400)

    results = [None] * len(_PHOTO_STYLES)
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as pool:
        future_map = {
            pool.submit(_call_stability, api_key, resized_bytes, style): style["index"]
            for style in _PHOTO_STYLES
        }
        for future in concurrent.futures.as_completed(future_map):
            result = future.result()
            results[result["index"]] = result

    return Response({"results": results})


# ── Meesho Inventory ───────────────────────────────────────────────────────────

_INV_COL_MAP = {
    "serial no":           "serial_no",
    "serial_no":           "serial_no",
    "catalog name":        "catalog_name",
    "catalog_name":        "catalog_name",
    "catalog id":          "catalog_id",
    "catalog_id":          "catalog_id",
    "product name":        "product_name",
    "product_name":        "product_name",
    "product id":          "product_id",
    "product_id":          "product_id",
    "style id":            "style_id",
    "style_id":            "style_id",
    "variation id":        "variation_id",
    "variation_id":        "variation_id",
    "variation":           "variation",
    "stock":               "stock_type",
    "stock_type":          "stock_type",
    "system stock count":  "system_stock_count",
    "system_stock_count":  "system_stock_count",
    "your stock count":    "seller_stock_count",
    "seller_stock_count":  "seller_stock_count",
}

_DESCRIPTION_ROW_MARKERS = {
    "row identifier", "catalog name", "catalog id", "product id/style id",
}


def _safe_inv_int(val, default=None):
    try:
        v = int(float(str(val).strip()))
        return v
    except Exception:
        return default


@api_view(["POST"])
@parser_classes([MultiPartParser])
def meesho_inventory_upload(request, business_id):
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        if file.name.lower().endswith((".xlsx", ".xls")):
            import openpyxl as _opxl
            file.seek(0)
            wb = _opxl.load_workbook(file, read_only=True, data_only=True)
            # Find the sheet whose name contains "fill this" (case-insensitive)
            sheet_name = None
            for name in wb.sheetnames:
                if "fill this" in name.lower():
                    sheet_name = name
                    break
            wb.close()
            file.seek(0)
            df = pd.read_excel(file, sheet_name=sheet_name, dtype=str)
        else:
            df = pd.read_csv(file, dtype=str)
    except Exception as exc:
        return Response({"error": f"Could not read file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

    df.columns = [str(c).strip().lower() for c in df.columns]
    df.rename(columns=_INV_COL_MAP, inplace=True)

    if "serial_no" not in df.columns:
        return Response({"error": "Missing required column: SERIAL NO"}, status=status.HTTP_400_BAD_REQUEST)

    def _is_description_row(row):
        val = str(row.get("serial_no", "")).strip().lower()
        return val in _DESCRIPTION_ROW_MARKERS or not val.lstrip("-").lstrip().replace(".", "").isdigit()

    df = df[~df.apply(_is_description_row, axis=1)].reset_index(drop=True)

    created = updated = skipped = 0
    with transaction.atomic():
        for _, row in df.iterrows():
            sno     = _safe_inv_int(row.get("serial_no"))
            cat_id  = _safe_inv_int(row.get("catalog_id"))
            prod_id = _safe_inv_int(row.get("product_id"))
            if sno is None or cat_id is None or prod_id is None:
                skipped += 1
                continue

            raw_seller = row.get("seller_stock_count", "")
            seller_count = _safe_inv_int(raw_seller) if str(raw_seller).strip() not in ("", "nan") else None

            defaults = {
                "catalog_name":       str(row.get("catalog_name", "")).strip(),
                "catalog_id":         cat_id,
                "product_name":       str(row.get("product_name", "")).strip(),
                "product_id":         prod_id,
                "style_id":           str(row.get("style_id", "")).strip(),
                "variation_id":       _safe_inv_int(row.get("variation_id")),
                "variation":          str(row.get("variation", "")).strip(),
                "stock_type":         str(row.get("stock_type", "ALL")).strip().upper() or "ALL",
                "system_stock_count": _safe_inv_int(row.get("system_stock_count"), 0),
                "seller_stock_count": seller_count,
            }
            _, was_created = MeeshoInventory.objects.update_or_create(
                business=business, serial_no=sno, defaults=defaults,
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return Response(
        {"success": True, "created": created, "updated": updated, "skipped": skipped},
        status=status.HTTP_201_CREATED,
    )


@api_view(["GET", "PATCH"])
def meesho_inventory_list(request, business_id):
    business = get_authorized_business(request, business_id)
    if request.method == "PATCH":
        updates = request.data if isinstance(request.data, list) else []
        count = 0
        with transaction.atomic():
            for item in updates:
                pk  = item.get("id")
                val = item.get("seller_stock_count")
                if pk is None:
                    continue
                MeeshoInventory.objects.filter(pk=pk, business=business).update(
                    seller_stock_count=_safe_inv_int(val)
                )
                count += 1
        return Response({"updated": count})

    qs = MeeshoInventory.objects.filter(business=business).order_by("serial_no")
    if request.GET.get("low_stock") == "1":
        qs = qs.filter(system_stock_count__lt=100)

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(catalog_name__icontains=search) |
            DQ(product_name__icontains=search) |
            DQ(style_id__icontains=search)
        )

    data = list(qs.values(
        "id", "serial_no", "catalog_name", "catalog_id",
        "product_name", "product_id", "style_id",
        "variation_id", "variation", "stock_type",
        "system_stock_count", "seller_stock_count",
    ))
    total           = MeeshoInventory.objects.filter(business=business).count()
    low_stock_count = MeeshoInventory.objects.filter(business=business, system_stock_count__lt=100).count()
    return Response({"items": data, "total": total, "low_stock_count": low_stock_count})


@api_view(["GET"])
def meesho_inventory_download(request, business_id):
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    business = get_authorized_business(request, business_id)
    qs = MeeshoInventory.objects.filter(business=business).order_by("serial_no")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory-Fill this"

    headers = [
        "SERIAL NO", "CATALOG NAME", "CATALOG ID", "PRODUCT NAME",
        "PRODUCT ID", "STYLE ID", "VARIATION ID", "VARIATION",
        "STOCK", "SYSTEM STOCK COUNT", "YOUR STOCK COUNT",
    ]
    desc_row = [
        "Row identifier", "Catalog name", "Catalog id", "Product name",
        "Product id", "Product ID/Style ID", "Variation id", "Variation",
        "Stock type (IN_STOCK / OUT_OF_STOCK / ALL)",
        "Current system stock count",
        "Edit this (keep empty if no change in stock)",
    ]
    ws.append(headers)
    ws.append(desc_row)

    for item in qs:
        # YOUR STOCK COUNT: filled only where the seller has set a new value,
        # empty otherwise so Meesho skips those rows
        your_stock = item.seller_stock_count if item.seller_stock_count is not None else ""
        ws.append([
            item.serial_no,
            item.catalog_name,
            item.catalog_id,
            item.product_name,
            item.product_id,
            item.style_id,
            item.variation_id,
            item.variation,
            item.stock_type,
            item.system_stock_count,
            your_stock,
        ])

    col_widths = [10, 28, 14, 50, 14, 30, 14, 14, 10, 22, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="meesho_inventory_updated.xlsx"'
    return response


# ── Meesho Price Update ────────────────────────────────────────────────────────

@api_view(["GET", "PATCH"])
def meesho_price_update_list(request, business_id):
    """
    GET  — return all inventory items enriched with profit data.
           ?max_profit=<n>  filters to SKUs with avg_profit_per_unit < n
           ?q=<search>      filters by catalog/product/style
    PATCH — bulk-save price updates: [{inventory_id, new_msp, new_wdrp, new_mrp}, ...]
    """
    business = get_authorized_business(request, business_id)
    if request.method == "PATCH":
        updates = request.data if isinstance(request.data, list) else []
        inv_ids = [u.get("inventory_id") for u in updates if u.get("inventory_id") is not None]
        inv_ids = list({int(i) for i in inv_ids})

        inventory_map = {
            inv.id: inv
            for inv in MeeshoInventory.objects.filter(business=business, id__in=inv_ids)
        }
        existing_updates = {
            pu.inventory_id: pu
            for pu in MeeshoPriceUpdate.objects.filter(business=business, inventory_id__in=inv_ids)
        }

        to_create = []
        to_update = []
        touched_ids = set()

        for item in updates:
            inv_id = item.get("inventory_id")
            if inv_id is None:
                continue
            inv_id = int(inv_id)
            inv = inventory_map.get(inv_id)
            if not inv:
                continue

            payload = {}
            for field in ("new_msp", "new_wdrp", "new_mrp"):
                raw = item.get(field)
                payload[field] = safe_decimal(raw) if raw not in (None, "", "null") else None

            existing = existing_updates.get(inv_id)
            if existing:
                existing.new_msp = payload["new_msp"]
                existing.new_wdrp = payload["new_wdrp"]
                existing.new_mrp = payload["new_mrp"]
                to_update.append(existing)
            else:
                to_create.append(MeeshoPriceUpdate(business=business, inventory=inv, **payload))
            touched_ids.add(inv_id)

        with transaction.atomic():
            if to_create:
                MeeshoPriceUpdate.objects.bulk_create(to_create, batch_size=500)
            if to_update:
                MeeshoPriceUpdate.objects.bulk_update(to_update, ["new_msp", "new_wdrp", "new_mrp"], batch_size=500)

        return Response({"updated": len(touched_ids)})

    # ── GET ──────────────────────────────────────────────────────────────────
    from django.db.models import Avg, Count, Sum as DjSum

    # Per-SKU delivery stats from OrderPayment
    delivered_agg = (
        OrderPayment.objects
        .filter(business=business, live_order_status__iexact="delivered")
        .values("supplier_sku")
        .annotate(
            avg_settlement=Avg("final_settlement_amount"),
            delivery_count=Count("id"),
        )
    )
    delivered_map = {r["supplier_sku"]: r for r in delivered_agg}

    # Current listing price per SKU
    listing_agg = (
        OrderPayment.objects
        .filter(business=business)
        .exclude(listing_price_incl_taxes=None)
        .values("supplier_sku")
        .annotate(last_price=Avg("listing_price_incl_taxes"))
    )
    listing_map = {r["supplier_sku"]: float(r["last_price"] or 0) for r in listing_agg}

    # Cost from FinalPrice
    fp_map = _SkuMap((fp.sku_id, float(fp.final_price or 0))
                     for fp in FinalPrice.objects.filter(business=business).only("sku_id", "final_price"))

    # Existing price updates
    pu_map = {pu.inventory_id: pu
              for pu in MeeshoPriceUpdate.objects.filter(business=business).select_related("inventory")}

    qs = MeeshoInventory.objects.filter(business=business).order_by("serial_no")

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(catalog_name__icontains=search) |
            DQ(product_name__icontains=search) |
            DQ(style_id__icontains=search)
        )

    max_profit_str = request.GET.get("max_profit", "").strip()
    max_profit = None
    try:
        if max_profit_str != "":
            max_profit = float(max_profit_str)
    except ValueError:
        pass

    results = []
    for inv in qs:
        sid = inv.style_id.strip() if inv.style_id else ""
        d   = delivered_map.get(sid, {})

        avg_settlement = float(d.get("avg_settlement") or 0)
        delivery_count = int(d.get("delivery_count") or 0)
        cost           = fp_map.get(sid, None)
        current_price  = listing_map.get(sid, None)

        avg_profit = None
        if cost is not None and avg_settlement:
            avg_profit = round(avg_settlement - cost, 2)

        # Apply profit threshold filter
        if max_profit is not None:
            if avg_profit is None or avg_profit >= max_profit:
                continue

        pu = pu_map.get(inv.id)
        results.append({
            "inventory_id":    inv.id,
            "serial_no":       inv.serial_no,
            "catalog_name":    inv.catalog_name,
            "catalog_id":      inv.catalog_id,
            "product_name":    inv.product_name,
            "product_id":      inv.product_id,
            "style_id":        inv.style_id,
            "variation":       inv.variation,
            "variation_id":    inv.variation_id,
            "system_stock":    inv.system_stock_count,
            "delivery_count":  delivery_count,
            "avg_settlement":  round(avg_settlement, 2) if avg_settlement else None,
            "cost":            cost,
            "current_price":   current_price,
            "avg_profit":      avg_profit,
            # saved price updates
            "new_msp":         float(pu.new_msp)  if pu and pu.new_msp  is not None else None,
            "new_wdrp":        float(pu.new_wdrp) if pu and pu.new_wdrp is not None else None,
            "new_mrp":         float(pu.new_mrp)  if pu and pu.new_mrp  is not None else None,
        })

    return Response({
        "items": results,
        "total": len(results),
    })


@api_view(["GET"])
def meesho_price_update_download(request, business_id):
    """
    Download the Meesho price update sheet.
    Only rows where new_msp has been set are included.
    Columns match Meesho's bulk price update template exactly.
    """
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl

    business = get_authorized_business(request, business_id)
    rows = (
        MeeshoPriceUpdate.objects
        .filter(business=business, new_msp__isnull=False)
        .select_related("inventory")
        .order_by("inventory__serial_no")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MSP Update"

    headers = [
        "Catalog ID",
        "Product ID",
        "Variation Name",
        "Variation ID",
        "New Meesho Selling Price (MSP)",
        "Wrong/Defective Return Price (WDRP) (if catalog is part of WDRP)",
        "New Maximum Retail Price (MRP) (Optional)",
    ]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for pu in rows:
        inv = pu.inventory
        ws.append([
            inv.catalog_id,
            inv.product_id,
            inv.variation or "Free Size",
            inv.variation_id,
            float(pu.new_msp),
            float(pu.new_wdrp) if pu.new_wdrp is not None else "",
            float(pu.new_mrp)  if pu.new_mrp  is not None else "",
        ])

    col_widths = [14, 14, 20, 14, 32, 50, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = 'attachment; filename="meesho_price_update.xlsx"'
    return resp


# ═══════════════════════════════════════════════════════════════════════════
#  Expenses: packaging / other expense invoices + daily transport charges
# ═══════════════════════════════════════════════════════════════════════════

def _month_range(month):
    """'YYYY-MM' -> (first_of_month, first_of_next_month) as ISO strings, or (None, None)."""
    if not month:
        return None, None
    y, m = (int(x) for x in month.split("-"))
    start = f"{y:04d}-{m:02d}-01"
    ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
    end = f"{ny:04d}-{nm:02d}-01"
    return start, end


def _expense_invoice_to_dict(inv):
    """Serialize an ExpenseInvoice (with prefetched items) + computed totals."""
    _q = Decimal("0.01")
    total = Decimal("0")
    by_category = {}
    items = []
    for it in inv.items.all():
        amt = ((it.quantity or Decimal("0")) * (it.unit_rate or Decimal("0"))).quantize(_q)
        total += amt
        by_category[it.category] = by_category.get(it.category, Decimal("0")) + amt
        items.append({
            "id": it.id,
            "description": it.description,
            "category": it.category,
            "quantity": str(it.quantity),
            "unit_rate": str(it.unit_rate),
            "amount": str(amt),
        })
    return {
        "id": inv.id,
        "invoice_no": inv.invoice_no,
        "vendor": inv.vendor,
        "title": inv.title,
        "date": str(inv.date),
        "notes": inv.notes,
        "items": items,
        "total_amount": str(total),
        "by_category": {k: str(v) for k, v in by_category.items()},
        "created_at": inv.created_at.isoformat(),
    }


@api_view(["GET", "POST"])
def expense_invoices_list(request, business_id):
    business = get_authorized_business(request, business_id)

    if request.method == "GET":
        qs = ExpenseInvoice.objects.filter(business=business).prefetch_related("items")
        month = request.GET.get("month", "")
        search = request.GET.get("search", "").strip()
        start, end = _month_range(month)
        if start:
            qs = qs.filter(date__gte=start, date__lt=end)
        if request.GET.get("date_from"):
            qs = qs.filter(date__gte=request.GET["date_from"])
        if request.GET.get("date_to"):
            qs = qs.filter(date__lte=request.GET["date_to"])
        if search:
            qs = qs.filter(
                DQ(vendor__icontains=search) | DQ(title__icontains=search) |
                DQ(invoice_no__icontains=search) | DQ(items__description__icontains=search)
            ).distinct()
        invoices = [_expense_invoice_to_dict(i) for i in qs.order_by("-date", "-created_at")]
        grand_total = sum((Decimal(i["total_amount"]) for i in invoices), Decimal("0"))
        cat_total = {}
        for i in invoices:
            for k, v in i["by_category"].items():
                cat_total[k] = cat_total.get(k, Decimal("0")) + Decimal(v)
        return Response({
            "results": invoices,
            "total": len(invoices),
            "grand_total": str(grand_total),
            "by_category": {k: str(v) for k, v in cat_total.items()},
        })

    # POST — create invoice with nested line items
    data = request.data
    with transaction.atomic():
        inv = ExpenseInvoice.objects.create(
            business=business,
            invoice_no=data.get("invoice_no", ""),
            vendor=data.get("vendor", ""),
            title=data.get("title", ""),
            date=data["date"],
            notes=data.get("notes", ""),
        )
        for it in data.get("items", []):
            if not (it.get("description") or "").strip():
                continue
            ExpenseInvoiceItem.objects.create(
                business=business,
                invoice=inv,
                description=it.get("description", ""),
                category=it.get("category", "packaging"),
                quantity=Decimal(str(it.get("quantity") or 0)),
                unit_rate=Decimal(str(it.get("unit_rate") or 0)),
            )
    inv = ExpenseInvoice.objects.prefetch_related("items").get(pk=inv.pk, business=business)
    return Response(_expense_invoice_to_dict(inv), status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "DELETE"])
def expense_invoice_detail(request, business_id, invoice_id):
    business = get_authorized_business(request, business_id)
    try:
        inv = ExpenseInvoice.objects.prefetch_related("items").get(pk=invoice_id, business=business)
    except ExpenseInvoice.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "GET":
        return Response(_expense_invoice_to_dict(inv))

    if request.method == "DELETE":
        inv.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # PUT — replace fields + items
    data = request.data
    with transaction.atomic():
        inv.invoice_no = data.get("invoice_no", inv.invoice_no)
        inv.vendor = data.get("vendor", inv.vendor)
        inv.title = data.get("title", inv.title)
        inv.date = data.get("date", inv.date)
        inv.notes = data.get("notes", inv.notes)
        inv.save()
        inv.items.all().delete()
        for it in data.get("items", []):
            if not (it.get("description") or "").strip():
                continue
            ExpenseInvoiceItem.objects.create(
                business=business,
                invoice=inv,
                description=it.get("description", ""),
                category=it.get("category", "packaging"),
                quantity=Decimal(str(it.get("quantity") or 0)),
                unit_rate=Decimal(str(it.get("unit_rate") or 0)),
            )
    inv = ExpenseInvoice.objects.prefetch_related("items").get(pk=inv.pk, business=business)
    return Response(_expense_invoice_to_dict(inv))


@api_view(["GET", "POST"])
def transport_charges_list(request, business_id):
    business = get_authorized_business(request, business_id)

    if request.method == "GET":
        qs = TransportCharge.objects.filter(business=business)
        month = request.GET.get("month", "")
        start, end = _month_range(month)
        if start:
            qs = qs.filter(date__gte=start, date__lt=end)
        if request.GET.get("date_from"):
            qs = qs.filter(date__gte=request.GET["date_from"])
        if request.GET.get("date_to"):
            qs = qs.filter(date__lte=request.GET["date_to"])
        rows = list(qs.order_by("-date", "-created_at"))
        total = sum((r.amount for r in rows), Decimal("0"))
        return Response({
            "results": [
                {"id": r.id, "date": str(r.date), "amount": str(r.amount), "note": r.note}
                for r in rows
            ],
            "total": len(rows),
            "total_amount": str(total),
        })

    # POST — add a daily transport charge
    data = request.data
    r = TransportCharge.objects.create(
        business=business,
        date=data["date"],
        amount=Decimal(str(data.get("amount") or 0)),
        note=data.get("note", ""),
    )
    return Response({"id": r.id, "date": str(r.date), "amount": str(r.amount), "note": r.note},
                    status=status.HTTP_201_CREATED)


@api_view(["PUT", "DELETE"])
def transport_charge_detail(request, business_id, charge_id):
    business = get_authorized_business(request, business_id)
    try:
        r = TransportCharge.objects.get(pk=charge_id, business=business)
    except TransportCharge.DoesNotExist:
        return Response({"error": "Not found"}, status=404)
    if request.method == "DELETE":
        r.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    data = request.data
    r.date = data.get("date", r.date)
    if "amount" in data:
        r.amount = Decimal(str(data.get("amount") or 0))
    r.note = data.get("note", r.note)
    r.save()
    return Response({"id": r.id, "date": str(r.date), "amount": str(r.amount), "note": r.note})


@api_view(["GET"])
def expenses_summary(request, business_id):
    """Combined expense totals (invoices + transport) for a period."""
    business = get_authorized_business(request, business_id)
    month = request.GET.get("month", "")
    date_from = request.GET.get("date_from", "")
    date_to   = request.GET.get("date_to", "")
    start, end = _month_range(month)

    inv_items = ExpenseInvoiceItem.objects.filter(business=business)
    transport = TransportCharge.objects.filter(business=business)
    if start:
        inv_items = inv_items.filter(invoice__date__gte=start, invoice__date__lt=end)
        transport = transport.filter(date__gte=start, date__lt=end)
    if date_from:
        inv_items = inv_items.filter(invoice__date__gte=date_from)
        transport = transport.filter(date__gte=date_from)
    if date_to:
        inv_items = inv_items.filter(invoice__date__lte=date_to)
        transport = transport.filter(date__lte=date_to)

    packaging = Decimal("0")
    other = Decimal("0")
    for it in inv_items.only("category", "quantity", "unit_rate"):
        amt = ((it.quantity or Decimal("0")) * (it.unit_rate or Decimal("0"))).quantize(Decimal("0.01"))
        if it.category == "packaging":
            packaging += amt
        else:
            other += amt
    transport_total = transport.aggregate(s=Sum("amount"))["s"] or Decimal("0")
    grand = packaging + other + transport_total
    return Response({
        "packaging": str(packaging),
        "other": str(other),
        "transport": str(transport_total),
        "grand_total": str(grand),
    })


@api_view(["GET"])
def tax_check(request, business_id):
    """Compare the tax % set on each SKU (FinalPrice.tax_percent) against the
    GST % Meesho actually reports on that SKU's payments
    (OrderPayment.product_gst_percent), and flag mismatches.

    Query params: status = mismatch|match|no_data|unset|all (default all),
    search (sku / parent).
    """
    from collections import defaultdict, Counter
    business = get_authorized_business(request, business_id)
    status_filter = request.GET.get("status", "all")
    search = request.GET.get("search", "").strip().lower()

    # Meesho GST per SKU (bucketed to whole-percent) from payment rows.
    gst_counts = defaultdict(Counter)
    for sku, gst in (
        OrderPayment.objects
        .filter(business=business)
        .exclude(product_gst_percent__isnull=True)
        .exclude(supplier_sku__isnull=True).exclude(supplier_sku="")
        .values_list("supplier_sku", "product_gst_percent")
        .iterator()
    ):
        gst_counts[sku][int(round(float(gst)))] += 1

    results = []
    counts = {"priced": 0, "matched": 0, "mismatched": 0, "no_data": 0, "unset": 0}

    for fp in FinalPrice.objects.filter(business=business).select_related("parent"):
        counts["priced"] += 1
        set_tax = fp.tax_percent
        counter = gst_counts.get(fp.sku_id)
        parent = fp.parent.item_id if fp.parent_id else None

        if not counter:
            row_status = "no_data"
            meesho = None
            values = []
            order_count = 0
        else:
            meesho = counter.most_common(1)[0][0]
            values = sorted(counter.keys())
            order_count = sum(counter.values())
            if set_tax is None:
                row_status = "unset"
            elif int(set_tax) == meesho:
                row_status = "match"
            else:
                row_status = "mismatch"

        counts_key = {"no_data": "no_data", "match": "matched", "mismatch": "mismatched", "unset": "unset"}[row_status]
        counts[counts_key] += 1

        row = {
            "sku_id": fp.sku_id,
            "parent": parent,
            "set_tax": set_tax,
            "meesho_gst": meesho,
            "meesho_values": values,
            "multiple_meesho": len(values) > 1,
            "diff": (int(set_tax) - meesho) if (set_tax is not None and meesho is not None) else None,
            "order_count": order_count,
            "status": row_status,
        }
        if status_filter != "all" and row_status != status_filter:
            continue
        if search and search not in fp.sku_id.lower() and not (parent and search in parent.lower()):
            continue
        results.append(row)

    # SKUs Meesho charges GST on but that have no FinalPrice row (unpriced).
    priced_skus = {
        _sku_key(s) for s in
        FinalPrice.objects.filter(business=business).values_list("sku_id", flat=True)
    }
    unpriced_with_orders = sum(1 for s in gst_counts if _sku_key(s) not in priced_skus)

    # Mismatches first, biggest |diff| first.
    order_rank = {"mismatch": 0, "unset": 1, "no_data": 2, "match": 3}
    results.sort(key=lambda r: (order_rank.get(r["status"], 9), -(abs(r["diff"]) if r["diff"] is not None else 0)))

    return Response({
        "kpi": {
            "priced_skus": counts["priced"],
            "matched": counts["matched"],
            "mismatched": counts["mismatched"],
            "no_meesho_data": counts["no_data"],
            "unset": counts["unset"],
            "unpriced_with_orders": unpriced_with_orders,
        },
        "results": results,
    })


# ══════════════════════════════════════════════════════════════════════════════
# Return deliveries — returns physically received back, and their 7-day claims
# ══════════════════════════════════════════════════════════════════════════════

# Header names in the Meesho "Returns → Completed/Delivered" CSV → model fields.
_RETURN_COL_MAP = {
    "product name":           "product_name",
    "sku":                    "sku",
    "variation":              "variation",
    "meesho pid":             "meesho_pid",
    "category":               "category",
    "qty":                    "qty",
    "order number":           "order_no",
    "suborder number":        "suborder_no",
    "sub order number":       "suborder_no",
    "dispatch date":          "dispatch_date",
    "return created date":    "return_created_date",
    "type of return":         "type_of_return",
    "sub type":               "sub_type",
    "delivered date":         "delivered_date",
    "courier partner":        "courier_partner",
    "awb number":             "awb_number",
    "tracking link":          "tracking_link",
    "proof of delivery":      "proof_of_delivery",
    "return price type":      "return_price_type",
    "return reason":          "return_reason",
    "detailed return reason": "detailed_return_reason",
    "otp verified at":        "otp_verified_at",
}

# Cells Meesho writes to mean "nothing here".
_RETURN_BLANK_VALUES = {"", "nan", "none", "null", "na", "n/a", "-", "--"}

# Claim fields a client may write, and the statuses that mean the claim has
# already been dealt with (so no countdown applies any more).
_CLAIM_WRITABLE = {
    "claim_status", "claim_amount", "claim_reference", "claim_notes", "verify_result",
    "packet_id",
}


def _check_packet(business, row, scanned):
    """
    Verify a scanned packet id against what we know went out for this sub-order.

    Compared against, in order: the return's own reverse AWB, then the outbound
    label's AWB, then the sub-order number itself — because what is printed on a
    returned parcel varies by courier, and any of the three is a legitimate match.

    Matching is loose on purpose: scanners pick up carrier prefixes and separators
    that aren't part of the number, so both sides are reduced to their
    alphanumerics and a containment test either way counts as a hit. A strict
    equality test would flag genuine matches as mismatches, and a mismatch here
    tells the operator to stop and look — it must not cry wolf.
    """
    def norm(value):
        return "".join(ch for ch in str(value or "").upper() if ch.isalnum())

    scanned_n = norm(scanned)
    if not scanned_n:
        return ReturnDelivery.PACKET_UNKNOWN, ""

    label = LabelOrder.objects.filter(business=business, order_id=row.suborder_no).first()
    candidates = [
        row.awb_number,
        label.awb_number if label else None,
        row.suborder_no,
        row.order_no,
    ]

    for candidate in candidates:
        cand_n = norm(candidate)
        if not cand_n:
            continue
        if scanned_n == cand_n or scanned_n in cand_n or cand_n in scanned_n:
            return ReturnDelivery.PACKET_MATCH, str(candidate)

    # Nothing to compare against is materially different from a real mismatch:
    # one is missing data, the other is a wrong parcel in your hand.
    if not any(norm(c) for c in candidates):
        return ReturnDelivery.PACKET_UNKNOWN, ""

    return ReturnDelivery.PACKET_MISMATCH, str(next((c for c in candidates if norm(c)), ""))
_CLAIM_DECIDED  = (
    ReturnDelivery.CLAIM_RAISED,
    ReturnDelivery.CLAIM_APPROVED,
    ReturnDelivery.CLAIM_REJECTED,
)

# A scanned code shorter than this is too ambiguous to fuzzy-match — a stray
# keystroke shouldn't pull back ten unrelated returns.
_MIN_PARTIAL_SCAN_LEN = 6


def _return_text(val, blank_as_empty=True):
    """Sheet cell → clean string. Meesho's literal 'NA' placeholders become ''."""
    s = safe_str(val) or ""
    if blank_as_empty and s.strip().lower() in _RETURN_BLANK_VALUES:
        return ""
    return s.strip()


def _find_return_header_row(lines):
    """
    The export starts with 6 lines of supplier metadata before the real header,
    and Meesho has moved that block around before — so locate the header by its
    content instead of hardcoding a skiprows count.
    """
    for idx, line in enumerate(lines[:60]):
        low = line.lower()
        if "suborder number" in low and "delivered date" in low:
            return idx
    return None


def _read_return_sheet(file):
    """Parse the uploaded returns export into a DataFrame with mapped columns."""
    name = (file.name or "").lower()

    if name.endswith((".xlsx", ".xls")):
        file.seek(0)
        raw = pd.read_excel(file, header=None, dtype=str)
        header_idx = None
        for idx, row in raw.iterrows():
            joined = " ".join(str(c).lower() for c in row.tolist() if c is not None)
            if "suborder number" in joined and "delivered date" in joined:
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("Could not find the header row (expected 'Suborder Number' and 'Delivered Date').")
        df = raw.iloc[header_idx + 1:].copy()
        df.columns = [str(c) for c in raw.iloc[header_idx].tolist()]
    else:
        blob = file.read()
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = blob.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = blob.decode("utf-8", errors="replace")

        lines = text.splitlines()
        header_idx = _find_return_header_row(lines)
        if header_idx is None:
            raise ValueError("Could not find the header row (expected 'Suborder Number' and 'Delivered Date').")
        df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])), dtype=str)

    df.columns = [str(c).strip().lower() for c in df.columns]
    df.rename(columns=_RETURN_COL_MAP, inplace=True)
    return df


@api_view(["POST"])
@parser_classes([MultiPartParser])
def return_deliveries_upload(request, business_id):
    """
    Upload the Meesho "Returns → Completed/Delivered" export (CSV or Excel).

    Keyed on (suborder_no, awb_number), so re-uploading an overlapping export
    updates rows in place. Claim fields are deliberately left out of the update
    payload — an upload can never overwrite claim work already recorded.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = _read_return_sheet(file)
    except Exception as exc:
        return Response({"error": f"Could not read file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

    if "suborder_no" not in df.columns:
        return Response(
            {"error": "Missing required column: Suborder Number"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created = updated = skipped = 0
    no_delivered_date = 0
    seen_keys = set()

    with transaction.atomic():
        for _, row in df.iterrows():
            suborder = _return_text(row.get("suborder_no"))
            if not suborder:
                skipped += 1
                continue

            awb = _return_text(row.get("awb_number"))
            key = (suborder, awb)
            if key in seen_keys:
                # Same return listed twice inside one file — keep the first.
                skipped += 1
                continue
            seen_keys.add(key)

            delivered = safe_date(row.get("delivered_date"))
            if delivered is None:
                no_delivered_date += 1

            qty = safe_int(row.get("qty"))

            defaults = {
                "order_no":               _return_text(row.get("order_no")),
                "product_name":           _return_text(row.get("product_name")),
                "sku":                    _return_text(row.get("sku")),
                "variation":              _return_text(row.get("variation")),
                "meesho_pid":             _return_text(row.get("meesho_pid")),
                "category":               _return_text(row.get("category")),
                "qty":                    1 if qty is None else max(0, qty),
                "dispatch_date":          safe_date(row.get("dispatch_date")),
                "return_created_date":    safe_date(row.get("return_created_date")),
                "delivered_date":         delivered,
                "type_of_return":         _return_text(row.get("type_of_return")),
                "sub_type":               _return_text(row.get("sub_type")),
                "courier_partner":        _return_text(row.get("courier_partner")),
                "tracking_link":          _return_text(row.get("tracking_link")),
                "proof_of_delivery":      _return_text(row.get("proof_of_delivery")),
                "return_price_type":      _return_text(row.get("return_price_type")),
                "return_reason":          _return_text(row.get("return_reason")),
                "detailed_return_reason": _return_text(row.get("detailed_return_reason")),
                "otp_verified_at":        safe_datetime(row.get("otp_verified_at")),
            }

            _, was_created = ReturnDelivery.objects.update_or_create(
                business=business, suborder_no=suborder, awb_number=awb,
                defaults=defaults,
            )
            if was_created:
                created += 1
            else:
                updated += 1

    return Response(
        {
            "success": True,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "without_delivered_date": no_delivered_date,
        },
        status=status.HTTP_201_CREATED,
    )


def _return_claim_stats(business):
    """
    Claim workload for the whole business — deliberately never narrowed by the
    period filter, so a claim about to expire can't hide behind the selected
    month. Counted in the database (not in Python) so it stays cheap as the
    returns table grows.
    """
    base   = ReturnDelivery.objects.filter(business=business)
    today  = timezone.localdate()
    window = ReturnDelivery.CLAIM_WINDOW_DAYS
    warn   = ReturnDelivery.WARN_FROM_DAY

    by_status = {
        row["claim_status"]: row["n"]
        for row in base.values("claim_status").annotate(n=Count("id"))
    }

    # Day-of-window boundaries, as delivered_date ranges:
    #   deadline today          → delivered exactly `window` days ago
    #   1–2 days left (day 5–6) → delivered between `window` and `warn` days ago
    #   window closed           → delivered more than `window` days ago
    last_day_on   = today - timedelta(days=window)
    warn_start_on = today - timedelta(days=window - 1)   # day 6 → 1 day left
    warn_end_on   = today - timedelta(days=warn)         # day 5 → 2 days left

    def _count(statuses, **date_filters):
        return base.filter(claim_status__in=statuses, **date_filters).count()

    required = (ReturnDelivery.CLAIM_REQUIRED,)
    open_all = ReturnDelivery.OPEN_CLAIM_STATUSES

    stats = {
        "total":          base.count(),
        "unreviewed":     by_status.get(ReturnDelivery.CLAIM_UNREVIEWED, 0),
        "claim_required": by_status.get(ReturnDelivery.CLAIM_REQUIRED, 0),
        "not_required":   by_status.get(ReturnDelivery.CLAIM_NOT_REQUIRED, 0),
        "claim_raised":   sum(by_status.get(s, 0) for s in _CLAIM_DECIDED),

        # Flagged as needing a claim, still not raised:
        "expiring": _count(required,
                           delivered_date__gte=warn_start_on,
                           delivered_date__lte=warn_end_on),
        "last_day": _count(required, delivered_date=last_day_on),
        "expired":  _count(required, delivered_date__lt=last_day_on),

        # Never reviewed and already at day 5 or beyond — nobody has even
        # decided whether these need a claim.
        "unreviewed_expiring": _count(
            (ReturnDelivery.CLAIM_UNREVIEWED,), delivered_date__lte=warn_end_on
        ),
        "open_total": base.filter(claim_status__in=open_all).count(),
    }

    # Everything that needs someone to act today.
    stats["needs_attention"] = stats["expiring"] + stats["last_day"]
    return stats


def _order_context(business, suborder_no):
    """
    What we know about this sub-order from the rest of the app, so the person
    at the scanning desk can cross-check the parcel in their hands: what was
    shipped (label) and what Meesho settled / already paid as a claim.
    """
    context = {"label": None, "payments": None}

    label = LabelOrder.objects.filter(business=business, order_id=suborder_no).first()
    if label:
        context["label"] = {
            "sku": label.sku,
            "qty": label.qty,
            "size": label.size,
            "color": label.color,
            "courier_name": label.courier_name,
            "awb_number": label.awb_number,
            "customer_name": label.customer_name,
            "customer_city": label.customer_city,
            "customer_state": label.customer_state,
            "customer_pincode": label.customer_pincode,
            "order_date": label.order_date,
        }

    rows = list(
        OrderPayment.objects.filter(business=business, sub_order_no=suborder_no)
        .order_by("payment_date", "live_order_status")
    )
    if rows:
        context["payments"] = {
            "row_count": len(rows),
            "statuses": [r.live_order_status for r in rows if r.live_order_status],
            "supplier_sku": next((r.supplier_sku for r in rows if r.supplier_sku), None),
            "quantity": next((r.quantity for r in rows if r.quantity), None),
            "listing_price": next(
                (float(r.listing_price_incl_taxes) for r in rows if r.listing_price_incl_taxes is not None), None
            ),
            "net_settlement": round(sum(float(r.final_settlement_amount or 0) for r in rows), 2),
            "claims_paid": round(sum(float(r.claims or 0) for r in rows), 2),
            "return_shipping_charge": round(sum(float(r.return_shipping_charge or 0) for r in rows), 2),
            "last_payment_date": max((r.payment_date for r in rows if r.payment_date), default=None),
        }

    return context


@api_view(["GET"])
def return_deliveries_list(request, business_id):
    """
    Paginated list of received returns.

    Query params:
      q            — matches suborder / order no / AWB / SKU / product name
      date_from / date_to  — on delivered_date
      claim_status — one of the CLAIM_* values, or "open" for anything still owed
      urgency      — "attention" (expiring or last day), "expired", "unreviewed"
      type         — "customer" | "rto"
      page, page_size
    """
    business = get_authorized_business(request, business_id)

    qs = ReturnDelivery.objects.filter(business=business)

    date_from = request.GET.get("date_from", "").strip()
    date_to   = request.GET.get("date_to", "").strip()
    if date_from:
        qs = qs.filter(delivered_date__gte=date_from)
    if date_to:
        qs = qs.filter(delivered_date__lte=date_to)

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(suborder_no__icontains=search) |
            DQ(order_no__icontains=search) |
            DQ(awb_number__icontains=search) |
            DQ(sku__icontains=search) |
            DQ(product_name__icontains=search)
        )

    claim_status = request.GET.get("claim_status", "").strip().upper()
    if claim_status == "OPEN":
        qs = qs.filter(claim_status__in=ReturnDelivery.OPEN_CLAIM_STATUSES)
    elif claim_status:
        qs = qs.filter(claim_status=claim_status)

    return_type = request.GET.get("type", "").strip().lower()
    if return_type == "customer":
        qs = qs.filter(type_of_return__icontains="customer")
    elif return_type == "rto":
        qs = qs.filter(type_of_return__icontains="rto")

    # Urgency depends on today's date vs delivered_date, so translate it into a
    # delivered_date range rather than filtering in Python over the whole table.
    urgency = request.GET.get("urgency", "").strip().lower()
    today  = timezone.localdate()
    window = ReturnDelivery.CLAIM_WINDOW_DAYS
    warn   = ReturnDelivery.WARN_FROM_DAY
    if urgency == "attention":
        # Day WARN_FROM_DAY..window of the claim window, claim still owed.
        qs = qs.filter(
            claim_status__in=ReturnDelivery.OPEN_CLAIM_STATUSES,
            delivered_date__gte=today - timedelta(days=window),
            delivered_date__lte=today - timedelta(days=warn),
        )
    elif urgency == "expired":
        qs = qs.filter(
            claim_status__in=ReturnDelivery.OPEN_CLAIM_STATUSES,
            delivered_date__lt=today - timedelta(days=window),
        )
    elif urgency == "unreviewed":
        qs = qs.filter(claim_status=ReturnDelivery.CLAIM_UNREVIEWED)

    total = qs.count()

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 50))))
    except (TypeError, ValueError):
        page_size = 50

    start = (page - 1) * page_size
    rows  = qs[start:start + page_size]

    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": ReturnDeliverySerializer(rows, many=True).data,
        "stats": _return_claim_stats(business),
        "claim_window_days": window,
    })


@api_view(["GET"])
def return_delivery_lookup(request, business_id):
    """
    Barcode-scanner endpoint: resolve a scanned AWB / sub-order / order number
    to the return(s) it belongs to, with cross-check context for each.

    A scanned *order* number can cover several sub-orders, so this always
    returns a list and lets the UI pick when there's more than one.
    """
    business = get_authorized_business(request, business_id)

    raw = request.GET.get("q", "")
    # Scanners often append whitespace/newlines, and some encode the AWB with
    # surrounding junk characters — keep it to what a code can actually contain.
    code = raw.strip().strip("\r\n\t ").strip('"').strip("'")
    if not code:
        return Response({"error": "Nothing scanned."}, status=status.HTTP_400_BAD_REQUEST)

    def _search(scope_business):
        """Most specific match first, so a code that is both an AWB and a
        sub-order prefix resolves to the AWB."""
        base = ReturnDelivery.objects.filter(business=scope_business)
        rows = list(base.filter(awb_number__iexact=code))
        if rows:
            return rows, "awb"
        rows = list(base.filter(suborder_no__iexact=code))
        if rows:
            return rows, "suborder"
        rows = list(base.filter(order_no__iexact=code))
        if rows:
            return rows, "order"
        if len(code) >= _MIN_PARTIAL_SCAN_LEN:
            # Fall back to a contains search — handles scanners that prepend a
            # carrier prefix, and lets a partially-read code still resolve.
            rows = list(
                base.filter(
                    DQ(awb_number__icontains=code) |
                    DQ(suborder_no__icontains=code) |
                    DQ(order_no__icontains=code)
                )[:10]
            )
            if rows:
                return rows, "partial"
        return [], None

    # Selected business first; a returned parcel for a sibling business still
    # resolves rather than reading as unknown, because one desk receives for all
    # of them. Only businesses this user already belongs to are searched.
    owner = business
    rows, matched_by = _search(business)
    if not rows:
        for sibling in accessible_businesses(request, include=business):
            if sibling.pk == business.pk:
                continue
            rows, matched_by = _search(sibling)
            if rows:
                owner = sibling
                break

    if not rows:
        return Response({
            "found": False,
            "scanned": code,
            "message": "No received return matches this code. Upload the latest "
                       "returns export, or check that the parcel has been marked "
                       "delivered on Meesho.",
        }, status=status.HTTP_404_NOT_FOUND)

    matches = []
    for row in rows:
        data = ReturnDeliverySerializer(row).data
        data["order_context"] = _order_context(owner, row.suborder_no)
        matches.append(data)

    return Response({
        "found": True,
        "scanned": code,
        "matched_by": matched_by,
        "match_count": len(matches),
        # Which business the parcel actually belongs to — the UI badges it when
        # it isn't the one currently selected.
        "cross_business": owner.pk != business.pk,
        "business_id": owner.pk,
        "business_name": owner.name,
        "matches": matches,
    })


@api_view(["GET", "PATCH"])
def return_delivery_detail(request, business_id, pk):
    """Read one return, or record claim / verification progress against it."""
    business = get_authorized_business(request, business_id)

    try:
        row = ReturnDelivery.objects.get(pk=pk, business=business)
    except ReturnDelivery.DoesNotExist:
        return Response({"error": "Return not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "PATCH":
        payload = request.data if isinstance(request.data, dict) else {}
        unknown = set(payload) - _CLAIM_WRITABLE
        if unknown:
            return Response(
                {"error": f"Not editable: {', '.join(sorted(unknown))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        now = timezone.now()

        if "claim_status" in payload:
            new_status = str(payload["claim_status"] or "").strip().upper()
            valid = {c[0] for c in ReturnDelivery.CLAIM_STATUS_CHOICES}
            if new_status not in valid:
                return Response(
                    {"error": f"Invalid claim_status. Expected one of: {', '.join(sorted(valid))}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            row.claim_status = new_status

            # Stamp when the claim was flagged, and when it was actually raised.
            if new_status == ReturnDelivery.CLAIM_REQUIRED and not row.claim_marked_at:
                row.claim_marked_at = now
            if new_status in _CLAIM_DECIDED:
                if not row.claim_marked_at:
                    row.claim_marked_at = now
                if not row.claim_raised_at:
                    row.claim_raised_at = now
            else:
                # Reopened / reset — the claim is owed again, so drop the
                # raised stamp or the countdown would stay hidden.
                row.claim_raised_at = None
            if new_status == ReturnDelivery.CLAIM_UNREVIEWED:
                row.claim_marked_at = None

        if "claim_amount" in payload:
            raw_amount = payload["claim_amount"]
            row.claim_amount = None if raw_amount in (None, "") else safe_decimal(raw_amount)

        if "claim_reference" in payload:
            row.claim_reference = (str(payload["claim_reference"] or "")).strip()[:150]

        if "claim_notes" in payload:
            row.claim_notes = str(payload["claim_notes"] or "").strip()

        if "verify_result" in payload:
            verify = str(payload["verify_result"] or "").strip().upper()
            valid_verify = {c[0] for c in ReturnDelivery.VERIFY_CHOICES}
            if verify and verify not in valid_verify:
                return Response(
                    {"error": f"Invalid verify_result. Expected one of: {', '.join(sorted(valid_verify))}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            row.verify_result = verify
            row.verified_at = now if verify else None

        if "packet_id" in payload:
            scanned = _clean_scan_code(payload["packet_id"])[:150]
            row.packet_id = scanned
            if scanned:
                row.packet_check, row.packet_matched_against = _check_packet(business, row, scanned)
                row.packet_scanned_at = now
            else:
                # Cleared — drop the verdict too, or a stale MATCH would keep
                # vouching for a packet that is no longer recorded.
                row.packet_check = ""
                row.packet_matched_against = ""
                row.packet_scanned_at = None

        row.save()

    data = ReturnDeliverySerializer(row).data
    data["order_context"] = _order_context(business, row.suborder_no)
    return Response(data)


# ══════════════════════════════════════════════════════════════════════════════
# Order scanning — record outgoing parcels at the desk and track their status
# ══════════════════════════════════════════════════════════════════════════════

# The only fields a client may change on a recorded scan. Everything else is
# either identity or scan bookkeeping: letting the UI rewrite those would make
# the log unreliable, which is the one thing it exists to be.
_SCAN_WRITABLE = {"status", "notes"}

_SCAN_STATUSES = {c[0] for c in ScannedOrder.STATUS_CHOICES}

# Meesho's own live_order_status strings, normalised to the buckets the UI
# colours. Anything unrecognised is passed through as-is rather than dropped.
_MEESHO_STATUS_BUCKETS = (
    ("DELIVERED", "delivered"),
    ("RTO",       "rto"),
    ("RETURN",    "return"),
    ("CANCEL",    "cancelled"),
    ("SHIP",      "shipped"),
    ("DOOR",      "shipped"),      # "Door Step Exchanged" and friends
)


def _meesho_bucket(raw):
    """Which coarse bucket a Meesho status string falls into."""
    s = (raw or "").upper()
    for needle, bucket in _MEESHO_STATUS_BUCKETS:
        if needle in s:
            return bucket
    return "other" if s else None


def _local_day_start(day):
    """
    The aware datetime at which a local calendar day begins.

    Used instead of filtering a DateTimeField with `__date`: on MySQL that
    compiles to CONVERT_TZ(), which returns NULL unless the server's timezone
    tables have been loaded (`mysql_tzinfo_to_sql`) — and they are not on this
    deployment, so such a filter silently matches *nothing* rather than erroring.
    An explicit half-open datetime range is correct on every backend.
    """
    return timezone.make_aware(datetime.combine(day, time.min))


def _parse_day(raw):
    """A "YYYY-MM-DD" query param as a date, or None when absent/unparseable."""
    text = (raw or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def _clean_scan_code(raw):
    """
    Scanners append whitespace/newlines and sometimes wrap the value in quotes.
    Trim to what a barcode can actually contain before matching on it.
    """
    return (raw or "").strip().strip("\r\n\t ").strip('"').strip("'")


def _resolve_scan_code_pooled(request, business, code):
    """
    Resolve a scanned code against every business the user can reach.

    (owning_business, snapshot). The selected business is tried first so the
    common case costs exactly one lookup and never changes owner unnecessarily;
    only an unrecognised code falls through to the siblings.

    Why the *owning* business is returned rather than always the selected one:
    one desk packs for several businesses, and filing a Rudam 2 parcel under
    Rudam because that happened to be the dropdown value would put the row where
    nobody looks for it. The scan lands with its order, and the UI says so.
    """
    snapshot = _resolve_scan_code(business, code)
    if snapshot["matched_from"] != ScannedOrder.MATCH_NONE:
        return business, snapshot

    for sibling in accessible_businesses(request, include=business):
        if sibling.pk == business.pk:
            continue
        other = _resolve_scan_code(sibling, code)
        if other["matched_from"] != ScannedOrder.MATCH_NONE:
            return sibling, other

    # Nothing anywhere — log it against the business the operator had selected.
    return business, snapshot


def _resolve_scan_code(business, code):
    """
    Work out which order a scanned code belongs to, and snapshot what we know
    about it.

    Sources in order of richness: the shipping label (has courier, customer and
    AWB), then the payments sheet, then the orders sheet. An unmatched code is
    not an error — it is recorded as-is so the scan is never lost, and it will
    read as "not recognised" in the table.
    """
    snapshot = {
        "sub_order_no": code,
        "awb_number": "",
        "matched_from": ScannedOrder.MATCH_NONE,
        "sku": "", "product_name": "", "size": "", "qty": 1,
        "courier_name": "", "payment_type": "",
        "customer_name": "", "customer_city": "", "customer_state": "",
        "customer_pincode": "", "order_date": None,
    }

    labels = LabelOrder.objects.filter(business=business)
    # AWB first: it is what is actually printed as a barcode on the parcel.
    label = (
        labels.filter(awb_number__iexact=code).first()
        or labels.filter(order_id__iexact=code).first()
    )
    if label is None and len(code) >= _MIN_PARTIAL_SCAN_LEN:
        # A partially-read code, or one the courier prefixes — still resolvable.
        label = labels.filter(
            DQ(awb_number__icontains=code) | DQ(order_id__icontains=code)
        ).first()

    if label:
        snapshot.update({
            "sub_order_no": label.order_id,
            "awb_number": label.awb_number or "",
            "matched_from": ScannedOrder.MATCH_LABEL,
            "sku": label.sku or "",
            "size": label.size or "",
            "qty": label.qty or 1,
            "courier_name": label.courier_name or "",
            "payment_type": label.payment_type or "",
            "customer_name": label.customer_name or "",
            "customer_city": label.customer_city or "",
            "customer_state": label.customer_state or "",
            "customer_pincode": label.customer_pincode or "",
            "order_date": label.order_date,
        })
        return snapshot

    payment = (
        OrderPayment.objects.filter(business=business, sub_order_no__iexact=code)
        .order_by("-order_date")
        .first()
    )
    if payment:
        snapshot.update({
            "sub_order_no": payment.sub_order_no,
            "matched_from": ScannedOrder.MATCH_PAYMENT,
            "sku": payment.supplier_sku or "",
            "product_name": payment.product_name or "",
            "qty": payment.quantity or 1,
            "order_date": timezone.localtime(payment.order_date).date() if payment.order_date else None,
        })
        return snapshot

    order = (
        Order.objects.filter(business=business, sub_order_no__iexact=code)
        .order_by("-order_date")
        .first()
    )
    if order:
        snapshot.update({
            "sub_order_no": order.sub_order_no,
            "matched_from": ScannedOrder.MATCH_ORDER,
            "sku": order.sku or "",
            "product_name": order.product_name or "",
            "size": order.size or "",
            "qty": order.quantity or 1,
            "customer_state": order.customer_state or "",
            "order_date": order.order_date,
        })

    return snapshot


def _meesho_status_map(business, sub_order_nos):
    """
    Meesho's own status for a batch of sub-orders, in two queries rather than
    two per row.

    This is the half of "what happened to it" we can't know at the desk: it only
    becomes available once the payment / orders sheet for that period is
    uploaded. Absent from the map means Meesho hasn't reported on it yet.
    """
    nos = [n for n in sub_order_nos if n]
    if not nos:
        return {}

    resolved = {}

    # Orders sheet is the coarser source — fill from it first so the payments
    # sheet (which carries the live status verbatim) can overwrite it.
    for row in Order.objects.filter(
        business=business, sub_order_no__in=nos
    ).values("sub_order_no", "reason_for_credit_entry", "order_date").order_by("order_date"):
        raw = row["reason_for_credit_entry"]
        if not raw:
            continue
        resolved[row["sub_order_no"]] = {
            "status": raw.replace("_", " ").title(),
            "bucket": _meesho_bucket(raw),
            "source": "orders",
            "as_of": row["order_date"],
        }

    for row in OrderPayment.objects.filter(
        business=business, sub_order_no__in=nos, live_order_status__isnull=False
    ).exclude(live_order_status="").values(
        "sub_order_no", "live_order_status", "payment_date"
    ).order_by("payment_date"):
        resolved[row["sub_order_no"]] = {
            "status": row["live_order_status"],
            "bucket": _meesho_bucket(row["live_order_status"]),
            "source": "payments",
            "as_of": row["payment_date"],
        }

    return resolved


def _attach_meesho_status(business, rows):
    """
    Hang the resolved Meesho status on each row for the serializer to pick up.

    Rows are grouped by their *own* business rather than resolved against the
    passed one: a pooled list mixes businesses, and looking a Rudam 2 sub-order
    up in Rudam's orders would report "not reported yet" for something we
    actually know. One query pair per business present, not per row.
    """
    rows = list(rows)
    by_business = {}
    for row in rows:
        by_business.setdefault(row.business_id, []).append(row)

    for business_id, group in by_business.items():
        status_map = _meesho_status_map(
            business if business_id == business.pk else Business.objects.get(pk=business_id),
            [r.sub_order_no for r in group],
        )
        for row in group:
            row.meesho_status = status_map.get(row.sub_order_no)
    return rows


# Most-advanced-wins ordering. "Did it leave the building" is monotonic: once a
# parcel has shipped, no later row makes it un-shipped. CANCELLED outranks
# NOT_SHIPPED because a cancelled order is settled, not still waiting.
_SHIP_PRIORITY = {
    ScannedOrder.SHIP_SHIPPED: 3,
    ScannedOrder.SHIP_CANCELLED: 2,
    ScannedOrder.SHIP_NOT_SHIPPED: 1,
    ScannedOrder.SHIP_UNKNOWN: 0,
}


def _status_raws(businesses, sub_orders):
    """
    Every status string recorded against these sub-orders, with the business it
    came from — two queries per business rather than per sub-order.
    """
    subs = [s for s in sub_orders if s]
    if not subs:
        return []

    found = []
    for b in businesses:
        for raw in (
            Order.objects.filter(business=b, sub_order_no__in=subs)
            .values_list("reason_for_credit_entry", flat=True)
        ):
            if raw:
                found.append((raw, b, "orders"))
        for raw in (
            OrderPayment.objects.filter(business=b, sub_order_no__in=subs)
            .exclude(live_order_status__isnull=True).exclude(live_order_status="")
            .values_list("live_order_status", flat=True)
        ):
            found.append((raw, b, "payments"))
    return found


def _suborders_for_codes(businesses, codes):
    """
    Sub-order numbers reachable from a set of AWB / packet codes.

    Two bridges exist and both are needed: the shipping label maps AWB →
    sub-order for outbound parcels, and a return maps its reverse AWB →
    sub-order. A scan that only ever carried an AWB is otherwise a dead end,
    which is the main reason orders sat at "no status".
    """
    wanted = {c for c in codes if c}
    if not wanted:
        return set()

    subs = set()
    for b in businesses:
        subs.update(
            LabelOrder.objects.filter(business=b, awb_number__in=wanted)
            .values_list("order_id", flat=True)
        )
        subs.update(
            ReturnDelivery.objects.filter(business=b, awb_number__in=wanted)
            .values_list("suborder_no", flat=True)
        )
    return subs


def _returned_suborders(businesses, sub_orders):
    """
    Which of these sub-orders have a return on file.

    A parcel that came back is proof it went out, whatever the Orders sheet says
    — and for an order whose sheet was never uploaded, this may be the only
    evidence we hold.
    """
    subs = [s for s in sub_orders if s]
    if not subs:
        return set()
    found = set()
    for b in businesses:
        found.update(
            ReturnDelivery.objects.filter(business=b, suborder_no__in=subs)
            .values_list("suborder_no", flat=True)
        )
    return found


def _resolve_ship_status(business, sub_order_no, status_hint=None, row=None, scope=None):
    """
    Whether this order actually shipped, per everything we hold.

    Returns (ship_status, raw_value, source).

    Resolution widens only as far as it has to, so the common case stays at two
    queries and only genuinely unknown orders pay for the rest:

      1. the sub-order, in its own business
      2. the sub-order, in every business the user can reach — the parcel may be
         filed under one business while its Orders sheet was uploaded to another
      3. the AWB and the raw scanned code, mapped to sub-orders via labels and
         returns, then looked up again across all of them
      4. a return on file for any candidate, which proves it shipped

    A sub-order legitimately has several Orders rows — SHIPPED, then RTO_LOCKED,
    then RTO_COMPLETE — and in this data they all carry the *same* order_date, so
    "the latest row" is decided only by upload time. Picking one row would let a
    re-uploaded old export flip a shipped parcel back to READY_TO_SHIP, so every
    status found is classified and the most-advanced bucket wins.
    """
    own = [business]
    candidates = {sub_order_no}
    if row is not None:
        candidates.add(row.sub_order_no)

    def _best(raws):
        best, best_raw, best_src = ScannedOrder.SHIP_UNKNOWN, "", ""
        for raw, biz, table in raws:
            bucket = ScannedOrder.classify_ship_status(raw)
            if _SHIP_PRIORITY[bucket] > _SHIP_PRIORITY[best]:
                best, best_raw = bucket, (raw or "")
                best_src = table if biz.pk == business.pk else f"{table} · {biz.name}"
        return best, best_raw, best_src

    # 1 — the hint the caller already has, plus this business's own sheets.
    raws = []
    if status_hint and status_hint.get("status"):
        raws.append((status_hint["status"], business, "orders"))
    raws += _status_raws(own, candidates)
    best, best_raw, best_src = _best(raws)
    if best != ScannedOrder.SHIP_UNKNOWN:
        return best, best_raw, best_src

    # Everything below is the widening path, reached only when we'd otherwise
    # have to report "no status".
    if scope is None:
        scope = own
    siblings = [b for b in scope if b.pk != business.pk]

    # 2 — the same sub-order, in the other businesses.
    if siblings:
        best, best_raw, best_src = _best(_status_raws(siblings, candidates))
        if best != ScannedOrder.SHIP_UNKNOWN:
            return best, best_raw, best_src

    # 3 — reach the sub-order through the AWB or the raw code.
    codes = {sub_order_no}
    if row is not None:
        codes.update({row.awb_number, row.scanned_code})
    extra = _suborders_for_codes(scope, codes) - candidates
    if extra:
        candidates |= extra
        best, best_raw, best_src = _best(_status_raws(scope, extra))
        if best != ScannedOrder.SHIP_UNKNOWN:
            return best, best_raw, f"{best_src} · via AWB"

    # 4 — it came back, so it must have gone out.
    if _returned_suborders(scope, candidates):
        return ScannedOrder.SHIP_SHIPPED, "RETURNED", "return on file"

    return ScannedOrder.SHIP_UNKNOWN, "", ""


def _refresh_ship_status(business, row, status_hint=None, save=True, scope=None):
    """
    Re-resolve and store a row's shipping status.

    Worth re-running rather than trusting the stored value: a parcel scanned
    before its sheet was uploaded is UNKNOWN, and only becomes answerable later.
    """
    ship, raw, source = _resolve_ship_status(
        business, row.sub_order_no, status_hint, row=row, scope=scope
    )
    changed = (row.ship_status != ship) or (row.ship_status_raw != raw)
    row.ship_status = ship
    row.ship_status_raw = raw[:100]
    row.ship_source = (source or "")[:100]
    row.ship_checked_at = timezone.now()
    if save:
        row.save(update_fields=[
            "ship_status", "ship_status_raw", "ship_source", "ship_checked_at", "updated_at",
        ])
    return changed


def _scanned_order_stats(business):
    """
    Scan workload for the whole business — counted in the database so it stays
    cheap as the log grows, and never narrowed by the period filter so nothing
    on hold can hide behind the selected month.
    """
    base = ScannedOrder.objects.filter(business=business)
    today = timezone.localdate()

    by_status = {
        row["status"]: row["n"]
        for row in base.values("status").annotate(n=Count("id"))
    }

    return {
        "total":       base.count(),
        "scanned":     by_status.get(ScannedOrder.STATUS_SCANNED, 0),
        "packed":      by_status.get(ScannedOrder.STATUS_PACKED, 0),
        "dispatched":  by_status.get(ScannedOrder.STATUS_DISPATCHED, 0),
        "on_hold":     by_status.get(ScannedOrder.STATUS_ON_HOLD, 0),
        "issue":       by_status.get(ScannedOrder.STATUS_ISSUE, 0),
        "cancelled":   by_status.get(ScannedOrder.STATUS_CANCELLED, 0),
        "open":        base.filter(status__in=ScannedOrder.OPEN_STATUSES).count(),
        "needs_attention": base.filter(status__in=ScannedOrder.ATTENTION_STATUSES).count(),
        "unrecognised":    base.filter(matched_from=ScannedOrder.MATCH_NONE).count(),
        "scanned_today":   base.filter(
            last_scanned_at__gte=_local_day_start(today),
            last_scanned_at__lt=_local_day_start(today + timedelta(days=1)),
        ).count(),
        "rescanned":       base.filter(scan_count__gt=1).count(),

        # Shipping cross-check: scanned by us, but what does Meesho say?
        "ship_shipped":     base.filter(ship_status=ScannedOrder.SHIP_SHIPPED).count(),
        "ship_not_shipped": base.filter(ship_status=ScannedOrder.SHIP_NOT_SHIPPED).count(),
        "ship_cancelled":   base.filter(ship_status=ScannedOrder.SHIP_CANCELLED).count(),
        "ship_unknown":     base.filter(ship_status=ScannedOrder.SHIP_UNKNOWN).count(),
        "ship_problem":     base.filter(ship_status__in=[
            ScannedOrder.SHIP_NOT_SHIPPED, ScannedOrder.SHIP_CANCELLED,
        ]).count(),
    }


@api_view(["GET", "POST"])
def scanned_orders_list(request, business_id):
    """
    GET — the scan log, paginated, with each row's current Meesho status attached.

      q          — matches sub-order / AWB / scanned code / SKU / product / customer
      status     — a STATUS_* value, or "OPEN" for anything still in our hands
      matched    — "yes" | "no" (whether the code resolved to a known order)
      date_from / date_to — on the scan date
      page, page_size

    POST — record a scan: {"code": "...", "status": "...", "notes": "..."}.
      Re-scanning an order already in the log bumps its scan_count and returns
      already_scanned=true instead of creating a second row, so the desk is told
      about the duplicate rather than silently logging it twice.
    """
    business = get_authorized_business(request, business_id)

    if request.method == "POST":
        payload = request.data if isinstance(request.data, dict) else {}
        code = _clean_scan_code(payload.get("code") or payload.get("q"))
        if not code:
            return Response({"error": "Nothing scanned."}, status=status.HTTP_400_BAD_REQUEST)
        if len(code) > 200:
            return Response({"error": "That code is too long to be a barcode."},
                            status=status.HTTP_400_BAD_REQUEST)

        new_status = str(payload.get("status") or ScannedOrder.STATUS_SCANNED).strip().upper()
        if new_status not in _SCAN_STATUSES:
            return Response(
                {"error": f"Invalid status. Expected one of: {', '.join(sorted(_SCAN_STATUSES))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Pooled: a code belonging to a sibling business resolves and is filed
        # against that business, not the one in the dropdown.
        owner, snapshot = _resolve_scan_code_pooled(request, business, code)
        cross_business = owner.pk != business.pk
        now = timezone.now()

        existing = ScannedOrder.objects.filter(
            business=owner, sub_order_no=snapshot["sub_order_no"]
        ).first()

        if existing:
            # Already logged. Bump the scan, and only move the status when the
            # caller actually asked for one — a plain re-scan must not quietly
            # reset a parcel that was already marked dispatched.
            existing.scan_count += 1
            existing.last_scanned_at = now
            existing.scanned_code = code
            existing.scanned_by = request.user
            fields = ["scan_count", "last_scanned_at", "scanned_code", "scanned_by", "updated_at"]

            if payload.get("status") and new_status != existing.status:
                existing.status = new_status
                existing.status_updated_at = now
                fields += ["status", "status_updated_at"]
            if "notes" in payload:
                existing.notes = str(payload["notes"] or "").strip()
                fields.append("notes")

            existing.save(update_fields=fields)
            row = _attach_meesho_status(business, [existing])[0]
            # Re-check on every scan: the sheet that answers this may have been
            # uploaded since the parcel was first seen. Resolved against the
            # owning business, which for a pooled scan is not the selected one.
            _refresh_ship_status(owner, row, row.meesho_status,
                                 scope=accessible_businesses(request, include=owner))
            return Response({
                "already_scanned": True,
                "scan_count": row.scan_count,
                "matched_from": row.matched_from,
                "ship_status": row.ship_status,
                "cross_business": cross_business,
                "business_id": owner.pk,
                "business_name": owner.name,
                "row": ScannedOrderSerializer(row).data,
                "stats": _scanned_order_stats(business),
            })

        row = ScannedOrder.objects.create(
            business=owner,
            scanned_code=code,
            status=new_status,
            status_updated_at=now if new_status != ScannedOrder.STATUS_SCANNED else None,
            notes=str(payload.get("notes") or "").strip(),
            last_scanned_at=now,
            scanned_by=request.user,
            **snapshot,
        )
        row = _attach_meesho_status(business, [row])[0]
        _refresh_ship_status(owner, row, row.meesho_status,
                             scope=accessible_businesses(request, include=owner))
        return Response({
            "already_scanned": False,
            "scan_count": 1,
            "matched_from": row.matched_from,
            "ship_status": row.ship_status,
            "cross_business": cross_business,
            "business_id": owner.pk,
            "business_name": owner.name,
            "row": ScannedOrderSerializer(row).data,
            "stats": _scanned_order_stats(business),
        }, status=status.HTTP_201_CREATED)

    # ?pool=1 widens the list to every business the user belongs to. Off by
    # default so the tab still means "this business" unless asked otherwise —
    # but a scan filed against a sibling (see _resolve_scan_code_pooled) is
    # invisible without it, which is exactly when you want it on.
    pooled = request.GET.get("pool", "").strip() in ("1", "true", "yes")
    if pooled:
        scope = accessible_businesses(request, include=business)
        qs = ScannedOrder.objects.filter(business__in=scope)
    else:
        qs = ScannedOrder.objects.filter(business=business)
    qs = qs.select_related("scanned_by", "business")

    # Half-open datetime bounds rather than __date — see _local_day_start.
    date_from = _parse_day(request.GET.get("date_from"))
    date_to   = _parse_day(request.GET.get("date_to"))
    if date_from:
        qs = qs.filter(last_scanned_at__gte=_local_day_start(date_from))
    if date_to:
        qs = qs.filter(last_scanned_at__lt=_local_day_start(date_to + timedelta(days=1)))

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(sub_order_no__icontains=search) |
            DQ(scanned_code__icontains=search) |
            DQ(awb_number__icontains=search) |
            DQ(sku__icontains=search) |
            DQ(product_name__icontains=search) |
            DQ(customer_name__icontains=search)
        )

    wanted = request.GET.get("status", "").strip().upper()
    if wanted == "OPEN":
        qs = qs.filter(status__in=ScannedOrder.OPEN_STATUSES)
    elif wanted == "ATTENTION":
        qs = qs.filter(status__in=ScannedOrder.ATTENTION_STATUSES)
    elif wanted:
        qs = qs.filter(status=wanted)

    matched = request.GET.get("matched", "").strip().lower()
    if matched == "no":
        qs = qs.filter(matched_from=ScannedOrder.MATCH_NONE)
    elif matched == "yes":
        qs = qs.exclude(matched_from=ScannedOrder.MATCH_NONE)

    # ?ship=shipped | not_shipped | cancelled | unknown | problem
    ship = request.GET.get("ship", "").strip().upper()
    if ship == "PROBLEM":
        qs = qs.filter(ship_status__in=[
            ScannedOrder.SHIP_NOT_SHIPPED, ScannedOrder.SHIP_CANCELLED,
        ])
    elif ship in {c[0] for c in ScannedOrder.SHIP_STATUS_CHOICES}:
        qs = qs.filter(ship_status=ship)

    total = qs.count()

    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 50))))
    except (TypeError, ValueError):
        page_size = 50

    start = (page - 1) * page_size
    rows  = _attach_meesho_status(business, qs[start:start + page_size])

    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": ScannedOrderSerializer(rows, many=True).data,
        "stats": _scanned_order_stats(business),
        "statuses": [{"value": v, "label": l} for v, l in ScannedOrder.STATUS_CHOICES],
    })


@api_view(["GET", "PATCH", "DELETE"])
def scanned_order_detail(request, business_id, pk):
    """Read one recorded scan, update its status/notes, or delete a mis-scan."""
    business = get_authorized_business(request, business_id)

    try:
        row = ScannedOrder.objects.select_related("scanned_by").get(pk=pk, business=business)
    except ScannedOrder.DoesNotExist:
        return Response({"error": "Scan not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        row.delete()
        return Response({"deleted": True, "stats": _scanned_order_stats(business)})

    if request.method == "PATCH":
        payload = request.data if isinstance(request.data, dict) else {}
        unknown = set(payload) - _SCAN_WRITABLE
        if unknown:
            return Response(
                {"error": f"Not editable: {', '.join(sorted(unknown))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if "status" in payload:
            new_status = str(payload["status"] or "").strip().upper()
            if new_status not in _SCAN_STATUSES:
                return Response(
                    {"error": f"Invalid status. Expected one of: {', '.join(sorted(_SCAN_STATUSES))}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if new_status != row.status:
                row.status = new_status
                row.status_updated_at = timezone.now()

        if "notes" in payload:
            row.notes = str(payload["notes"] or "").strip()

        row.save()

    row = _attach_meesho_status(business, [row])[0]
    data = ScannedOrderSerializer(row).data
    data["stats"] = _scanned_order_stats(business)
    return Response(data)


@api_view(["POST"])
def scanned_orders_refresh_ship(request, business_id):
    """
    Re-check every scan's shipping status against the sheets we now hold.

    Needed because a parcel is usually scanned *before* the sheet that reports on
    it exists, so a freshly-scanned row is legitimately UNKNOWN. Uploading this
    week's Orders export is what turns those into real answers, and nothing else
    would notice.

    Only rows that could still change are touched: a SHIPPED row is terminal for
    this purpose, so re-resolving thousands of them every time would be waste.
    """
    business = get_authorized_business(request, business_id)

    qs = ScannedOrder.objects.filter(business=business).exclude(
        ship_status=ScannedOrder.SHIP_SHIPPED
    )
    if request.data.get("all"):
        qs = ScannedOrder.objects.filter(business=business)

    rows = list(qs)
    status_map = _meesho_status_map(business, [r.sub_order_no for r in rows])
    # Resolved with the full pool available, so a row whose Orders sheet was
    # uploaded under a different business stops reading as "no status".
    scope = accessible_businesses(request, include=business)

    changed = 0
    for row in rows:
        if _refresh_ship_status(business, row, status_map.get(row.sub_order_no), scope=scope):
            changed += 1

    return Response({
        "checked": len(rows),
        "changed": changed,
        "stats": _scanned_order_stats(business),
    })


@api_view(["POST"])
def scanned_orders_bulk_status(request, business_id):
    """
    Move a list of recorded scans to one status — the "select the day's pile and
    mark it all dispatched" action.
    """
    business = get_authorized_business(request, business_id)
    payload  = request.data if isinstance(request.data, dict) else {}

    ids = payload.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return Response({"error": "Give a non-empty list of scan ids."},
                        status=status.HTTP_400_BAD_REQUEST)

    new_status = str(payload.get("status") or "").strip().upper()
    if new_status not in _SCAN_STATUSES:
        return Response(
            {"error": f"Invalid status. Expected one of: {', '.join(sorted(_SCAN_STATUSES))}"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    updated = ScannedOrder.objects.filter(business=business, id__in=ids).exclude(
        status=new_status
    ).update(status=new_status, status_updated_at=timezone.now())

    return Response({
        "updated": updated,
        "status": new_status,
        "stats": _scanned_order_stats(business),
    })


# ══════════════════════════════════════════════════════════════════════════════
# Worker tasks + wallet — paid piecework for the team
# ══════════════════════════════════════════════════════════════════════════════

def _is_admin(user):
    """
    Who may create, review and pay.

    Workers are ordinary business_user accounts (no separate role), so the
    dividing line is super_admin. Everything below treats a non-admin as a
    worker who can only ever see and act on their own tasks — enforced here in
    the queryset, not just hidden in the UI.
    """
    return getattr(user, "role", None) == User.ROLE_SUPER_ADMIN


def _credit(task, kind, amount, by_user, earner, note="", listing=None):
    """
    Write one ledger line, or nothing if there's nothing to pay.

    `earner` is explicit rather than read off the task: a task now has several
    assignees, so who gets paid depends on who did that particular listing.

    Returns the entry (or None). Callers guard against double-payment with the
    *_credited_at stamps; this stays dumb on purpose so "has it been paid" has
    exactly one owner.
    """
    amount = safe_decimal(amount) or Decimal("0")
    if amount == 0:
        return None
    return WalletEntry.objects.create(
        business=task.business,
        user=earner,
        task=task,
        listing=listing,
        kind=kind,
        amount=amount,
        note=note,
        created_by=by_user,
    )


def _wallet_totals(qs):
    """(earned, settled, pending) for a set of ledger entries."""
    earned = qs.aggregate(t=Sum("amount"))["t"] or Decimal("0")
    settled = qs.filter(settlement__isnull=False).aggregate(t=Sum("amount"))["t"] or Decimal("0")
    return float(earned), float(settled), float(earned - settled)


def _resolve_rate(business, platform, task_type, user=None):
    """
    The rate to use for a new task: that specific worker's override if the
    admin has set one, else the business-wide standing rate. Same lookup for
    both — a missing user-specific row falls straight through to today's
    (business, platform, task_type) rate, so a business with no per-user rates
    set behaves exactly as before.
    """
    if user is not None:
        rate = PlatformRate.objects.filter(
            business=business, platform=platform, task_type=task_type, user=user
        ).first()
        if rate:
            return rate
    return PlatformRate.objects.filter(
        business=business, platform=platform, task_type=task_type, user__isnull=True
    ).first()


@api_view(["GET", "POST"])
def worker_tasks_list(request, business_id):
    """
    GET  — tasks. Admins see everything for the business; a worker sees only
           their own, whatever they ask for.
           Filters: status, task_type, assigned_to (admin only), q
    POST — create and assign a task (admin only).
    """
    business = get_authorized_business(request, business_id)
    admin = _is_admin(request.user)

    if request.method == "POST":
        if not admin:
            return Response({"error": "Only an admin can create tasks."},
                            status=status.HTTP_403_FORBIDDEN)
        payload = request.data if isinstance(request.data, dict) else {}

        task_type = str(payload.get("task_type") or "").strip().upper()
        if task_type not in {c[0] for c in WorkerTask.TYPE_CHOICES}:
            return Response({"error": "task_type must be LISTING or RETURN_CLAIM."},
                            status=status.HTTP_400_BAD_REQUEST)

        raw_ids = payload.get("assignees") or payload.get("assigned_to")
        if not isinstance(raw_ids, list):
            raw_ids = [raw_ids] if raw_ids else []
        assignees = list(User.objects.filter(pk__in=[i for i in raw_ids if i]))
        if not assignees:
            return Response({"error": "Pick at least one person for this task."},
                            status=status.HTTP_400_BAD_REQUEST)

        platform = str(payload.get("platform") or WorkerTask.PLATFORM_MEESHO).strip().upper()
        if platform not in {c[0] for c in WorkerTask.PLATFORM_CHOICES}:
            return Response({"error": "Unknown platform."}, status=status.HTTP_400_BAD_REQUEST)

        # The rate is whatever is currently set for this platform, unless the
        # caller overrides it. A task with exactly one assignee prefers that
        # person's own standing rate (if the admin set one) over the business
        # default; a shared task with several people still pays one flat rate,
        # since there's nowhere on a single task to hold a rate per assignee.
        # Copied onto the task either way, so a later rate change can never
        # re-price work that has already been handed out.
        rate = _resolve_rate(business, platform, task_type,
                             user=assignees[0] if len(assignees) == 1 else None)
        reward = safe_decimal(payload.get("reward_amount"))
        if reward is None:
            reward = rate.reward_amount if rate else Decimal("0")
        bonus = safe_decimal(payload.get("bonus_amount"))
        if bonus is None:
            bonus = rate.bonus_amount if rate else Decimal("0")
        if reward < 0 or bonus < 0:
            return Response({"error": "Amounts can't be negative."},
                            status=status.HTTP_400_BAD_REQUEST)

        parent_sku = None
        parent_item_id = str(payload.get("parent_sku_item_id") or "").strip()
        if parent_item_id:
            try:
                parent_sku = ParentItemPrice.objects.get(business=business, item_id=parent_item_id)
            except ParentItemPrice.DoesNotExist:
                return Response({"error": f"No parent SKU '{parent_item_id}' in this business."},
                                status=status.HTTP_400_BAD_REQUEST)

        listing_template = None
        template_id = payload.get("listing_template")
        if template_id:
            try:
                listing_template = ListingTemplate.objects.get(pk=template_id, business=business)
            except (ListingTemplate.DoesNotExist, ValueError, TypeError):
                return Response({"error": "That listing template doesn't exist."},
                                status=status.HTTP_400_BAD_REQUEST)

        task = WorkerTask.objects.create(
            business=business,
            task_type=task_type,
            platform=platform,
            title=str(payload.get("title") or "").strip()[:200],
            source_link=str(payload.get("source_link") or "").strip(),
            instructions=str(payload.get("instructions") or "").strip(),
            suborder_no=str(payload.get("suborder_no") or "").strip()[:100],
            created_by=request.user,
            reward_amount=reward,
            bonus_amount=bonus if task_type == WorkerTask.TYPE_RETURN_CLAIM else Decimal("0"),
            listing_defaults=payload.get("listing_defaults") or {},
            parent_sku=parent_sku,
            listing_template=listing_template,
        )
        task.assignees.set(assignees)
        return Response(WorkerTaskSerializer(task, context={"request": request}).data, status=status.HTTP_201_CREATED)

    qs = WorkerTask.objects.filter(business=business).select_related(
        "created_by", "reviewed_by", "business"
    ).prefetch_related("wallet_entries", "assignees", "listings")

    if not admin:
        qs = qs.filter(assignees=request.user)
    elif request.GET.get("assigned_to"):
        qs = qs.filter(assignees__id=request.GET["assigned_to"])

    platform = request.GET.get("platform", "").strip().upper()
    if platform:
        qs = qs.filter(platform=platform)

    wanted = request.GET.get("status", "").strip().upper()
    if wanted == "OPEN":
        qs = qs.filter(status__in=WorkerTask.OPEN_STATUSES)
    elif wanted:
        qs = qs.filter(status=wanted)

    task_type = request.GET.get("task_type", "").strip().upper()
    if task_type:
        qs = qs.filter(task_type=task_type)

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(title__icontains=search) |
            DQ(submitted_sku__icontains=search) |
            DQ(suborder_no__icontains=search) |
            DQ(submitted_reference__icontains=search)
        )

    rows = list(qs[:500])
    by_status = {
        r["status"]: r["n"]
        for r in qs.values("status").annotate(n=Count("id"))
    }

    # The wallet summary alongside the tasks, scoped the same way — a worker
    # sees their own balance, an admin sees the whole business's liability.
    ledger = WalletEntry.objects.filter(business=business)
    if not admin:
        ledger = ledger.filter(user=request.user)
    earned, settled, pending = _wallet_totals(ledger)

    return Response({
        "is_admin": admin,
        "total": len(rows),
        "results": WorkerTaskSerializer(rows, many=True, context={"request": request}).data,
        "stats": {
            "assigned":  by_status.get(WorkerTask.STATUS_ASSIGNED, 0),
            "submitted": by_status.get(WorkerTask.STATUS_SUBMITTED, 0),
            "approved":  by_status.get(WorkerTask.STATUS_APPROVED, 0),
            "rejected":  by_status.get(WorkerTask.STATUS_REJECTED, 0),
            "awaiting_bonus": qs.filter(
                task_type=WorkerTask.TYPE_RETURN_CLAIM,
                status=WorkerTask.STATUS_APPROVED,
                bonus_credited_at__isnull=True,
            ).exclude(bonus_amount=0).count(),
            "earned": earned,
            "settled": settled,
            "pending": pending,
        },
    })


@api_view(["GET", "PATCH", "DELETE"])
def worker_task_detail(request, business_id, pk):
    """Read one task; an admin may edit its brief or delete it before work starts."""
    business = get_authorized_business(request, business_id)
    admin = _is_admin(request.user)

    try:
        task = WorkerTask.objects.select_related(
            "created_by", "reviewed_by", "business"
        ).prefetch_related("assignees", "listings").get(pk=pk, business=business)
    except WorkerTask.DoesNotExist:
        return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

    if not admin and not task.assignees.filter(pk=request.user.pk).exists():
        return Response({"error": "That task isn't yours."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == "DELETE":
        if not admin:
            return Response({"error": "Only an admin can delete a task."},
                            status=status.HTTP_403_FORBIDDEN)
        # Deleting a paid task is allowed: the ledger survives it. WalletEntry
        # keeps the money with the worker and nulls its task link, so balances
        # and settlement history are unaffected — only the brief disappears.
        paid = float(task.wallet_entries.aggregate(t=Sum("amount"))["t"] or 0)
        task.delete()
        return Response({"deleted": True, "wallet_entries_kept": paid > 0, "amount_kept": paid})

    if request.method == "PATCH":
        if not admin:
            return Response({"error": "Only an admin can edit a task."},
                            status=status.HTTP_403_FORBIDDEN)
        editable = {"title", "source_link", "instructions", "suborder_no",
                    "reward_amount", "bonus_amount", "assignees", "platform", "is_paused",
                    "listing_defaults", "parent_sku_item_id", "listing_template"}
        payload = request.data if isinstance(request.data, dict) else {}
        unknown = set(payload) - editable
        if unknown:
            return Response({"error": f"Not editable: {', '.join(sorted(unknown))}"},
                            status=status.HTTP_400_BAD_REQUEST)

        for field in ("title", "source_link", "instructions", "suborder_no"):
            if field in payload:
                setattr(task, field, str(payload[field] or "").strip())
        for field in ("reward_amount", "bonus_amount"):
            if field in payload:
                value = safe_decimal(payload[field]) or Decimal("0")
                if value < 0:
                    return Response({"error": "Amounts can't be negative."},
                                    status=status.HTTP_400_BAD_REQUEST)
                setattr(task, field, value)
        if "listing_defaults" in payload:
            defaults = payload["listing_defaults"]
            if not isinstance(defaults, dict):
                return Response({"error": "listing_defaults must be an object."},
                                status=status.HTTP_400_BAD_REQUEST)
            task.listing_defaults = defaults
        if "platform" in payload:
            platform = str(payload["platform"] or "").strip().upper()
            if platform not in {c[0] for c in WorkerTask.PLATFORM_CHOICES}:
                return Response({"error": "Unknown platform."}, status=status.HTTP_400_BAD_REQUEST)
            task.platform = platform
        if "is_paused" in payload:
            paused = bool(payload["is_paused"])
            task.is_paused = paused
            task.paused_at = timezone.now() if paused else None
        if "parent_sku_item_id" in payload:
            parent_item_id = str(payload["parent_sku_item_id"] or "").strip()
            if not parent_item_id:
                task.parent_sku = None
            else:
                try:
                    task.parent_sku = ParentItemPrice.objects.get(business=business, item_id=parent_item_id)
                except ParentItemPrice.DoesNotExist:
                    return Response({"error": f"No parent SKU '{parent_item_id}' in this business."},
                                    status=status.HTTP_400_BAD_REQUEST)
        if "listing_template" in payload:
            template_id = payload["listing_template"]
            if not template_id:
                task.listing_template = None
            else:
                try:
                    task.listing_template = ListingTemplate.objects.get(pk=template_id, business=business)
                except (ListingTemplate.DoesNotExist, ValueError, TypeError):
                    return Response({"error": "That listing template doesn't exist."},
                                    status=status.HTTP_400_BAD_REQUEST)
        task.save()
        if "assignees" in payload:
            ids = payload["assignees"] if isinstance(payload["assignees"], list) else [payload["assignees"]]
            people = list(User.objects.filter(pk__in=[i for i in ids if i]))
            if not people:
                return Response({"error": "A task needs at least one person."},
                                status=status.HTTP_400_BAD_REQUEST)
            task.assignees.set(people)

    return Response(WorkerTaskSerializer(task, context={"request": request}).data)


@api_view(["POST"])
def worker_task_submit(request, business_id, pk):
    """
    The worker reporting the job done — the SKU they listed, or the claim they
    raised. Re-submitting a rejected task is allowed and puts it back in the
    review queue, because "fix it and send it again" is the normal path.
    """
    business = get_authorized_business(request, business_id)

    try:
        task = WorkerTask.objects.select_related("business").get(pk=pk, business=business)
    except WorkerTask.DoesNotExist:
        return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

    if not task.assignees.filter(pk=request.user.pk).exists() and not _is_admin(request.user):
        return Response({"error": "That task isn't yours."}, status=status.HTTP_403_FORBIDDEN)
    if task.status == WorkerTask.STATUS_APPROVED:
        return Response({"error": "This task is already approved."},
                        status=status.HTTP_400_BAD_REQUEST)

    payload = request.data if isinstance(request.data, dict) else {}
    sku = str(payload.get("submitted_sku") or "").strip()[:300]
    reference = str(payload.get("submitted_reference") or "").strip()[:200]

    if task.task_type == WorkerTask.TYPE_LISTING and not sku:
        return Response({"error": "Enter the SKU id of the listing you created."},
                        status=status.HTTP_400_BAD_REQUEST)

    task.submitted_sku = sku
    task.submitted_reference = reference
    task.submitted_note = str(payload.get("submitted_note") or "").strip()
    task.submitted_at = timezone.now()
    task.submitted_by = request.user
    task.status = WorkerTask.STATUS_SUBMITTED
    # A resubmission is a fresh ask, so the previous verdict shouldn't linger.
    task.review_comment = ""
    task.reviewed_by = None
    task.reviewed_at = None
    task.save()

    return Response(WorkerTaskSerializer(task, context={"request": request}).data)


@api_view(["POST"])
def worker_task_review(request, business_id, pk):
    """
    Approve or reject submitted work. Approving is what pays:

      LISTING       → the task's reward
      RETURN_CLAIM  → the raise fee now; the Meesho-approval bonus is released
                      later by the claim sheet (see _release_claim_bonuses)

    Rejecting pays nothing and records why. Money already credited is never
    clawed back automatically — use a manual adjustment, so a reversal is always
    a deliberate, attributable act.
    """
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"error": "Only an admin can review work."},
                        status=status.HTTP_403_FORBIDDEN)

    try:
        task = WorkerTask.objects.select_related("business").get(pk=pk, business=business)
    except WorkerTask.DoesNotExist:
        return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

    payload = request.data if isinstance(request.data, dict) else {}
    decision = str(payload.get("decision") or "").strip().upper()
    if decision not in ("APPROVE", "REJECT"):
        return Response({"error": "decision must be APPROVE or REJECT."},
                        status=status.HTTP_400_BAD_REQUEST)

    now = timezone.now()
    task.review_comment = str(payload.get("comment") or "").strip()
    task.reviewed_by = request.user
    task.reviewed_at = now
    credited = None

    if decision == "REJECT":
        task.status = WorkerTask.STATUS_REJECTED
        task.save()
        return Response({"task": WorkerTaskSerializer(task, context={"request": request}).data, "credited": None})

    task.status = WorkerTask.STATUS_APPROVED
    # Listing pay happens per SKU (see task_listing_review); only claim work is
    # paid at the task level, and it goes to whoever submitted it.
    if task.task_type == WorkerTask.TYPE_RETURN_CLAIM and task.reward_credited_at is None:
        earner = task.submitted_by or task.assignees.first()
        entry = _credit(task, WalletEntry.KIND_CLAIM_RAISED, task.reward_amount,
                        request.user, earner,
                        note=task.title or task.suborder_no) if earner else None
        if entry:
            task.reward_credited_at = now
            credited = float(entry.amount)
    task.save()

    return Response({
        "task": WorkerTaskSerializer(task, context={"request": request}).data,
        "credited": credited,
        "awaiting_bonus": task.awaiting_bonus,
    })


def _release_claim_bonuses(business, user=None):
    """
    Pay the Meesho-approval bonus on claim tasks whose ticket now reads Approved.

    Called after a ticket sheet import, which is the only thing that can tell us
    Meesho accepted a claim. Matching is on sub-order, and the task must already
    be approved by you — the bonus rewards a claim that stuck, not one that was
    never accepted internally.
    """
    pending = WorkerTask.objects.filter(
        business=business,
        task_type=WorkerTask.TYPE_RETURN_CLAIM,
        status=WorkerTask.STATUS_APPROVED,
        bonus_credited_at__isnull=True,
    ).exclude(bonus_amount=0).exclude(suborder_no="")

    if not pending:
        return 0, 0.0

    approved_subs = set(
        ClaimTicket.objects.filter(
            business=business,
            ticket_status=ClaimTicket.STATUS_APPROVED,
            suborder_no__in=[t.suborder_no for t in pending],
        ).values_list("suborder_no", flat=True)
    )

    now = timezone.now()
    count, total = 0, Decimal("0")
    for task in pending:
        if task.suborder_no not in approved_subs:
            continue
        earner = task.submitted_by or task.assignees.first()
        if earner is None:
            continue
        entry = _credit(task, WalletEntry.KIND_CLAIM_BONUS, task.bonus_amount, user, earner,
                        note=f"Meesho approved claim for {task.suborder_no}")
        if entry:
            task.bonus_credited_at = now
            task.save(update_fields=["bonus_credited_at", "updated_at"])
            count += 1
            total += entry.amount
    return count, float(total)


@api_view(["GET"])
def wallet_summary(request, business_id):
    """
    Balances. An admin gets one row per worker plus the ledger; a worker gets
    only their own, because a wallet is nobody else's business.
    """
    business = get_authorized_business(request, business_id)
    admin = _is_admin(request.user)

    entries = WalletEntry.objects.filter(business=business).select_related(
        "user", "task", "business"
    )
    if not admin:
        entries = entries.filter(user=request.user)

    per_user = []
    grouped = entries.values("user_id", "user__username").annotate(
        earned=Sum("amount"), n=Count("id"),
    ).order_by("user__username")
    for row in grouped:
        user_entries = entries.filter(user_id=row["user_id"])
        earned, settled, pending = _wallet_totals(user_entries)
        per_user.append({
            "user_id": row["user_id"],
            "username": row["user__username"],
            "entries": row["n"],
            "earned": earned,
            "settled": settled,
            "pending": pending,
        })

    earned, settled, pending = _wallet_totals(entries)
    return Response({
        "is_admin": admin,
        "totals": {"earned": earned, "settled": settled, "pending": pending},
        "per_user": per_user,
        "entries": WalletEntrySerializer(entries[:300], many=True).data,
        "settlements": WalletSettlementSerializer(
            WalletSettlement.objects.filter(business=business).select_related("user", "created_by")[:100]
            if admin else
            WalletSettlement.objects.filter(business=business, user=request.user).select_related("user", "created_by")[:100],
            many=True,
        ).data,
    })


@api_view(["POST"])
def wallet_settle(request, business_id):
    """
    Record a payout: stamps this worker's unsettled entries as covered.

    The amount is computed from the entries rather than taken from the client —
    paying against a number the browser supplied would let a stale page settle
    the wrong balance.
    """
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"error": "Only an admin can settle payments."},
                        status=status.HTTP_403_FORBIDDEN)

    payload = request.data if isinstance(request.data, dict) else {}
    try:
        worker = User.objects.get(pk=payload.get("user_id"))
    except (User.DoesNotExist, ValueError, TypeError):
        return Response({"error": "Pick a worker to settle."},
                        status=status.HTTP_400_BAD_REQUEST)

    with transaction.atomic():
        outstanding = (
            WalletEntry.objects
            .select_for_update()
            .filter(business=business, user=worker, settlement__isnull=True)
        )
        total = outstanding.aggregate(t=Sum("amount"))["t"] or Decimal("0")
        if total <= 0:
            return Response({"error": "Nothing outstanding for that worker."},
                            status=status.HTTP_400_BAD_REQUEST)

        settlement = WalletSettlement.objects.create(
            business=business,
            user=worker,
            amount=total,
            paid_on=_parse_day(payload.get("paid_on")) or timezone.localdate(),
            method=str(payload.get("method") or "").strip()[:60],
            reference=str(payload.get("reference") or "").strip()[:150],
            note=str(payload.get("note") or "").strip(),
            created_by=request.user,
        )
        covered = outstanding.update(settlement=settlement)

    return Response({
        "settlement": WalletSettlementSerializer(settlement).data,
        "entries_covered": covered,
    })


@api_view(["POST"])
def wallet_adjust(request, business_id):
    """
    A manual ledger line — a bonus, a correction, or a clawback (negative).

    Exists so that fixing a mistake is a recorded event with a reason, rather
    than someone editing a balance.
    """
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"error": "Only an admin can adjust a wallet."},
                        status=status.HTTP_403_FORBIDDEN)

    payload = request.data if isinstance(request.data, dict) else {}
    try:
        worker = User.objects.get(pk=payload.get("user_id"))
    except (User.DoesNotExist, ValueError, TypeError):
        return Response({"error": "Pick a worker."}, status=status.HTTP_400_BAD_REQUEST)

    amount = safe_decimal(payload.get("amount"))
    if amount is None or amount == 0:
        return Response({"error": "Enter a non-zero amount."},
                        status=status.HTTP_400_BAD_REQUEST)
    note = str(payload.get("note") or "").strip()
    if not note:
        return Response({"error": "Say why — an unexplained adjustment is worse than none."},
                        status=status.HTTP_400_BAD_REQUEST)

    entry = WalletEntry.objects.create(
        business=business, user=worker, kind=WalletEntry.KIND_ADJUSTMENT,
        amount=amount, note=note, created_by=request.user,
    )
    return Response(WalletEntrySerializer(entry).data, status=status.HTTP_201_CREATED)


# India's GST slabs. Fixed by law rather than by us, so they are a closed list —
# a free-text tax field is how you end up with 5.5% on an invoice.
GST_SLABS = ["0", "0.25", "3", "5", "12", "18", "28"]

_LISTING_WRITABLE = {
    "sku_id", "listing_price", "wrong_defective_price", "mrp",
    "min_settlement_amount", "sku_prefix", "hsn_code", "tax_percent", "notes",
}

# What a worker is allowed to send. Prices, HSN, tax and the required prefix are
# decided by the admin on the task and inherited by every SKU — a worker typing
# their own price would make the catalogue unreliable, and these are commercial
# decisions rather than data entry. Enforced here, not merely hidden in the UI.
_LISTING_WORKER_WRITABLE = {"sku_id", "notes"}
_LISTING_MONEY = {"listing_price", "wrong_defective_price", "mrp", "min_settlement_amount"}


@api_view(["GET"])
def listing_reference(request, business_id):
    """
    The dropdown data for a listing form: GST slabs, and the HSN codes this
    business actually uses with the rate it has actually charged against each.

    Derived from your own GST filings rather than a bundled HSN table: a generic
    list would be thousands of irrelevant codes, and any rate in it would be our
    guess. What you have invoiced is the honest answer for what you sell.
    """
    business = get_authorized_business(request, business_id)

    # The tax lines carry both the HSN and the rate charged, so the suggestion
    # is simply the rate this business has most often filed under that code.
    tally = {}
    for hsn, rate in (
        GstTransaction.objects.filter(business=business).exclude(hsn_code="")
        .values_list("hsn_code", "gst_rate")
    ):
        bucket = tally.setdefault(hsn, {"count": 0, "rates": {}})
        bucket["count"] += 1
        if rate is not None:
            key = str(int(rate)) if rate == int(rate) else str(rate)
            bucket["rates"][key] = bucket["rates"].get(key, 0) + 1

    # Codes that appear on invoices but never on a tax line still belong in the
    # dropdown — they're things this business genuinely sells.
    for hsn in (
        GstInvoiceDetail.objects.filter(business=business).exclude(hsn="")
        .values_list("hsn", flat=True).distinct()
    ):
        tally.setdefault(hsn, {"count": 0, "rates": {}})

    hsn_codes = []
    for hsn, info in sorted(tally.items(), key=lambda kv: -kv[1]["count"]):
        common = max(info["rates"].items(), key=lambda kv: kv[1])[0] if info["rates"] else None
        hsn_codes.append({"hsn": hsn, "used": info["count"], "suggested_tax_percent": common})

    return Response({
        "gst_slabs": GST_SLABS,
        "hsn_codes": hsn_codes,
        "platforms": [{"value": v, "label": l} for v, l in WorkerTask.PLATFORM_CHOICES],
    })


@api_view(["POST"])
def task_listing_add(request, business_id, pk):
    """
    Add one SKU to a listing task.

    Refuses a duplicate SKU outright — the same SKU listed twice is either a
    slip or two workers claiming the same job, and either way the tasker needs
    telling immediately rather than at review time. The check spans the whole
    business, not just this task, because that is the scope in which a SKU is
    meant to be unique.
    """
    business = get_authorized_business(request, business_id)

    try:
        task = WorkerTask.objects.get(pk=pk, business=business)
    except WorkerTask.DoesNotExist:
        return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

    admin = _is_admin(request.user)
    if not admin and not task.assignees.filter(pk=request.user.pk).exists():
        return Response({"error": "That task isn't yours."}, status=status.HTTP_403_FORBIDDEN)
    if task.task_type != WorkerTask.TYPE_LISTING:
        return Response({"error": "Only listing tasks take SKUs."},
                        status=status.HTTP_400_BAD_REQUEST)
    if task.is_paused:
        return Response({"error": "This task is paused — no new SKUs can be added."},
                        status=status.HTTP_400_BAD_REQUEST)

    payload = request.data if isinstance(request.data, dict) else {}
    sku_id = str(payload.get("sku_id") or "").strip()[:300]
    if not sku_id:
        return Response({"error": "Enter the SKU id."}, status=status.HTTP_400_BAD_REQUEST)

    clash = TaskListing.objects.filter(business=business, sku_id__iexact=sku_id).first()
    if clash:
        who = clash.created_by.username if clash.created_by_id else "someone"
        return Response({
            "error": f"SKU '{sku_id}' is already listed"
                     + (f" on task #{clash.task_id} by {who}." if clash.task_id else "."),
            "duplicate_of": {"task_id": clash.task_id, "listing_id": clash.pk, "by": who},
        }, status=status.HTTP_409_CONFLICT)

    if not admin:
        forbidden = set(payload) - _LISTING_WORKER_WRITABLE
        if forbidden:
            return Response(
                {"error": "Those details are set by the admin on the task: "
                          + ", ".join(sorted(forbidden))},
                status=status.HTTP_403_FORBIDDEN,
            )

    # Anything the worker leaves blank falls back to what you set on the task,
    # so the values you decided are applied even if the form is rushed.
    defaults = task.listing_defaults or {}
    fields = {}
    for key in _LISTING_WRITABLE - {"sku_id"}:
        raw = payload.get(key)
        if raw in (None, ""):
            raw = defaults.get(key)
        if raw in (None, ""):
            continue
        if key in _LISTING_MONEY or key == "tax_percent":
            fields[key] = safe_decimal(raw)
        else:
            fields[key] = str(raw).strip()

    listing = TaskListing.objects.create(
        business=business, task=task, created_by=request.user, sku_id=sku_id, **fields
    )
    return Response(TaskListingSerializer(listing, context={"request": request}).data, status=status.HTTP_201_CREATED)


@api_view(["GET"])
def worker_task_generate_sku(request, business_id, pk):
    """
    Suggest the next SKU id for this task: <prefix><next number>, zero-padded to
    3 digits (widening automatically past 999). The prefix is whatever the
    admin set on the task (listing_defaults.sku_prefix), or ?prefix= for an
    ad-hoc one; the next number is one past the highest already used with that
    prefix anywhere in the business, checked against both listed and catalogued
    SKUs so it can't suggest something already taken by either.

    Pure suggestion — nothing is written here, so nothing can race. The
    duplicate check in task_listing_add is still what actually guards a save;
    if two people generate the same number at once, the second one's Add just
    gets a 409 and regenerates.
    """
    business = get_authorized_business(request, business_id)
    try:
        task = WorkerTask.objects.get(pk=pk, business=business)
    except WorkerTask.DoesNotExist:
        return Response({"error": "Task not found."}, status=status.HTTP_404_NOT_FOUND)

    admin = _is_admin(request.user)
    if not admin and not task.assignees.filter(pk=request.user.pk).exists():
        return Response({"error": "That task isn't yours."}, status=status.HTTP_403_FORBIDDEN)

    prefix = str(request.GET.get("prefix") or (task.listing_defaults or {}).get("sku_prefix") or "").strip()
    if not prefix:
        return Response({"error": "Set a starting SKU prefix on this task first."},
                        status=status.HTTP_400_BAD_REQUEST)

    # Prefix match is cheap at the DB (indexed LIKE); the strict "prefix + only
    # digits" check runs in Python on that much smaller set.
    pattern = re.compile(r"^" + re.escape(prefix) + r"(\d+)$", re.IGNORECASE)
    candidates = (
        set(TaskListing.objects.filter(business=business, sku_id__istartswith=prefix)
            .values_list("sku_id", flat=True))
        | set(FinalPrice.objects.filter(business=business, sku_id__istartswith=prefix)
              .values_list("sku_id", flat=True))
    )
    highest = 0
    for sku_id in candidates:
        m = pattern.match((sku_id or "").strip())
        if m:
            highest = max(highest, int(m.group(1)))

    suggested = f"{prefix}{highest + 1:03d}"
    return Response({"sku_id": suggested, "prefix": prefix})


@api_view(["GET", "PATCH", "DELETE"])
def task_listing_detail(request, business_id, pk):
    """Edit or remove one SKU. A paused task, or an approved listing, is frozen."""
    business = get_authorized_business(request, business_id)
    admin = _is_admin(request.user)

    try:
        listing = TaskListing.objects.select_related("task", "created_by", "reviewed_by").get(
            pk=pk, business=business
        )
    except TaskListing.DoesNotExist:
        return Response({"error": "Listing not found."}, status=status.HTTP_404_NOT_FOUND)

    mine = listing.created_by_id == request.user.pk
    if not admin and not mine:
        return Response({"error": "That listing isn't yours."}, status=status.HTTP_403_FORBIDDEN)

    if request.method == "GET":
        return Response(TaskListingSerializer(listing, context={"request": request}).data)

    frozen = listing.task.is_paused or listing.status == TaskListing.STATUS_APPROVED
    if frozen and not admin:
        reason = ("This task is paused — its listings are read-only."
                  if listing.task.is_paused else "This listing is approved and can't be changed.")
        return Response({"error": reason}, status=status.HTTP_400_BAD_REQUEST)

    if request.method == "DELETE":
        if listing.wallet_entries.exists():
            return Response({"error": "This listing has already been paid for."},
                            status=status.HTTP_400_BAD_REQUEST)
        listing.delete()
        return Response({"deleted": True})

    payload = request.data if isinstance(request.data, dict) else {}
    allowed = _LISTING_WRITABLE if admin else _LISTING_WORKER_WRITABLE
    unknown = set(payload) - allowed
    if unknown:
        return Response(
            {"error": ("Those details are set by the admin on the task: "
                       if not admin else "Not editable: ") + ", ".join(sorted(unknown))},
            status=status.HTTP_403_FORBIDDEN if not admin else status.HTTP_400_BAD_REQUEST,
        )

    if "sku_id" in payload:
        new_sku = str(payload["sku_id"] or "").strip()[:300]
        if not new_sku:
            return Response({"error": "SKU id can't be empty."}, status=status.HTTP_400_BAD_REQUEST)
        if TaskListing.objects.filter(business=business, sku_id__iexact=new_sku).exclude(pk=listing.pk).exists():
            return Response({"error": f"SKU '{new_sku}' is already listed."},
                            status=status.HTTP_409_CONFLICT)
        listing.sku_id = new_sku

    for key in _LISTING_WRITABLE - {"sku_id"}:
        if key not in payload:
            continue
        if key in _LISTING_MONEY or key == "tax_percent":
            setattr(listing, key, safe_decimal(payload[key]))
        else:
            setattr(listing, key, str(payload[key] or "").strip())

    # An edit is a fresh ask, so a previous rejection shouldn't stand.
    if listing.status == TaskListing.STATUS_REJECTED:
        listing.status = TaskListing.STATUS_PENDING
        listing.review_comment = ""
    listing.save()
    return Response(TaskListingSerializer(listing, context={"request": request}).data)


@api_view(["POST"])
def task_listing_review(request, business_id, pk):
    """
    Approve or reject one SKU. Approving is what pays, at the task's rate,
    frozen onto the listing so a later rate change can't rewrite history.
    """
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"error": "Only an admin can review listings."},
                        status=status.HTTP_403_FORBIDDEN)

    try:
        listing = TaskListing.objects.select_related("task", "created_by").get(
            pk=pk, business=business
        )
    except TaskListing.DoesNotExist:
        return Response({"error": "Listing not found."}, status=status.HTTP_404_NOT_FOUND)

    payload = request.data if isinstance(request.data, dict) else {}
    decision = str(payload.get("decision") or "").strip().upper()
    if decision not in ("APPROVE", "REJECT"):
        return Response({"error": "decision must be APPROVE or REJECT."},
                        status=status.HTTP_400_BAD_REQUEST)

    now = timezone.now()
    listing.review_comment = str(payload.get("comment") or "").strip()
    listing.reviewed_by = request.user
    listing.reviewed_at = now
    credited = None

    if decision == "REJECT":
        listing.status = TaskListing.STATUS_REJECTED
        listing.save()
        return Response({"listing": TaskListingSerializer(listing, context={"request": request}).data, "credited": None})

    listing.status = TaskListing.STATUS_APPROVED

    # Everything approval does — joining the catalogue (and syncing to the
    # task's parent SKU, if any) and paying out — happens in one transaction,
    # so a failure partway through can't leave a listing marked "approved"
    # without its SKU actually in the catalogue, or paid without being linked.
    catalog_created = False
    parent_linked = False
    with transaction.atomic():
        # Approving is what makes the SKU real: it joins the pricing catalogue
        # so the rest of the app (pricing, inventory, SKU analysis) can see it.
        # Only ever done here, so a SKU can't enter the catalogue without your
        # sign-off.
        if listing.added_to_catalog_at is None:
            if listing.task.parent_sku_id:
                # Linked to a parent: reuse the same bulk-link path the pricing
                # tab uses, so the SKU is created *and* immediately inherits the
                # parent's price/tax/packaging instead of landing unpriced.
                result = _bulk_link_skus_to_parent(
                    business=business, parent=listing.task.parent_sku, sku_ids=[listing.sku_id],
                )
                catalog_created = bool(result.get("created"))
                parent_linked = result.get("failed", 0) == 0
            else:
                # Deliberately only the tax rate is carried over. FinalPrice.item_price
                # is the *purchase* cost, and the listing's price is what it sells
                # for — copying one into the other would quietly corrupt every
                # profit figure.
                _, catalog_created = FinalPrice.objects.get_or_create(
                    business=business, sku_id=listing.sku_id,
                    defaults={"tax_percent": int(listing.tax_percent) if listing.tax_percent is not None else None},
                )
            listing.added_to_catalog_at = now

        if listing.reward_credited_at is None:
            listing.reward_amount = listing.task.reward_amount
            entry = _credit(listing.task, WalletEntry.KIND_LISTING, listing.reward_amount,
                            request.user, listing.created_by,
                            note=f"{listing.sku_id} ({listing.task.platform})", listing=listing)
            if entry:
                listing.reward_credited_at = now
                credited = float(entry.amount)

        listing.save()

    return Response({
        "listing": TaskListingSerializer(listing, context={"request": request}).data,
        "credited": credited,
        "catalog_sku_created": catalog_created,
        "parent_linked": parent_linked,
    })


@api_view(["GET", "PUT"])
def platform_rates(request, business_id):
    """
    The standing rate per platform, and optionally a per-worker override on top
    of it (see PlatformRate.user). Read by everyone (a worker should be able to
    see what work is worth, including their own override); only an admin can
    change it, and only an admin can see someone else's override.
    """
    business = get_authorized_business(request, business_id)
    admin = _is_admin(request.user)

    if request.method == "PUT":
        if not admin:
            return Response({"error": "Only an admin can set rates."},
                            status=status.HTTP_403_FORBIDDEN)
        rows = request.data.get("rates") if isinstance(request.data, dict) else None
        if not isinstance(rows, list):
            return Response({"error": "Send a list of rates."}, status=status.HTTP_400_BAD_REQUEST)

        for row in rows:
            platform = str(row.get("platform") or "").strip().upper()
            task_type = str(row.get("task_type") or WorkerTask.TYPE_LISTING).strip().upper()
            if platform not in {c[0] for c in WorkerTask.PLATFORM_CHOICES}:
                continue
            # A row without a user is the business default; one with a user_id
            # is that specific person's override, checked first when a task is
            # created for them (see _resolve_rate).
            user = None
            raw_user_id = row.get("user_id")
            if raw_user_id:
                user = User.objects.filter(pk=raw_user_id, memberships__business=business).first()
                if user is None:
                    return Response({"error": f"No such worker (id {raw_user_id}) in this business."},
                                    status=status.HTTP_400_BAD_REQUEST)
            reward = safe_decimal(row.get("reward_amount")) or Decimal("0")
            bonus = safe_decimal(row.get("bonus_amount")) or Decimal("0")
            if reward < 0 or bonus < 0:
                return Response({"error": "Rates can't be negative."},
                                status=status.HTTP_400_BAD_REQUEST)
            PlatformRate.objects.update_or_create(
                business=business, platform=platform, task_type=task_type, user=user,
                defaults={"reward_amount": reward, "bonus_amount": bonus,
                          "updated_by": request.user},
            )

    rates = PlatformRate.objects.filter(business=business).select_related("updated_by", "user")
    if not admin:
        # A worker sees the business rate and their own override, never
        # anyone else's — same boundary as the wallet and the task roster.
        rates = rates.filter(DQ(user__isnull=True) | DQ(user=request.user))
    return Response({
        "results": PlatformRateSerializer(rates, many=True).data,
        "platforms": [{"value": v, "label": l} for v, l in WorkerTask.PLATFORM_CHOICES],
        "task_types": [{"value": v, "label": l} for v, l in WorkerTask.TYPE_CHOICES],
    })


@api_view(["GET", "POST"])
def task_documents(request, business_id):
    """The how-to list. Everyone reads it; an admin maintains it."""
    business = get_authorized_business(request, business_id)

    if request.method == "POST":
        if not _is_admin(request.user):
            return Response({"error": "Only an admin can add documents."},
                            status=status.HTTP_403_FORBIDDEN)
        payload = request.data if isinstance(request.data, dict) else {}
        title = str(payload.get("title") or "").strip()[:200]
        if not title:
            return Response({"error": "Give the document a title."},
                            status=status.HTTP_400_BAD_REQUEST)
        doc = TaskDocument.objects.create(
            business=business,
            title=title,
            url=str(payload.get("url") or "").strip(),
            description=str(payload.get("description") or "").strip(),
            platform=str(payload.get("platform") or "").strip().upper()[:20],
            sort_order=int(payload.get("sort_order") or 0),
            created_by=request.user,
        )
        return Response(TaskDocumentSerializer(doc).data, status=status.HTTP_201_CREATED)

    docs = TaskDocument.objects.filter(business=business).select_related("created_by")
    return Response({"results": TaskDocumentSerializer(docs, many=True).data})


@api_view(["PATCH", "DELETE"])
def task_document_detail(request, business_id, pk):
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"error": "Only an admin can change documents."},
                        status=status.HTTP_403_FORBIDDEN)
    try:
        doc = TaskDocument.objects.get(pk=pk, business=business)
    except TaskDocument.DoesNotExist:
        return Response({"error": "Document not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        doc.delete()
        return Response({"deleted": True})

    payload = request.data if isinstance(request.data, dict) else {}
    for field in ("title", "url", "description", "platform"):
        if field in payload:
            setattr(doc, field, str(payload[field] or "").strip())
    if "sort_order" in payload:
        doc.sort_order = int(payload["sort_order"] or 0)
    doc.save()
    return Response(TaskDocumentSerializer(doc).data)


@api_view(["GET"])
def worker_list(request, business_id):
    """Who tasks can be assigned to — members of this business, admins aside."""
    business = get_authorized_business(request, business_id)
    if not _is_admin(request.user):
        return Response({"results": []})

    users = User.objects.filter(
        memberships__business=business
    ).distinct().order_by("username")
    return Response({"results": [
        {"id": u.pk, "username": u.username, "role": u.role} for u in users
    ]})


# ══════════════════════════════════════════════════════════════════════════════
# Claim tickets — the Meesho supplier-panel ticket export
# ══════════════════════════════════════════════════════════════════════════════

_CLAIM_COL_MAP = {
    "s no":              "s_no",
    "sno":               "s_no",
    "product name":      "product_name",
    "sku":               "sku",
    "variation":         "variation",
    "qty":               "qty",
    "meesho pid":        "meesho_pid",
    "order number":      "order_no",
    "suborder number":   "suborder_no",
    "sub order number":  "suborder_no",
    "ticket id":         "ticket_id",
    "ticket status":     "ticket_status",
    "created date":      "created_date",
    "issue":             "issue",
    "last update":       "last_update",
    "reopen validity":   "reopen_validity",
    "cpp flag":          "cpp_flag",
}

# Money in the prose: "Rs 511.56", "Rs. 356.26", "Rs 1,024.50".
_CLAIM_AMOUNT_RE = re.compile(r"Rs\.?\s*([\d,]+(?:\.\d+)?)", re.IGNORECASE)
_CLAIM_TXN_RE = re.compile(r"transaction\s*id\s*[:\-]?\s*([A-Za-z0-9._/-]+)", re.IGNORECASE)
# "was done on 05 Aug 2026" / "has been scheduled on 2026-07-15"
_CLAIM_PAID_ON_RE = re.compile(r"was\s+done\s+on\s+([0-9]{1,2}\s+\w+\s+[0-9]{4}|[0-9]{4}-[0-9]{2}-[0-9]{2})", re.IGNORECASE)
_CLAIM_SCHED_ON_RE = re.compile(r"scheduled\s+on\s+([0-9]{1,2}\s+\w+\s+[0-9]{4}|[0-9]{4}-[0-9]{2}-[0-9]{2})", re.IGNORECASE)

_CLAIM_STATUS_MAP = {
    "open": ClaimTicket.STATUS_OPEN,
    "approved": ClaimTicket.STATUS_APPROVED,
    "rejected": ClaimTicket.STATUS_REJECTED,
}

# A ticket outcome maps onto the claim state we track per return, so uploading
# the sheet moves the return's claim forward without anyone retyping it.
_TICKET_TO_CLAIM_STATUS = {
    ClaimTicket.STATUS_OPEN:     ReturnDelivery.CLAIM_RAISED,
    ClaimTicket.STATUS_APPROVED: ReturnDelivery.CLAIM_APPROVED,
    ClaimTicket.STATUS_REJECTED: ReturnDelivery.CLAIM_REJECTED,
}


def _parse_claim_outcome(last_update):
    """
    (claim_amount, transaction_id, settled_on, payment_state) out of the prose.

    Only the *paid* wording yields a transaction id, so the presence of one is
    what distinguishes money that moved from money that is merely promised —
    which is the distinction that matters when reconciling against a bank
    statement.
    """
    text = strip_html(last_update)
    if not text:
        return None, "", None, ClaimTicket.PAY_NONE

    amount = None
    m = _CLAIM_AMOUNT_RE.search(text)
    if m:
        amount = safe_decimal(m.group(1).replace(",", ""))

    txn = ""
    m = _CLAIM_TXN_RE.search(text)
    if m:
        txn = m.group(1).strip().rstrip(".")

    settled_on, state = None, ClaimTicket.PAY_NONE
    m = _CLAIM_PAID_ON_RE.search(text)
    if m:
        settled_on, state = safe_date(m.group(1)), ClaimTicket.PAY_PAID
    else:
        m = _CLAIM_SCHED_ON_RE.search(text)
        if m:
            settled_on, state = safe_date(m.group(1)), ClaimTicket.PAY_SCHEDULED
    # A transaction id with no parseable date still means it was paid.
    if state == ClaimTicket.PAY_NONE and txn:
        state = ClaimTicket.PAY_PAID

    return amount, txn, settled_on, state


def _find_claim_header_row(lines):
    for idx, line in enumerate(lines[:60]):
        low = line.lower()
        if "ticket id" in low and "ticket status" in low:
            return idx
    return None


def _read_claim_sheet(file):
    """Parse the uploaded ticket export into a DataFrame with mapped columns."""
    name = (file.name or "").lower()

    if name.endswith((".xlsx", ".xls")):
        file.seek(0)
        raw = pd.read_excel(file, header=None, dtype=str)
        header_idx = None
        for idx, row in raw.iterrows():
            joined = " ".join(str(c).lower() for c in row.tolist() if c is not None)
            if "ticket id" in joined and "ticket status" in joined:
                header_idx = idx
                break
        if header_idx is None:
            raise ValueError("Could not find the header row (expected 'Ticket ID' and 'Ticket Status').")
        df = raw.iloc[header_idx + 1:].copy()
        df.columns = [str(c) for c in raw.iloc[header_idx].tolist()]
    else:
        blob = file.read()
        text = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                text = blob.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            text = blob.decode("utf-8", errors="replace")

        lines = text.splitlines()
        header_idx = _find_claim_header_row(lines)
        if header_idx is None:
            raise ValueError("Could not find the header row (expected 'Ticket ID' and 'Ticket Status').")
        # The Last Update column contains embedded newlines inside quotes, so the
        # whole remaining text is handed to the CSV parser rather than split by
        # line — splitting would tear those cells apart.
        df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])), dtype=str)

    df.columns = [str(c).strip().lower() for c in df.columns]
    df.rename(columns=_CLAIM_COL_MAP, inplace=True)
    return df


def _link_ticket_to_return(business, ticket, apply_status=True):
    """
    Attach a ticket to its return and, optionally, move the return's claim state.

    A sub-order can have more than one ReturnDelivery row (a second return leg
    travels back on a different AWB), so the most recent is chosen — that is the
    leg a freshly-raised ticket refers to.

    Claim notes are never touched: they are hand-written at the desk and the
    sheet has no business overwriting them.
    """
    if not ticket.suborder_no:
        return False

    ret = (
        ReturnDelivery.objects.filter(business=business, suborder_no=ticket.suborder_no)
        .order_by("-delivered_date", "-uploaded_at")
        .first()
    )
    if ret is None:
        return False

    ticket.linked_return = ret

    if not apply_status:
        return True

    new_claim = _TICKET_TO_CLAIM_STATUS.get(ticket.ticket_status)
    fields = []
    if new_claim and ret.claim_status != new_claim:
        ret.claim_status = new_claim
        fields.append("claim_status")
        if not ret.claim_marked_at:
            ret.claim_marked_at = ticket.created_at_meesho or timezone.now()
            fields.append("claim_marked_at")
        # The ticket existing at all means the claim was raised on Meesho.
        if not ret.claim_raised_at:
            ret.claim_raised_at = ticket.created_at_meesho or timezone.now()
            fields.append("claim_raised_at")

    # The sheet is authoritative on money and on the ticket reference.
    if ticket.claim_amount is not None and ret.claim_amount != ticket.claim_amount:
        ret.claim_amount = ticket.claim_amount
        fields.append("claim_amount")
    if ticket.ticket_id and ret.claim_reference != ticket.ticket_id:
        ret.claim_reference = ticket.ticket_id
        fields.append("claim_reference")

    if fields:
        ret.save(update_fields=fields + ["updated_at"])
    return True


def _claim_stats(business):
    """Claim workload and money, counted in the database."""
    base = ClaimTicket.objects.filter(business=business)
    today = timezone.localdate()

    by_status = {
        r["ticket_status"]: r["n"]
        for r in base.values("ticket_status").annotate(n=Count("id"))
    }

    def money(**flt):
        return float(base.filter(**flt).aggregate(t=Sum("claim_amount"))["t"] or 0)

    approved = base.filter(ticket_status=ClaimTicket.STATUS_APPROVED)
    return {
        "total":    base.count(),
        "open":     by_status.get(ClaimTicket.STATUS_OPEN, 0),
        "approved": by_status.get(ClaimTicket.STATUS_APPROVED, 0),
        "rejected": by_status.get(ClaimTicket.STATUS_REJECTED, 0),

        "amount_approved":  money(ticket_status=ClaimTicket.STATUS_APPROVED),
        "amount_paid":      money(payment_state=ClaimTicket.PAY_PAID),
        "amount_scheduled": money(payment_state=ClaimTicket.PAY_SCHEDULED),

        "paid_count":      approved.filter(payment_state=ClaimTicket.PAY_PAID).count(),
        "scheduled_count": approved.filter(payment_state=ClaimTicket.PAY_SCHEDULED).count(),

        # Rejections you can still argue, and the ones whose window has closed.
        "reopenable": base.filter(
            ticket_status=ClaimTicket.STATUS_REJECTED, reopen_validity__gte=today
        ).count(),
        "reopen_expired": base.filter(
            ticket_status=ClaimTicket.STATUS_REJECTED, reopen_validity__lt=today
        ).count(),

        # Tickets we can't tie to a return — almost always a missing returns export.
        "unlinked": base.filter(linked_return__isnull=True).count(),
    }


@api_view(["POST"])
@parser_classes([MultiPartParser])
def claim_tickets_upload(request, business_id):
    """
    Upload the Meesho supplier-panel ticket export (CSV or Excel).

    Keyed on ticket id, so re-uploading an overlapping date range updates rows
    instead of duplicating them. Each ticket is then linked to its return and the
    return's claim status is moved to match — which is the point of the upload.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        df = _read_claim_sheet(file)
    except Exception as exc:
        return Response({"error": f"Could not read file: {exc}"}, status=status.HTTP_400_BAD_REQUEST)

    if "ticket_id" not in df.columns:
        return Response({"error": "Missing required column: Ticket ID"},
                        status=status.HTTP_400_BAD_REQUEST)

    created = updated = skipped = linked = stale = 0
    seen = set()

    with transaction.atomic():
        for _, row in df.iterrows():
            ticket_id = _return_text(row.get("ticket_id"))
            if not ticket_id:
                skipped += 1
                continue
            if ticket_id in seen:
                skipped += 1          # listed twice in one file — keep the first
                continue
            seen.add(ticket_id)

            raw_status = _return_text(row.get("ticket_status"))
            status_key = _CLAIM_STATUS_MAP.get(raw_status.strip().lower(), ClaimTicket.STATUS_OTHER)
            last_update = _return_text(row.get("last_update"), blank_as_empty=False)
            amount, txn, settled_on, pay_state = _parse_claim_outcome(last_update)

            qty_raw = _return_text(row.get("qty"))
            try:
                qty = int(float(qty_raw)) if qty_raw else None
            except ValueError:
                qty = None

            payload = {
                "suborder_no":       _return_text(row.get("suborder_no")),
                "order_no":          _return_text(row.get("order_no")),
                "product_name":      _return_text(row.get("product_name")),
                "sku":               _return_text(row.get("sku")),
                "variation":         _return_text(row.get("variation")),
                "meesho_pid":        _return_text(row.get("meesho_pid")),
                "qty":               qty,
                "ticket_status":     status_key,
                "ticket_status_raw": raw_status[:60],
                "issue":             _return_text(row.get("issue"))[:255],
                "created_at_meesho": safe_datetime(row.get("created_date")),
                "last_update":       last_update or "",
                "reopen_validity":   safe_date(_return_text(row.get("reopen_validity")) or None),
                "cpp_flag":          _return_text(row.get("cpp_flag"))[:40],
                "claim_amount":      amount,
                "transaction_id":    txn[:80],
                "settled_on":        settled_on,
                "payment_state":     pay_state,
            }

            existing = ClaimTicket.objects.filter(
                business=business, ticket_id=ticket_id
            ).first()

            # Guard against an older export reverting a decided ticket.
            #
            # The panel exports carry no "last modified" column, so two files
            # covering overlapping ranges are indistinguishable by content — yet
            # they disagree: in the real exports, 21 tickets read Open in the
            # older file and Approved/Rejected in the newer one, never the
            # reverse. Uploading the older file second would therefore throw away
            # the outcome *and* its amount, and drag the linked return's claim
            # status back to "raised". A decision is not something a stale sheet
            # gets to undo, so keep it and report the row as stale instead.
            if (
                existing
                and existing.ticket_status in (ClaimTicket.STATUS_APPROVED, ClaimTicket.STATUS_REJECTED)
                and status_key == ClaimTicket.STATUS_OPEN
            ):
                stale += 1
                continue

            ticket, was_created = ClaimTicket.objects.update_or_create(
                business=business, ticket_id=ticket_id, defaults=payload,
            )
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

            if _link_ticket_to_return(business, ticket):
                linked += 1
                ticket.save(update_fields=["linked_return", "updated_at"])

    # An imported ticket reading Approved is the only signal that Meesho
    # accepted a claim, so this is the moment worker bonuses can be released.
    bonus_count, bonus_total = _release_claim_bonuses(business, request.user)

    return Response({
        "success": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        # Rows an older export would have reverted — kept as-is on purpose.
        "stale_kept": stale,
        "linked": linked,
        "unlinked": created + updated - linked,
        "bonuses_released": bonus_count,
        "bonus_total": bonus_total,
        "stats": _claim_stats(business),
    })


@api_view(["POST"])
def claim_tickets_relink(request, business_id):
    """
    Re-run linking for tickets that never found a return.

    Needed because the two exports arrive independently: tickets uploaded before
    the matching returns sheet have nothing to attach to, and only a later pass
    can join them up.
    """
    business = get_authorized_business(request, business_id)

    qs = ClaimTicket.objects.filter(business=business, linked_return__isnull=True)
    if request.data.get("all"):
        qs = ClaimTicket.objects.filter(business=business)

    linked = 0
    for ticket in qs:
        if _link_ticket_to_return(business, ticket):
            ticket.save(update_fields=["linked_return", "updated_at"])
            linked += 1

    return Response({"checked": qs.count(), "linked": linked, "stats": _claim_stats(business)})


@api_view(["GET"])
def claim_tickets_list(request, business_id):
    """
    Paginated claim tickets.

      q           — ticket id / sub-order / order no / SKU / issue
      ticket_status — OPEN | APPROVED | REJECTED, or REOPENABLE for arguable rejections
      payment     — PAID | SCHEDULED | NONE
      issue       — exact issue text
      linked      — "no" for tickets with no matching return
      date_from / date_to — on the ticket's created date
      page, page_size
    """
    business = get_authorized_business(request, business_id)

    qs = ClaimTicket.objects.filter(business=business).select_related("linked_return")

    date_from = _parse_day(request.GET.get("date_from"))
    date_to   = _parse_day(request.GET.get("date_to"))
    if date_from:
        qs = qs.filter(created_at_meesho__gte=_local_day_start(date_from))
    if date_to:
        qs = qs.filter(created_at_meesho__lt=_local_day_start(date_to + timedelta(days=1)))

    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(
            DQ(ticket_id__icontains=search) |
            DQ(suborder_no__icontains=search) |
            DQ(order_no__icontains=search) |
            DQ(sku__icontains=search) |
            DQ(issue__icontains=search) |
            DQ(transaction_id__icontains=search)
        )

    wanted = request.GET.get("ticket_status", "").strip().upper()
    if wanted == "REOPENABLE":
        qs = qs.filter(ticket_status=ClaimTicket.STATUS_REJECTED,
                       reopen_validity__gte=timezone.localdate())
    elif wanted:
        qs = qs.filter(ticket_status=wanted)

    payment = request.GET.get("payment", "").strip().upper()
    if payment in {c[0] for c in ClaimTicket.PAYMENT_STATE_CHOICES}:
        qs = qs.filter(payment_state=payment)

    issue = request.GET.get("issue", "").strip()
    if issue:
        qs = qs.filter(issue=issue)

    if request.GET.get("linked", "").strip().lower() == "no":
        qs = qs.filter(linked_return__isnull=True)

    total = qs.count()
    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 50))))
    except (TypeError, ValueError):
        page_size = 50

    start = (page - 1) * page_size
    rows = qs[start:start + page_size]

    # The issue list drives a filter dropdown; taken from the data so a new
    # Meesho issue type appears without a code change.
    issues = list(
        ClaimTicket.objects.filter(business=business)
        .exclude(issue="").values_list("issue", flat=True).distinct().order_by("issue")
    )

    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "results": ClaimTicketSerializer(rows, many=True).data,
        "stats": _claim_stats(business),
        "issues": issues,
    })


# ══════════════════════════════════════════════════════════════════════════════
# Listing templates — synced from the Meesho browser extension
# ══════════════════════════════════════════════════════════════════════════════

# The extension's export file is a plain object keyed by its own local template
# id: {"tpl_123_ab": {name, fields, labels, sourceUrl, ...}}. Server ids differ
# from those local ids, so imports are reconciled on name — see _import_templates.
_TEMPLATE_WRITABLE = {"name", "fields", "labels", "source_url"}


def _template_payload(raw):
    """
    Normalise one incoming template into our field names.

    The extension speaks camelCase (`sourceUrl`) because that is what it stores
    in chrome.storage; accepting both spellings means an export file taken from
    an older build still imports.
    """
    if not isinstance(raw, dict):
        return None
    name = str(raw.get("name") or "").strip()
    fields = raw.get("fields")
    if not name or not isinstance(fields, dict):
        return None
    return {
        "name": name[:200],
        "fields": fields,
        "labels": raw.get("labels") if isinstance(raw.get("labels"), dict) else {},
        "source_url": str(raw.get("source_url") or raw.get("sourceUrl") or ""),
    }


def _import_templates(business, user, incoming):
    """
    Upsert a batch of templates by name, returning (created, updated, skipped).

    Matching on name rather than id is deliberate: the id in an export file came
    from whichever browser produced it, so importing on a second machine would
    otherwise duplicate every template.
    """
    created = updated = skipped = 0
    for raw in incoming:
        payload = _template_payload(raw)
        if payload is None:
            skipped += 1
            continue
        existing = ListingTemplate.objects.filter(
            business=business, name__iexact=payload["name"]
        ).first()
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            existing.updated_by = user
            existing.save()
            updated += 1
        else:
            ListingTemplate.objects.create(
                business=business, created_by=user, updated_by=user, **payload
            )
            created += 1
    return created, updated, skipped


def _incoming_template_list(data):
    """
    Accept every shape the extension might send: a bare list, the export object
    keyed by local id, or {"templates": <either>}.
    """
    if isinstance(data, dict) and "templates" in data:
        data = data["templates"]
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return list(data.values())
    return []


@api_view(["GET", "POST"])
def listing_templates_list(request, business_id):
    """
    GET  — every template for this business (newest first), optionally ?q= filtered.
           Pass ?full=0 to omit field values when only the list is needed.
    POST — save a template. An existing template with the same name is updated in
           place, so the extension's "Save as template" is idempotent per name.
    """
    business = get_authorized_business(request, business_id)

    if request.method == "POST":
        payload = _template_payload(request.data if isinstance(request.data, dict) else {})
        if payload is None:
            return Response(
                {"error": "Send at least a name and a fields object."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existing = ListingTemplate.objects.filter(
            business=business, name__iexact=payload["name"]
        ).first()
        if existing:
            for key, value in payload.items():
                setattr(existing, key, value)
            existing.updated_by = request.user
            existing.save()
            row, created = existing, False
        else:
            row = ListingTemplate.objects.create(
                business=business, created_by=request.user, updated_by=request.user, **payload
            )
            created = True

        return Response(
            {"created": created, "template": ListingTemplateSerializer(row).data},
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
        )

    qs = ListingTemplate.objects.filter(business=business).select_related(
        "created_by", "updated_by"
    )
    search = request.GET.get("q", "").strip()
    if search:
        qs = qs.filter(DQ(name__icontains=search) | DQ(source_url__icontains=search))

    rows = list(qs)
    data = ListingTemplateSerializer(rows, many=True).data
    if request.GET.get("full") == "0":
        for item in data:
            item.pop("fields", None)
            item.pop("labels", None)

    return Response({"total": len(data), "results": data})


@api_view(["GET", "PATCH", "DELETE"])
def listing_template_detail(request, business_id, pk):
    """Read, rename/edit, or delete one template."""
    business = get_authorized_business(request, business_id)

    try:
        row = ListingTemplate.objects.select_related("created_by", "updated_by").get(
            pk=pk, business=business
        )
    except ListingTemplate.DoesNotExist:
        return Response({"error": "Template not found."}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "DELETE":
        row.delete()
        return Response({"deleted": True})

    if request.method == "PATCH":
        payload = request.data if isinstance(request.data, dict) else {}
        # Accept camelCase from the extension, then reject anything unknown so a
        # typo fails loudly instead of being silently dropped.
        if "sourceUrl" in payload:
            payload = {**payload, "source_url": payload.pop("sourceUrl")}
        unknown = set(payload) - _TEMPLATE_WRITABLE
        if unknown:
            return Response(
                {"error": f"Not editable: {', '.join(sorted(unknown))}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = ListingTemplateSerializer(row, data=payload, partial=True)
        serializer.is_valid(raise_exception=True)

        new_name = serializer.validated_data.get("name")
        if new_name and ListingTemplate.objects.filter(
            business=business, name__iexact=new_name
        ).exclude(pk=row.pk).exists():
            return Response(
                {"error": f'Another template is already called "{new_name}".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer.save(updated_by=request.user)
        row = serializer.instance

    return Response(ListingTemplateSerializer(row).data)


@api_view(["POST"])
def listing_templates_import(request, business_id):
    """
    Bulk-import templates from an extension export file (or another browser).
    Existing names are updated rather than duplicated.
    """
    business = get_authorized_business(request, business_id)

    incoming = _incoming_template_list(request.data)
    if not incoming:
        return Response(
            {"error": "No templates found in that file."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    created, updated, skipped = _import_templates(business, request.user, incoming)
    return Response({
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "total": ListingTemplate.objects.filter(business=business).count(),
    })


@api_view(["GET"])
def listing_templates_export(request, business_id):
    """
    Every template as a JSON file, in the same shape the extension's own export
    produces — so a file from here can be imported straight back into any browser
    running the extension, and vice versa.
    """
    business = get_authorized_business(request, business_id)

    payload = {}
    for row in ListingTemplate.objects.filter(business=business):
        key = f"tpl_{row.id}"
        payload[key] = {
            "id": key,
            "name": row.name,
            "fields": row.fields or {},
            "labels": row.labels or {},
            "sourceUrl": row.source_url or "",
            "createdAt": int(row.created_at.timestamp() * 1000),
            "updatedAt": int(row.updated_at.timestamp() * 1000),
        }

    stamp = timezone.localdate().isoformat()
    response = Response(payload)
    response["Content-Disposition"] = (
        f'attachment; filename="meesho-templates-{business.id}-{stamp}.json"'
    )
    return response


# ══════════════════════════════════════════════════════════════════════════════
# Expense sheet — export to Excel and import back
# ══════════════════════════════════════════════════════════════════════════════

# Sheet 1 is one row per invoice *line item*, with the invoice's own fields
# repeated on each of its rows — that's the shape people actually want to read
# and edit in Excel. Rows are tied back to existing records by the ID column.
_EXPENSE_HEADERS = [
    "Invoice ID", "Date", "Invoice No", "Vendor", "Title",
    "Item Description", "Category", "Quantity", "Unit Rate", "Amount", "Notes",
]
_TRANSPORT_HEADERS = ["Charge ID", "Date", "Amount", "Note"]

# Header text → the key used internally, so a re-import tolerates case and
# spacing differences after a trip through Excel.
_EXPENSE_COL_MAP = {
    "invoice id": "invoice_id",
    "date": "date",
    "invoice no": "invoice_no",
    "invoice number": "invoice_no",
    "vendor": "vendor",
    "title": "title",
    "item description": "description",
    "description": "description",
    "category": "category",
    "quantity": "quantity",
    "qty": "quantity",
    "unit rate": "unit_rate",
    "rate": "unit_rate",
    "amount": "amount",          # computed on import, never trusted from the sheet
    "notes": "notes",
}
_TRANSPORT_COL_MAP = {
    "charge id": "charge_id",
    "id": "charge_id",
    "date": "date",
    "amount": "amount",
    "note": "note",
    "notes": "note",
}

_VALID_EXPENSE_CATEGORIES = {c[0] for c in ExpenseInvoiceItem.CATEGORY_CHOICES}


def _expense_category(raw):
    """Accept either the stored key ("packaging") or the display label."""
    v = (safe_str(raw) or "").strip().lower()
    if v in _VALID_EXPENSE_CATEGORIES:
        return v
    for key, label in ExpenseInvoiceItem.CATEGORY_CHOICES:
        if v == label.lower():
            return key
    if "packag" in v:
        return "packaging"
    return "other" if v else "packaging"


@api_view(["GET"])
def expenses_export(request, business_id):
    """
    Download every expense invoice (with line items) and transport charge as a
    two-sheet Excel workbook. The same file can be edited and fed back to
    /expenses/import/ — the ID columns are what make that round-trip update
    existing rows instead of duplicating them.

    Honours the same date filters as the list endpoints so you can export just
    one period.
    """
    from io import BytesIO
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    business = get_authorized_business(request, business_id)

    inv_qs = ExpenseInvoice.objects.filter(business=business).prefetch_related("items")
    tc_qs  = TransportCharge.objects.filter(business=business)

    month = request.GET.get("month", "")
    start, end = _month_range(month)
    if start:
        inv_qs = inv_qs.filter(date__gte=start, date__lt=end)
        tc_qs  = tc_qs.filter(date__gte=start, date__lt=end)
    if request.GET.get("date_from"):
        inv_qs = inv_qs.filter(date__gte=request.GET["date_from"])
        tc_qs  = tc_qs.filter(date__gte=request.GET["date_from"])
    if request.GET.get("date_to"):
        inv_qs = inv_qs.filter(date__lte=request.GET["date_to"])
        tc_qs  = tc_qs.filter(date__lte=request.GET["date_to"])

    head_font = Font(bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="6D28D9")
    money_fmt = "#,##0.00"

    wb = openpyxl.Workbook()

    # ── Sheet 1: expenses ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Expenses"
    ws.append(_EXPENSE_HEADERS)
    for col, _ in enumerate(_EXPENSE_HEADERS, start=1):
        c = ws.cell(row=1, column=col)
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    rows = 0
    for inv in inv_qs.order_by("-date", "-created_at"):
        items = list(inv.items.all())
        if not items:
            # Keep empty invoices in the sheet so they survive a round-trip.
            items = [None]
        for it in items:
            qty  = (it.quantity if it else None) or Decimal("0")
            rate = (it.unit_rate if it else None) or Decimal("0")
            ws.append([
                inv.id,
                inv.date,
                inv.invoice_no or "",
                inv.vendor or "",
                inv.title or "",
                (it.description if it else ""),
                (it.get_category_display() if it else ""),
                float(qty),
                float(rate),
                float((qty * rate).quantize(Decimal("0.01"))),
                inv.notes or "",
            ])
            rows += 1

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        for col in (8, 9, 10):
            ws.cell(row=r, column=col).number_format = money_fmt

    # ── Sheet 2: transport charges ───────────────────────────────────────────
    ws2 = wb.create_sheet("Transport")
    ws2.append(_TRANSPORT_HEADERS)
    for col, _ in enumerate(_TRANSPORT_HEADERS, start=1):
        c = ws2.cell(row=1, column=col)
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    tc_rows = 0
    for tc in tc_qs.order_by("-date", "-created_at"):
        ws2.append([tc.id, tc.date, float(tc.amount or 0), tc.note or ""])
        tc_rows += 1
    for r in range(2, ws2.max_row + 1):
        ws2.cell(row=r, column=2).number_format = "yyyy-mm-dd"
        ws2.cell(row=r, column=3).number_format = money_fmt

    # ── Sheet 3: how the round-trip works ────────────────────────────────────
    ws3 = wb.create_sheet("How to use")
    for line in [
        ["Editing and re-importing this workbook"],
        [""],
        ["1. Edit any cell except the ID and Amount columns."],
        ["2. Amount is recalculated on import as Quantity x Unit Rate — editing it has no effect."],
        ["3. Leave 'Invoice ID' as-is to update that invoice. Clear it (or add a new row"],
        ["   without one) to create a new invoice."],
        ["4. Line items belonging to one invoice are the rows that share its Invoice ID."],
        ["   For a brand new invoice spanning several item rows, leave Invoice ID blank on"],
        ["   all of them and give them the same Date + Invoice No + Vendor so they group."],
        ["5. Deleting a line-item row removes that line from its invoice on import."],
        ["6. Deleting a whole invoice's rows does NOT delete the invoice — remove it in the app."],
        ["7. Category accepts 'Packaging Material' or 'Other Expense'."],
        ["8. The Transport sheet works the same way, keyed on 'Charge ID'."],
    ]:
        ws3.append(line)
    ws3.column_dimensions["A"].width = 90
    ws3.cell(row=1, column=1).font = Font(bold=True, size=12)

    # Readable column widths.
    for sheet, headers in ((ws, _EXPENSE_HEADERS), (ws2, _TRANSPORT_HEADERS)):
        for idx, name in enumerate(headers, start=1):
            longest = len(name)
            for r in range(2, min(sheet.max_row, 400) + 1):
                v = sheet.cell(row=r, column=idx).value
                longest = max(longest, len(str(v)) if v is not None else 0)
            sheet.column_dimensions[get_column_letter(idx)].width = min(max(longest + 2, 10), 46)
        sheet.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = (month or "all").replace("/", "-")
    filename = f"expenses_{business.name.replace(' ', '_')}_{label}.xlsx"
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["X-Expense-Rows"] = str(rows)
    resp["X-Transport-Rows"] = str(tc_rows)
    return resp


def _read_expense_sheet(file, sheet_name, col_map):
    """Read one sheet into a list of dicts keyed by our internal names."""
    file.seek(0)
    try:
        df = pd.read_excel(file, sheet_name=sheet_name, dtype=object)
    except ValueError:
        return None   # sheet absent — caller decides whether that's fatal
    df.columns = [str(c).strip().lower() for c in df.columns]
    df.rename(columns=col_map, inplace=True)
    return df


@api_view(["POST"])
@parser_classes([MultiPartParser])
def expenses_import(request, business_id):
    """
    Import the expense workbook produced by /expenses/export/.

    Rows carrying an ID update that record; rows without one create a new
    record. Nothing is ever deleted at the invoice/charge level — removing a
    row from the sheet only drops that *line item* from its invoice. That
    asymmetry is deliberate: an accidental deletion in Excel shouldn't wipe
    expense history.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)

    name = (file.name or "").lower()
    if not name.endswith((".xlsx", ".xls")):
        return Response(
            {"error": "Please upload the .xlsx workbook produced by Export."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        exp_df = _read_expense_sheet(file, "Expenses", _EXPENSE_COL_MAP)
        tc_df  = _read_expense_sheet(file, "Transport", _TRANSPORT_COL_MAP)
    except Exception as exc:
        return Response({"error": f"Could not read the workbook: {exc}"},
                        status=status.HTTP_400_BAD_REQUEST)

    if exp_df is None and tc_df is None:
        return Response(
            {"error": "Workbook has neither an 'Expenses' nor a 'Transport' sheet."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = {
        "invoices_created": 0, "invoices_updated": 0, "items_written": 0,
        "transport_created": 0, "transport_updated": 0,
        "skipped_rows": 0, "warnings": [],
    }

    def _add_warning(msg):
        if msg not in result["warnings"] and len(result["warnings"]) < 12:
            result["warnings"].append(msg)

    with transaction.atomic():
        # ── Expenses ─────────────────────────────────────────────────────────
        if exp_df is not None and "date" in exp_df.columns:
            # Group rows into invoices: by existing id when present, otherwise
            # by the natural key so several new item rows form one invoice.
            groups = {}
            order = []
            for _, row in exp_df.iterrows():
                inv_id = safe_int(row.get("invoice_id"))
                date   = safe_date(row.get("date"))
                desc   = (safe_str(row.get("description")) or "").strip()

                if date is None and inv_id is None:
                    if desc:
                        result["skipped_rows"] += 1
                        _add_warning("Skipped row(s) with no date.")
                    continue

                key = ("id", inv_id) if inv_id is not None else (
                    "new",
                    str(date),
                    (safe_str(row.get("invoice_no")) or "").strip().lower(),
                    (safe_str(row.get("vendor")) or "").strip().lower(),
                    (safe_str(row.get("title")) or "").strip().lower(),
                )
                if key not in groups:
                    groups[key] = {"header": row, "items": []}
                    order.append(key)
                if desc:
                    groups[key]["items"].append(row)

            for key in order:
                grp    = groups[key]
                header = grp["header"]
                date   = safe_date(header.get("date"))

                fields = {
                    "invoice_no": (safe_str(header.get("invoice_no")) or "")[:100],
                    "vendor":     (safe_str(header.get("vendor")) or "")[:255],
                    "title":      (safe_str(header.get("title")) or "")[:255],
                    "notes":      safe_str(header.get("notes")) or "",
                }

                if key[0] == "id":
                    inv = ExpenseInvoice.objects.filter(pk=key[1], business=business).first()
                    if inv is None:
                        # An id from a different business (or a deleted row) must
                        # not silently reach across the tenant boundary.
                        _add_warning(
                            f"Invoice ID {key[1]} does not belong to this business — "
                            f"imported as a new invoice instead."
                        )
                        inv = None
                    if inv is not None:
                        if date is not None:
                            inv.date = date
                        for f, v in fields.items():
                            setattr(inv, f, v)
                        inv.save()
                        result["invoices_updated"] += 1
                    else:
                        if date is None:
                            result["skipped_rows"] += 1
                            continue
                        inv = ExpenseInvoice.objects.create(business=business, date=date, **fields)
                        result["invoices_created"] += 1
                else:
                    if date is None:
                        result["skipped_rows"] += 1
                        continue
                    inv = ExpenseInvoice.objects.create(business=business, date=date, **fields)
                    result["invoices_created"] += 1

                # The sheet is the source of truth for this invoice's lines.
                inv.items.all().delete()
                for it in grp["items"]:
                    qty  = safe_decimal(it.get("quantity"))
                    rate = safe_decimal(it.get("unit_rate"))
                    ExpenseInvoiceItem.objects.create(
                        business=business,
                        invoice=inv,
                        description=(safe_str(it.get("description")) or "")[:500],
                        category=_expense_category(it.get("category")),
                        quantity=qty if qty is not None else Decimal("0"),
                        unit_rate=rate if rate is not None else Decimal("0"),
                    )
                    result["items_written"] += 1
        elif exp_df is not None:
            _add_warning("'Expenses' sheet has no Date column — it was ignored.")

        # ── Transport charges ────────────────────────────────────────────────
        if tc_df is not None and "date" in tc_df.columns:
            for _, row in tc_df.iterrows():
                date   = safe_date(row.get("date"))
                amount = safe_decimal(row.get("amount"))
                if date is None or amount is None:
                    if date is not None or amount is not None:
                        result["skipped_rows"] += 1
                        _add_warning("Skipped transport row(s) missing a date or amount.")
                    continue
                note = (safe_str(row.get("note")) or "")[:255]
                cid  = safe_int(row.get("charge_id"))

                tc = TransportCharge.objects.filter(pk=cid, business=business).first() if cid else None
                if tc:
                    tc.date, tc.amount, tc.note = date, amount, note
                    tc.save()
                    result["transport_updated"] += 1
                else:
                    if cid:
                        _add_warning(
                            f"Charge ID {cid} does not belong to this business — "
                            f"imported as a new charge instead."
                        )
                    TransportCharge.objects.create(
                        business=business, date=date, amount=amount, note=note
                    )
                    result["transport_created"] += 1
        elif tc_df is not None:
            _add_warning("'Transport' sheet has no Date column — it was ignored.")

    result["success"] = True
    return Response(result, status=status.HTTP_200_OK)


# ══════════════════════════════════════════════════════════════════════════════
# GST — monthly liability from Meesho's TCS exports, plus rate-mismatch checks
# ══════════════════════════════════════════════════════════════════════════════

# The first two digits of a GSTIN are the state code. Only needed to decide
# whether a line is intra-state (CGST+SGST) or inter-state (IGST), which is the
# split GSTR-3B asks for.
_GST_STATE_CODES = {
    "01": "JAMMU AND KASHMIR", "02": "HIMACHAL PRADESH", "03": "PUNJAB",
    "04": "CHANDIGARH", "05": "UTTARAKHAND", "06": "HARYANA", "07": "DELHI",
    "08": "RAJASTHAN", "09": "UTTAR PRADESH", "10": "BIHAR", "11": "SIKKIM",
    "12": "ARUNACHAL PRADESH", "13": "NAGALAND", "14": "MANIPUR", "15": "MIZORAM",
    "16": "TRIPURA", "17": "MEGHALAYA", "18": "ASSAM", "19": "WEST BENGAL",
    "20": "JHARKHAND", "21": "ODISHA", "22": "CHATTISGARH", "23": "MADHYA PRADESH",
    "24": "GUJARAT", "26": "DADRA AND NAGAR HAVELI AND DAMAN AND DIU",
    "27": "MAHARASHTRA", "29": "KARNATAKA", "30": "GOA", "31": "LAKSHADWEEP",
    "32": "KERALA", "33": "TAMIL NADU", "34": "PUDUCHERRY", "35": "ANDAMAN AND NICOBAR ISLANDS",
    "36": "TELANGANA", "37": "ANDHRA PRADESH", "38": "LADAKH",
}

# Columns of the TCS sales / sales-return exports → model fields.
_GST_TCS_COL_MAP = {
    "sub_order_num": "sub_order_num", "suborder_num": "sub_order_num",
    "order_date": "order_date",
    "manifest_date": "manifest_date",
    "cancel_return_date": "cancel_return_date",
    "hsn_code": "hsn_code", "hsn": "hsn_code",
    "quantity": "quantity",
    "gst_rate": "gst_rate",
    "total_taxable_sale_value": "total_taxable_sale_value",
    "tax_amount": "tax_amount",
    "total_invoice_value": "total_invoice_value",
    "taxable_shipping": "taxable_shipping",
    "end_customer_state_new": "end_customer_state", "end_customer_state": "end_customer_state",
    "gstin": "supplier_gstin",
    "eco_tcs_gstin": "eco_tcs_gstin",
    "supplier_id": "supplier_id",
    "sup_name": "sup_name",
    "financial_year": "financial_year",
    "month_number": "month_number",
}

_GST_INVOICE_COL_MAP = {
    "type": "doc_type",
    "order date": "order_date",
    "suborder no.": "suborder_no", "suborder no": "suborder_no",
    "sub order no.": "suborder_no", "sub_order_no": "suborder_no",
    "product description": "product_description",
    "hsn": "hsn",
    "invoice no.": "invoice_no", "invoice no": "invoice_no",
}

_GST_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"]


def _gst_read_any_sheet(file, col_map):
    """
    Read the first sheet of the workbook and map its headers.

    The TCS exports name their sheet after the supplier id ("3564327"), so the
    sheet name can't be hardcoded — always take the first one.
    """
    file.seek(0)
    df = pd.read_excel(file, sheet_name=0, dtype=object)
    df.columns = [str(c).strip().lower() for c in df.columns]
    df.rename(columns=col_map, inplace=True)
    return df


def _gst_detect_kind(df):
    """
    Work out which of the three exports this is from its columns, so the user
    doesn't have to pick the file type (and can't pick it wrongly).
    """
    cols = set(df.columns)
    if {"doc_type", "invoice_no", "suborder_no"} & cols and "total_taxable_sale_value" not in cols:
        return "INVOICE_DETAILS"
    if "total_taxable_sale_value" in cols:
        # The return export is the one carrying a cancellation date.
        return "RETURN" if "cancel_return_date" in cols else "SALE"
    return None


@api_view(["POST"])
@parser_classes([MultiPartParser])
def gst_upload(request, business_id):
    """
    Upload any of Meesho's three GST exports — tcs_sales, tcs_sales_return, or
    Tax invoice details. The file type is detected from its columns.

    An upload **replaces** every stored row for the periods present in the file
    rather than merging. That matches how tax filing works: if Meesho reissues a
    corrected export for a month, you want the corrected set, not the union of
    old and new rows.
    """
    business = get_authorized_business(request, business_id)
    file = request.FILES.get("file")
    if not file:
        return Response({"error": "No file provided."}, status=status.HTTP_400_BAD_REQUEST)
    if not (file.name or "").lower().endswith((".xlsx", ".xls", ".csv")):
        return Response({"error": "Upload the .xlsx file exactly as downloaded from Meesho."},
                        status=status.HTTP_400_BAD_REQUEST)

    try:
        if (file.name or "").lower().endswith(".csv"):
            file.seek(0)
            df = pd.read_csv(file, dtype=object)
            df.columns = [str(c).strip().lower() for c in df.columns]
            df.rename(columns={**_GST_TCS_COL_MAP, **_GST_INVOICE_COL_MAP}, inplace=True)
        else:
            df = _gst_read_any_sheet(file, {**_GST_TCS_COL_MAP, **_GST_INVOICE_COL_MAP})
    except Exception as exc:
        return Response({"error": f"Could not read the file: {exc}"},
                        status=status.HTTP_400_BAD_REQUEST)

    kind = _gst_detect_kind(df)
    if kind is None:
        return Response(
            {"error": "Unrecognised file. Expected one of Meesho's tcs_sales, "
                      "tcs_sales_return or Tax invoice details exports."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # ── Tax invoice details ──────────────────────────────────────────────────
    if kind == "INVOICE_DETAILS":
        rows, periods, skipped = [], set(), 0
        for _, r in df.iterrows():
            suborder = (safe_str(r.get("suborder_no")) or "").strip()
            invoice  = (safe_str(r.get("invoice_no")) or "").strip()
            when     = safe_datetime(r.get("order_date"))
            if not suborder or not invoice or when is None:
                skipped += 1
                continue
            periods.add((when.year, when.month))
            rows.append(GstInvoiceDetail(
                business=business,
                doc_type=(safe_str(r.get("doc_type")) or "").strip()[:30],
                order_date=when,
                suborder_no=suborder[:100],
                product_description=safe_str(r.get("product_description")) or "",
                hsn=(safe_str(r.get("hsn")) or "").strip()[:20],
                invoice_no=invoice[:100],
                period_year=when.year,
                period_month=when.month,
            ))

        with transaction.atomic():
            removed = 0
            for (y, m) in periods:
                removed += GstInvoiceDetail.objects.filter(
                    business=business, period_year=y, period_month=m
                ).delete()[0]
            GstInvoiceDetail.objects.bulk_create(rows, batch_size=1000)

        return Response({
            "success": True, "file_type": "Tax invoice details",
            "rows_imported": len(rows), "rows_replaced": removed, "skipped": skipped,
            "periods": sorted(f"{y}-{m:02d}" for y, m in periods),
        }, status=status.HTTP_201_CREATED)

    # ── TCS sales / sales returns ────────────────────────────────────────────
    rows, periods, skipped = [], set(), 0
    for _, r in df.iterrows():
        suborder = (safe_str(r.get("sub_order_num")) or "").strip()
        fy = safe_int(r.get("financial_year"))
        mn = safe_int(r.get("month_number"))
        if not suborder or fy is None or mn is None:
            skipped += 1
            continue
        periods.add((fy, mn))
        rows.append(GstTransaction(
            business=business,
            kind=kind,
            financial_year=fy,
            month_number=mn,
            sub_order_num=suborder[:100],
            order_date=safe_date(r.get("order_date")),
            manifest_date=safe_date(r.get("manifest_date")),
            cancel_return_date=safe_date(r.get("cancel_return_date")),
            hsn_code=(safe_str(r.get("hsn_code")) or "").strip()[:20],
            quantity=safe_int(r.get("quantity")) or 0,
            gst_rate=safe_decimal(r.get("gst_rate")) or Decimal("0"),
            total_taxable_sale_value=safe_decimal(r.get("total_taxable_sale_value")) or Decimal("0"),
            tax_amount=safe_decimal(r.get("tax_amount")) or Decimal("0"),
            total_invoice_value=safe_decimal(r.get("total_invoice_value")) or Decimal("0"),
            taxable_shipping=safe_decimal(r.get("taxable_shipping")) or Decimal("0"),
            end_customer_state=(safe_str(r.get("end_customer_state")) or "").strip()[:100],
            supplier_gstin=(safe_str(r.get("supplier_gstin")) or "").strip()[:20],
            eco_tcs_gstin=(safe_str(r.get("eco_tcs_gstin")) or "").strip()[:20],
            supplier_id=(safe_str(r.get("supplier_id")) or "").strip()[:50],
            sup_name=(safe_str(r.get("sup_name")) or "").strip()[:255],
        ))

    with transaction.atomic():
        removed = 0
        for (fy, mn) in periods:
            removed += GstTransaction.objects.filter(
                business=business, kind=kind, financial_year=fy, month_number=mn
            ).delete()[0]
        GstTransaction.objects.bulk_create(rows, batch_size=1000)

    return Response({
        "success": True,
        "file_type": "TCS sales" if kind == "SALE" else "TCS sales returns",
        "rows_imported": len(rows), "rows_replaced": removed, "skipped": skipped,
        "periods": sorted(f"FY{fy} M{mn:02d}" for fy, mn in periods),
    }, status=status.HTTP_201_CREATED)


def _gst_period_filter(request, qs):
    """Narrow a GstTransaction queryset to the requested filing period."""
    fy = request.GET.get("financial_year") or request.GET.get("fy")
    mn = request.GET.get("month_number") or request.GET.get("month")
    if fy:
        qs = qs.filter(financial_year=safe_int(fy))
    if mn:
        qs = qs.filter(month_number=safe_int(mn))
    return qs


def _d(v):
    return Decimal(str(v or 0))


@api_view(["GET"])
def gst_periods(request, business_id):
    """Which filing periods have data, so the tab can offer a month picker."""
    business = get_authorized_business(request, business_id)
    rows = (GstTransaction.objects.filter(business=business)
            .values("financial_year", "month_number", "kind")
            .annotate(n=Count("id")))
    periods = {}
    for r in rows:
        key = (r["financial_year"], r["month_number"])
        p = periods.setdefault(key, {
            "financial_year": r["financial_year"],
            "month_number": r["month_number"],
            "label": f"{_GST_MONTH_NAMES[r['month_number']] if 1 <= r['month_number'] <= 12 else r['month_number']} {r['financial_year']}",
            "sale_lines": 0, "return_lines": 0,
        })
        p["sale_lines" if r["kind"] == GstTransaction.KIND_SALE else "return_lines"] = r["n"]
    out = sorted(periods.values(), key=lambda p: (-p["financial_year"], -p["month_number"]))
    return Response({"periods": out})


@api_view(["GET"])
def gst_summary(request, business_id):
    """
    The GST position for a filing period: output tax on sales, less returns,
    broken down by rate and by place of supply.

    `tcs_rate` (default 0.5%) only drives the informational TCS-credit line —
    Meesho's export does not state the TCS it collected, so that figure is an
    estimate, never a filed number.
    """
    business = get_authorized_business(request, business_id)
    qs = _gst_period_filter(request, GstTransaction.objects.filter(business=business))

    tcs_rate = safe_decimal(request.GET.get("tcs_rate")) 
    if tcs_rate is None:
        tcs_rate = Decimal("0.5")

    supplier_state = ""
    first = qs.exclude(supplier_gstin="").values_list("supplier_gstin", flat=True).first()
    if first:
        supplier_state = _GST_STATE_CODES.get(str(first)[:2], "")

    totals = {
        "SALE":   {"lines": 0, "qty": 0, "taxable": Decimal("0"), "tax": Decimal("0"), "invoice": Decimal("0"), "shipping": Decimal("0")},
        "RETURN": {"lines": 0, "qty": 0, "taxable": Decimal("0"), "tax": Decimal("0"), "invoice": Decimal("0"), "shipping": Decimal("0")},
    }
    by_rate  = {}
    by_hsn   = {}
    by_state = {}
    intra = {"taxable": Decimal("0"), "tax": Decimal("0")}
    inter = {"taxable": Decimal("0"), "tax": Decimal("0")}
    recompute_mismatch = []

    for t in qs.iterator():
        bucket = totals[t.kind]
        bucket["lines"] += 1
        bucket["qty"] += t.quantity or 0
        bucket["taxable"]  += _d(t.total_taxable_sale_value)
        bucket["tax"]      += _d(t.tax_amount)
        bucket["invoice"]  += _d(t.total_invoice_value)
        bucket["shipping"] += _d(t.taxable_shipping)

        sign = Decimal("-1") if t.kind == GstTransaction.KIND_RETURN else Decimal("1")
        s_taxable = _d(t.total_taxable_sale_value) * sign
        s_tax     = _d(t.tax_amount) * sign

        rate_key = str(_d(t.gst_rate).quantize(Decimal("0.01")))
        r = by_rate.setdefault(rate_key, {
            "gst_rate": float(_d(t.gst_rate)), "lines": 0,
            "sale_taxable": Decimal("0"), "sale_tax": Decimal("0"),
            "return_taxable": Decimal("0"), "return_tax": Decimal("0"),
        })
        r["lines"] += 1
        if t.kind == GstTransaction.KIND_SALE:
            r["sale_taxable"] += _d(t.total_taxable_sale_value)
            r["sale_tax"]     += _d(t.tax_amount)
        else:
            r["return_taxable"] += _d(t.total_taxable_sale_value)
            r["return_tax"]     += _d(t.tax_amount)

        h = by_hsn.setdefault(t.hsn_code or "—", {
            "hsn": t.hsn_code or "—", "lines": 0, "qty": 0,
            "net_taxable": Decimal("0"), "net_tax": Decimal("0"), "rates": set(),
        })
        h["lines"] += 1
        h["qty"] += (t.quantity or 0) * (1 if t.kind == GstTransaction.KIND_SALE else -1)
        h["net_taxable"] += s_taxable
        h["net_tax"]     += s_tax
        h["rates"].add(float(_d(t.gst_rate)))

        st = by_state.setdefault(t.end_customer_state or "—", {
            "state": t.end_customer_state or "—", "lines": 0,
            "net_taxable": Decimal("0"), "net_tax": Decimal("0"),
        })
        st["lines"] += 1
        st["net_taxable"] += s_taxable
        st["net_tax"]     += s_tax

        target = intra if (supplier_state and (t.end_customer_state or "").upper() == supplier_state) else inter
        target["taxable"] += s_taxable
        target["tax"]     += s_tax

        # Sanity check on the file itself: taxable x rate should equal the tax.
        expected = (_d(t.total_taxable_sale_value) * _d(t.gst_rate) / Decimal("100"))
        if abs(expected - _d(t.tax_amount)) > Decimal("0.05"):
            if len(recompute_mismatch) < 50:
                recompute_mismatch.append({
                    "sub_order_num": t.sub_order_num,
                    "kind": t.kind,
                    "gst_rate": float(_d(t.gst_rate)),
                    "taxable": float(_d(t.total_taxable_sale_value)),
                    "tax_in_file": float(_d(t.tax_amount)),
                    "tax_recomputed": float(expected.quantize(Decimal("0.01"))),
                    "difference": float((expected - _d(t.tax_amount)).quantize(Decimal("0.01"))),
                })

    q2 = Decimal("0.01")
    net_taxable = totals["SALE"]["taxable"] - totals["RETURN"]["taxable"]
    net_tax     = totals["SALE"]["tax"]     - totals["RETURN"]["tax"]
    tcs_credit  = (net_taxable * tcs_rate / Decimal("100")).quantize(q2)

    def money(d):
        return str(Decimal(d).quantize(q2))

    rate_rows = []
    for key in sorted(by_rate, key=lambda k: float(k)):
        r = by_rate[key]
        nt  = r["sale_taxable"] - r["return_taxable"]
        ntx = r["sale_tax"] - r["return_tax"]
        rate_rows.append({
            "gst_rate": r["gst_rate"], "lines": r["lines"],
            "sale_taxable": money(r["sale_taxable"]), "sale_tax": money(r["sale_tax"]),
            "return_taxable": money(r["return_taxable"]), "return_tax": money(r["return_tax"]),
            "net_taxable": money(nt), "net_tax": money(ntx),
        })

    hsn_rows = sorted(
        [{"hsn": h["hsn"], "lines": h["lines"], "qty": h["qty"],
          "net_taxable": money(h["net_taxable"]), "net_tax": money(h["net_tax"]),
          "rates": sorted(h["rates"])} for h in by_hsn.values()],
        key=lambda x: -Decimal(x["net_taxable"]),
    )
    state_rows = sorted(
        [{"state": s["state"], "lines": s["lines"],
          "net_taxable": money(s["net_taxable"]), "net_tax": money(s["net_tax"])}
         for s in by_state.values()],
        key=lambda x: -Decimal(x["net_taxable"]),
    )

    return Response({
        "supplier_state": supplier_state,
        "sales":   {**{k: (money(v) if isinstance(v, Decimal) else v) for k, v in totals["SALE"].items()}},
        "returns": {**{k: (money(v) if isinstance(v, Decimal) else v) for k, v in totals["RETURN"].items()}},
        "net": {
            "taxable": money(net_taxable),
            "tax": money(net_tax),
            "gst_payable": money(net_tax),
        },
        "place_of_supply": {
            "intra_state": {"taxable": money(intra["taxable"]), "tax": money(intra["tax"]),
                            "cgst": money(intra["tax"] / 2), "sgst": money(intra["tax"] / 2)},
            "inter_state": {"taxable": money(inter["taxable"]), "igst": money(inter["tax"])},
        },
        "tcs": {
            "rate": float(tcs_rate),
            "estimated_credit": money(tcs_credit),
            "note": "Estimated: Meesho's export does not state the TCS it collected. "
                    "Confirm against your GSTR-2X / cash ledger before offsetting.",
        },
        "by_rate": rate_rows,
        "by_hsn": hsn_rows,
        "by_state": state_rows,
        "file_arithmetic_issues": recompute_mismatch,
    })


@api_view(["GET"])
def gst_mismatches(request, business_id):
    """
    Where the GST Meesho filed disagrees with what this catalogue expects.

    Two independent checks:

    1. **Rate vs your SKU pricing** — the rate on the filed line against the
       `tax_percent` configured for that SKU, with the resulting tax
       difference. Requires the SKU, which comes from the payment sheet via the
       sub-order, so lines whose sub-order isn't in Order Payments can't be
       checked.
    2. **One HSN, several rates** — the same HSN code filed at more than one
       rate in the period. An HSN has a single statutory rate, so this is a
       filing inconsistency worth explaining before it is queried.
    """
    business = get_authorized_business(request, business_id)
    qs = _gst_period_filter(request, GstTransaction.objects.filter(business=business))

    # sub-order → SKU, from the payment sheet.
    sku_of = {}
    for so, sku in (OrderPayment.objects
                    .filter(business=business)
                    .exclude(supplier_sku__isnull=True).exclude(supplier_sku="")
                    .values_list("sub_order_no", "supplier_sku").iterator()):
        sku_of.setdefault(so, sku)

    # SKU → the rate configured in pricing.
    configured = _SkuMap(
        FinalPrice.objects.filter(business=business)
        .exclude(tax_percent__isnull=True).values_list("sku_id", "tax_percent")
    )

    groups = {}
    unmatched_suborder = 0
    unpriced_sku = 0
    checked = 0
    hsn_rates = {}

    for t in qs.iterator():
        hsn_rates.setdefault(t.hsn_code or "—", {})
        bucket = hsn_rates[t.hsn_code or "—"].setdefault(float(_d(t.gst_rate)), {
            "lines": 0, "net_taxable": Decimal("0"),
        })
        bucket["lines"] += 1
        bucket["net_taxable"] += _d(t.total_taxable_sale_value) * (
            Decimal("-1") if t.kind == GstTransaction.KIND_RETURN else Decimal("1")
        )

        sku = sku_of.get(t.sub_order_num)
        if not sku:
            unmatched_suborder += 1
            continue
        if sku not in configured:
            unpriced_sku += 1
            continue
        checked += 1

        filed = _d(t.gst_rate)
        mine  = _d(configured[sku])
        if abs(filed - mine) <= Decimal("0.001"):
            continue

        key = (sku, str(filed), str(mine))
        g = groups.setdefault(key, {
            "sku": sku, "hsn": t.hsn_code, "filed_rate": float(filed), "configured_rate": float(mine),
            "lines": 0, "net_taxable": Decimal("0"), "tax_filed": Decimal("0"),
            "sub_orders": [],
        })
        sign = Decimal("-1") if t.kind == GstTransaction.KIND_RETURN else Decimal("1")
        g["lines"] += 1
        g["net_taxable"] += _d(t.total_taxable_sale_value) * sign
        g["tax_filed"]   += _d(t.tax_amount) * sign
        if len(g["sub_orders"]) < 5:
            g["sub_orders"].append(t.sub_order_num)

    q2 = Decimal("0.01")
    rate_rows = []
    total_diff = Decimal("0")
    for g in groups.values():
        at_configured = (g["net_taxable"] * Decimal(str(g["configured_rate"])) / Decimal("100"))
        diff = at_configured - g["tax_filed"]
        total_diff += diff
        rate_rows.append({
            "sku": g["sku"], "hsn": g["hsn"],
            "filed_rate": g["filed_rate"], "configured_rate": g["configured_rate"],
            "lines": g["lines"],
            "net_taxable": str(g["net_taxable"].quantize(q2)),
            "tax_filed": str(g["tax_filed"].quantize(q2)),
            "tax_at_configured_rate": str(at_configured.quantize(q2)),
            "difference": str(diff.quantize(q2)),
            "direction": "overcharged" if diff < 0 else "undercharged",
            "sample_sub_orders": g["sub_orders"],
        })
    rate_rows.sort(key=lambda r: -abs(Decimal(r["difference"])))

    hsn_rows = []
    for hsn, rates in hsn_rates.items():
        if len(rates) < 2:
            continue
        hsn_rows.append({
            "hsn": hsn,
            "rates": [
                {"gst_rate": r, "lines": v["lines"], "net_taxable": str(v["net_taxable"].quantize(q2))}
                for r, v in sorted(rates.items())
            ],
            "total_lines": sum(v["lines"] for v in rates.values()),
        })
    hsn_rows.sort(key=lambda h: -h["total_lines"])

    return Response({
        "rate_mismatches": rate_rows,
        "rate_mismatch_total_difference": str(total_diff.quantize(q2)),
        "hsn_rate_conflicts": hsn_rows,
        "coverage": {
            "lines_checked": checked,
            "lines_without_matching_sub_order": unmatched_suborder,
            "lines_on_unpriced_skus": unpriced_sku,
        },
    })


# ══════════════════════════════════════════════════════════════════════════════
# Label customers — identity is name + the whole address, not the pincode alone
# ══════════════════════════════════════════════════════════════════════════════

def _norm_customer_text(v):
    """Collapse whitespace and case so 'A  B' and 'a b' group together."""
    return " ".join((v or "").split()).strip().lower()


def _customer_key(row):
    """
    Group key for one physical customer.

    The older duplicate report keyed on (pincode, city, state), which lumped
    together every unrelated household in a pincode. Identity here is the
    person's name *plus* the full street address, so a genuine repeat buyer is
    distinguished from a neighbour.
    """
    return (
        _norm_customer_text(row.get("customer_name")),
        _norm_customer_text(row.get("customer_address")),
        _norm_customer_text(row.get("customer_city")),
        _norm_customer_text(row.get("customer_state")),
        (row.get("customer_pincode") or "").strip(),
    )


# Statuses grouped the way the rest of the app reads them.
_CUST_RETURN_STATUSES    = {"RETURN", "RETURNED"}
_CUST_RTO_STATUSES       = {"RTO", "RTO_COMPLETE", "RTO_LOCKED", "RTO_OFD"}
_CUST_DELIVERED_STATUSES = {"DELIVERED", "DOOR_STEP_EXCHANGED"}


def _customer_status_maps(business, order_ids):
    """sub_order_no → (status, settlement, payment_date, qty) from the payment sheet."""
    out = {}
    if not order_ids:
        return out
    for p in (OrderPayment.objects
              .filter(business=business, sub_order_no__in=order_ids)
              .values("sub_order_no", "live_order_status", "final_settlement_amount",
                      "payment_date", "quantity")
              .iterator()):
        cur = out.setdefault(p["sub_order_no"], {
            "statuses": set(), "settlement": Decimal("0"),
            "payment_date": None, "quantity": None,
        })
        if p["live_order_status"]:
            cur["statuses"].add(p["live_order_status"].upper())
        cur["settlement"] += Decimal(str(p["final_settlement_amount"] or 0))
        if p["payment_date"] and (cur["payment_date"] is None or p["payment_date"] > cur["payment_date"]):
            cur["payment_date"] = p["payment_date"]
        if cur["quantity"] is None and p["quantity"]:
            cur["quantity"] = p["quantity"]
    return out


def _classify_customer_status(statuses):
    """One headline status per order, using the same precedence as the P&L."""
    if statuses & _CUST_RETURN_STATUSES:
        return "RETURN"
    if statuses & _CUST_RTO_STATUSES:
        return "RTO"
    if statuses & _CUST_DELIVERED_STATUSES:
        return "DELIVERED"
    for s in statuses:
        if "CANCEL" in s:
            return "CANCELLED"
    if statuses:
        return sorted(statuses)[0]
    return "PENDING"


@api_view(["GET"])
def label_customers(request, business_id):
    """
    Every customer seen in shipping labels, keyed on name + full address, with
    their complete ordering record.

    Query params:
      date_from / date_to — restrict which customers appear (by label upload
                            date); their history is always all-time, so a
                            repeat buyer's full record is visible
      q                   — search name / address / city / pincode / SKU
      min_orders          — only customers with at least this many orders (default 1)
      sort                — orders | qty | returns | return_rate | last | name
      page / page_size
    """
    business = get_authorized_business(request, business_id)

    date_from = request.GET.get("date_from", "").strip()
    date_to   = request.GET.get("date_to", "").strip()
    search    = _norm_customer_text(request.GET.get("q", ""))
    sort      = (request.GET.get("sort") or "orders").strip().lower()
    try:
        min_orders = max(1, int(request.GET.get("min_orders", 1)))
    except (TypeError, ValueError):
        min_orders = 1
    try:
        page = max(1, int(request.GET.get("page", 1)))
    except (TypeError, ValueError):
        page = 1
    try:
        page_size = min(200, max(1, int(request.GET.get("page_size", 40))))
    except (TypeError, ValueError):
        page_size = 40

    _FIELDS = ("order_id", "customer_name", "customer_address", "customer_city",
               "customer_state", "customer_pincode", "sku", "size", "qty",
               "courier_name", "awb_number", "payment_type", "order_date",
               "uploaded_date")

    all_rows = list(LabelOrder.objects.filter(business=business).values(*_FIELDS))

    # Which customers fall inside the requested window — their history stays
    # all-time so you can see everything they've ever ordered.
    in_window = set()
    if date_from or date_to:
        for r in all_rows:
            d = r["uploaded_date"]
            if date_from and (not d or str(d) < date_from):
                continue
            if date_to and (not d or str(d) > date_to):
                continue
            in_window.add(_customer_key(r))
    else:
        in_window = None   # no window filter

    groups = {}
    for r in all_rows:
        key = _customer_key(r)
        if in_window is not None and key not in in_window:
            continue
        groups.setdefault(key, []).append(r)

    if not groups:
        return Response({"total": 0, "page": page, "page_size": page_size, "results": [],
                         "totals": {"customers": 0, "orders": 0, "repeat_customers": 0}})

    status_map = _customer_status_maps(
        business, [r["order_id"] for rows in groups.values() for r in rows]
    )
    blocked_set = set(
        BlockedCustomer.objects.filter(business=business, is_active=True)
        .values_list("customer_name", "customer_pincode")
    )

    # SKU → parent name, so the customer view speaks the same language as the
    # labels table.
    all_skus = {r["sku"] for rows in groups.values() for r in rows if r["sku"]}
    sku_to_parent = _SkuMap(
        FinalPrice.objects.filter(business=business, sku_id__in=all_skus)
        .select_related("parent").values_list("sku_id", "parent__item_id")
    ) if all_skus else _SkuMap()

    customers = []
    for key, rows in groups.items():
        rows.sort(key=lambda r: (str(r["uploaded_date"] or ""), r["order_id"]), reverse=True)
        newest = rows[0]

        counts = {"DELIVERED": 0, "RETURN": 0, "RTO": 0, "CANCELLED": 0, "PENDING": 0, "OTHER": 0}
        total_qty = 0
        settlement = Decimal("0")
        orders = []
        for r in rows:
            info = status_map.get(r["order_id"], {})
            st = _classify_customer_status(info.get("statuses") or set())
            counts[st if st in counts else "OTHER"] += 1
            qty = r["qty"] or info.get("quantity") or 1
            total_qty += qty
            settlement += Decimal(str(info.get("settlement") or 0))
            orders.append({
                "order_id":     r["order_id"],
                "sku":          r["sku"],
                "parent_sku":   sku_to_parent.get(r["sku"]) or None,
                "size":         r["size"],
                "qty":          qty,
                "status":       st,
                "courier_name": r["courier_name"],
                "awb_number":   r["awb_number"],
                "payment_type": r["payment_type"],
                "order_date":   str(r["order_date"]) if r["order_date"] else None,
                "label_date":   str(r["uploaded_date"]) if r["uploaded_date"] else None,
                "payment_date": str(info["payment_date"]) if info.get("payment_date") else None,
                "settlement":   str(Decimal(str(info.get("settlement") or 0)).quantize(Decimal("0.01"))),
            })

        n = len(rows)
        if n < min_orders:
            continue

        bad = counts["RETURN"] + counts["RTO"]
        settled = counts["DELIVERED"] + bad
        customers.append({
            "customer_name":    newest["customer_name"],
            "customer_address": newest["customer_address"],
            "customer_city":    newest["customer_city"],
            "customer_state":   newest["customer_state"],
            "customer_pincode": newest["customer_pincode"],
            "order_count":      n,
            "total_qty":        total_qty,
            "delivered":        counts["DELIVERED"],
            "returned":         counts["RETURN"],
            "rto":              counts["RTO"],
            "cancelled":        counts["CANCELLED"],
            "pending":          counts["PENDING"],
            # Share of *concluded* orders that came back — pending ones would
            # otherwise flatter the rate.
            "return_rate":      round(bad / settled, 4) if settled else 0.0,
            "net_settlement":   str(settlement.quantize(Decimal("0.01"))),
            "first_ordered":    str(min(r["uploaded_date"] for r in rows if r["uploaded_date"])) if any(r["uploaded_date"] for r in rows) else None,
            "last_ordered":     str(max(r["uploaded_date"] for r in rows if r["uploaded_date"])) if any(r["uploaded_date"] for r in rows) else None,
            "distinct_skus":    sorted({r["sku"] for r in rows if r["sku"]}),
            "is_blocked":       (newest["customer_name"], newest["customer_pincode"]) in blocked_set,
            "orders":           orders,
        })

    if search:
        def hit(c):
            hay = " ".join([
                _norm_customer_text(c["customer_name"]),
                _norm_customer_text(c["customer_address"]),
                _norm_customer_text(c["customer_city"]),
                _norm_customer_text(c["customer_state"]),
                (c["customer_pincode"] or "").lower(),
                " ".join(s.lower() for s in c["distinct_skus"]),
            ])
            return search in hay
        customers = [c for c in customers if hit(c)]

    sorters = {
        "orders":      lambda c: (-c["order_count"], c["customer_name"] or ""),
        "qty":         lambda c: (-c["total_qty"], c["customer_name"] or ""),
        "returns":     lambda c: (-(c["returned"] + c["rto"]), -c["order_count"]),
        "return_rate": lambda c: (-c["return_rate"], -c["order_count"]),
        "last":        lambda c: (c["last_ordered"] or "", ),
        "name":        lambda c: (_norm_customer_text(c["customer_name"]), ),
    }
    key_fn = sorters.get(sort, sorters["orders"])
    customers.sort(key=key_fn, reverse=(sort == "last"))

    total = len(customers)
    start = (page - 1) * page_size
    return Response({
        "total": total,
        "page": page,
        "page_size": page_size,
        "sort": sort,
        "results": customers[start:start + page_size],
        "totals": {
            "customers": total,
            "orders": sum(c["order_count"] for c in customers),
            "repeat_customers": sum(1 for c in customers if c["order_count"] > 1),
            "blocked": sum(1 for c in customers if c["is_blocked"]),
        },
    })
